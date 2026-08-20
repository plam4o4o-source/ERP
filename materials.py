# -*- coding: utf-8 -*-
"""Справочник материали (ABB part ID → описание → нето тегло кг/бр).

Заявка: „Прикачвам ти файл със съответния материал, описанието на
материала и килограми на всеки материал. От файла с килограмите
автоматично да се извличат съответните килограми във фактурата“ + „да не
се зарежда всеки път файла с материалите; като се зареди веднъж, да си
остава зареден в програмата“.

Затова справочникът НЕ се качва при всяка фактура, а се внася ВЕДНЪЖ в
таблица `materials` в базата (виж db.SCHEMA) и оттам нататък стои. При
издаване на фактура въведеният код на материала се търси в тази таблица и
теглото/описанието се попълват автоматично (виж routes_invoices.py).
Повторно качване на нов файл просто ОБНОВЯВА вече заредените редове по
код и добавя новите — старите редове, които ги няма в новия файл, се
запазват (частично обновяване на ценоразпис не бива да изтрива
исторически материали, ползвани от вече издадени фактури).

Теглото се пази като ТЕКСТ, а не като REAL — по същата причина, по която
и количествата/теглата в самите документи са текст: стойностите идват от
Excel файл на трета страна, могат да са с различен десетичен разделител
или с празни/нечислови клетки, а програмата само ги ПОКАЗВА обратно, не
смята с тях (общото тегло се смята на момента, толерантно, виж
parse_number по-долу).
"""
import io
import itertools
import math
import zipfile

# Одит (12.08.2026, находка №19): _parse_decimal е СЪЩАТА стриктна
# валидация, ползвана навсякъде другаде за количество/тегло/цена
# (appcore.py) — вижте parse_number по-долу за пълния разказ защо
# локалната реализация тук преди беше по-малко строга.
from appcore import _parse_decimal

#: Заглавия на колоните в подадения Excel файл, по които се разпознават
#: трите нужни колони. Търси се точно съвпадение след смъкване до малки
#: букви и изчистване на празните места/нови редове (заглавието на
#: колоната с теглото в оригиналния файл е на два реда: "Net weight\n[KG/pc]").
_CODE_HEADERS = ("abb part id", "part id", "material code", "code", "материал", "код")
_DESC_HEADERS = ("description", "material description", "описание")
_WEIGHT_HEADERS = ("net weight [kg/pc]", "net weight[kg/pc]", "net weight",
                   "weight [kg/pc]", "weight", "kg/pc", "тегло")

# Одит (19.08.2026, находка №38, дребна): справочникът не беше получил
# НИТО ЕДНА от защитите, които двата други Excel импорта (routes_pallet_
# extra/routes_invoices) вече имат от находка №18 (16.08) — нито таван на
# редовете, нито предупреждение за обединени клетки, нито съобщение „кой
# ред е приет за заглавен“. Стойностите са СЪЩИТЕ като там нарочно: един и
# същ оператор качва и трите файла и не бива да получава три различни
# поведения при иначе еднакви файлове.
_HEADER_SCAN_ROWS = 10
_MAX_IMPORT_DATA_ROWS = 5000


def _cellstr(v):
    """Клетка към низ, без излишно „.0“ за цели числа, записани като float
    (същата помощна функция като в routes_pallet_extra._cellstr — Excel
    връща всички числа като float, а кодовете на материали се ползват като
    текстови ключове)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _norm_header(v):
    """Заглавие на колона към сравним вид: малки букви, без нови редове и
    без повтарящи се празни места."""
    return " ".join(_cellstr(v).lower().split())


def parse_number(value):
    """Толерантно четене на число от текстово поле (тегло, количество,
    цена) — приема и десетична запетая, и точка, и празно. Връща None при
    непарсваема/празна стойност, за да може извикващият да реши какво да
    показва (обичайно „—“ вместо 0).

    Одит (12.08.2026, находка №19): преди тази поправка тук имаше отделна,
    ПО-МАЛКО строга реализация от appcore._parse_decimal (единствената
    употреба на числа в проекта иначе) — приемаше `nan`/`inf`/научна
    нотация без отхвърляне (float("nan") и float("1e10") не гърмят), и
    махаше само обикновен интервал (`.replace(" ", "")`), не и
    non-breaking space (\\xa0, чест при копиране от Excel). Резултат:
    развален ред можеше да доведе до буквално `net_weight="nan"`,
    записано в справочника и после показано на бланка на фактура. Сега
    делегира на СЪЩАТА regex-базирана валидация като останалия проект —
    вижте appcore._parse_decimal/_DECIMAL_RE."""
    return _parse_decimal(value)


def _fmt_weight(value):
    """Тегло към текст за запис в справочника — реже безсмисленото
    „плаващо“ опашче на float-овете от Excel (0.087135000000001) до 6
    знака.

    Одит (16.08.2026, находка №26, регресия от находка №19): parse_number
    (делегиращо на appcore._parse_decimal) отхвърля nan/inf (целта на
    находка №19) И научна нотация (регексът няма поддръжка на „e“) — а
    преди онази поправка И ДВАТА случая падаха към суровия
    `_cellstr(value)`: `nan`/`inf` се записваше БУКВАЛНО в справочника, а
    малки тегла (openpyxl връща числови клетки като float; Python
    сериализира |x|<1e-4 в научна нотация, напр. 8.7135e-05) ставаха низ,
    който СЛЕД ТОВА навсякъде другаде се ОТХВЪРЛЯ от _parse_decimal —
    теглото „изчезва“ от изчисленията на фактурата.

    Одит (19.08.2026, находка №28а): цялата логика се премести в
    _weight_cell по-долу, защото извикващият вече има нужда и от втория
    ѝ изход — „клетката беше непразна, но неизползваема“ (за брояча „X
    реда с неразпознато тегло“). Тук остава само тънката обвивка."""
    return _weight_cell(value)[0]


def _weight_cell(value):
    """Одит (19.08.2026, находка №28а, средна): поправката на находка №26
    (16.08) покри само `float('nan')` — тоест ЧИСЛОВА клетка. ТЕКСТОВА
    клетка със същото съдържание („nan“, „N/A“, „—“, „#VALUE!“, каквото
    Excel/BI износът напише вместо липсваща стойност) падаше в последния
    клон към суровия `_cellstr(value)` и се записваше БУКВАЛНО в
    справочника; оттам се попълва автоматично в поле „Net weight“ на
    фактура, показва се на официалната бланка и се отхвърля от
    _parse_decimal при сумирането — тоест ред с видимо тегло „nan“, който
    не участва в общото тегло. Отделно ОТРИЦАТЕЛНО тегло се приемаше
    безмълвно (килограми под нула няма).

    Затова тук всяка стойност, която не е разпознаваемо НЕОТРИЦАТЕЛНО
    число, се записва като ПРАЗНО тегло (същото поведение като липсваща
    клетка — материалът остава в справочника, теглото се въвежда ръчно),
    а извикващият получава сигнал, за да преброи такива редове и да
    съобщи „X реда с неразпознато тегло“ вместо да мълчи.

    Връща (текст, разпознато): `разпознато` е False САМО когато клетката е
    непразна, но неизползваема — празната клетка не е грешка."""
    number = parse_number(value)
    if number is None:
        if isinstance(value, float):
            # Числова клетка, отхвърлена от строгия регекс: nan/inf (виж
            # находка №26) или научна нотация при много малко тегло
            # (8.7135e-05) — второто е напълно валидно и се форматира тук.
            if math.isnan(value) or math.isinf(value):
                return "", False
            if value < 0:
                return "", False
            text = "%.6f" % value
            text = text.rstrip("0").rstrip(".")
            return text or "0", True
        # Генуинно нечислова стойност (текст в клетката) — вече НЕ се
        # записва сурова, вижте разказа по-горе.
        return "", (_cellstr(value) == "")
    if number < 0:
        return "", False
    text = "%.6f" % number
    text = text.rstrip("0").rstrip(".")
    return text or "0", True


def xlsx_has_merged_cells(file_bytes):
    """Одит (19.08.2026, находка №38): огледално на
    routes_pallet_extra._xlsx_has_merged_cells — обединена клетка връща
    стойност САМО в горния ляв ъгъл на диапазона, всички останали клетки от
    него се четат като None, тоест кодове/тегла могат тихо да „изчезнат“ от
    заредения справочник. Проверката чете суровия XML на листовете
    (`<mergeCell `), защото openpyxl в `read_only=True` изобщо не излага
    `worksheet.merged_cells`."""
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            for name in zf.namelist():
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                    if b"<mergeCell " in zf.read(name):
                        return True
    except Exception:
        return False
    return False


def parse_catalog_xlsx(file_bytes, stats=None):
    """Чете качения Excel файл със справочника и връща списък от
    (код, описание, тегло) тройки, ИЛИ None ако файлът не съдържа
    разпознаваеми колони.

    Разпознава колоните по ЗАГЛАВИЕ (не по позиция), за да продължи да
    работи, ако ABB разместят колоните в следваща версия на ценоразписа.
    Търси заглавния ред сред първите 10 реда — реалните файлове често имат
    празни/декоративни редове най-отгоре.

    Редове без код се пропускат. Редове с код, но без тегло или без
    описание, се ЗАПАЗВАТ (в подадения файл има 845 без тегло и 105 без
    описание) — за тях просто няма какво да се попълни автоматично във
    фактурата и полето остава за ръчно въвеждане, което е по-добре от това
    материалът изобщо да липсва от справочника.

    Одит (19.08.2026, находка №38): `stats` е НЕЗАДЪЛЖИТЕЛЕН речник, който
    се допълва с числата за съобщенията към оператора — `header_row`
    (1-базиран ред, на който е намерено заглавието), `truncated`,
    `bad_weights`, `duplicate_codes`, `merged_cells`. САМИТЕ съобщения се
    съставят в routes_materials.materials_import, а не тук: този модул се
    ползва и извън заявка (тестове/скриптове), където flask_babel.gettext
    няма контекст, а маркиран за превод низ, скрит зад собствена обвивка,
    просто нямаше да бъде извлечен от `pybabel extract` и щеше да остане
    тихо непреведен (точно дефектът от находка №13).

    Параметърът е незадължителен нарочно — десетки съществуващи извиквания
    (и тестове) ползват само върнатия списък, а разширяването на върнатата
    стойност до тъпъл би ги счупило всичките, без да добави нищо."""
    from openpyxl import load_workbook

    if stats is None:
        stats = {}
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    for ws in wb.worksheets:
        # Одит (19.08.2026, находки №38 и №14): чете се ограничен брой
        # редове (заглавие + таван на данните + един за откриване на
        # орязване), вместо целият лист — виж _MAX_IMPORT_DATA_ROWS и
        # огледалния коментар в routes_pallet_extra._read_limited_rows.
        rows = ws.iter_rows(values_only=True)
        header_row = None
        cols = None
        for i, row in enumerate(rows):
            if i > _HEADER_SCAN_ROWS:
                break
            headers = [_norm_header(c) for c in row]
            code_i = desc_i = weight_i = None
            for j, h in enumerate(headers):
                if code_i is None and h in _CODE_HEADERS:
                    code_i = j
                elif desc_i is None and h in _DESC_HEADERS:
                    desc_i = j
                elif weight_i is None and h in _WEIGHT_HEADERS:
                    weight_i = j
            if code_i is not None and (desc_i is not None or weight_i is not None):
                header_row = i
                cols = (code_i, desc_i, weight_i)
                break
        if header_row is None:
            continue

        code_i, desc_i, weight_i = cols
        out = []
        bad_weights = 0
        seen_codes = {}
        duplicate_codes = 0
        # +1 ред над тавана — само за да се разбере, че файлът има още
        # данни (орязване), без да се чете целият лист.
        data_rows = itertools.islice(rows, _MAX_IMPORT_DATA_ROWS + 1)
        truncated = False
        for row in data_rows:  # итераторът вече е СЛЕД заглавния ред
            if len(out) >= _MAX_IMPORT_DATA_ROWS:
                truncated = True
                break
            code = _cellstr(row[code_i]) if code_i < len(row) else ""
            if not code:
                continue
            desc = _cellstr(row[desc_i]) if desc_i is not None and desc_i < len(row) else ""
            weight, ok = (_weight_cell(row[weight_i])
                          if weight_i is not None and weight_i < len(row) else ("", True))
            if not ok:
                bad_weights += 1
            # Одит (19.08.2026, находка №38): дублиран код в САМИЯ файл
            # тихо презаписваше първия (различно тегло!), а съобщението
            # накрая рапортуваше „1 нови и 1 обновени“ — все едно става
            # дума за два различни материала. Сравнението е по горен
            # регистър, за да улови и разминаване само по регистър (виж
            # находка №28б в replace_catalog по-долу).
            key = code.upper()
            if key in seen_codes:
                duplicate_codes += 1
            seen_codes[key] = True
            out.append((code, desc, weight))
        if not truncated and next(rows, None) is not None:
            truncated = True
        if out:
            stats.update({
                "header_row": header_row + 1,
                "truncated": truncated,
                "max_rows": _MAX_IMPORT_DATA_ROWS,
                "bad_weights": bad_weights,
                "duplicate_codes": duplicate_codes,
                "merged_cells": xlsx_has_merged_cells(file_bytes),
                "sheet": ws.title,
            })
            return out
    return None


def replace_catalog(con, entries, stats=None):
    """Записва прочетените редове в базата — обновява вече съществуващите
    по код и добавя новите, в ЕДНА транзакция.

    Нарочно НЕ трие таблицата преди това: нов ценоразпис обичайно покрива
    само част от материалите, а вече издадени фактури сочат към кодове,
    които трябва да продължат да се разпознават. Връща (нови, обновени).

    Одит (19.08.2026, находка №28б, средна): съпоставянето с вече
    заредените кодове ставаше ТОЧНО (`code in existing`), а `materials.code`
    е PRIMARY KEY със същото точно сравнение — тоест код, който се
    различава от вече записания САМО по регистър („abc-77“ срещу
    „ABC-77“), се вмъкваше като ВТОРИ ред. Оттам нататък `lookup` (точно
    съвпадение) и `lookup_many` (`UPPER(code) IN …`, слива вариантите и
    печели последният) връщаха РАЗЛИЧНО тегло за един и същ материал в
    една и съща фактура. Сега съпоставянето е по горен регистър и записът
    отива в СЪЩЕСТВУВАЩИЯ ред (пази се неговият изписан код), а броят на
    такива случаи се връща в `stats["case_conflicts"]`, за да може
    routes_materials да предупреди оператора „X кода се различават само по
    регистър“. Вижте и db._m010_materials_code_case_insensitive."""
    if stats is None:
        stats = {}
    # {код с ГОРЕН регистър: изписването, с което кодът вече е в базата}
    existing = {}
    for r in con.execute("SELECT code FROM materials"):
        existing[(r["code"] or "").upper()] = r["code"]
    added = updated = case_conflicts = 0
    for code, desc, weight in entries:
        key = code.upper()
        stored = existing.get(key)
        if stored is None:
            added += 1
            existing[key] = code
            stored = code
        else:
            updated += 1
            if stored != code:
                case_conflicts += 1
        con.execute(
            "INSERT INTO materials (code, description, net_weight, updated_at)"
            " VALUES (?, ?, ?, datetime('now','localtime'))"
            " ON CONFLICT(code) DO UPDATE SET"
            " description = excluded.description,"
            " net_weight = excluded.net_weight,"
            " updated_at = excluded.updated_at",
            (stored, desc, weight),
        )
    con.commit()
    stats["case_conflicts"] = case_conflicts
    return added, updated


def code_candidates(code):
    """Кандидатите за търсене на един код в справочника, ПО РЕД НА
    ПРЕДПОЧИТАНИЕ: първо пълният код както е подаден, после с махната по
    една „опашка“ след тире, отдясно наляво.

    Заявка: „референция 1TGB110025P1204-RAS да се вмъкне като
    1TGB110025P1204, за да може да се заредят автоматично килограмите“ —
    справките за поръчки понякога добавят суфикс към ABB кода (напр.
    „-RAS“), който го няма в ценоразписа.

    Защо стъпка по стъпка ОТДЯСНО, а не просто рязане при първото тире: в
    реалния ценоразпис има 26 кода, които САМИ съдържат тире (напр.
    „3ACD5282AA842-1“, „1TGB110381P0024COD-68“) — за
    „3ACD5282AA842-1-RAS“ правилният кандидат е „3ACD5282AA842-1“, не
    „3ACD5282AA842“. Пълният код винаги е първи, така истински код с тире
    от ценоразписа никога не губи от своя орязан вариант."""
    code = (code or "").strip()
    if not code:
        return []
    out = [code]
    while "-" in code:
        code = code.rsplit("-", 1)[0].strip()
        if code:
            out.append(code)
    return out


def lookup(con, code):
    """Един материал по код, или None. Кодът се търси както е въведен и в
    горен регистър (операторите често пишат кода на ръка и регистърът се
    разминава с този във файла); ако пълният код липсва, се пробват и
    вариантите с махната опашка след тире (виж code_candidates)."""
    for candidate in code_candidates(code):
        row = con.execute("SELECT * FROM materials WHERE code = ?", (candidate,)).fetchone()
        if row is None:
            row = con.execute("SELECT * FROM materials WHERE UPPER(code) = UPPER(?)",
                              (candidate,)).fetchone()
        if row is not None:
            return row
    return None


def lookup_many(con, codes):
    """Няколко материала наведнъж → {търсен код: ред} за намерените.

    Ползва се при зареждане на цяла палетна карта във фактура (десетки
    реда) — партидни заявки вместо по една на ред. Заявката е с UPPER(),
    за да хване и разминат регистър, а резултатът се връща с ОРИГИНАЛНО
    подадения код като ключ, за да може извикващият да го съпостави с реда
    си без допълнително нормализиране. За всеки код се пробват и
    вариантите с махната опашка след тире (code_candidates) — печели
    НАЙ-ДЪЛГОТО съвпадение, тоест пълният код има предимство пред орязания.
    """
    wanted = [c for c in ((c or "").strip() for c in codes) if c]
    if not wanted:
        return {}
    candidates_by_code = {c: code_candidates(c) for c in wanted}
    all_candidates = []
    seen = set()
    for cands in candidates_by_code.values():
        for cand in cands:
            key = cand.upper()
            if key not in seen:
                seen.add(key)
                all_candidates.append(cand)

    by_upper = {}
    # Разбива на партиди — SQLite има ограничение за брой параметри (999 по
    # подразбиране в по-старите билдове), а палетна карта може да е голяма.
    for start in range(0, len(all_candidates), 500):
        chunk = all_candidates[start:start + 500]
        # placeholders е само поредица от „?“, изчислена от БРОЯ елементи —
        # самите кодове никога не влизат в SQL текста, а се подават като
        # bound параметри на реда отдолу. Същият шаблон като
        # db.get_unload_points_map.
        placeholders = ",".join("?" for _ in chunk)
        rows = con.execute(
            "SELECT * FROM materials WHERE UPPER(code) IN (%s)" % placeholders,  # nosec B608 -- само „?“ плейсхолдъри по брой; стойностите са bound параметри (виж коментара по-горе)
            [c.upper() for c in chunk],
        ).fetchall()
        for r in rows:
            by_upper[r["code"].upper()] = r

    found = {}
    for code in wanted:
        for candidate in candidates_by_code[code]:
            row = by_upper.get(candidate.upper())
            if row is not None:
                found[code] = row
                break
    return found


def count(con):
    """Брой заредени материали — показва се в раздел „Материали“."""
    return con.execute("SELECT COUNT(*) AS c FROM materials").fetchone()["c"]


def search(con, query, limit=200):
    """Търсене по код ИЛИ описание (за екрана на раздел „Материали“).
    Празна заявка връща първите `limit` реда по код."""
    query = (query or "").strip()
    if not query:
        return con.execute(
            "SELECT * FROM materials ORDER BY code LIMIT ?", (limit,)).fetchall()
    # В7: ci_contains (db._ci_contains) — вижте routes_documents.py.
    return con.execute(
        "SELECT * FROM materials WHERE ci_contains(code, ?) OR ci_contains(description, ?)"
        " ORDER BY code LIMIT ?",
        (query, query, limit),
    ).fetchall()
