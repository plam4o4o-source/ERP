# -*- coding: utf-8 -*-
"""ПачоЛогистик — логистична програма за издаване на ЧМР товарителници,
опаковъчни листи и палетни карти с баркодове.

Стартиране:  python app.py  →  http://127.0.0.1:5000
Първоначален вход: потребител "admin", парола "admin123" (сменете я!).
"""
import hmac
import io
import json
import os
import secrets
import sys
import threading
import time
import webbrowser
from datetime import date, datetime, timedelta
from functools import wraps

# Компилираната .exe версия се билдва без конзолен прозорец (--windowed),
# затова sys.stdout/sys.stderr са None — обикновен print() би гръмнал.
# Пренасочваме извеждането към лог файл до .exe-то в такъв случай. Иначе
# (стартиране от изходния код, или билд с конзола) конзолата на Windows
# често е с кодировка cp1252/cp866 и гърми при кирилица — принуждаваме UTF-8.
if getattr(sys, "frozen", False):
    _base_dir_early = os.path.dirname(os.path.abspath(sys.executable))
else:
    _base_dir_early = os.path.dirname(os.path.abspath(__file__))

if sys.stdout is None or sys.stderr is None:
    _log_file = open(os.path.join(_base_dir_early, "pacho_startup.log"),
                     "a", encoding="utf-8", errors="replace", buffering=1)
    sys.stdout = _log_file
    sys.stderr = _log_file
else:
    for _stream in (sys.stdout, sys.stderr):
        if hasattr(_stream, "reconfigure"):
            try:
                _stream.reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass

from flask import (Flask, abort, flash, redirect, render_template, request,
                   Response, send_file, session, url_for)
from markupsafe import Markup
from werkzeug.security import check_password_hash, generate_password_hash

import backup
import branding
import config as appconfig
import db
import desktop
import jsonutil
import login_guard
import remote_tunnel
import updater
from barcode128 import code128_svg
from icons import render_icon
from version import __version__

APP_NAME = "ПачоЛогистик"
MIN_PASSWORD_LENGTH = 8  # прилага се еднакво във всички пътища за задаване на парола

# В компилираната .exe версия шаблоните и стиловете са разопаковани
# във временната папка на PyInstaller (sys._MEIPASS).
if getattr(sys, "frozen", False):
    _bundle = sys._MEIPASS
    app = Flask(__name__,
                template_folder=os.path.join(_bundle, "templates"),
                static_folder=os.path.join(_bundle, "static"))
else:
    app = Flask(__name__)
app.secret_key = db.get_secret_key()
app.json.ensure_ascii = False

# Сесийни бисквитки: HttpOnly пречи на JS да ги прочете (namelijk при XSS —
# виж и поправката на H2 по-долу), SameSite=Lax пречи на браузъра да я
# изпрати при заявка, започната от чужд сайт (базова CSRF защита в допълнение
# към явния токен по-долу). SESSION_COOKIE_SECURE НЕ се задава тук: по
# подразбиране програмата се ползва по обикновено HTTP (127.0.0.1 или LAN);
# Secure=True би направило бисквитката невидима за самия локален достъп.
# Когато отдалеченият достъп е през Cloudflare тунела, връзката браузър↔
# тунел вече Е https — тунелът е публичният HTTPS вход, вижте remote_tunnel.py.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
)

# Ако базата данни още не съществува локално (чисто нова инсталация) и е
# зададена GitHub синхронизация в pacho_config.json, изтегляме автоматично
# последната запазена база — за да могат нови служители да заредят вече
# съществуващите клиенти/документи веднага, без ръчна настройка.
if not os.path.exists(db.DB_PATH):
    _boot_cfg = appconfig.load_config()
    if _boot_cfg.get("gh_owner") and _boot_cfg.get("gh_repo") and _boot_cfg.get("gh_token"):
        try:
            backup.pull_db(
                _boot_cfg["gh_owner"], _boot_cfg["gh_repo"], _boot_cfg["gh_token"],
                _boot_cfg.get("gh_branch", "main") or "main",
                _boot_cfg.get("gh_path", "pacho_logistic.db") or "pacho_logistic.db",
                db.DB_PATH,
            )
        except Exception:
            pass  # без интернет или друга грешка — просто тръгваме с нова база

db.init_db()


# ---------------------------------------------------------------- помощни

@app.context_processor
def inject_globals():
    return {
        "APP_NAME": APP_NAME,
        "APP_VERSION": __version__,
        "current_year": date.today().year,
        "today": date.today().isoformat(),
        "has_logo": branding.logo_path() is not None,
    }


@app.template_filter("barcode")
def barcode_filter(code, height=55, responsive=False):
    return Markup(code128_svg(code, height=height, responsive=responsive))


app.add_template_global(render_icon, name="icon")


@app.after_request
def _sync_after_write(response):
    """След всяка успешна POST/PUT/DELETE заявка (нов документ, клиент,
    служител, настройка) насрочваме автоматична синхронизация с GitHub
    (ако е включена) — обединена с кратко забавяне, за да не се качва база
    данни при всяко единично поле, а веднъж след кратка пауза в работата."""
    if request.method in ("POST", "PUT", "DELETE") and response.status_code < 400:
        try:
            backup.mark_dirty(appconfig.load_config)
        except Exception:
            pass
    return response


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login", next=request.path))
        if session.get("role") != "admin":
            abort(403)
        return view(*args, **kwargs)
    return wrapped


# ---------------------------------------------------------------- CSRF защита
# Всяка POST/PUT/PATCH/DELETE заявка трябва да носи токен, съвпадащ с този в
# сесията на потребителя — иначе заявка, стартирана от чужда страница (напр.
# скрита форма/картинка на злонамерен сайт, докато служителят е логнат тук),
# не може да предизвика реално действие (създаване на admin, изтриване на
# документ/клиент и т.н.). Токенът се генерира лениво (при първото четене)
# и се пази в сесията; шаблоните го вграждат чрез {{ csrf_token() }}.
def _get_csrf_token():
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_hex(16)
        session["_csrf_token"] = token
    return token


app.add_template_global(_get_csrf_token, name="csrf_token")

_CSRF_UNSAFE_METHODS = ("POST", "PUT", "PATCH", "DELETE")


@app.before_request
def _check_csrf():
    if request.method not in _CSRF_UNSAFE_METHODS:
        return None
    expected = session.get("_csrf_token")
    sent = request.form.get("csrf_token") or request.headers.get("X-CSRFToken")
    if not expected or not sent or not hmac.compare_digest(str(sent), str(expected)):
        abort(400, description=(
            "Невалидна или изтекла сесия на формата (CSRF защита). "
            "Презаредете страницата и опитайте отново."
        ))
    return None


# ---------------------------------------------------------- задължителна смяна на парола
# Прилага се към акаунти с users.must_change_password = 1 (първоначалният
# 'admin' със засятата парола 'admin123', и всеки служител, на когото друг
# администратор е задал/нулирал паролата — виж admin_user_new/
# admin_user_password по-долу). Пренасочва навсякъде другаде към „Смяна на
# парола“, докато служителят не си зададе собствена.
_PASSWORD_CHANGE_EXEMPT_ENDPOINTS = {"change_password", "logout", "static", "barcode_svg"}


@app.before_request
def _enforce_password_change():
    if "user_id" not in session or not session.get("must_change_password"):
        return None
    if request.endpoint and request.endpoint not in _PASSWORD_CHANGE_EXEMPT_ENDPOINTS:
        flash("Първо задайте нова парола, преди да продължите.")
        return redirect(url_for("change_password"))
    return None


def form_data(exclude=("csrf_token", "items_json")):
    """Всички полета от формата като речник (за съхранение в JSON)."""
    return {k: v.strip() for k, v in request.form.items() if k not in exclude}


def load_clients(con):
    return con.execute("SELECT * FROM clients ORDER BY name COLLATE NOCASE").fetchall()


def clients_json(clients, con=None):
    """JSON списък с клиентите за автопопълване във формите. Ако е подаден
    отворен con (ПРЕДИ да се затвори — виж cmr_new), вгражда за всеки
    клиент и списъка му с пунктове за разтоварване (unload_points), за да
    може ЧМР формата да ги предложи за избор без допълнителна заявка.

    Резултатът се вгражда directно в <script> блок в шаблоните (с |safe —
    виж cmr_form.html и др.), затова минава през
    jsonutil.dumps_for_inline_script вместо обикновен json.dumps: иначе
    име/адрес на клиент, съдържащ "</script><script>...", би прекъснало
    блока и изпълнило произволен JS за всеки, отворил формата (stored XSS)."""
    data = [dict(c) for c in clients]
    points_map = (db.get_unload_points_map(con, [c["id"] for c in data])
                  if con is not None and data else {})
    for c in data:
        c["unload_points"] = [
            {k: p.get(k, "") for k in ("label", "address", "city", "postcode", "country")}
            for p in points_map.get(c["id"], [])
        ]
    return jsonutil.dumps_for_inline_script(data)


def parse_items():
    """Редовете от таблицата с артикули, подадени като JSON от формата."""
    raw = request.form.get("items_json", "[]")
    try:
        items = json.loads(raw)
    except ValueError:
        items = []
    return items if isinstance(items, list) else []


def save_document(con, doc_type, data):
    number, year, seq, barcode = db.next_number(con, doc_type)
    data["number"] = number
    data["barcode"] = barcode
    cur = con.execute(
        "INSERT INTO documents (doc_type, number, year, seq, barcode, data, created_by)"
        " VALUES (?, ?, ?, ?, ?, ?, ?)",
        (doc_type, number, year, seq, barcode,
         json.dumps(data, ensure_ascii=False), session["user_id"]),
    )
    con.commit()
    return cur.lastrowid


# ---------------------------------------------------------------- предварителен преглед
# Прегледите се показват през POST (формата подава още незаписаните данни).
# Ако страницата се рендира директно като отговор на този POST, презареждане
# ѝ (F5), връщане/възстановяване на раздел от браузъра, или Windows
# автоматично възстановяване на затворен прозорец, кара браузъра да се
# опита да ПОВТОРИ същата POST заявка — което дава „Повторно изпращане на
# формуляра?“ или направо ERR_CACHE_MISS (наблюдавано от потребител при
# преглед на палетни карти през Edge, след restart на приложението).
# Затова тук ползваме POST → съхрани → пренасочи → GET: POST-ът пази
# данните временно на сървъра под случаен токен и пренасочва към обикновен
# GET адрес, който само ги чете — презареждане/връщане назад там е напълно
# безопасно, защото вече не е обвързано с еднократна POST заявка.
_preview_store = {}
_PREVIEW_TTL = 1800  # 30 минути — достатъчно за преглед, без да трупа памет за постоянно


def _cleanup_previews():
    now = time.time()
    for token in [t for t, (exp, _k, _p) in _preview_store.items() if exp < now]:
        del _preview_store[token]


def _store_preview(kind, payload):
    _cleanup_previews()
    token = secrets.token_urlsafe(16)
    _preview_store[token] = (time.time() + _PREVIEW_TTL, kind, payload)
    return token


def _get_preview(token, kind):
    """Чете преглед по токен, БЕЗ да го трие — трябва да остане валиден за
    многократно презареждане/връщане назад, докато не изтече (_PREVIEW_TTL),
    иначе първото презареждане би счупило точно проблема, който поправяме."""
    _cleanup_previews()
    entry = _preview_store.get(token)
    if entry is None or entry[1] != kind:
        return None
    return entry[2]


def render_preview(doc_type, data):
    """Приема POST-а с still-незаписаните данни на формата, пази ги временно
    на сървъра и пренасочва към GET адрес, който показва документа както ще
    изглежда при печат — БЕЗ да го запазва в базата и БЕЗ да изразходва
    пореден номер. GET адресът е безопасен за презареждане/връщане назад."""
    token = _store_preview("doc", (doc_type, data))
    return redirect(url_for("preview_document", token=token))


@app.route("/preview/<token>")
@login_required
def preview_document(token):
    payload = _get_preview(token, "doc")
    if payload is None:
        flash("Прегледът е изтекъл или вече е използван — генерирайте го отново от формата.")
        return redirect(url_for("dashboard"))
    doc_type, data = payload
    draft_doc = {
        "id": 0,
        "doc_type": doc_type,
        "number": "ПРЕДВАРИТЕЛЕН ПРЕГЛЕД / DRAFT",
        "barcode": "DRAFT-PREVIEW",
        "author": session.get("full_name") or session.get("username"),
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    return render_template(PRINT_TEMPLATES[doc_type], doc=draft_doc, d=data,
                           copies=1, preview=True, label_format=False)


def fetch_document(con, doc_id):
    row = con.execute(
        "SELECT d.*, u.full_name AS author FROM documents d"
        " LEFT JOIN users u ON u.id = d.created_by WHERE d.id = ?",
        (doc_id,),
    ).fetchone()
    if row is None:
        abort(404)
    return row, json.loads(row["data"])


# ---------------------------------------------------------------- вход/изход

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        locked, wait_seconds = login_guard.is_locked_out(username)
        if locked:
            wait_minutes = max(1, (wait_seconds + 59) // 60)
            error = ("Твърде много неуспешни опити за вход. Опитайте отново след "
                     "около %d мин." % wait_minutes)
        else:
            con = db.get_db()
            user = con.execute(
                "SELECT * FROM users WHERE username = ?", (username,)
            ).fetchone()
            con.close()
            if user and user["active"] and check_password_hash(user["password_hash"], password):
                login_guard.clear(username)
                con2 = db.get_db()
                theme = db.get_user_theme(con2, user["id"])
                con2.close()
                session.clear()
                session["user_id"] = user["id"]
                session["username"] = user["username"]
                session["full_name"] = user["full_name"]
                session["role"] = user["role"]
                session["theme"] = theme
                session["must_change_password"] = bool(user["must_change_password"])
                target = request.args.get("next") or url_for("dashboard")
                if not target.startswith("/"):
                    target = url_for("dashboard")
                return redirect(target)
            login_guard.register_failure(username)
            error = "Грешно потребителско име или парола, или акаунтът е деактивиран."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    flash("Излязохте от системата.")
    return redirect(url_for("login"))


# ---------------------------------------------------------------- табло и сканиране

@app.route("/")
@login_required
def dashboard():
    con = db.get_db()
    recent = con.execute(
        "SELECT d.*, u.full_name AS author FROM documents d"
        " LEFT JOIN users u ON u.id = d.created_by"
        " ORDER BY d.id DESC LIMIT 10"
    ).fetchall()
    counts = {t: con.execute(
        "SELECT COUNT(*) AS c FROM documents WHERE doc_type = ? AND year = ?",
        (t, date.today().year),
    ).fetchone()["c"] for t in db.DOC_TYPES}
    con.close()
    return render_template("dashboard.html", recent=recent, counts=counts,
                           doc_types=db.DOC_TYPES,
                           update=updater.check_cached(),
                           recent_docs_meta=[json.loads(r["data"]) for r in recent])


@app.route("/scan", methods=["POST"])
@login_required
def scan():
    """Зареждане на документ чрез сканиран баркод (или въведен номер)."""
    code = request.form.get("code", "").strip()
    con = db.get_db()
    doc = con.execute("SELECT id FROM documents WHERE barcode = ?", (code,)).fetchone()
    if doc is None:
        # опит и по номер, напр. "0001/2026"
        doc = con.execute(
            "SELECT id FROM documents WHERE number = ? ORDER BY id DESC", (code,)
        ).fetchone()
    con.close()
    if doc is None:
        flash("Няма документ с баркод „%s“." % code)
        return redirect(url_for("dashboard"))
    return redirect(url_for("view_document", doc_id=doc["id"]))


# ---------------------------------------------------------------- документи (списък/преглед)

@app.route("/docs")
@login_required
def documents():
    doc_type = request.args.get("type", "")
    query = request.args.get("q", "").strip()
    sql = ("SELECT d.*, u.full_name AS author FROM documents d"
           " LEFT JOIN users u ON u.id = d.created_by WHERE 1=1")
    params = []
    if doc_type in db.DOC_TYPES:
        sql += " AND d.doc_type = ?"
        params.append(doc_type)
    if query:
        sql += " AND (d.number LIKE ? OR d.barcode LIKE ? OR d.data LIKE ?)"
        like = "%" + query + "%"
        params += [like, like, like]
    sql += " ORDER BY d.id DESC LIMIT 300"
    con = db.get_db()
    docs = con.execute(sql, params).fetchall()
    con.close()
    metas = [json.loads(d["data"]) for d in docs]
    return render_template("documents.html", docs=docs, metas=metas,
                           doc_types=db.DOC_TYPES, sel_type=doc_type, q=query)


PRINT_TEMPLATES = {
    "cmr": "cmr_print.html",
    "packing": "packing_print.html",
    "pallet": "pallet_print.html",
    "waybill": "waybill_print.html",
    "dualuse": "dualuse_print.html",
    "export_it": "export_it_print.html",
}

FORM_TEMPLATES = {
    "cmr": "cmr_form.html",
    "packing": "packing_form.html",
    "pallet": "pallet_form.html",
    "waybill": "waybill_form.html",
    "dualuse": "dualuse_form.html",
    "export_it": "export_it_form.html",
}

# Типове документи с редове артикули (JSON масив items) — общо за
# edit_document (кой да парсва items_json при редакция) и другите места,
# които трябва да различат "документ с таблица" от "документ без таблица".
ITEM_DOC_TYPES = ("packing", "pallet", "waybill", "dualuse", "export_it")


@app.route("/doc/<int:doc_id>")
@login_required
def view_document(doc_id):
    con = db.get_db()
    row, data = fetch_document(con, doc_id)
    con.close()
    copies = request.args.get("copies", type=int) or 1
    label_format = request.args.get("format") == "label"
    return render_template(PRINT_TEMPLATES[row["doc_type"]], doc=row, d=data,
                           copies=min(copies, 5), preview=False,
                           label_format=label_format)


@app.route("/doc/<int:doc_id>/edit", methods=["GET", "POST"])
@login_required
def edit_document(doc_id):
    """Редакция на вече издаден документ — номерът, баркодът, годината и
    поредността се пазят непроменени (не се преиздава нов номер); само
    съдържанието (data) се обновява. Ползва СЪЩИТЕ форми, както при
    издаване, предварително попълнени с текущите стойности."""
    con = db.get_db()
    row, data = fetch_document(con, doc_id)
    doc_type = row["doc_type"]
    if doc_type not in FORM_TEMPLATES:
        con.close()
        abort(404)

    if request.method == "POST":
        new_data = form_data()
        if doc_type in ITEM_DOC_TYPES:
            new_data["items"] = parse_items()
            if "items_format" in data:
                new_data["items_format"] = data["items_format"]
        # номерът/баркодът се пазят от оригинала — редакцията не преиздава нов номер
        new_data["number"] = row["number"]
        new_data["barcode"] = row["barcode"]
        con.execute("UPDATE documents SET data = ? WHERE id = ?",
                    (json.dumps(new_data, ensure_ascii=False), doc_id))
        con.commit()
        con.close()
        flash("Документ № %s е обновен." % row["number"])
        return redirect(url_for("view_document", doc_id=doc_id))

    clients = load_clients(con)
    settings = db.get_settings(con)
    ctx = {
        "clients": clients,
        "clients_json": clients_json(clients, con) if doc_type == "cmr" else clients_json(clients),
        "s": settings,
        "edit_doc": row,
        "edit_data": data,
    }
    if doc_type in ITEM_DOC_TYPES:
        ctx["items"] = data.get("items", [])
    con.close()
    return render_template(FORM_TEMPLATES[doc_type], **ctx)


# ---------------------------------------------------------------- износ в Excel (.xlsx)
# Данните на всеки документ (номер, страни, стоки и т.н.) + редовете
# артикули (ако има) — в удобен за отваряне в Excel файл. PDF не се
# генерира отделно — печатните шаблони вече поддържат "Save as PDF" през
# диалога за печат на браузъра (вграден във Windows/Chromium, работи offline,
# без нужда от допълнителни компоненти в самата програма).

_XLSX_FIELDS = {
    "cmr": [
        ("Дата на съставяне", "established_date"), ("Място на съставяне", "established_place"),
        ("Изпращач", "sender_name"), ("Адрес изпращач", "sender_address"),
        ("Град изпращач", "sender_city"), ("Държава изпращач", "sender_country"),
        ("Получател", "consignee_name"), ("Адрес получател", "consignee_address"),
        ("Град получател", "consignee_city"), ("Държава получател", "consignee_country"),
        ("Разтоварен пункт", "place_delivery"), ("Товарен пункт", "place_loading"),
        ("Дата на натоварване", "date_loading"), ("Приложени документи", "attached_docs"),
        ("Марки и номера", "marks"), ("Брой колети", "packages"), ("Вид на опаковката", "packing"),
        ("Вид на стоката", "goods"), ("Статистически №", "stat_no"),
        ("Бруто тегло, кг", "weight"), ("Обем, м³", "volume"),
        ("Указания на изпращача", "sender_instructions"), ("Плащане на превоза", "payment_instructions"),
        ("Наложен платеж", "cod"), ("Специални споразумения", "special_agreements"),
        ("Превозвач", "carrier"), ("Последващи превозвачи", "successive_carriers"),
        ("Рег. № влекач", "truck_reg"), ("Рег. № ремарке", "trailer_reg"), ("Шофьор", "driver"),
        ("Резерви на превозвача", "reservations"),
    ],
    "packing": [
        ("Дата", "doc_date"), ("Изпращач", "sender_name"), ("Адрес изпращач", "sender_address"),
        ("Получател", "receiver_name"), ("Адрес получател", "receiver_address"),
        ("Град получател", "receiver_city"), ("Държава получател", "receiver_country"),
        ("Фактура №", "invoice_no"), ("Поръчка №", "order_no"),
        ("Общо колети", "total_packages"), ("Общо нето, кг", "total_net"), ("Общо бруто, кг", "total_gross"),
        ("Забележки", "notes"),
    ],
    "pallet": [
        ("Дата", "doc_date"), ("Палет №", "pallet_no"), ("Тип палет", "pallet_type"),
        ("Изпращач", "sender_name"), ("Клиент", "client_name"), ("Адрес клиент", "client_address"),
        ("Град клиент", "client_city"), ("Държава клиент", "client_country"),
        ("Брой кашони", "boxes"), ("Нето, кг", "net"), ("Бруто, кг", "gross"), ("Височина, см", "height"),
        ("Свързано ЧМР №", "ref_cmr"), ("Забележки", "notes"),
    ],
    "waybill": [
        ("Издадена в", "established_place"), ("Издадена на", "established_date"),
        ("Изпращач", "sender_name"), ("Адрес изпращач", "sender_address"),
        ("Превозвач", "carrier_name"), ("Адрес превозвач", "carrier_address"),
        ("Получател", "consignee_name"), ("Адрес получател", "consignee_address"),
        ("Град получател", "consignee_city"), ("Държава получател", "consignee_country"),
        ("Място на натоварване", "place_loading"), ("Дата на натоварване", "date_loading"),
        ("Място на разтоварване", "place_delivery"), ("Дата на разтоварване", "date_delivery"),
        ("Пробег, км", "mileage"),
        ("Опасен товар — клас", "dangerous_class"), ("Опасен товар — наименование", "dangerous_name"),
        ("Придружител на товара", "escort_name"), ("Брой придружители", "escort_count"),
        ("Превозна цена", "transport_price"), ("Допълнителни разходи", "extra_costs"),
        ("Марка на автомобила", "vehicle_make"), ("Модел на автомобила", "vehicle_model"),
        ("Рег. № на автомобила", "vehicle_reg"), ("Пътен лист №", "route_sheet_no"),
        ("Инструкции на превозвача", "carrier_instructions"),
        ("Натоварване — дата", "loading_date"), ("Натоварване — от час", "loading_from"),
        ("Натоварване — до час", "loading_to"),
        ("Разтоварване — дата", "unloading_date"), ("Разтоварване — от час", "unloading_from"),
        ("Разтоварване — до час", "unloading_to"),
        ("Забележка", "notes"),
    ],
    "dualuse": [
        ("Дата", "doc_date"), ("Износител", "sender_name"), ("ЕИК/ЕГН", "sender_eik"),
        ("Фактура/и №", "invoice_numbers"), ("Дата на фактурата", "invoice_date"),
        ("Държава на износ", "destination_country"), ("Място на съставяне", "place"),
        ("Декларатор", "declarant_name"), ("Длъжност", "declarant_position"),
    ],
    "export_it": [
        ("Дата", "doc_date"), ("Декларатор", "declarant_name"),
        ("Пълномощник на", "represented_company"), ("Фактура №", "invoice_no"),
        ("Износител", "exporter_company"), ("Получател", "receiver_name"),
        ("Ref. ЧМР №", "ref_cmr"), ("Място на съставяне", "place"),
    ],
}

_XLSX_ITEM_COLUMNS = {
    "packing": [("description", "Описание"), ("qty", "Количество"), ("packing", "Опаковка"),
               ("net", "Нето, кг"), ("gross", "Бруто, кг")],
    "pallet_generic": [("code", "Артикул/код"), ("description", "Описание"),
                       ("qty", "Количество"), ("weight", "Тегло, кг")],
    "pallet_orders": [("order_no", "Поръчка №"), ("pos", "Позиция"), ("reference", "Референция"),
                      ("reference_desc", "Описание"), ("qty", "Количество")],
    "waybill": [("description", "Наименование"), ("packing", "Опаковка"), ("marks", "Маркировка/номера"),
               ("weight", "Тегло, кг"), ("qty", "Брой")],
}


@app.route("/doc/<int:doc_id>/export.xlsx")
@login_required
def export_document_xlsx(doc_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    con = db.get_db()
    row, data = fetch_document(con, doc_id)
    con.close()
    doc_type = row["doc_type"]
    title = db.DOC_TYPES.get(doc_type, {}).get("title", doc_type)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Документ"

    bold = Font(bold=True)
    ws.append(["%s № %s" % (title, row["number"])])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append(["Баркод", row["barcode"]])
    ws.cell(row=2, column=1).font = bold
    ws.append([])

    for label, key in _XLSX_FIELDS.get(doc_type, []):
        ws.append([label, data.get(key, "")])
        ws.cell(row=ws.max_row, column=1).font = bold

    items = data.get("items") or []
    if items:
        if doc_type == "pallet":
            cols = _XLSX_ITEM_COLUMNS["pallet_orders" if data.get("items_format") == "orders"
                                      else "pallet_generic"]
        else:
            cols = _XLSX_ITEM_COLUMNS.get(doc_type, [])
        if cols:
            ws.append([])
            header_row = ws.max_row + 1
            ws.append([label for _key, label in cols])
            for c in range(1, len(cols) + 1):
                ws.cell(row=header_row, column=c).font = bold
            for it in items:
                ws.append([it.get(key, "") for key, _label in cols])

    for col_cells in ws.columns:
        lengths = [len(str(c.value)) for c in col_cells if c.value is not None]
        width = max(lengths) + 2 if lengths else 10
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(width, 10), 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "%s_%s.xlsx" % (doc_type, row["number"].replace("/", "-"))
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument"
                              ".spreadsheetml.sheet")


@app.route("/doc/<int:doc_id>/delete", methods=["POST"])
@admin_required
def delete_document(doc_id):
    con = db.get_db()
    con.execute("DELETE FROM documents WHERE id = ?", (doc_id,))
    con.commit()
    con.close()
    flash("Документът е изтрит.")
    return redirect(url_for("documents"))


@app.route("/barcode/<code>.svg")
@login_required
def barcode_svg(code):
    return Response(code128_svg(code), mimetype="image/svg+xml")


# ---------------------------------------------------------------- ЧМР

@app.route("/cmr/new", methods=["GET", "POST"])
@login_required
def cmr_new():
    con = db.get_db()
    if request.method == "POST":
        data = form_data()
        doc_id = save_document(con, "cmr", data)
        con.close()
        flash("ЧМР № %s е издадено и запазено в базата данни." % data["number"])
        return redirect(url_for("view_document", doc_id=doc_id))
    clients = load_clients(con)
    settings = db.get_settings(con)
    cj = clients_json(clients, con)  # con все още отворен — вгражда unload_points
    con.close()
    return render_template("cmr_form.html", clients=clients,
                           clients_json=cj, s=settings)


@app.route("/cmr/preview", methods=["POST"])
@login_required
def cmr_preview():
    return render_preview("cmr", form_data())


# ---------------------------------------------------------------- Опаковъчен лист

@app.route("/packing/new", methods=["GET", "POST"])
@login_required
def packing_new():
    con = db.get_db()
    if request.method == "POST":
        data = form_data()
        data["items"] = parse_items()
        doc_id = save_document(con, "packing", data)
        con.close()
        flash("Опаковъчен лист № %s е издаден и запазен." % data["number"])
        return redirect(url_for("view_document", doc_id=doc_id))
    clients = load_clients(con)
    settings = db.get_settings(con)
    con.close()
    return render_template("packing_form.html", clients=clients,
                           clients_json=clients_json(clients), s=settings,
                           items=[])


@app.route("/packing/preview", methods=["POST"])
@login_required
def packing_preview():
    data = form_data()
    data["items"] = parse_items()
    return render_preview("packing", data)


@app.route("/packing/pull-pallet", methods=["POST"])
@login_required
def packing_pull_pallet():
    """Издърпва обобщен ред (съдържание + нето/бруто тегло) от вече
    издадена палетна карта по нейния номер или баркод, за добавяне в
    опаковъчния лист — без ръчно преписване на данните."""
    code = request.form.get("code", "").strip()
    if not code:
        return {"ok": False, "error": "Въведете номер или баркод на палетна карта."}
    con = db.get_db()
    row = con.execute(
        "SELECT * FROM documents WHERE doc_type = 'pallet' AND (barcode = ? OR number = ?)"
        " ORDER BY id DESC LIMIT 1",
        (code, code),
    ).fetchone()
    if row is None:
        other = con.execute(
            "SELECT doc_type FROM documents WHERE barcode = ? OR number = ?"
            " ORDER BY id DESC LIMIT 1",
            (code, code),
        ).fetchone()
        con.close()
        if other is not None:
            title = db.DOC_TYPES.get(other["doc_type"], {}).get("title", other["doc_type"])
            return {"ok": False, "error": "Намереният документ не е палетна карта (%s)." % title}
        return {"ok": False, "error": "Няма документ с номер/баркод „%s“." % code}
    con.close()

    d = json.loads(row["data"])
    items = d.get("items") or []
    if d.get("items_format") == "orders":
        labels = [it.get("reference_desc") or it.get("reference") or it.get("order_no") or ""
                 for it in items]
    else:
        labels = [it.get("description") or it.get("code") or "" for it in items]
    labels = [l for l in labels if l]
    summary = ", ".join(labels[:3])
    if len(labels) > 3:
        summary += " и още %d" % (len(labels) - 3)
    description = "Палет %s" % (d.get("pallet_no") or row["number"])
    if summary:
        description += " — " + summary

    return {
        "ok": True,
        "number": row["number"],
        "row": {
            "description": description,
            "qty": d.get("boxes") or str(len(items)) or "1",
            "packing": "Палет",
            "net": d.get("net", ""),
            "gross": d.get("gross", ""),
        },
    }


# ---------------------------------------------------------------- Палетна карта

@app.route("/pallet/new", methods=["GET", "POST"])
@login_required
def pallet_new():
    con = db.get_db()
    if request.method == "POST":
        data = form_data()
        data["items"] = parse_items()
        doc_id = save_document(con, "pallet", data)
        con.close()
        flash("Палетна карта № %s е издадена и запазена." % data["number"])
        return redirect(url_for("view_document", doc_id=doc_id))
    clients = load_clients(con)
    settings = db.get_settings(con)
    con.close()
    return render_template("pallet_form.html", clients=clients,
                           clients_json=clients_json(clients), s=settings,
                           items=[])


@app.route("/pallet/preview", methods=["POST"])
@login_required
def pallet_preview():
    data = form_data()
    data["items"] = parse_items()
    return render_preview("pallet", data)


@app.route("/pallet/import", methods=["POST"])
@login_required
def pallet_import():
    """Импорт на редове за палетна карта от Excel файл (.xlsx).

    Очаквани колони: Артикул/код | Описание | Количество | Тегло (кг).
    Първият ред се пропуска, ако изглежда като заглавен.
    """
    from openpyxl import load_workbook

    file = request.files.get("excel_file")
    if not file or not file.filename:
        flash("Моля, изберете Excel файл (.xlsx).")
        return redirect(url_for("pallet_new"))
    try:
        wb = load_workbook(io.BytesIO(file.read()), data_only=True)
    except Exception:
        flash("Файлът не може да бъде прочетен. Уверете се, че е валиден .xlsx файл.")
        return redirect(url_for("pallet_new"))

    ws = wb.worksheets[0]
    items = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = ["" if c is None else str(c).strip() for c in row[:4]]
        cells += [""] * (4 - len(cells))
        if not any(cells):
            continue
        if i == 0 and _looks_like_header(cells):
            continue
        items.append({"code": cells[0], "description": cells[1],
                      "qty": cells[2], "weight": cells[3]})

    if not items:
        flash("Във файла не бяха намерени редове с данни.")
        return redirect(url_for("pallet_new"))

    con = db.get_db()
    clients = load_clients(con)
    settings = db.get_settings(con)
    con.close()
    flash("Заредени са %d реда от „%s“. Прегледайте и издайте картата." %
          (len(items), file.filename))
    return render_template("pallet_form.html", clients=clients,
                           clients_json=clients_json(clients), s=settings,
                           items=items)


def _looks_like_header(cells):
    joined = " ".join(cells).lower()
    keywords = ("артикул", "код", "описание", "колич", "тегло",
                "code", "item", "description", "qty", "quantity", "weight")
    return any(k in joined for k in keywords)


def _cellstr(v):
    """Клетка към низ, без излишно „.0“ за цели числа, записани като float."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _parse_order_export(ws):
    """Разпознава експортен файл на поръчки (колони Due Date, Order No, Pos,
    Project, Reference, Reference Desc, Open Qty, Unit, Stock, <номер на
    палетна карта>) и групира редовете по последната колона — всеки различен
    номер там става отделна палетна карта. Reference и Reference Desc се
    оставят празни за конкретен ред, ако липсват там (или изобщо няма такива
    колони във файла) — не се попълват с друга стойност. Връща
    {номер: [items]} подредени по реда на поява, или None ако форматът не е
    разпознат."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    header = [_cellstr(c) for c in rows[0]]
    header_lower = [h.lower() for h in header]

    def find_col(*names):
        for name in names:
            for i, h in enumerate(header_lower):
                if h == name:
                    return i
        return None

    col_order = find_col("order no", "order number", "orderno")
    col_pos = find_col("pos", "position")
    col_ref = find_col("reference")
    col_ref_desc = find_col("reference desc", "reference description", "ref desc")
    col_qty = find_col("open qty", "qty", "quantity")
    if col_order is None or col_qty is None:
        return None

    # Групиращата колона е последната без заглавие (примерният файл я оставя
    # безименна) — резервно, ако всички колони имат заглавие, вземаме
    # последната изобщо.
    group_col = None
    for i in range(len(header) - 1, -1, -1):
        if header[i] == "":
            group_col = i
            break
    if group_col is None:
        group_col = len(header) - 1

    def cell(row, i):
        if i is None or i >= len(row):
            return ""
        return _cellstr(row[i])

    groups = {}
    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue
        order_no = cell(row, col_order)
        if not order_no:
            continue
        group_raw = row[group_col] if group_col < len(row) else None
        try:
            group = int(group_raw)
        except (TypeError, ValueError):
            group = 1
        groups.setdefault(group, []).append({
            "order_no": order_no,
            "pos": cell(row, col_pos),
            "reference": cell(row, col_ref),
            "reference_desc": cell(row, col_ref_desc),
            "qty": cell(row, col_qty),
        })
    return groups if groups else None


@app.route("/pallet/sample.xlsx")
@login_required
def pallet_sample():
    """Примерен Excel файл за импорт на палетна карта."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Палетна карта"
    ws.append(["Артикул/код", "Описание", "Количество", "Тегло (кг)"])
    ws.append(["ART-001", "Кашон резервни части", 10, 125.5])
    ws.append(["ART-002", "Кутия крепежни елементи", 4, 38])
    for col, width in zip("ABCD", (16, 40, 14, 14)):
        ws.column_dimensions[col].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="primeren_palet_import.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument"
                              ".spreadsheetml.sheet")


@app.route("/pallet/bulk-import", methods=["POST"])
@login_required
def pallet_bulk_import():
    """Импорт от справка за поръчки (Order No, Pos, Reference, Open Qty) —
    редовете се разделят автоматично в отделни палетни карти по последната
    колона на файла (номер на палет)."""
    from openpyxl import load_workbook

    file = request.files.get("excel_file")
    if not file or not file.filename:
        flash("Моля, изберете Excel файл (.xlsx).")
        return redirect(url_for("pallet_new"))
    try:
        wb = load_workbook(io.BytesIO(file.read()), data_only=True)
    except Exception:
        flash("Файлът не може да бъде прочетен. Уверете се, че е валиден .xlsx файл.")
        return redirect(url_for("pallet_new"))

    groups = _parse_order_export(wb.worksheets[0])
    if not groups:
        flash("Файлът не съдържа разпознаваеми колони (Order No, Pos, Reference, "
             "Open Qty) или редове за импорт.")
        return redirect(url_for("pallet_new"))

    con = db.get_db()
    clients = load_clients(con)
    settings = db.get_settings(con)
    con.close()
    ordered = sorted(groups.items())
    flash("Открити са %d палетни карти (%d реда общо) от „%s“. Прегледайте и издайте." %
          (len(ordered), sum(len(v) for _, v in ordered), file.filename))
    return render_template("pallet_bulk_review.html", clients=clients,
                           clients_json=clients_json(clients), s=settings,
                           groups=ordered)


def _collect_bulk_pallet_drafts():
    """Чете подадените от прегледа за bulk импорт полета (общи за
    партидата + поотделно за всеки палет — тип, кашони, нето, бруто,
    височина) и връща списък от речници с данните за всяка карта, БЕЗ да
    ги записва в базата. Ползва се и от прегледа (без запис), и от
    реалното издаване."""
    shared_fields = ("sender_name", "sender_city", "client_name", "client_address",
                     "client_city", "client_country", "doc_date", "ref_cmr", "notes")
    shared = {k: request.form.get(k, "").strip() for k in shared_fields}
    per_card_fields = ("pallet_type", "boxes", "net", "gross", "height")
    group_ids = [g for g in request.form.get("groups", "").split(",") if g.strip()]

    drafts = []
    for g in group_ids:
        raw = request.form.get("items_json_%s" % g, "[]")
        try:
            items = json.loads(raw)
        except ValueError:
            items = []
        items = [it for it in items if isinstance(it, dict) and
                 any((it.get(k) or "").strip() if isinstance(it.get(k), str) else it.get(k)
                     for k in ("order_no", "pos", "reference", "reference_desc", "qty"))]
        if not items:
            continue
        data = dict(shared)
        for f in per_card_fields:
            data[f] = request.form.get("%s_%s" % (f, g), "").strip()
        data["items"] = items
        data["items_format"] = "orders"
        data["pallet_no"] = "%s от %s" % (g, len(group_ids))
        drafts.append(data)
    return drafts


@app.route("/pallet/bulk-preview", methods=["POST"])
@login_required
def pallet_bulk_preview():
    """Предварителен преглед на всички палетни карти от прегледа за bulk
    импорт, точно както ще изглеждат при печат — БЕЗ да се записват в
    базата и БЕЗ да се изразходват номера. POST → съхрани → пренасочи →
    GET, за да е безопасно презареждане/връщане назад на страницата (виж
    _store_preview по-горе)."""
    drafts = _collect_bulk_pallet_drafts()
    if not drafts:
        flash("Няма палетни карти за преглед (всички редове са празни).")
        return redirect(url_for("pallet_new"))
    token = _store_preview("bulk_pallet", drafts)
    return redirect(url_for("pallet_bulk_preview_view", token=token))


@app.route("/pallet/bulk-preview/<token>")
@login_required
def pallet_bulk_preview_view(token):
    drafts = _get_preview(token, "bulk_pallet")
    if drafts is None:
        flash("Прегледът е изтекъл или вече е използван — заредете файла отново.")
        return redirect(url_for("pallet_new"))
    return render_template("pallet_bulk_preview.html", drafts=drafts)


@app.route("/pallet/bulk-issue", methods=["POST"])
@login_required
def pallet_bulk_issue():
    """Издава наведнъж всички палетни карти от прегледа за импорт от
    справка за поръчки. Изпращач/клиент/дата/бележки са общи за цялата
    партида, но размерите и теглото на всеки палет (тип, кашони, нето,
    бруто, височина) се задават и записват отделно за всяка карта."""
    drafts = _collect_bulk_pallet_drafts()
    if not drafts:
        flash("Няма палетни карти за издаване (всички редове са празни).")
        return redirect(url_for("pallet_new"))

    con = db.get_db()
    created = []
    for data in drafts:
        doc_id = save_document(con, "pallet", data)
        created.append((data["number"], doc_id))
    con.close()

    flash("Издадени и запазени %d палетни карти: %s" %
         (len(created), ", ".join(num for num, _ in created)))
    return redirect(url_for("pallet_bulk_result",
                            ids=",".join(str(doc_id) for _, doc_id in created)))


@app.route("/pallet/bulk-result")
@login_required
def pallet_bulk_result():
    """Преглед на току-що издадените палетни карти преди печат — списък с
    бърз линк към всяка, за да се провери всяка карта, преди да се
    разпечата."""
    ids = [int(x) for x in request.args.get("ids", "").split(",") if x.strip().isdigit()]
    con = db.get_db()
    docs = []
    for doc_id in ids:
        row = con.execute(
            "SELECT d.*, u.full_name AS author FROM documents d"
            " LEFT JOIN users u ON u.id = d.created_by WHERE d.id = ?",
            (doc_id,),
        ).fetchone()
        if row is not None:
            docs.append((row, json.loads(row["data"])))
    con.close()
    return render_template("pallet_bulk_result.html", docs=docs)


# ---------------------------------------------------------------- Товарителница (вътрешен превоз)
# Бланка по образеца от Наредба № 33 на МТСИТ за обществен автомобилен
# превоз на товари в страната — различна от международната ЧМР (CMR):
# едноезична, с автомобил/пътен лист/инструкции на превозвача вместо
# митнически данни. Номерът/баркодът се генерират от самата програма
# (не се следи отделен сериен № на предпечатана бланка).

@app.route("/waybill/new", methods=["GET", "POST"])
@login_required
def waybill_new():
    con = db.get_db()
    if request.method == "POST":
        data = form_data()
        data["items"] = parse_items()
        doc_id = save_document(con, "waybill", data)
        con.close()
        flash("Товарителница № %s е издадена и запазена." % data["number"])
        return redirect(url_for("view_document", doc_id=doc_id))
    clients = load_clients(con)
    settings = db.get_settings(con)
    con.close()
    return render_template("waybill_form.html", clients=clients,
                           clients_json=clients_json(clients), s=settings,
                           items=[])


@app.route("/waybill/preview", methods=["POST"])
@login_required
def waybill_preview():
    data = form_data()
    data["items"] = parse_items()
    return render_preview("waybill", data)


# ---------------------------------------------------------------- Декларация за двойна употреба

@app.route("/dualuse/new", methods=["GET", "POST"])
@login_required
def dualuse_new():
    con = db.get_db()
    if request.method == "POST":
        data = form_data()
        data["items"] = parse_items()
        doc_id = save_document(con, "dualuse", data)
        con.close()
        flash("Декларация за двойна употреба № %s е издадена и запазена." % data["number"])
        return redirect(url_for("view_document", doc_id=doc_id))
    clients = load_clients(con)
    settings = db.get_settings(con)
    con.close()
    return render_template("dualuse_form.html", clients=clients,
                           clients_json=clients_json(clients), s=settings, items=[])


@app.route("/dualuse/preview", methods=["POST"])
@login_required
def dualuse_preview():
    data = form_data()
    data["items"] = parse_items()
    return render_preview("dualuse", data)


# ---------------------------------------------------------------- Декларация за износ (Италия)

@app.route("/export-it/new", methods=["GET", "POST"])
@login_required
def export_it_new():
    con = db.get_db()
    if request.method == "POST":
        data = form_data()
        data["items"] = parse_items()
        doc_id = save_document(con, "export_it", data)
        con.close()
        flash("Декларация за износ № %s е издадена и запазена." % data["number"])
        return redirect(url_for("view_document", doc_id=doc_id))
    clients = load_clients(con)
    settings = db.get_settings(con)
    con.close()
    return render_template("export_it_form.html", clients=clients,
                           clients_json=clients_json(clients), s=settings, items=[])


@app.route("/export-it/preview", methods=["POST"])
@login_required
def export_it_preview():
    data = form_data()
    data["items"] = parse_items()
    return render_preview("export_it", data)


# ---------------------------------------------------------------- адресна книга

@app.route("/clients")
@login_required
def clients_list():
    con = db.get_db()
    clients = load_clients(con)
    con.close()
    return render_template("clients.html", clients=clients)


@app.route("/clients/new", methods=["GET", "POST"])
@app.route("/clients/<int:client_id>/edit", methods=["GET", "POST"])
@login_required
def client_edit(client_id=None):
    con = db.get_db()
    client = None
    if client_id is not None:
        client = con.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
        if client is None:
            con.close()
            abort(404)
    if request.method == "POST":
        fields = ("name", "address", "city", "postcode", "country", "eik",
                  "vat", "phone", "email", "contact")
        values = [request.form.get(f, "").strip() for f in fields]
        if not values[0]:
            flash("Името на фирмата е задължително.")
        else:
            if client is None:
                cur = con.execute(
                    "INSERT INTO clients (%s) VALUES (%s)"
                    % (", ".join(fields), ", ".join("?" * len(fields))),
                    values,
                )
                new_client_id = cur.lastrowid
            else:
                con.execute(
                    "UPDATE clients SET %s WHERE id = ?"
                    % ", ".join(f + " = ?" for f in fields),
                    values + [client_id],
                )
                new_client_id = client_id
            try:
                unload_points = json.loads(request.form.get("unload_points_json", "[]"))
            except ValueError:
                unload_points = []
            db.save_unload_points(con, new_client_id,
                                  unload_points if isinstance(unload_points, list) else [])
            con.commit()
            con.close()
            flash("Клиентът е запазен в адресната книга.")
            return redirect(url_for("clients_list"))
    unload_points = db.get_unload_points(con, client_id) if client_id is not None else []
    con.close()
    return render_template("client_form.html", client=client,
                           unload_points=[dict(p) for p in unload_points])


@app.route("/clients/<int:client_id>/delete", methods=["POST"])
@login_required
def client_delete(client_id):
    con = db.get_db()
    con.execute("DELETE FROM clients WHERE id = ?", (client_id,))
    con.commit()
    con.close()
    flash("Клиентът е изтрит от адресната книга.")
    return redirect(url_for("clients_list"))


# ---------------------------------------------------------------- фирма изпращач

@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings_page():
    con = db.get_db()
    if request.method == "POST":
        keys = ("sender_name", "sender_address", "sender_city", "sender_postcode",
                "sender_country", "sender_eik", "sender_vat", "sender_phone",
                "sender_email", "sender_person")
        db.save_settings(con, {k: request.form.get(k, "").strip() for k in keys})
        con.commit()
        con.close()
        flash("Данните на фирмата изпращач са запазени.")
        return redirect(url_for("settings_page"))
    s = db.get_settings(con)
    con.close()
    return render_template("settings.html", s=s)


@app.route("/settings/logo", methods=["POST"])
@login_required
def settings_logo_upload():
    file = request.files.get("logo_file")
    if not file or not file.filename:
        flash("Моля, изберете файл с изображение.")
        return redirect(url_for("settings_page"))
    try:
        branding.save_logo(file)
        flash("Логото на фирмата е качено успешно.")
    except ValueError as exc:
        flash("Логото не бе прието: %s" % exc)
    return redirect(url_for("settings_page"))


@app.route("/settings/logo/remove", methods=["POST"])
@login_required
def settings_logo_remove():
    branding.remove_logo()
    flash("Логото на фирмата е премахнато.")
    return redirect(url_for("settings_page"))


@app.route("/logo.img")
@login_required
def company_logo_image():
    path = branding.logo_path()
    if path is None:
        abort(404)
    return send_file(path, mimetype=branding.logo_mimetype(path))


# ---------------------------------------------------------------- лични настройки (тема)

@app.route("/my-settings", methods=["GET", "POST"])
@login_required
def my_settings():
    con = db.get_db()
    if request.method == "POST":
        theme = request.form.get("theme", db.DEFAULT_THEME)
        if theme not in db.THEMES:
            theme = db.DEFAULT_THEME
        db.save_user_settings(con, session["user_id"], {"theme": theme})
        con.commit()
        con.close()
        session["theme"] = theme
        flash("Настройките са запазени.")
        return redirect(url_for("my_settings"))
    current_theme = db.get_user_theme(con, session["user_id"])
    ctx = {"themes": db.THEMES, "current_theme": current_theme}
    if session.get("role") == "admin":
        # Системните настройки (мрежа/архив/GitHub синхронизация) се
        # показват на същата страница, видими само за администратори.
        ctx.update(s=db.get_settings(con), cfg=appconfig.load_config(),
                  db_path=db.DB_PATH, sync=backup.sync_status())
    con.close()
    return render_template("my_settings.html", **ctx)


# ---------------------------------------------------------------- системни настройки (админ,
# показвани вградени в „Настройки“ — вижте my_settings по-горе)

@app.route("/admin/system", methods=["GET", "POST"])
@admin_required
def system_settings():
    if request.method == "GET":
        return redirect(url_for("my_settings"))
    con = db.get_db()
    form = request.form.get("form")
    if form == "network":
        appconfig.save_config({
            "db_path": request.form.get("db_path", "").strip(),
            "network_mode": request.form.get("network_mode") == "on",
            "network_port": int(request.form.get("network_port") or 5000),
        })
        flash("Мрежовите настройки са запазени. Рестартирайте програмата, "
             "за да влязат в сила.")
    elif form == "backup_folder":
        db.save_settings(con, {
            "backup_folder": request.form.get("backup_folder", "").strip(),
            "backup_auto": "on" if request.form.get("backup_auto") == "on" else "",
        })
        con.commit()
        flash("Настройките за локален/мрежов архив са запазени.")
    elif form == "backup_github":
        # GitHub данните се пазят в pacho_config.json (не в базата), за да
        # може нова инсталация да ги прочете и да изтегли базата ПРЕДИ тя
        # изобщо да съществува локално.
        current = appconfig.load_config()
        appconfig.save_config({
            "gh_owner": request.form.get("gh_owner", "").strip(),
            "gh_repo": request.form.get("gh_repo", "").strip(),
            "gh_branch": request.form.get("gh_branch", "main").strip() or "main",
            "gh_path": request.form.get("gh_path", "pacho_logistic.db").strip()
                      or "pacho_logistic.db",
            "gh_token": request.form.get("gh_token", "").strip() or current.get("gh_token", ""),
            "gh_auto_sync": request.form.get("gh_auto_sync") == "on",
        })
        flash("Настройките за GitHub синхронизация са запазени.")
    con.close()
    return redirect(url_for("my_settings"))


@app.route("/admin/system/backup-now", methods=["POST"])
@admin_required
def system_backup_now():
    con = db.get_db()
    folder = db.get_settings(con).get("backup_folder", "").strip()
    con.close()
    try:
        path = backup.local_backup(folder)
        flash("Резервно копие е записано: %s" % path)
    except Exception as exc:
        flash("Архивирането е неуспешно: %s" % exc)
    return redirect(url_for("my_settings"))


@app.route("/admin/system/backup-github-now", methods=["POST"])
@admin_required
def system_backup_github_now():
    cfg = appconfig.load_config()
    try:
        url = backup.github_backup(
            cfg.get("gh_owner", ""), cfg.get("gh_repo", ""), cfg.get("gh_token", ""),
            cfg.get("gh_branch", "main") or "main",
            cfg.get("gh_path", "pacho_logistic.db") or "pacho_logistic.db",
        )
        backup._sync_state["dirty"] = False
        backup._sync_state["last_synced_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        backup._sync_state["last_error"] = None
        flash("Базата данни е качена в GitHub успешно.%s" % (" " + url if url else ""))
    except Exception as exc:
        flash("Синхронизацията с GitHub е неуспешна: %s" % exc)
    return redirect(url_for("my_settings"))


@app.route("/admin/system/pull-now", methods=["POST"])
@admin_required
def system_pull_now():
    """Ръчно изтегляне на базата данни от GitHub (замества текущата
    локална база!) — за възстановяване или преминаване към споделените
    данни на друга инсталация."""
    cfg = appconfig.load_config()
    ok, err = backup.pull_db(
        cfg.get("gh_owner", ""), cfg.get("gh_repo", ""), cfg.get("gh_token", ""),
        cfg.get("gh_branch", "main") or "main",
        cfg.get("gh_path", "pacho_logistic.db") or "pacho_logistic.db",
        db.DB_PATH,
    )
    if ok:
        flash("Базата данни е изтеглена от GitHub. Рестартирайте програмата, "
             "за да заредите новите данни.")
    else:
        flash("Изтеглянето от GitHub е неуспешно: %s" % err)
    return redirect(url_for("my_settings"))


# ---------------------------------------------------------------- отдалечен достъп (сканиране с телефон)

@app.route("/admin/system/remote-start", methods=["POST"])
@admin_required
def system_remote_start():
    port = int(appconfig.load_config().get("network_port") or 5000)
    remote_tunnel.start(port)
    flash("Стартира се отдалечен достъп… изчакайте няколко секунди, статусът "
         "по-долу ще се обнови автоматично.")
    return redirect(url_for("my_settings"))


@app.route("/admin/system/remote-stop", methods=["POST"])
@admin_required
def system_remote_stop():
    remote_tunnel.stop()
    flash("Отдалеченият достъп е спрян.")
    return redirect(url_for("my_settings"))


@app.route("/admin/system/remote-status")
@admin_required
def system_remote_status():
    return remote_tunnel.status()


# ---------------------------------------------------------------- админ панел

@app.route("/admin/users")
@admin_required
def admin_users():
    con = db.get_db()
    users = con.execute("SELECT * FROM users ORDER BY username").fetchall()
    con.close()
    return render_template("admin_users.html", users=users)


@app.route("/admin/users/new", methods=["POST"])
@admin_required
def admin_user_new():
    username = request.form.get("username", "").strip()
    full_name = request.form.get("full_name", "").strip()
    password = request.form.get("password", "")
    role = "admin" if request.form.get("role") == "admin" else "employee"
    if not username or not password:
        flash("Потребителско име и парола са задължителни.")
        return redirect(url_for("admin_users"))
    if len(password) < MIN_PASSWORD_LENGTH:
        flash("Паролата трябва да е поне %d символа." % MIN_PASSWORD_LENGTH)
        return redirect(url_for("admin_users"))
    con = db.get_db()
    exists = con.execute("SELECT 1 FROM users WHERE username = ?", (username,)).fetchone()
    if exists:
        flash("Вече има служител с потребителско име „%s“." % username)
    else:
        # must_change_password=1: администраторът вече знае тази парола
        # (той я е въвел тук), затова не е лична тайна на служителя —
        # задължаваме смяна при първия му вход.
        con.execute(
            "INSERT INTO users"
            " (username, password_hash, full_name, role, active, must_change_password)"
            " VALUES (?, ?, ?, ?, 1, 1)",
            (username, generate_password_hash(password), full_name, role),
        )
        con.commit()
        flash("Служителят „%s“ е добавен. Ще трябва да смени паролата при първия вход." % username)
    con.close()
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/toggle", methods=["POST"])
@admin_required
def admin_user_toggle(user_id):
    if user_id == session["user_id"]:
        flash("Не можете да деактивирате собствения си акаунт.")
        return redirect(url_for("admin_users"))
    con = db.get_db()
    con.execute("UPDATE users SET active = 1 - active WHERE id = ?", (user_id,))
    con.commit()
    con.close()
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/password", methods=["POST"])
@admin_required
def admin_user_password(user_id):
    password = request.form.get("password", "")
    if not password:
        flash("Въведете нова парола.")
        return redirect(url_for("admin_users"))
    if len(password) < MIN_PASSWORD_LENGTH:
        flash("Паролата трябва да е поне %d символа." % MIN_PASSWORD_LENGTH)
        return redirect(url_for("admin_users"))
    con = db.get_db()
    # must_change_password=1 по същата причина, както при admin_user_new —
    # администраторът, не служителят, е избрал тази парола.
    con.execute(
        "UPDATE users SET password_hash = ?, must_change_password = 1 WHERE id = ?",
        (generate_password_hash(password), user_id))
    con.commit()
    con.close()
    flash("Паролата е сменена. Служителят ще трябва да я смени при следващия си вход.")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@admin_required
def admin_user_delete(user_id):
    if user_id == session["user_id"]:
        flash("Не можете да изтриете собствения си акаунт.")
        return redirect(url_for("admin_users"))
    con = db.get_db()
    con.execute("UPDATE documents SET created_by = NULL WHERE created_by = ?", (user_id,))
    con.execute("DELETE FROM users WHERE id = ?", (user_id,))
    con.commit()
    con.close()
    flash("Служителят е изтрит.")
    return redirect(url_for("admin_users"))


# ---------------------------------------------------------------- обновяване

@app.route("/update/check")
@login_required
def update_check():
    """Ръчна проверка за нова версия в GitHub Releases."""
    try:
        info = updater.check_for_update()
    except Exception as exc:
        flash("Проверката за обновяване е неуспешна: %s" % updater.describe_error(exc))
        return redirect(url_for("dashboard"))
    updater._cache["info"] = info
    updater._cache["last_error"] = None
    updater._cache["time"] = time.time()
    if info["available"]:
        flash("Налична е нова версия %s (текущата е %s)." % (info["latest"], info["current"]))
    else:
        flash("Използвате най-новата версия (%s)." % info["current"])
    return redirect(url_for("dashboard"))


@app.route("/update/install", methods=["POST"])
@login_required
def update_install():
    """Изтегля новата версия и рестартира програмата."""
    try:
        info = updater.check_for_update()
    except Exception as exc:
        flash("Проверката за обновяване е неуспешна: %s" % updater.describe_error(exc))
        return redirect(url_for("dashboard"))
    if not info["available"]:
        flash("Вече използвате най-новата версия (%s)." % info["current"])
        return redirect(url_for("dashboard"))
    try:
        updater.install_update(info["download"], info.get("expected_sha256"))
    except Exception as exc:
        flash("Обновяването е неуспешно: %s" % updater.describe_error(exc))
        return redirect(url_for("dashboard"))
    return render_template("updating.html", latest=info["latest"])


# ---------------------------------------------------------------- смяна на парола

@app.route("/password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        current = request.form.get("current", "")
        new = request.form.get("new", "")
        repeat = request.form.get("repeat", "")
        con = db.get_db()
        user = con.execute("SELECT * FROM users WHERE id = ?",
                           (session["user_id"],)).fetchone()
        if not check_password_hash(user["password_hash"], current):
            flash("Текущата парола е грешна.")
        elif len(new) < MIN_PASSWORD_LENGTH:
            flash("Новата парола трябва да е поне %d символа." % MIN_PASSWORD_LENGTH)
        elif new != repeat:
            flash("Двете нови пароли не съвпадат.")
        else:
            con.execute(
                "UPDATE users SET password_hash = ?, must_change_password = 0 WHERE id = ?",
                (generate_password_hash(new), session["user_id"]))
            con.commit()
            con.close()
            session["must_change_password"] = False
            flash("Паролата е сменена успешно.")
            return redirect(url_for("dashboard"))
        con.close()
    return render_template("change_password.html", forced=session.get("must_change_password", False))


def _get_backup_settings():
    con = db.get_db()
    s = db.get_settings(con)
    con.close()
    return s


if __name__ == "__main__":
    _cfg = appconfig.load_config()
    _host = "0.0.0.0" if _cfg.get("network_mode") else "127.0.0.1"
    _port = int(_cfg.get("network_port") or 5000)
    _local_url = "http://127.0.0.1:%d" % _port

    # Фоновият архивиращ таймер винаги стартира; сам проверява дали е
    # зададена папка за архив в „Системни настройки“ и иначе не прави нищо.
    backup.start_auto_backup(_get_backup_settings)

    # Истински автоматично обновяване: проверява и инсталира новата версия
    # във фонов режим, без потребителят да трябва да натиска бутон.
    # Пропуска се, ако тази инсталация в момента е централен сървър за
    # други компютри в офиса (мрежов режим) — там рестарт би прекъснал
    # работата на всички останали неочаквано, затова остава ръчно.
    updater.start_auto_update_loop(lambda: _host == "0.0.0.0")

    if getattr(sys, "frozen", False):
        # Истинско настолно приложение: Flask сървърът работи във фонова
        # нишка в СЪЩИЯ процес, а прозорецът е вграден (pywebview/WebView2)
        # — без изобщо да се стартира отделен браузър процес. При неуспех
        # (напр. WebView2 липсва) пада към Chrome/Edge в режим „приложение“,
        # а като последна мярка — обикновен браузър.
        server_thread = threading.Thread(
            target=lambda: app.run(host=_host, port=_port, debug=False,
                                   use_reloader=False),
            daemon=True,
        )
        server_thread.start()
        print("%s v%s — %s (настолен режим)" % (APP_NAME, __version__, _local_url))
        opened_native = desktop.run_native_window(
            _local_url, title="%s v%s" % (APP_NAME, __version__))
        if not opened_native:
            if not desktop.open_app_window(_local_url):
                webbrowser.open(_local_url)
            server_thread.join()
        else:
            os._exit(0)
    else:
        print("%s v%s — %s%s" % (
            APP_NAME, __version__, _local_url,
            " (мрежов режим — достъпно и от други компютри в мрежата)" if _host == "0.0.0.0" else ""))
        app.run(host=_host, port=_port, debug=False)
