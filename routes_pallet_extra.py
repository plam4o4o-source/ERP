# -*- coding: utf-8 -*-
"""Допълнителни хендлъри около палетни карти и опаковъчни листи: издърпване
на обобщен ред от палетна карта в опаковъчен лист, bulk импорт от справка за
поръчки, плюс предварителен преглед и масово издаване на bulk-внесените
карти. Извлечено от app.py (Фаза 3) без промяна в поведението."""
import io
import itertools
import json
import zipfile

from flask import flash, redirect, render_template, request, url_for
from flask_babel import gettext as _

import applog
import db
from appcore import (CLIENT_EMBED_LIMIT, XlsxTooLargeError, ensure_xlsx_within_limits, _get_preview,
                     _store_preview, clients_json,
                     count_clients, get_db, load_clients, login_required,
                     negative_item_rows, pallet_total_qty, safe_json_data,
                     save_document, unparsable_item_rows)

# Одит (16.08.2026, находка №18, средна): вижте _parse_order_export по-долу
# за пълния разказ — сканира се само ограничен брой редове за заглавие, а
# импортът се ограничава до тук зададения максимум редове данни, за да не
# може прекалено голям качен файл да изчерпи паметта на процеса.
_HEADER_SCAN_ROWS = 10
_MAX_IMPORT_DATA_ROWS = 5000


def _read_limited_rows(ws):
    """Одит (19.08.2026, находка №14, висока): таванът от находка №18
    (_MAX_IMPORT_DATA_ROWS) не пазеше НИТО паметта, НИТО времето —
    коментарът твърдеше, че `read_only=True` пести памет, но следващият ред
    правеше `list(ws.iter_rows(...))`, тоест ЦЕЛИЯТ лист се материализираше
    в паметта и чак СЛЕД това се режеше на 5000 реда. Измерено срещу файл
    от 9 MB с 300 000 реда: 27 792 ms и +148 MB RSS в един waitress работен
    процес (при таванa MAX_CONTENT_LENGTH от 25 MB това е ~70 сек и
    ~400 MB) — двама оператори наведнъж стигат до забиване на офисния
    компютър, а накрая програмата любезно съобщава, че е прочела само
    първите 5000 реда.

    Затова тук се четат САМО толкова реда, колкото изобщо могат да бъдат
    използвани: _HEADER_SCAN_ROWS (заглавието се търси най-много до 10-ия
    ред) + _MAX_IMPORT_DATA_ROWS данни, плюс ЕДИН допълнителен ред. Този
    един допълнителен ред служи само за откриване, че файлът съдържа още
    данни — предупреждението за орязване се вдига, ако итераторът НЕ е
    изчерпан, вместо да се сравнява дължината на вече прочетен цял лист.

    Връща (rows, exhausted): `exhausted` е True, когато листът е прочетен
    докрай (тоест НЯМА орязване)."""
    it = ws.iter_rows(values_only=True)
    limit = _HEADER_SCAN_ROWS + _MAX_IMPORT_DATA_ROWS + 1
    rows = list(itertools.islice(it, limit))
    exhausted = next(it, None) is None if len(rows) == limit else True
    return rows, exhausted


def _xlsx_has_formulas(file_bytes):
    """Одит (19.08.2026, находка №39, дребна): `data_only=True` връща
    кешираната стойност на формулна клетка — но файл, ГЕНЕРИРАН от външна
    библиотека (а не записан от самия Excel), няма такъв кеш и всяка
    формулна клетка се чете като None. Типично за автоматично генерирани
    справки: ред с `=2+3` в „Open Qty“ се внасяше с `qty: ''`, а
    съобщението гласеше „Открити са 1 палетни карти (1 реда общо)“ — на вид
    напълно успешен импорт с празни количества.

    Проверката е огледална на _xlsx_has_merged_cells по-долу и по същата
    причина чете суровия XML на листовете директно (елементът `<f>` вътре
    в `<c>` е самата формула), вместо да прави ВТОРИ пълен прочит на
    работната книга с `data_only=False` — евтино дори за голям файл и не
    връща паметта, спестена от находка №14 по-горе."""
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            for name in zf.namelist():
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                    xml = zf.read(name)
                    if b"<f>" in xml or b"<f " in xml:
                        return True
    except Exception:
        return False
    return False


#: Одит (19.08.2026, находка №40, дребна): при заглавен ред след 10-ия
#: съобщението беше ТОЧНО същото като при напълно грешен файл, така че
#: операторът нямаше как да се сети, че просто трябва да махне
#: декоративните редове отгоре. Допълнението се долепя към двете
#: съобщения „Файлът не съдържа разпознаваеми колони…“ (тук и в
#: routes_invoices).
def _header_row_hint():
    return _("Заглавният ред трябва да е в първите %d реда на листа.") % _HEADER_SCAN_ROWS


def _formula_hint():
    """Одит (19.08.2026, находка №39): текстът, който обяснява защо иначе
    валиден файл се внася с празни количества (виж _xlsx_has_formulas)."""
    return _("Файлът съдържа формули без запазени стойности — отворете го и го "
             "запишете от Excel, след което опитайте отново.")


def _parse_sheets(wb, parser):
    """Одит (19.08.2026, находка №29, средна): двата Excel импорта четяха
    САМО `wb.worksheets[0]`, докато materials.parse_catalog_xlsx в СЪЩИЯ
    проект отдавна обхожда всички листове. Реален файл с декоративен лист
    „Инфо“ отпред и лист „Данни“ с валидните колони (дори когато „Данни“ е
    активният лист!) отказваше с „Файлът не съдържа разпознаваеми колони“
    при напълно валиден файл.

    Обхожда листовете по ред и връща резултата от ПЪРВИЯ с разпознати
    колони: (parsed, warnings, sheet_title). При нито един разпознат лист
    връща (None, warnings-от-първия-лист, None) — предупрежденията на
    неразпознатите листове (напр. „заглавието е на ред 3“) не се показват,
    за да не обяснява програмата подробности за декоративен лист."""
    sheets = list(wb.worksheets)
    for ws in sheets:
        parsed, warnings = parser(ws)
        if parsed:
            if len(sheets) > 1:
                warnings.insert(0, _("Данните са прочетени от лист „%(sheet)s“ "
                                     "(файлът съдържа %(count)d листа).")
                                % {"sheet": ws.title, "count": len(sheets)})
            return parsed, warnings, ws.title
    return None, [], None


def _xlsx_has_merged_cells(file_bytes):
    """Одит (16.08.2026, находка №18): openpyxl в `read_only=True` режим
    (виж load_workbook по-долу) НЕ излага `worksheet.merged_cells` изобщо
    (AttributeError) — затова проверката тук чете директно суровия XML на
    листовете вътре в .xlsx (ZIP архив), без да минава през openpyxl,
    евтино дори за голям файл (само за наличие на `<mergeCell `, не пълен
    разбор). Обединена клетка връща стойност САМО в горния ляв ъгъл на
    диапазона — всички останали клетки от диапазона се четат като None,
    което може тихо да „изгуби“ данни от импортирания файл, ако заглавие/
    ред попада точно върху такъв диапазон; предупреждаваме потребителя,
    вместо да се преструваме, че не забелязваме."""
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            for name in zf.namelist():
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                    if b"<mergeCell " in zf.read(name):
                        return True
    except Exception:
        # Невалиден/повреден архив — самият load_workbook по-долу ще
        # хвърли собствена, по-конкретна грешка; тук просто не съобщаваме
        # лъжливо предупреждение за обединени клетки.
        return False
    return False


def register(app):
    app.add_url_rule("/packing/pull-pallet", "packing_pull_pallet",
                     packing_pull_pallet, methods=["POST"])
    app.add_url_rule("/pallet/bulk-import", "pallet_bulk_import",
                     pallet_bulk_import, methods=["POST"])
    app.add_url_rule("/pallet/bulk-preview", "pallet_bulk_preview",
                     pallet_bulk_preview, methods=["POST"])
    app.add_url_rule("/pallet/bulk-preview/<token>", "pallet_bulk_preview_view",
                     pallet_bulk_preview_view)
    app.add_url_rule("/pallet/bulk-review/restore/<token>", "pallet_bulk_review_restore",
                     pallet_bulk_review_restore)
    app.add_url_rule("/pallet/bulk-issue", "pallet_bulk_issue",
                     pallet_bulk_issue, methods=["POST"])
    app.add_url_rule("/pallet/bulk-result", "pallet_bulk_result", pallet_bulk_result)
    app.add_url_rule("/pallet/bulk-print", "pallet_bulk_print", pallet_bulk_print)


@login_required
def packing_pull_pallet():
    """Издърпва обобщен ред (съдържание + нето/бруто тегло) от вече
    издадена палетна карта по нейния номер или баркод, за добавяне в
    опаковъчния лист — без ръчно преписване на данните."""
    code = request.form.get("code", "").strip()
    if not code:
        return {"ok": False, "error": _("Въведете номер или баркод на палетна карта.")}
    con = get_db()
    row = con.execute(
        "SELECT * FROM documents WHERE doc_type = 'pallet' AND (barcode = ? OR number = ?)"
        " ORDER BY id DESC LIMIT 1",
        (code, code),
    ).fetchone()
    if row is None:
        other = con.execute(
            "SELECT doc_type FROM documents WHERE barcode = ? OR number = ?"
            " ORDER BY id DESC LIMIT 1",
            (code, code),
        ).fetchone()
        if other is not None:
            title = db.DOC_TYPES.get(other["doc_type"], {}).get("title", other["doc_type"])
            return {"ok": False, "error": _("Намереният документ не е палетна карта (%s).") % title}
        return {"ok": False, "error": _("Няма документ с номер/баркод „%s“.") % code}

    d = safe_json_data(row["data"])
    # Одит (01.09.2026, девети одит, находка №6): огледално на
    # routes_invoices.invoice_pull_pallet — виж пълното обяснение там.
    items = [it for it in (d.get("items") or []) if isinstance(it, dict)]
    if d.get("items_format") == "orders":
        labels = [it.get("reference_desc") or it.get("reference") or it.get("order_no") or ""
                 for it in items]
    else:
        labels = [it.get("description") or it.get("code") or "" for it in items]
    labels = [l for l in labels if l]
    summary = ", ".join(labels[:3])
    if len(labels) > 3:
        # Дребни (одит): голи низове без _() — при интерфейс на EN/TR
        # излизаха на български независимо от избрания език.
        summary += _(" и още %d") % (len(labels) - 3)
    description = _("Палет %s") % (d.get("pallet_no") or row["number"])
    if summary:
        description += " — " + summary

    return {
        "ok": True,
        "number": row["number"],
        # Одит (находка С12, нисък риск): по-рано тук стоеше d.get("net", "")
        # — палетната карта отдавна НЯМА поле „net“ (заменено с „Общ брой“,
        # виж appcore.pallet_total_qty), затова полето беше ВИНАГИ празно,
        # без операторът да разбира защо. "note" обяснява изрично защо
        # нето теглото трябва да се въведе ръчно, вместо мълчаливо празно
        # поле да изглежда като грешка в самата програма.
        "note": _("Палетната карта не пази нето тегло — попълнете го ръчно."),
        "row": {
            "description": description,
            "qty": pallet_total_qty(items) or str(len(items)) or "1",
            "packing": _("Палет"),
            "gross": d.get("gross", ""),
        },
    }


def _cellstr(v):
    """Клетка към низ, без излишно „.0“ за цели числа, записани като float."""
    if v is None:
        return ""
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        # Одит (02.09.2026, десети одит, находка №4): дробният float се
        # връщаше СУРОВ (`str(v)`), тоест точно както IEEE754 го пази.
        # Клетка с формула в Excel („Open Qty“ = разлика от две числа)
        # редовно държи 2.9000000000000004, а `fmt_num(value)` с
        # decimals=None НАРОЧНО пази въведената точност — тоест този запис
        # с 16 знака се отпечатваше буквално в колоната за количество на
        # палетна карта / търговска фактура, документ за клиента и за
        # митницата. Отделно беше и спусъкът на находка №1: един такъв ред
        # разваляше цялата жива сума на екрана.
        # `materials._weight_cell` решава същия проблем за теглата от
        # 19.08.2026 („%.6f“ + отрязване на нулите) — тук е същото, но с
        # предпазна клауза: ако закръглянето би превърнало ненулева
        # стойност в „0“ (напр. 8.7e-09), се връща суровият запис, за да
        # остане редът разпознат от `unparsable_item_rows` и операторът да
        # получи предупреждение, вместо тихо да види количество нула.
        text = ("%.6f" % v).rstrip("0").rstrip(".")
        if text in ("", "0", "-0") and v != 0:
            return str(v).strip()
        return text or "0"
    return str(v).strip()


def _parse_group_numbers(raw):
    """Парсва стойността на групиращата колона в _parse_order_export —
    обикновено едно цяло число, но може да съдържа няколко номера,
    разделени с „+“ (напр. „1+3+4“), ако редът принадлежи физически на
    няколко палетни карти едновременно (материалът е физически наличен и в
    двете/трите). Връща списък от int групи; непарсваемо/празно съдържание
    пада към [1] (по подразбиране всичко отива в карта № 1).

    Одит (12.08.2026, находка №7, high): дублирани номера в стойността
    (напр. „1+1“ — печатна грешка или copy-paste в изходния файл) преди
    тази поправка добавяха СЪЩИЯ артикул ДВА ПЪТИ в СЪЩАТА карта —
    удвоено количество, дублиран ред на бланката, без нищо да сигнализира
    проблема. Дубликатите се премахват тук, като се пази реда на първа
    поява (напр. „1+2+1“ → [1, 2], не [1, 2, 1])."""
    if raw is None:
        return [1]
    parts = str(raw).split("+")
    nums = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        try:
            n = int(float(p))
        except (TypeError, ValueError):
            continue
        if n not in nums:
            nums.append(n)
    return nums if nums else [1]


def _parse_order_export(ws):
    """Разпознава експортен файл на поръчки (колони Order No, Pos,
    Reference, Reference Desc, Open Qty — плюс произволни други,
    напр. Due Date/Project/Unit/Stock/Company, и номер на палетна карта в
    последната колона) и групира редовете по последната колона — всеки
    различен номер там става отделна палетна карта. Ако последната колона
    съдържа няколко номера, разделени с "+" (напр. "1+3+4"), редът се
    добавя КЪМ ВСЯКА от изброените карти (виж _parse_group_numbers) — за
    материал, физически наличен в повече от една карта.

    Пазят се САМО тези 5 колони (заявка: „палетна карта да зарежда само
    информацията от следните колони и да съдържа само тях... Order No,
    Pos, Reference, Reference Desc, Open Qty“) — всички останали колони на
    файла (Due Date, Project, Unit, Stock и т.н.) се игнорират нарочно,
    дори да присъстват. Липсваща/неразпозната от тези 5 колона просто
    остава празна за съответното поле, не се попълва с друга стойност.

    Връща (groups, warnings): `groups` е {номер: [items]} подредени по
    реда на поява, или None ако форматът не е разпознат; `warnings` е
    списък от низове за flash (напр. орязване при твърде много редове —
    виж находка №18)."""
    warnings = []
    # Одит (19.08.2026, находка №14): вместо `list(ws.iter_rows(...))` —
    # виж _read_limited_rows по-горе за пълния разказ (целият файл влизаше
    # в паметта ПРЕДИ рязането на 5000 реда).
    rows, exhausted = _read_limited_rows(ws)
    if not rows:
        return None, warnings

    # Одит (16.08.2026, находка №18, средна): преди тази поправка ЗАГЛАВНИЯТ
    # РЕД се приемаше БЕЗУСЛОВНО за rows[0] — реален износ от ERP/BI
    # системи често има допълнителен ред отгоре (заглавие на справката,
    # дата на генериране, лого и т.н.), заради което истинските заглавия
    # на колоните никога не се сравняваха с очакваните имена и целият внос
    # отказваше с „файлът не съдържа разпознаваеми колони“. Сега се
    # сканират първите _HEADER_SCAN_ROWS реда и се взема ПЪРВИЯТ, в който
    # намираме И ДВЕТЕ задължителни колони (Order No, Open Qty) — вместо
    # сляпо да предполагаме позиция 0.
    def find_col_in(header_lower, *names):
        for name in names:
            for i, h in enumerate(header_lower):
                if h == name:
                    return i
        return None

    header_idx = 0
    header = [_cellstr(c) for c in (rows[0] or [])]
    header_lower = [h.lower() for h in header]
    for idx in range(min(_HEADER_SCAN_ROWS, len(rows))):
        candidate = [_cellstr(c) for c in (rows[idx] or [])]
        candidate_lower = [h.lower() for h in candidate]
        if (find_col_in(candidate_lower, "order no", "order number", "orderno") is not None
                and find_col_in(candidate_lower, "open qty", "qty", "quantity") is not None):
            header_idx, header, header_lower = idx, candidate, candidate_lower
            break
    if header_idx > 0:
        warnings.append(_("Заглавният ред е открит на ред %d от файла (пропуснати са "
                          "%d реда над него) — проверете дали разпознатите данни са "
                          "правилни.") % (header_idx + 1, header_idx))

    data_rows = rows[header_idx + 1:]
    # Одит (19.08.2026, находка №14): орязването се разпознава по това, че
    # итераторът НЕ е изчерпан (или че прочетените редове вече надхвърлят
    # тавана) — по-рано тук се сравняваше дължината на СПИСЪК с целия лист.
    if len(data_rows) > _MAX_IMPORT_DATA_ROWS or not exhausted:
        warnings.append(_("Файлът съдържа повече от %d реда данни — заредени са само "
                          "първите %d, останалите са пропуснати.")
                        % (_MAX_IMPORT_DATA_ROWS, _MAX_IMPORT_DATA_ROWS))
        data_rows = data_rows[:_MAX_IMPORT_DATA_ROWS]

    def find_col(*names):
        return find_col_in(header_lower, *names)

    def cell_has_value(row, i):
        """Дали клетка i от този ред е реално попълнена (не None/празен
        низ след трим) — ползва се само за откриване на групиращата
        колона (находка №6), затова минималната проверка е достатъчна."""
        if row is None or i is None or i >= len(row):
            return False
        return _cellstr(row[i]) != ""

    col_order = find_col("order no", "order number", "orderno")
    col_pos = find_col("pos", "position")
    col_ref = find_col("reference")
    col_ref_desc = find_col("reference desc", "reference description", "ref desc")
    col_qty = find_col("open qty", "qty", "quantity")
    if col_order is None or col_qty is None:
        return None, warnings

    # Групиращата колона е последната без заглавие (примерният файл я оставя
    # безименна) — резервно, ако всички колони имат заглавие, вземаме
    # последната изобщо.
    #
    # Одит (12.08.2026, находка №6, high): преди тази поправка тук се
    # вземаше просто НАЙ-ДЯСНАТА колона с празно заглавие, по ПОЗИЦИЯ —
    # без изобщо да се провери дали тя реално съдържа данни. Чест остатък
    # при износ от Excel е ОЩЕ една напълно празна колона след истинската
    # групираща (напр. неизползван диапазон, включен в експорта) — в този
    # случай кодът избираше именно нея (истински празна навсякъде),
    # `cell(row, group_col)` връщаше "" за всеки ред, `_parse_group_numbers`
    # падаше към подразбиращото се [1] за всеки ред — ВСИЧКИ редове
    # погрешно попадаха в карта №1, вместо да се разпределят по палети.
    # Сега сред колоните с празно заглавие се търси (от дясно наляво,
    # запазвайки предпочитанието към последната) ПЪРВАТА, която реално
    # съдържа поне една непразна стойност в данните — истински празните
    # колони отдясно на нея се прескачат.
    candidates = [i for i in range(len(header) - 1, -1, -1) if header[i] == ""]
    group_col = None
    for i in candidates:
        if any(cell_has_value(row, i) for row in data_rows):
            group_col = i
            break
    if group_col is None and candidates:
        # Нито една от безименните колони не съдържа данни (напр. файл без
        # реална групираща колона, но с празен остатъчен диапазон) — пази
        # старото поведение (последната безименна) вместо да гърми.
        group_col = candidates[0]
    if group_col is None:
        group_col = len(header) - 1

    def cell(row, i):
        if i is None or i >= len(row):
            return ""
        return _cellstr(row[i])

    groups = {}
    for row in data_rows:
        if row is None or all(c is None for c in row):
            continue
        order_no = cell(row, col_order)
        if not order_no:
            continue
        group_raw = row[group_col] if group_col < len(row) else None
        item = {
            "order_no": order_no,
            "pos": cell(row, col_pos),
            "reference": cell(row, col_ref),
            "reference_desc": cell(row, col_ref_desc),
            "qty": cell(row, col_qty),
        }
        # Един и същ ред може да принадлежи на няколко карти наведнъж
        # ("1+3" и т.н.) — добавяме СЪЩИЯ артикул към всяка от тях (не
        # копие — общите редакции по-нататък не мутират тези речници).
        for group in _parse_group_numbers(group_raw):
            groups.setdefault(group, []).append(item)
    return (groups if groups else None), warnings


@login_required
def pallet_bulk_import():
    """Импорт от справка за поръчки (Order No, Pos, Reference, Reference
    Desc, Open Qty — само тези колони се зареждат, останалите от файла се
    игнорират, вижте _parse_order_export) — редовете се разделят
    автоматично в отделни палетни карти по последната колона на файла
    (номер на палет)."""
    from openpyxl import load_workbook

    file = request.files.get("excel_file")
    if not file or not file.filename:
        flash(_("Моля, изберете Excel файл (.xlsx)."), "error")
        return redirect(url_for("pallet_new"))
    file_bytes = file.read()
    # Одит (31.08.2026, находка №7): таван на РАЗАРХИВИРАНИЯ размер, ПРЕДИ
    # каквото и да е четене на архива (вкл. помощните проверки по-долу и
    # самия load_workbook) — MAX_CONTENT_LENGTH пази само свития вход.
    try:
        ensure_xlsx_within_limits(file_bytes)
    except XlsxTooLargeError as exc:
        flash(str(exc), "error")
        return redirect(url_for("pallet_new"))
    # Одит (16.08.2026, находка №18): read_only=True пести памет за голям
    # файл (openpyxl не зарежда целия работен лист в паметта наведнъж) —
    # вижте _MAX_IMPORT_DATA_ROWS/_HEADER_SCAN_ROWS по-горе за
    # допълнителните защити (лимит на редовете, търсене на заглавния ред).
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:
        applog.log_exception("routes_pallet_extra: неуспешно четене на качен .xlsx файл")
        flash(_("Файлът не може да бъде прочетен. Уверете се, че е валиден .xlsx файл."), "error")
        return redirect(url_for("pallet_new"))
    if _xlsx_has_merged_cells(file_bytes):
        flash(_("Файлът съдържа обединени клетки — стойности извън първата клетка на "
                "обединен диапазон може да липсват след импорт. Проверете внимателно "
                "резултата по-долу."), "warning")

    # Одит (19.08.2026, находка №29): всички листове, не само първият —
    # виж _parse_sheets по-горе.
    parsed_groups, parse_warnings, _sheet = _parse_sheets(wb, _parse_order_export)
    for w in parse_warnings:
        flash(w, "warning")
    if not parsed_groups:
        # Одит (19.08.2026, находки №40 и №39): към общото съобщение се
        # добавя КЪДЕ се търси заглавният ред (иначе текстът е идентичен с
        # този при напълно грешен файл), а ако файлът съдържа формули без
        # запазени стойности — обяснението, че точно това е причината.
        msg = _("Файлът не съдържа разпознаваеми колони (Order No, Pos, Reference, "
                "Reference Desc, Open Qty) или редове за импорт.") + " " + _header_row_hint()
        if _xlsx_has_formulas(file_bytes):
            msg += " " + _formula_hint()
        flash(msg, "error")
        return redirect(url_for("pallet_new"))

    con = get_db()
    # Одит (19.08.2026, находка №25): вграждат се най-много
    # CLIENT_EMBED_LIMIT клиента — виж appcore.CLIENT_EMBED_LIMIT.
    clients = load_clients(con, CLIENT_EMBED_LIMIT)
    settings = db.get_settings(con)
    ordered = sorted(parsed_groups.items())
    # Одит (16.08.2026, находка №30): груповата структура за шаблона е
    # обща с pallet_bulk_review_restore по-долу (списък от речници с
    # group_no/items и по избор попълнени packaging_type/pallet_type/
    # height/gross) — тук нищо не се възстановява (чисто нов импорт),
    # затова per-card полетата остават празни (шаблонните им подразбиращи
    # се стойности си остават).
    groups_ctx = [{"group_no": g, "items": items} for g, items in ordered]
    # Одит (19.08.2026, находка №39): колоната „Open Qty“ е изцяло празна,
    # а файлът съдържа формули → количествата ги няма не защото файлът е
    # грешен, а защото няма кеширани стойности. Без това предупреждение
    # съобщението по-долу изглежда напълно успешно („Открити са 1 палетни
    # карти (1 реда общо)“) при импорт с нула използваеми количества.
    all_items = [it for _g, items in ordered for it in items]
    if all_items and not any((it.get("qty") or "").strip() for it in all_items) \
            and _xlsx_has_formulas(file_bytes):
        flash(_formula_hint(), "warning")
    flash(_("Открити са %d палетни карти (%d реда общо) от „%s“. Прегледайте и издайте.") %
          (len(ordered), sum(len(v) for _, v in ordered), file.filename), "success")
    return render_template("pallet_bulk_review.html", clients=clients,
                           clients_json=clients_json(clients),
                           clients_total=count_clients(con), s=settings,
                           groups=groups_ctx, shared=None)


@login_required
def pallet_bulk_review_restore(token):
    """Одит (16.08.2026, находка №30): „Назад към формата“ от
    pallet_bulk_preview_view (прегледа на РЪЧНО композирани/Excel-внесени
    палетни карти) водеше винаги към ПРАЗНА форма (url_for('pallet_new'))
    — самите данни СА запазени в _preview_store (kind „bulk_pallet“, виж
    pallet_bulk_preview() по-долу), само че никой досега не ги четеше
    обратно, за разлика от огледалния механизъм за ЕДИНИЧНИТЕ документи
    (appcore.render_preview/routes_documents.edit_document ?restore=,
    v3.61.1). Възстановява прегледа на екрана за bulk преглед от
    съхранените drafts — общите полета (изпращач/клиент/дата/бележки) и
    per-card полетата (вид опаковка/размери/бруто) на всяка карта, плюс
    самите редове."""
    drafts = _get_preview(token, "bulk_pallet")
    if drafts is None:
        flash(_("Прегледът е изтекъл или вече е използван — заредете файла отново."), "warning")
        return redirect(url_for("pallet_new"))
    con = get_db()
    # Одит (19.08.2026, находка №25) — виж pallet_bulk_import по-горе.
    clients = load_clients(con, CLIENT_EMBED_LIMIT)
    settings = db.get_settings(con)
    # Всеки draft носи ЕДНИ И СЪЩИ стойности на споделените полета (виж
    # _collect_bulk_pallet_drafts — `data = dict(shared)` за всеки draft) —
    # първият е представителен за всички.
    shared = drafts[0] if drafts else None
    groups_ctx = [
        {"group_no": idx, "items": d.get("items") or [],
         "packaging_type": d.get("packaging_type", ""), "pallet_type": d.get("pallet_type", ""),
         "height": d.get("height", ""), "gross": d.get("gross", "")}
        for idx, d in enumerate(drafts, start=1)
    ]
    flash(_("Възстановени са незаписаните данни от прегледа — %d палетни карти.")
         % len(groups_ctx), "info")
    return render_template("pallet_bulk_review.html", clients=clients,
                           clients_json=clients_json(clients),
                           clients_total=count_clients(con), s=settings,
                           groups=groups_ctx, shared=shared)


def _collect_bulk_pallet_drafts():
    """Чете подадените от прегледа за bulk импорт полета (общи за
    партидата + поотделно за всеки палет — тип, вид опаковка, бруто,
    височина) и връща списък от речници с данните за всяка карта, БЕЗ да
    ги записва в базата. Ползва се от ТРИ различни екрана: (1) прегледа за
    bulk импорт от справка за поръчки (pallet_bulk_review.html — редовете
    са във формат „orders“), (2) неговия предварителен преглед (без
    запис), и (3) ръчния композитор за няколко карти наведнъж на самата
    pallet_form.html (виж initPalletMultiCard в static/app.js — от заявката
    „палетна карта да съдържа в съдържание на палета Order No, Pos,
    Reference, Reference Desc, Qty“ и той е с колоните на формат „orders“;
    старият формат код/описание/кол./тегло се среща само при редакция на
    вече издадени стари карти).
    „Общ брой“ НЕ е сред тях — изчислява се на момента от items (виж
    appcore.pallet_total_qty), не се пази като отделно подадено поле."""
    shared_fields = ("sender_name", "sender_city", "client_name", "client_address",
                     "client_city", "client_country", "doc_date", "ref_cmr", "notes")
    shared = {k: request.form.get(k, "").strip() for k in shared_fields}
    per_card_fields = ("pallet_type", "packaging_type", "gross", "height")
    group_ids = [g for g in request.form.get("groups", "").split(",") if g.strip()]

    item_check_fields = ("due_date", "order_no", "pos", "project", "reference",
                         "reference_desc", "qty", "unit", "stock", "code", "description", "weight")
    drafts = []
    for g in group_ids:
        raw = request.form.get("items_json_%s" % g, "[]")
        try:
            items = json.loads(raw)
        except ValueError:
            items = []
        items = [it for it in items if isinstance(it, dict) and
                 any((it.get(k) or "").strip() if isinstance(it.get(k), str) else it.get(k)
                     for k in item_check_fields)]
        if not items:
            continue
        data = dict(shared)
        # items_format идва от самия екран — и pallet_bulk_review.html, и
        # ръчният композитор на pallet_form.html вече подават "orders"
        # (виж заявката в docstring-а по-горе). Липсващо поле пада към
        # "orders" — поведението на bulk-review преглед/Excel импорт
        # остава напълно непроменено.
        fmt = request.form.get("items_format_%s" % g, "orders").strip()
        data["items_format"] = fmt or "orders"
        for f in per_card_fields:
            data[f] = request.form.get("%s_%s" % (f, g), "").strip()
        data["items"] = items
        drafts.append(data)

    # Одит (находка С3, среден риск): преди поправката номерът се пресмяташе
    # ТУК, вътре в цикъла, като "%s от %s" % (g, len(group_ids)) — g е
    # номерът на групата ОТ ФАЙЛА (напр. 3, 4, 5 при палети, продължаващи
    # номерацията от предишна пратка), а len(group_ids) броеше ВСИЧКИ
    # подадени групи, включително тези, филтрирани по-горе като напълно
    # празни (if not items: continue). Резултат: файл с групи 3/4/5 даваше
    # карти „3 от 3“/„4 от 3“/„5 от 3“ вместо очакваното „1 от 3“/„2 от
    # 3“/„3 от 3“; а изчистена (празна) група сред тях правеше знаменателя
    # погрешно голям — „1 от 3“ и „3 от 3“ за само 2 РЕАЛНО издадени карти.
    # Номерът се печата на самата карта и отива при клиента, затова трябва
    # да отразява РЕАЛНАТА позиция сред РЕАЛНО издадените карти — броим
    # последователно (1, 2, 3, ...) само върху `drafts` (вече филтрирания
    # списък), а не върху суровия `group_ids`/номера на групата от файла.
    total = len(drafts)
    for idx, data in enumerate(drafts, start=1):
        # Дребни (одит): гол низ без _() — при интерфейс на EN/TR
        # печатната карта показваше „от“ независимо от избрания език.
        # Одит (19.08.2026, находка №13): msgid-ът мина от „%s от %s“ към
        # „{n} от {total}“, за да може СЪЩИЯТ низ да се ползва и от
        # многокартовата форма в static/app.js (palletNoOf) — иначе двата
        # пътя издават различно изписан номер на един и същи документ.
        # Заради Jinja преводът не може да остане с %-плейсхолдъри: там _()
        # е „newstyle“ и сама прилага `%` върху резултата (виж коментара
        # при js_i18n в templates/base.html).
        data["pallet_no"] = _("{n} от {total}").format(n=idx, total=total)
    return drafts


@login_required
def pallet_bulk_preview():
    """Предварителен преглед на всички палетни карти от прегледа за bulk
    импорт, точно както ще изглеждат при печат — БЕЗ да се записват в
    базата и БЕЗ да се изразходват номера. POST → съхрани → пренасочи →
    GET, за да е безопасно презареждане/връщане назад на страницата (виж
    _store_preview в appcore.py)."""
    drafts = _collect_bulk_pallet_drafts()
    if not drafts:
        flash(_("Няма палетни карти за преглед (всички редове са празни)."), "warning")
        return redirect(url_for("pallet_new"))
    token = _store_preview("bulk_pallet", drafts)
    return redirect(url_for("pallet_bulk_preview_view", token=token))


@login_required
def pallet_bulk_preview_view(token):
    drafts = _get_preview(token, "bulk_pallet")
    if drafts is None:
        flash(_("Прегледът е изтекъл или вече е използван — заредете файла отново."), "warning")
        return redirect(url_for("pallet_new"))
    # Одит (16.08.2026, находка №30): token се подава на шаблона, за да
    # може „Назад към формата“ да сочи към pallet_bulk_review_restore
    # (възстановява композираните карти) вместо към празна нова форма.
    return render_template("pallet_bulk_preview.html", drafts=drafts, token=token)


@login_required
def pallet_bulk_issue():
    """Издава наведнъж всички палетни карти от прегледа за импорт от
    справка за поръчки (или от ръчния композитор на pallet_form.html).
    Изпращач/клиент/дата/бележки са общи за цялата партида, но размерите
    и теглото на всеки палет (тип, вид опаковка, бруто, височина) се
    задават и записват отделно за всяка карта."""
    drafts = _collect_bulk_pallet_drafts()
    if not drafts:
        flash(_("Няма палетни карти за издаване (всички редове са празни)."), "warning")
        return redirect(url_for("pallet_new"))

    # Одит (19.08.2026, находки №43 и №7): ГРУПОВОТО издаване не правеше
    # НИТО ЕДНА от проверките, които единичното издаване и редакцията вече
    # правят — отрицателно количество или неразчитаемо число минаваше
    # напълно безшумно през този трети вход, макар че редът се вижда суров
    # на официалната карта, а изчезва от „Общ брой“. Проверката е за цялата
    # партида, с номер на картата, за да е ясно КОЯ карта да се погледне.
    for card_no, draft in enumerate(drafts, start=1):
        rows = negative_item_rows(draft.get("items"))
        if rows:
            flash(_("Внимание (карта №%(card)s): ред(ове) №%(rows)s съдържат "
                    "отрицателна стойност — не участват в „Общ брой“, но се "
                    "виждат на картата.")
                  % {"card": card_no, "rows": ", ".join(str(r) for r in rows)}, "warning")
        bad = unparsable_item_rows(draft.get("items"))
        if bad:
            flash(_("Внимание (карта №%(card)s): ред(ове) №%(rows)s съдържат "
                    "количество/тегло, което не може да бъде разчетено като число "
                    "— не участват в „Общ брой“, но се виждат на картата.")
                  % {"card": card_no, "rows": ", ".join(str(r) for r in bad)}, "warning")

    con = get_db()
    # Одит (находка В14, висок риск): преди поправката всеки save_document
    # тук commit-ваше ОТДЕЛНО (подразбиращото се поведение) — грешка по
    # средата на партида (напр. db.next_number блокирана от друг
    # едновременен процес, или неочакван ValueError в данните на конкретна
    # карта) оставяше ЧАСТ от партидата трайно записана в базата, а
    # останалата — изгубена, без ясен начин операторът да разбере кои
    # номера реално са издадени. Тук цялата партида е ЕДНА транзакция
    # (commit=False на всеки save_document + един общ commit/rollback накрая)
    # — или всички карти от партидата се записват, или НИТО ЕДНА.
    # Одит (находка В14, висок риск): преди поправката всеки save_document
    # тук commit-ваше ОТДЕЛНО (подразбиращото се поведение) — грешка по
    # средата на партида (напр. db.next_number блокирана от друг
    # едновременен процес, или неочакван ValueError в данните на конкретна
    # карта) оставяше ЧАСТ от партидата трайно записана в базата, а
    # останалата — изгубена, без ясен начин операторът да разбере кои
    # номера реално са издадени. Тук цялата партида е ЕДНА транзакция
    # (commit=False на всеки save_document + един общ commit/rollback накрая)
    # — или всички карти от партидата се записват, или НИТО ЕДНА.
    created = []
    try:
        for data in drafts:
            doc_id = save_document(con, "pallet", data, commit=False)
            created.append((data["number"], doc_id))
        con.commit()
    except db.NumberingExhaustedError as exc:
        # Одит (25.08.2026, находка №6): изчерпаната номерация се хващаше от
        # общия `except Exception` по-долу и операторът виждаше само
        # „Възникна грешка… опитайте отново“ — безполезно, защото повторният
        # опит ще удари СЪЩАТА изчерпана номерация. Съобщението на самото
        # изключение казва ТОЧНО причината (първите 1000 поредни номера са
        # заети, вероятно от ръчно въведени номера) — единичното издаване
        # (routes_documents) вече го показва; тук беше „непокритата половина“.
        # Партидата пак е all-or-nothing (rollback), затова добавяме и това.
        con.rollback()
        # Одит (01.09.2026, девети одит, находка №2): въведената партида се
        # ЗАПАЗВА, не се изхвърля. Дотук и двата error клона пращаха
        # оператора на празната /pallet/new — при N карти с до 5000 реда от
        # Excel импорт (плюс ръчните корекции по прегледа) това е
        # най-скъпата загуба на въведени данни в цялата програма, N пъти
        # по-скъпа от единичното издаване, което находка №4 (31.08) вече
        # пази. Механизмът е буквално в същия файл и вече работи:
        # pallet_bulk_preview пази drafts със _store_preview("bulk_pallet"),
        # а pallet_bulk_review_restore ги възстановява в прегледа.
        flash("%s %s" % (str(exc), _("Партидата е отменена изцяло — нищо не бе "
                                     "записано. Въведеното е запазено по-долу.")), "error")
        return redirect(url_for("pallet_bulk_review_restore",
                                token=_store_preview("bulk_pallet", drafts)))
    except Exception:
        con.rollback()
        applog.log_exception(
            "routes_pallet_extra.pallet_bulk_issue: грешка по средата на партида — "
            "цялата партида е върната назад (rollback), нищо не е записано")
        # Одит (01.09.2026, находка №2): същото запазване и тук — този клон
        # хваща и ОЧАКВАНАТА „database is locked“ при мрежов режим (друг
        # компютър пише в момента), при която повторният опит след секунди
        # е нормалният изход — но само ако въведеното още го има.
        flash(_("Възникна грешка при масовото издаване — нищо не бе записано "
               "(партидата е отменена изцяло, за да не останат частично издадени "
               "карти). Въведеното е запазено по-долу — опитайте отново."), "error")
        return redirect(url_for("pallet_bulk_review_restore",
                                token=_store_preview("bulk_pallet", drafts)))

    flash(_("Издадени и запазени %d палетни карти: %s") %
         (len(created), ", ".join(num for num, _ in created)), "success")
    return redirect(url_for("pallet_bulk_result",
                            ids=",".join(str(doc_id) for _, doc_id in created)))


def _fetch_pallet_docs_by_ids(con, ids_param):
    """Общо за pallet_bulk_result/pallet_bulk_print — чете ?ids=1,2,3 и
    връща списък от (doc_row, data) двойки, СЪЩАТА заявка и на двете
    места, за да не се разминат при бъдеща промяна."""
    ids = [int(x) for x in ids_param.split(",") if x.strip().isdigit()]
    docs = []
    for doc_id in ids:
        row = con.execute(
            "SELECT d.*, u.full_name AS author FROM documents d"
            " LEFT JOIN users u ON u.id = d.created_by WHERE d.id = ?",
            (doc_id,),
        ).fetchone()
        if row is not None:
            docs.append((row, safe_json_data(row["data"])))
    return docs


@login_required
def pallet_bulk_result():
    """Преглед на току-що издадените палетни карти преди печат — списък с
    бърз линк към всяка (за проверка поотделно), плюс бутон за печат на
    всички наведнъж (pallet_bulk_print)."""
    ids_param = request.args.get("ids", "")
    docs = _fetch_pallet_docs_by_ids(get_db(), ids_param)
    return render_template("pallet_bulk_result.html", docs=docs, ids_param=ids_param)


@login_required
def pallet_bulk_print():
    """Печат на няколко вече издадени палетни карти наведнъж, в ЕДИН
    документ (browser print обхваща всички едновременно) — заявка:
    „запазят картите да може да се принтират директно всичкия брой
    карти“, вместо да се отваря и печата всяка карта поотделно."""
    ids_param = request.args.get("ids", "")
    docs = _fetch_pallet_docs_by_ids(get_db(), ids_param)
    if not docs:
        flash(_("Няма намерени документи за печат."), "warning")
        return redirect(url_for("pallet_bulk_result", ids=ids_param))
    return render_template("pallet_bulk_print.html", docs=docs, ids_str=ids_param)
