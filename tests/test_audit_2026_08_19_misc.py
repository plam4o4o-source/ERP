# -*- coding: utf-8 -*-
"""Регресионни тестове за ДРЕБНИТЕ находки от одита на 19.08.2026:

* №30 (средна) — нетекстов тип в `pacho_config.json` убиваше приложението
  ПРИ СТАРТ (поправка №45 покри само `network_port`);
* №37 (дребна) — камерният скенер разпознаваше само `code_128`, нямаше
  подсказка при неразпознат код и не отваряше QR кода от собствената ни
  бланка (пълен адрес `…/p/<token>`, не номер за търсене);
* №46 (втора половина) — липсващият уникален индекс беше видим само в лог
  файла, който потребител на .exe никога не отваря;
* №47 (втора половина) — `.exe`, пуснат ДИРЕКТНО от мрежова споделена
  папка, включваше WAL върху SMB (BASE_DIR е на share-а);
* №50 (дребна) — авто-стоп таймерът не се отменяше, когато cloudflared
  умре сам (подвеждащ ред в лога до 2 часа по-късно), а
  `_terminate_process` оставяше зомби процес на POSIX;
* №52 (дребна) — `timeout-minutes: 20` в release.yml вече не е запас;
* информативните — мълчаливо рязане на дълъг текст при Excel износ, липса
  на таван за `_preview_store`, необвързан с потребител preview токен,
  липса на лимит за прикачени файлове и `open(лог, "a")` без try/except.
"""
import io
import os
import re
import subprocess
import sys
import threading
import time

import pytest

from conftest import get_csrf_token, post_with_csrf

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _read(*parts):
    with open(os.path.join(ROOT, *parts), encoding="utf-8") as fh:
        return fh.read()


# ================================================================ №30 — нетекстов тип в pacho_config.json

def _config_with(tmp_path, monkeypatch, payload):
    """Пренасочва config.CONFIG_PATH към временен файл с точно това
    съдържание (json.dump на подадения обект) и връща модула."""
    import json

    import config as appconfig
    path = os.path.join(str(tmp_path), "pacho_config.json")
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh)
    monkeypatch.setattr(appconfig, "CONFIG_PATH", path)
    return appconfig


@pytest.mark.parametrize("payload", [
    {"db_path": 12345},          # забравени кавички около пътя
    {"db_path": ["a"]},          # списък вместо низ
    {"db_path": None},
    # Бележка (25.08.2026): случаите gh_token/gh_owner/gh_branch отпаднаха
    # заедно с премахнатата GitHub синхронизация (тези ключове вече не са в
    # DEFAULTS и не се привеждат). Непознати ключове просто се игнорират.
    {"nepoznat_kljuch": True},   # непознат ключ не бива да чупи старта
])
def test_non_text_values_in_config_do_not_crash_the_start(payload, tmp_path, monkeypatch):
    """Одит (19.08.2026, находка №30): ръчната редакция на този файл е
    ДОКУМЕНТИРАНИЯТ bootstrap за мрежови инсталации. Преди поправката
    `{"db_path": 12345}` даваше `AttributeError: 'int' object has no
    attribute 'strip'` при самия ИМПОРТ на db.py, тоест в компилираната
    `--windowed` версия: тиха смърт без прозорец."""
    appconfig = _config_with(tmp_path, monkeypatch, payload)
    cfg = appconfig.load_config()          # не бива да хвърля
    resolved = appconfig.resolve_db_path(str(tmp_path))  # също (вика се при импорт на db.py)
    assert isinstance(resolved, str) and resolved
    for key, default in appconfig.DEFAULTS.items():
        if isinstance(default, str):
            assert isinstance(cfg[key], str), "%s остана нетекстов: %r" % (key, cfg[key])


def test_numeric_text_value_is_used_as_text_and_logged(tmp_path, monkeypatch, capsys):
    appconfig = _config_with(tmp_path, monkeypatch, {"db_path": 12345})
    cfg = appconfig.load_config()
    assert cfg["db_path"] == "12345"
    assert "db_path" in capsys.readouterr().out


def test_structured_value_falls_back_to_the_default_and_is_logged(tmp_path, monkeypatch, capsys):
    """Списък/речник няма смислен текстов еквивалент — пренебрегва се (с
    предупреждение), вместо да се запише „['a']“ като път до базата."""
    appconfig = _config_with(tmp_path, monkeypatch, {"db_path": ["a"]})
    cfg = appconfig.load_config()
    assert cfg["db_path"] == ""
    assert "db_path" in capsys.readouterr().out
    # без db_path пътят пада към подразбиращия се файл до програмата
    assert appconfig.resolve_db_path(str(tmp_path)).endswith("pacho_logistic.db")


def test_every_text_default_is_covered_by_the_coercion():
    """Пазач: ново текстово поле в DEFAULTS автоматично влиза в проверката
    (списъкът се строи от самите DEFAULTS, не се преписва на ръка)."""
    import config as appconfig
    text_defaults = {k for k, v in appconfig.DEFAULTS.items() if isinstance(v, str)}
    assert set(appconfig._TEXT_KEYS) == text_defaults
    assert "db_path" in appconfig._TEXT_KEYS


# ================================================================ №37 — камерният скенер

def test_camera_scanner_recognizes_qr_and_ean_besides_code_128():
    """Одит (19.08.2026, находка №37): операторът насочва телефона към QR
    кода, отпечатан на НАШАТА собствена бланка — при `formats:
    ["code_128"]` не се случваше нищо, безкрайно и без съобщение."""
    js = _read("static", "app.js")
    assert 'var wanted = ["qr_code", "code_128", "ean_13"];' in js
    assert 'formats: ["code_128"]' not in js, "старият едноформатен списък се е върнал"


def test_camera_scanner_filters_formats_by_browser_support():
    """Конструкторът на BarcodeDetector хвърля при неподдържан формат —
    тогава сканирането отново не би тръгнало ИЗОБЩО."""
    js = _read("static", "app.js")
    assert "window.BarcodeDetector.getSupportedFormats" in js
    assert "new window.BarcodeDetector()" in js, "липсва резервният вариант без formats"


def test_camera_scanner_shows_a_hint_after_ten_seconds_without_a_code():
    js = _read("static", "app.js")
    assert "function startCamHintTimer()" in js
    assert "}, 10000);" in js
    assert 't("camera_no_code_yet"' in js
    # таймерът се чисти при затваряне на модала, за да не изскача после
    assert "if (camHintTimer) { clearTimeout(camHintTimer); camHintTimer = null; }" in js


def test_scanned_public_qr_opens_the_document_instead_of_being_searched():
    """QR кодът от бланката съдържа ПЪЛЕН адрес (`https://хост/p/<token>`).
    Подаден както е в полето за търсене, той не намира нищо. Взимаме само
    токена и го отваряме на ТЕКУЩИЯ origin (виж и находка №21: отпечатаният
    хост може да е ефимерен tryclouflare адрес, който после е чужд)."""
    js = _read("static", "app.js")
    assert "function publicDocPath(value)" in js
    assert 'return m ? "/p/" + m[1] : null;' in js
    assert 'if (path) { stopCamera(); window.location.href = path; return; }' in js
    # чужд QR не се подава като код за търсене
    assert 't("camera_foreign_qr"' in js


def test_new_camera_strings_are_in_the_server_dictionary():
    """Иначе EN/TR потребителят вижда точно новите съобщения на български
    (тест-пазачът в tests/test_translations.py покрива общия случай)."""
    base = _read("templates", "base.html")
    assert "'camera_no_code_yet':" in base
    assert "'camera_foreign_qr':" in base


# ================================================================ №46 — видимо предупреждение за липсващия индекс

def _drop_unique_index(db_module):
    con = db_module.get_db()
    con.execute("DROP INDEX IF EXISTS idx_documents_type_year_number")
    con.commit()
    con.close()


def _insert_duplicate_documents(db_module):
    con = db_module.get_db()
    for barcode in ("DUP1", "DUP2"):
        con.execute(
            "INSERT INTO documents (doc_type, number, year, seq, barcode, data,"
            " created_at) VALUES ('cmr', '0005/2026', 2026, 5, ?, '{}', '2026-01-01')",
            (barcode,))
    con.commit()
    con.close()


def test_settings_page_warns_the_admin_while_the_unique_index_is_missing(admin_client, db_module):
    """Одит (19.08.2026, находка №46, втора половина): дотук единствената
    следа беше ред в лог файла — потребител на .exe никога не го отваря, а
    инсталацията през това време работи БЕЗ защитата срещу два документа с
    един и същ номер."""
    _drop_unique_index(db_module)
    _insert_duplicate_documents(db_module)
    body = admin_client.get("/my-settings").data.decode()
    assert "Защитата срещу повтарящи се номера на документи е ИЗКЛЮЧЕНА" in body
    assert "0005/2026" in body, "не се казва КОЙ номер да се почисти"


def test_no_warning_when_the_index_exists(admin_client, db_module):
    body = admin_client.get("/my-settings").data.decode()
    assert "Защитата срещу повтарящи се номера" not in body


def test_settings_page_retries_creating_the_index_after_the_cleanup(admin_client, db_module):
    """След като администраторът почисти дубликатите, индексът се създава
    при следващото отваряне на страницата — не се чака рестарт."""
    _drop_unique_index(db_module)
    _insert_duplicate_documents(db_module)
    assert "ИЗКЛЮЧЕНА" in admin_client.get("/my-settings").data.decode()

    con = db_module.get_db()
    con.execute("DELETE FROM documents WHERE barcode = 'DUP2'")
    con.commit()
    con.close()

    body = admin_client.get("/my-settings").data.decode()
    assert "Защитата срещу повтарящи се номера" not in body
    con = db_module.get_db()
    assert not db_module.unique_number_index_missing(con)
    con.close()


def test_employee_does_not_see_the_system_warning(employee_client, db_module):
    """Предупреждението е част от „Системни настройки“ — само за админ."""
    _drop_unique_index(db_module)
    _insert_duplicate_documents(db_module)
    body = employee_client.get("/my-settings").data.decode()
    assert "Защитата срещу повтарящи се номера" not in body


# ================================================================ №47 — WAL върху мрежов път

def test_unc_path_is_recognized_as_network(db_module):
    """Одит (19.08.2026, находка №47): `.exe`, пуснат ДИРЕКТНО от
    споделена папка (`\\\\SERVER\\share\\ПачоЛогистик\\pacho.exe`) —
    BASE_DIR е на share-а, значи подразбиращият се DB_PATH също, и старата
    проверка „базата на подразбиращото се място ли е“ включваше WAL точно
    върху SMB."""
    assert db_module._is_network_path(r"\\SERVER\share\pacho_logistic.db")
    assert db_module._is_network_path(r"\\?\UNC\SERVER\share\pacho_logistic.db")


@pytest.mark.skipif(os.name == "nt", reason=(
    "db._is_network_path клонира по os.name, НЕ по мокнатия _MOUNTS_PATH — "
    "на истински Windows функцията винаги минава по Windows клона "
    "(GetDriveTypeW), а /mnt/share... няма буква на диск, значи връща False "
    "веднага, независимо от съдържанието на _MOUNTS_PATH. POSIX клонът е "
    "непроверим, докато процесът реално не тече на POSIX — виж огледалния "
    "test_unc_path_is_recognized_as_network по-горе, който минава на "
    "всяка платформа (UNC low-level проверката е ПРЕДИ os.name клона)."))
def test_mounted_network_filesystem_is_recognized_on_posix(db_module, tmp_path, monkeypatch):
    mounts = tmp_path / "mounts"
    mounts.write_text(
        "/dev/sda1 / ext4 rw 0 0\n"
        "//server/share /mnt/share cifs rw 0 0\n"
        "server:/export /mnt/nfs nfs4 rw 0 0\n", encoding="utf-8")
    monkeypatch.setattr(db_module, "_MOUNTS_PATH", str(mounts))
    assert db_module._is_network_path("/mnt/share/pacho_logistic.db")
    assert db_module._is_network_path("/mnt/nfs/pacho_logistic.db")
    # най-дългата съвпадаща точка на монтиране печели — /mnt/shareX НЕ е /mnt/share
    assert not db_module._is_network_path("/mnt/shareX/pacho_logistic.db")
    assert not db_module._is_network_path("/home/ivan/pacho_logistic.db")


def test_is_network_path_never_raises(db_module, monkeypatch):
    """Евристика за диагностика: при съмнение връща False (губим само
    производителност), но НИКОГА не бива да събаря импорта на db.py."""
    monkeypatch.setattr(db_module, "_MOUNTS_PATH", "/няма/такъв/файл")
    assert db_module._is_network_path("/tmp/pacho_logistic.db") is False
    assert db_module._is_network_path(None) is False


def test_wal_decision_takes_the_network_check_into_account():
    """Пазач срещу връщане на старото условие (само сравнение с
    подразбиращия се път)."""
    source = _read("db.py")
    assert "and not _is_network_path(DB_PATH))" in source


# ================================================================ №50 — авто-стоп таймер и зомби процес

class _FakeProc:
    """Минимален заместител на Popen: изход без адрес и приключил процес."""

    def __init__(self, lines=()):
        self.stdout = io.StringIO("".join(lines))
        self.waited = 0

    def wait(self, timeout=None):
        self.waited += 1
        return 0


def test_auto_stop_timer_is_cancelled_when_cloudflared_dies_on_its_own():
    """Одит (19.08.2026, находка №50): дотук `_consume_output` нулираше
    process/status, но оставяше авто-стоп таймера жив — до 2 часа по-късно
    в лога влизаше „автоматично спиране на отдалечения достъп“ за тунел,
    паднал отдавна. Диагностиката след инцидент ставаше подвеждаща."""
    import remote_tunnel

    proc = _FakeProc()
    fired = []
    timer = threading.Timer(3600, lambda: fired.append(1))
    timer.daemon = True
    timer.start()
    try:
        with remote_tunnel._lock:
            remote_tunnel._state["process"] = proc
            remote_tunnel._state["status"] = "running"
            remote_tunnel._state["auto_stop_timer"] = timer
            gen_before = remote_tunnel._state["generation"]

        remote_tunnel._consume_output(proc)

        assert timer.finished.is_set(), "авто-стоп таймерът остана жив"
        assert remote_tunnel._state["auto_stop_timer"] is None
        assert remote_tunnel._state["generation"] > gen_before, (
            "поколението не е вдигнато — стар таймер още може да „стреля“")
        assert remote_tunnel._state["process"] is None
        assert fired == []
    finally:
        timer.cancel()
        with remote_tunnel._lock:
            remote_tunnel._state.update(process=None, status="stopped", url=None,
                                       error=None, auto_stop_timer=None)


def test_stale_auto_stop_timer_does_nothing_after_the_tunnel_died():
    """Същият сценарий, погледнат от таймера: дори да се събуди, вече не
    съвпада по поколение и не спира нищо."""
    import remote_tunnel

    proc = _FakeProc()
    with remote_tunnel._lock:
        remote_tunnel._state["process"] = proc
        remote_tunnel._state["status"] = "running"
        remote_tunnel._state["auto_stop_timer"] = None
        my_gen = remote_tunnel._state["generation"]
    try:
        remote_tunnel._consume_output(proc)
        remote_tunnel._auto_stop_if_still_current(my_gen)  # не бива да прави нищо
        assert remote_tunnel._state["status"] in ("error", "stopped")
    finally:
        with remote_tunnel._lock:
            remote_tunnel._state.update(process=None, status="stopped", url=None,
                                       error=None, auto_stop_timer=None)


@pytest.mark.skipif(os.name == "nt", reason="зомби процесите са POSIX явление")
def test_terminate_process_reaps_the_child_and_leaves_no_zombie():
    """Одит (19.08.2026, находка №50, свързано): `_terminate_process`
    нямаше `proc.wait()` — терминираният cloudflared оставаше ZOMBIE, а
    точно в клона за разминато поколение (веднага след Popen() в start())
    референцията се задържа, тоест дори GC не го прибираше."""
    import remote_tunnel

    proc = subprocess.Popen([sys.executable, "-c", "import time; time.sleep(30)"])
    try:
        remote_tunnel._terminate_process(proc)
        assert proc.returncode is not None, (
            "процесът не е прибран (wait) — остава зомби в таблицата на процесите")
    finally:
        if proc.returncode is None:  # застраховка, ако тестът се провали
            proc.kill()
            proc.wait()


# ================================================================ №52 — таванът на времето в release.yml

def test_release_workflow_timeout_matches_the_measured_runtime():
    """Одит (19.08.2026, находка №52): 20 мин. вече не е „многократен
    запас“ — измереното време на пълния не-e2e пакет на Linux е ~4.5 мин.
    при 821 теста, а Windows runner-ите са 2–4× по-бавни."""
    yml = _read(".github", "workflows", "release.yml")
    m = re.search(r"^\s*timeout-minutes:\s*(\d+)", yml, re.M)
    assert m, "стъпката с тестовете остана без timeout-minutes"
    assert int(m.group(1)) >= 40, "таванът пак е под измереното време на Windows"
    assert "ИЗМЕРЕНО" in yml, "обосновката трябва да цитира измерено време, не усещане"


# ================================================================ информативна: рязане при Excel износ

def test_overlong_text_is_truncated_visibly_and_logged(capsys):
    """Одит (19.08.2026, информативна находка): openpyxl реже низ над
    32 767 знака МЪЛЧАЛИВО още при присвояването (проверено: 40 000 знака
    влизат като 32 767, без изключение). Такъв низ е достижим през Excel
    импорта и отива при клиента/счетоводството непълен."""
    import routes_documents as rd

    value = rd._xlsx_safe_value("щ" * 40000)
    assert len(value) == rd._XLSX_MAX_CELL_LEN
    assert value.endswith(rd._XLSX_TRUNCATED_MARK)
    assert "отрязан" in capsys.readouterr().out


def test_normal_text_is_untouched_by_the_length_guard():
    import routes_documents as rd

    assert rd._xlsx_safe_value("Обикновен текст") == "Обикновен текст"
    assert rd._xlsx_safe_value(12.5) == 12.5
    assert rd._xlsx_safe_value(None) is None


def test_exported_cell_stays_within_the_xlsx_limit():
    """Целият път през _xlsx_append — точката, през която ВСИЧКИ износи в
    модула добавят ред."""
    import openpyxl

    import routes_documents as rd
    ws = openpyxl.Workbook().active
    rd._xlsx_append(ws, ["ok", "я" * 50000])
    assert ws["A1"].value == "ok"
    assert len(ws["B1"].value) == rd._XLSX_MAX_CELL_LEN
    assert "ОТРЯЗАН" in ws["B1"].value


# ================================================================ информативна: таван и собственик на прегледите

@pytest.fixture
def clean_preview_store():
    import appcore
    with appcore._preview_lock:
        appcore._preview_store.clear()
    yield appcore
    with appcore._preview_lock:
        appcore._preview_store.clear()


def test_preview_store_has_an_upper_bound_besides_the_ttl(clean_preview_store):
    """Одит (19.08.2026, информативна находка): дотук единственият таван
    беше TTL — групов преглед на 5 000 реда се пазеше 30 минути НА ТОКЕН,
    а нов токен се издава при всяко натискане на „Предварителен преглед“."""
    appcore = clean_preview_store
    tokens = [appcore._store_preview("doc", ("cmr", {"n": i}, None, None))
              for i in range(appcore._PREVIEW_MAX_ENTRIES + 5)]
    assert len(appcore._preview_store) <= appcore._PREVIEW_MAX_ENTRIES
    assert appcore._get_preview(tokens[0], "doc") is None, "най-старият трябваше да отпадне"
    assert appcore._get_preview(tokens[-1], "doc") is not None


def test_recently_read_preview_survives_the_eviction(clean_preview_store):
    """LRU, не FIFO: прегледът, който операторът в момента презарежда, не
    бива да отпадне заради по-нови токени."""
    appcore = clean_preview_store
    first = appcore._store_preview("doc", ("cmr", {"n": 0}, None, None))
    for i in range(appcore._PREVIEW_MAX_ENTRIES - 1):
        appcore._store_preview("doc", ("cmr", {"n": i + 1}, None, None))
    assert appcore._get_preview(first, "doc") is not None  # ползване → освежава
    for i in range(5):
        appcore._store_preview("doc", ("cmr", {"n": 100 + i}, None, None))
    assert appcore._get_preview(first, "doc") is not None


def _make_preview(client):
    resp = post_with_csrf(client, "/cmr/preview",
                         {"sender_name": "Тест", "consignee_name": "Клиент"},
                         csrf_source_url="/cmr/new", follow_redirects=False)
    assert resp.status_code == 302
    return resp.headers["Location"].rsplit("/", 1)[-1]


def test_preview_token_is_bound_to_the_user_who_created_it(admin_client, employee_client):
    """Одит (19.08.2026, информативна находка): прегледът съдържа ПЪЛНИТЕ
    данни на още неиздаден документ (получател, цени, бележки). Дотук всеки
    логнат служител, узнал чужд токен (адрес от историята на браузъра на
    общия компютър, изпратена връзка, лог на прокси), можеше да го отвори."""
    token = _make_preview(admin_client)
    assert admin_client.get("/preview/%s" % token).status_code == 200

    stolen = employee_client.get("/preview/%s" % token, follow_redirects=False)
    assert stolen.status_code == 302, "чужд токен не бива да се отваря"
    body = employee_client.get("/preview/%s" % token, follow_redirects=True).data.decode()
    assert "Клиент" not in body


def test_owner_can_still_reload_the_preview_many_times(admin_client):
    """Обвързването не бива да чупи самата причина за съществуването на
    токена — многократно презареждане/връщане назад от собственика."""
    token = _make_preview(admin_client)
    for _ in range(3):
        assert admin_client.get("/preview/%s" % token).status_code == 200
    assert admin_client.get("/cmr/new?restore=%s" % token).status_code == 200


# ================================================================ информативна: лимит на прикачените файлове

_PNG_BYTES = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
    b"\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\x00\x01"
    b"\x00\x00\x05\x00\x01\r\n-\xb4\x00\x00\x00\x00IEND\xaeB`\x82"
)


class _FakeUpload:
    def __init__(self, data, filename="скан.png"):
        self._data = data
        self.filename = filename

    def read(self):
        return self._data


def test_attachment_count_is_capped_per_document(admin_client, db_module):
    """Одит (19.08.2026, информативна находка): дотук нямаше НИКАКЪВ таван
    на броя/общия обем прикачени файлове към един документ — само на
    размера на всеки поотделно."""
    import attachments

    resp = post_with_csrf(admin_client, "/cmr/new",
                         {"sender_name": "Изпращач", "consignee_name": "Получател"},
                         csrf_source_url="/cmr/new", follow_redirects=False)
    doc_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])

    con = db_module.get_db()
    for _ in range(attachments.MAX_FILES):
        attachments.save_attachment(con, doc_id, _FakeUpload(_PNG_BYTES))
    with pytest.raises(ValueError) as exc:
        attachments.save_attachment(con, doc_id, _FakeUpload(_PNG_BYTES))
    con.close()
    assert str(attachments.MAX_FILES) in str(exc.value)
    assert "Изтрийте" in str(exc.value), "съобщението трябва да казва какво да се направи"

    # и през реалния маршрут потребителят получава ЯСНО съобщение, не 500
    token = get_csrf_token(admin_client, "/doc/%d" % doc_id)
    up = admin_client.post("/doc/%d/attachments" % doc_id,
                          data={"csrf_token": token,
                                "attachment": (io.BytesIO(_PNG_BYTES), "още.png")},
                          content_type="multipart/form-data",
                          follow_redirects=True)
    assert up.status_code == 200
    assert "максимумът" in up.data.decode()


def test_attachment_total_volume_is_capped(admin_client, db_module, monkeypatch):
    import attachments

    resp = post_with_csrf(admin_client, "/cmr/new",
                         {"sender_name": "Изпращач", "consignee_name": "Получател"},
                         csrf_source_url="/cmr/new", follow_redirects=False)
    doc_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])

    # смъкнат таван, за да не се пишат десетки MB на диска в теста
    monkeypatch.setattr(attachments, "MAX_TOTAL_SIZE", len(_PNG_BYTES) + 1)
    con = db_module.get_db()
    attachments.save_attachment(con, doc_id, _FakeUpload(_PNG_BYTES))
    with pytest.raises(ValueError) as exc:
        attachments.save_attachment(con, doc_id, _FakeUpload(_PNG_BYTES))
    con.close()
    assert "обем" in str(exc.value)


# ================================================================ информативна: лог файлът при старт

def _load_startup_log_helper():
    """Изпълнява САМО началото на app.py (до пренасочването на stdout) —
    импортът на целия модул би вдигнал приложението и би пипнал реалната
    база."""
    source = _read("app.py").split("if sys.stdout is None")[0]
    namespace = {"__file__": os.path.join(ROOT, "app.py")}
    exec(compile(source, "app.py", "exec"), namespace)  # nosec B102 -- собствен изходен код на проекта
    return namespace["_open_startup_log"]


def test_startup_log_falls_back_to_temp_when_the_folder_is_not_writable(tmp_path):
    """Одит (19.08.2026, информативна находка): `open(_log_path, "a")`
    беше без try/except — инсталация в папка без право на запис (Program
    Files, споделена папка само за четене, файл, заключен от антивирус)
    даваше PermissionError на ниво МОДУЛ: в `--windowed` режим програмата
    умира преди изобщо да е създаден прозорец."""
    import tempfile

    open_startup_log = _load_startup_log_helper()
    handle = open_startup_log(os.path.join(str(tmp_path), "няма-такава-папка"))
    try:
        assert handle is not None, "стартът не бива да зависи от лог файла"
        assert os.path.dirname(os.path.abspath(handle.name)) == \
            os.path.abspath(tempfile.gettempdir())
        handle.write("тест\n")
    finally:
        if handle is not None:
            handle.close()
            try:
                os.remove(handle.name)
            except OSError:
                pass


def test_startup_log_uses_the_program_folder_when_it_is_writable(tmp_path):
    open_startup_log = _load_startup_log_helper()
    handle = open_startup_log(str(tmp_path))
    try:
        # Одит (02.09.2026, находка №11): името вече носи отпечатък на
        # МАШИНАТА (виж app._startup_log_name) — важното тук е ПАПКАТА.
        assert os.path.dirname(handle.name) == str(tmp_path)
        assert os.path.basename(handle.name).startswith("pacho_startup")
    finally:
        handle.close()


def test_startup_log_open_is_guarded_in_the_source():
    """Пазач: самото извикване в app.py минава през защитената функция."""
    source = _read("app.py")
    assert "def _open_startup_log(base_dir):" in source
    assert "_log_file = _open_startup_log(_base_dir_early)" in source
    assert 'open(_log_path, "a"' not in source, "голото отваряне се е върнало"
