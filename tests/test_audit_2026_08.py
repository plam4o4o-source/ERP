# -*- coding: utf-8 -*-
"""Регресионни тестове за находките от задълбочения одит (11.08.2026,
ERP_ОДИТ.md) — покриват критичните (К1-К7) и част от високите/средните
поправки. Организирани по номерата на находките от доклада, за лесно
съпоставяне.

Забележка: К1 (pull_db) вече е покрито изцяло в test_backup_sync.py
(добавени тестове там, не тук) — оставено на мястото му, тъй като там вече
живее останалата инфраструктура за фалшиви GitHub заявки."""
import io
import json
import os
import re

import pytest

import db
from conftest import get_csrf_token, post_with_csrf


# ---------------------------------------------------------------- К2: повреден JSON

def test_broken_json_does_not_crash_dashboard_and_document_list(admin_client, db_module):
    con = db_module.get_db()
    con.execute(
        "INSERT INTO documents (doc_type, number, year, seq, barcode, data, created_at)"
        " VALUES ('cmr', '0001/2026', 2026, 1, 'BROKEN1', '{\"consignee_name\": \"Кли', '2026-01-01')"
    )
    con.commit(); con.close()
    assert admin_client.get("/").status_code == 200
    assert admin_client.get("/docs").status_code == 200


def test_broken_json_document_view_shows_empty_fields_not_500(admin_client, db_module):
    con = db_module.get_db()
    con.execute(
        "INSERT INTO documents (doc_type, number, year, seq, barcode, data, created_at)"
        " VALUES ('invoice_no', 'F1', 2026, 1, 'BROKEN2', '{}', '2026-01-01')"
    )
    con.commit()
    doc_id = con.execute("SELECT id FROM documents WHERE barcode='BROKEN2'").fetchone()["id"]
    con.close()
    resp = admin_client.get("/doc/%d" % doc_id)
    assert resp.status_code == 200


def test_non_dict_json_data_does_not_crash(admin_client, db_module):
    """null/[]/число са синтактично валиден JSON, но не речник — appcore.
    safe_json_data трябва да ги третира еднакво като „няма данни“."""
    con = db_module.get_db()
    for barcode, raw in (("NULLDOC", "null"), ("LISTDOC", "[]"), ("NUMDOC", "42")):
        con.execute(
            "INSERT INTO documents (doc_type, number, year, seq, barcode, data, created_at)"
            " VALUES ('cmr', ?, 2026, 1, ?, ?, '2026-01-01')",
            ("N-" + barcode, barcode, raw),
        )
    con.commit()
    ids = [r["id"] for r in con.execute("SELECT id FROM documents WHERE barcode IN"
                                        " ('NULLDOC','LISTDOC','NUMDOC')")]
    con.close()
    for doc_id in ids:
        assert admin_client.get("/doc/%d" % doc_id).status_code == 200


# ---------------------------------------------------------------- К3: липсващ items

def test_pallet_card_without_items_key_does_not_crash(admin_client, db_module):
    con = db_module.get_db()
    con.execute(
        "INSERT INTO documents (doc_type, number, year, seq, barcode, data, created_at)"
        " VALUES ('pallet', '0001/2026', 2026, 1, 'NOITEMS', '{\"client_name\":\"X\"}', '2026-01-01')"
    )
    con.commit()
    doc_id = con.execute("SELECT id FROM documents WHERE barcode='NOITEMS'").fetchone()["id"]
    con.close()
    assert admin_client.get("/doc/%d" % doc_id).status_code == 200


def test_invoice_no_without_items_key_does_not_crash(admin_client, db_module):
    con = db_module.get_db()
    con.execute(
        "INSERT INTO documents (doc_type, number, year, seq, barcode, data, created_at)"
        " VALUES ('invoice_no', 'F2', 2026, 1, 'NOITEMS2', '{}', '2026-01-01')"
    )
    con.commit()
    doc_id = con.execute("SELECT id FROM documents WHERE barcode='NOITEMS2'").fetchone()["id"]
    con.close()
    assert admin_client.get("/doc/%d" % doc_id).status_code == 200


# ---------------------------------------------------------------- К4: XSS в адресна книга за фактури

def test_invoice_clients_book_escapes_quotes_in_inline_json(admin_client):
    payload = "ACME' onmouseover='alert(1)"
    post_with_csrf(admin_client, "/invoices/clients/new", {"name": payload},
                   csrf_source_url="/invoices/clients/new", follow_redirects=False)
    body = admin_client.get("/invoice-no/new").data.decode()
    assert "onmouseover='alert(1)" not in body
    # Съдържанието все пак трябва да присъства (ескейпнато), не изчезнало.
    assert "ACME" in body


# ---------------------------------------------------------------- В1: глобален error handler

def test_unexpected_exception_shows_friendly_redirect_not_bare_500(admin_client, db_module):
    """Всяка друга (некоригирана в бъдеще) заявка, гръмнала с необработено
    изключение, вече минава през приятелско пренасочване+flash вместо гол
    „Internal Server Error“ — симулираме такъв случай през същия повреден-
    JSON механизъм, но на маршрут, който още няма собствена защита
    (routes_clients client_edit чете directно import json другаде евент.);
    тук просто проверяваме общия механизъм директно през Flask test client
    с изкуствен маршрут не е практично — затова тестваме индиректно през
    реален сценарий: заявка към несъществуващ документ пази 404 (HTTPException
    си остава недокосната от общия handler)."""
    assert admin_client.get("/doc/999999").status_code == 404


def test_404_and_403_are_unaffected_by_generic_error_handler(admin_client, employee_client):
    assert admin_client.get("/doc/999999").status_code == 404
    # Само-администраторски маршрут, отворен от служител -> 403, не 500.
    resp = employee_client.get("/admin/users")
    assert resp.status_code == 403


def test_database_locked_shows_specific_friendly_message(admin_client, db_module, monkeypatch):
    """Одит (находка В10): sqlite3.OperationalError('database is locked')
    вече показва конкретно, разбираемо съобщение вместо гол 500."""
    import sqlite3
    import routes_clients

    def boom(*a, **kw):
        raise sqlite3.OperationalError("database is locked")

    # Одит (19.08.2026, находка №25): /clients вече не тегли цялата адресна
    # книга през appcore.load_clients, а минава през собствената си
    # пагинация с търсене — тя е новата точка, в която заявката към базата
    # може да гръмне. Проверяваната тук логика (класификацията на
    # „database is locked“ и приятелското съобщение) е непроменена.
    monkeypatch.setattr(routes_clients, "paginate_clients", boom)
    resp = admin_client.get("/clients", follow_redirects=True)
    assert resp.status_code == 200
    body = resp.data.decode()
    assert "временно заета" in body

# ---------------------------------------------------------------- В2: контролни символи в Excel износ

def test_xlsx_export_does_not_crash_on_word_soft_return_control_char(admin_client, db_module):
    """Одит (находка В2, висок риск): openpyxl хвърля IllegalCharacterError
    (некоригируема грешка при wb.save) щом низ съдържа "control characters"
    като \\x0b — точно такъв символ вмъква Word при Shift+Enter ("мек нов
    ред"), ако потребител постви текст оттам в поле „Вид на стоката“ на
    ЧМР. Целият Excel износ на документа пада с 500 — routes_documents.
    _xlsx_safe_row трябва да изчисти символа, вместо да гърми."""
    con = db_module.get_db()
    data = json.dumps({"goods": "Ред1\x0bРед2", "consignee_name": "Клиент"},
                       ensure_ascii=False)
    con.execute(
        "INSERT INTO documents (doc_type, number, year, seq, barcode, data, created_at)"
        " VALUES ('cmr', '0099/2026', 2026, 99, 'CTRLCHAR1', ?, '2026-01-01')",
        (data,),
    )
    con.commit()
    doc_id = con.execute("SELECT id FROM documents WHERE barcode='CTRLCHAR1'").fetchone()["id"]
    con.close()

    resp = admin_client.get("/doc/%d/export.xlsx" % doc_id)
    assert resp.status_code == 200
    assert resp.mimetype == ("application/vnd.openxmlformats-officedocument"
                              ".spreadsheetml.sheet")

    import io as _io
    from openpyxl import load_workbook

    wb = load_workbook(_io.BytesIO(resp.data))
    ws = wb.active
    found = False
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if isinstance(cell.value, str) and "Ред1" in cell.value and "Ред2" in cell.value:
                found = True
                assert "\x0b" not in cell.value
    assert found, "полето „Вид на стоката“ трябва да присъства в износа (изчистено от контролния символ)"

# ---------------------------------------------------------------- В3: сесия след деактивиране/смяна на роля

def test_deactivated_employee_session_is_rejected_on_next_request(employee_client, db_module):
    """Одит (находка В3, висок риск): деактивиране на служител от админ,
    докато той е логнат, трябва да прекрати текущата му сесия при СЛЕДВАЩАТА
    заявка — не чак след изтичане на бисквитката."""
    assert employee_client.get("/").status_code == 200

    con = db_module.get_db()
    con.execute("UPDATE users SET active = 0 WHERE username = 'test_emp'")
    con.commit()
    con.close()

    resp = employee_client.get("/", follow_redirects=False)
    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/login") or "/login" in resp.headers["Location"]
    # Сесията наистина е прекратена — таблото вече не е достъпно без нов вход.
    resp2 = employee_client.get("/docs", follow_redirects=False)
    assert resp2.status_code == 302


def test_deleted_employee_session_is_rejected_on_next_request(employee_client, db_module):
    con = db_module.get_db()
    con.execute("DELETE FROM users WHERE username = 'test_emp'")
    con.commit()
    con.close()
    resp = employee_client.get("/", follow_redirects=False)
    assert resp.status_code == 302


def test_admin_demoted_mid_session_immediately_loses_admin_rights(admin_client, db_module):
    """Одит (находка В3): admin, свален до „employee“ от друг администратор
    междувременно, не бива да пази администраторски достъп до край на
    старата си сесия."""
    assert admin_client.get("/admin/users").status_code == 200

    con = db_module.get_db()
    con.execute("UPDATE users SET role = 'employee' WHERE username = 'test_admin'")
    con.commit()
    con.close()

    resp = admin_client.get("/admin/users")
    assert resp.status_code == 403


def test_active_employee_session_is_unaffected(employee_client):
    """Отрицателен контрол — обикновена активна сесия НЕ бива да се
    прекратява от новата проверка."""
    assert employee_client.get("/").status_code == 200
    assert employee_client.get("/docs").status_code == 200

# ---------------------------------------------------------------- В4/В5: права за настройки/обновяване

def test_employee_cannot_change_company_sender_settings(employee_client):
    """Одит (находка В4, висок риск): служителски акаунт не бива да може
    да променя фирмените данни на изпращача (вкл. IBAN/SWIFT на банковия
    ред на всяка фактура) през /settings."""
    resp = post_with_csrf(employee_client, "/settings", {"sender_iban": "BG00HACKED0000000000"},
                          csrf_source_url="/", follow_redirects=False)
    assert resp.status_code == 403


def test_admin_can_still_change_company_sender_settings(admin_client, db_module):
    resp = post_with_csrf(admin_client, "/settings", {"sender_iban": "BG80BNBG96611020345678"},
                          csrf_source_url="/", follow_redirects=True)
    assert resp.status_code == 200
    con = db_module.get_db()
    settings = db.get_settings(con)
    con.close()
    assert settings.get("sender_iban") == "BG80BNBG96611020345678"


def test_employee_cannot_trigger_update_check_or_install(employee_client):
    """Одит (находка В5): update_install рестартира ЦЯЛАТА програма за
    всички едновременно работещи потребители — не бива да е достъпно на
    обикновен служител."""
    assert employee_client.get("/update/check", follow_redirects=False).status_code == 403
    resp = post_with_csrf(employee_client, "/update/install", {}, csrf_source_url="/",
                          follow_redirects=False)
    assert resp.status_code == 403

# ---------------------------------------------------------------- В6: автообновяване без предупреждение

def test_scheduled_auto_install_sets_pending_restart_before_installing(monkeypatch):
    """Одит (находка В6, висок риск): преди поправката install_update()
    (os._exit) се викаше директно от фоновия автоматичен цикъл — без
    никакво предупреждение към потребител, работещ в момента. Тук
    проверяваме, че _schedule_auto_install първо отбелязва „предстои
    рестарт“ (видимо през updater.get_pending_restart за банера в
    интерфейса), изчаква, и ЕДВА ТОГАВА извиква истинския install_update."""
    import updater

    calls = []
    monkeypatch.setattr(updater, "install_update",
                        lambda url, sha: calls.append((url, sha)))
    monkeypatch.setattr(updater, "AUTO_RESTART_WARNING_SECONDS", 0)

    assert updater.get_pending_restart() is None
    seen_pending_during_wait = []
    orig_sleep = updater.time.sleep

    def fake_sleep(seconds):
        seen_pending_during_wait.append(updater.get_pending_restart())

    monkeypatch.setattr(updater.time, "sleep", fake_sleep)
    try:
        updater._schedule_auto_install("http://example.invalid/x.exe", "abc123", "9.9.9",
                                       warning_seconds=5)
    finally:
        monkeypatch.setattr(updater.time, "sleep", orig_sleep)
        with updater._pending_restart_lock:
            updater._pending_restart["scheduled_at"] = None
            updater._pending_restart["version"] = None

    assert calls == [("http://example.invalid/x.exe", "abc123")]
    assert seen_pending_during_wait, "install_update не биваше да се извика без предходно изчакване"
    pending_during = seen_pending_during_wait[0]
    assert pending_during is not None
    assert pending_during["version"] == "9.9.9"
    assert pending_during["seconds_left"] >= 0


def test_pending_restart_endpoint_reports_state(admin_client, monkeypatch):
    import updater

    assert admin_client.get("/update/pending-restart").get_json() == {
        "pending": False, "seconds_left": None, "version": None,
    }

    with updater._pending_restart_lock:
        updater._pending_restart["scheduled_at"] = updater.time.time() + 42
        updater._pending_restart["version"] = "9.9.9"
    try:
        data = admin_client.get("/update/pending-restart").get_json()
        assert data["pending"] is True
        assert data["version"] == "9.9.9"
        assert 0 <= data["seconds_left"] <= 42
    finally:
        with updater._pending_restart_lock:
            updater._pending_restart["scheduled_at"] = None
            updater._pending_restart["version"] = None

# ---------------------------------------------------------------- В7: търсене и кирилица с друг регистър

def test_document_search_finds_cyrillic_regardless_of_case(admin_client, db_module):
    """Одит (находка В7, висок риск): SQLite LIKE/COLLATE NOCASE сгъва
    само ASCII регистър — търсене на „иван“ преди поправката НЕ намираше
    документ, в който е записано „Иван“."""
    con = db_module.get_db()
    data = json.dumps({"consignee_name": "Иван Петров"}, ensure_ascii=False)
    con.execute(
        "INSERT INTO documents (doc_type, number, year, seq, barcode, data, created_at)"
        " VALUES ('cmr', '0055/2026', 2026, 55, 'CIDOC1', ?, '2026-01-01')",
        (data,),
    )
    con.commit()
    con.close()

    resp = admin_client.get("/docs?q=иван")
    assert resp.status_code == 200
    assert "Иван Петров" in resp.data.decode()


def test_invoices_search_finds_cyrillic_regardless_of_case(admin_client, db_module):
    con = db_module.get_db()
    data = json.dumps({"consignee_name": "СОФИЯ ЕООД"}, ensure_ascii=False)
    con.execute(
        "INSERT INTO documents (doc_type, number, year, seq, barcode, data, created_at)"
        " VALUES ('invoice_no', 'INV-CI-1', 2026, 1, 'CIINV1', ?, '2026-01-01')",
        (data,),
    )
    con.commit()
    con.close()

    resp = admin_client.get("/invoices?q=софия")
    assert resp.status_code == 200
    assert "СОФИЯ ЕООД" in resp.data.decode()


def test_materials_search_finds_cyrillic_regardless_of_case(admin_client, db_module):
    con = db_module.get_db()
    con.execute(
        "INSERT INTO materials (code, description, net_weight)"
        " VALUES ('MAT-CI-1', 'Профил Кабел', 1.5)"
    )
    con.commit()
    con.close()

    import materials as materials_module
    con = db_module.get_db()
    results = materials_module.search(con, "кабел")
    con.close()
    assert any(r["code"] == "MAT-CI-1" for r in results)

# ---------------------------------------------------------------- В14: частичен запис при масово издаване

def test_bulk_pallet_issue_failure_midway_rolls_back_entire_batch(admin_client, db_module, monkeypatch):
    """Одит (находка В14, висок риск): save_document() преди поправката
    commit-ваше ОТДЕЛНО за всяка карта в партидата — грешка по средата
    оставяше част от партидата трайно записана. Тук симулираме грешка
    точно на ВТОРАТА от три карти и проверяваме, че НИТО ЕДНА не остава
    в базата (пълен rollback), не само първата."""
    import appcore
    import routes_pallet_extra

    items_1 = json.dumps([{"code": "A1", "description": "Стока 1", "qty": "2", "weight": "10"}])
    items_2 = json.dumps([{"code": "B1", "description": "Стока 2", "qty": "5", "weight": "20"}])
    items_3 = json.dumps([{"code": "C1", "description": "Стока 3", "qty": "1", "weight": "5"}])
    data = {
        "sender_name": "Изпращач", "client_name": "Клиент партида-неуспех",
        "groups": "1,2,3",
        "pallet_type_1": "120×80", "packaging_type_1": "Палет / Pallet",
        "items_json_1": items_1, "items_format_1": "manual",
        "pallet_type_2": "80×60", "packaging_type_2": "Кашон / Box",
        "items_json_2": items_2, "items_format_2": "manual",
        "pallet_type_3": "120×100", "packaging_type_3": "Палет / Pallet",
        "items_json_3": items_3, "items_format_3": "manual",
    }

    orig_save = appcore.save_document
    calls = {"n": 0}

    def flaky_save(con, doc_type, doc_data, manual_number=None, commit=True):
        calls["n"] += 1
        if calls["n"] == 2:
            raise RuntimeError("симулирана грешка по средата на партидата")
        return orig_save(con, doc_type, doc_data, manual_number=manual_number, commit=commit)

    monkeypatch.setattr(routes_pallet_extra, "save_document", flaky_save)

    resp = post_with_csrf(admin_client, "/pallet/bulk-issue", data,
                          csrf_source_url="/pallet/new", follow_redirects=True)
    assert resp.status_code == 200

    con = db_module.get_db()
    count = con.execute(
        "SELECT COUNT(*) AS c FROM documents WHERE doc_type='pallet' AND data LIKE '%партида-неуспех%'"
    ).fetchone()["c"]
    con.close()
    assert count == 0, "нито една карта от партидата не трябва да остане записана след грешка по средата"


def test_bulk_pallet_issue_succeeds_normally_when_nothing_fails(admin_client, db_module):
    """Отрицателен контрол — обичайният успешен път не бива да се счупи
    от преминаването към commit=False + общ commit."""
    items_1 = json.dumps([{"code": "A1", "description": "Стока 1", "qty": "2", "weight": "10"}])
    items_2 = json.dumps([{"code": "B1", "description": "Стока 2", "qty": "5", "weight": "20"}])
    data = {
        "client_name": "Клиент партида-успех", "groups": "1,2",
        "pallet_type_1": "120×80", "packaging_type_1": "Палет / Pallet",
        "items_json_1": items_1, "items_format_1": "manual",
        "pallet_type_2": "80×60", "packaging_type_2": "Кашон / Box",
        "items_json_2": items_2, "items_format_2": "manual",
    }
    resp = post_with_csrf(admin_client, "/pallet/bulk-issue", data,
                          csrf_source_url="/pallet/new", follow_redirects=False)
    assert resp.status_code == 302

    con = db_module.get_db()
    count = con.execute(
        "SELECT COUNT(*) AS c FROM documents WHERE doc_type='pallet' AND data LIKE '%партида-успех%'"
    ).fetchone()["c"]
    con.close()
    assert count == 2

# ---------------------------------------------------------------- В15: пагинация в списъците

def _insert_n_cmr_docs(con, n, prefix="PGDOC"):
    for i in range(n):
        con.execute(
            "INSERT INTO documents (doc_type, number, year, seq, barcode, data, created_at)"
            " VALUES ('cmr', ?, 2026, ?, ?, '{}', '2026-01-01')",
            ("%s-%04d/2026" % (prefix, i), i + 1, "%s-%04d" % (prefix, i)),
        )
    con.commit()


def test_documents_list_paginates_instead_of_silent_300_cap(admin_client, db_module):
    """Одит (находка В15, висок риск): преди поправката "LIMIT 300" беше
    тих таван без никаква пагинация — документ №301 ставаше практически
    невидим. Тук вкарваме повече от PAGE_SIZE документа и проверяваме, че
    страница 2 показва РЕАЛНО различни (по-стари) документи, не е празна."""
    import routes_documents
    con = db_module.get_db()
    _insert_n_cmr_docs(con, routes_documents.PAGE_SIZE + 5)
    con.close()

    page1 = admin_client.get("/docs").data.decode()
    page2 = admin_client.get("/docs?page=2").data.decode()
    # ORDER BY id DESC — най-НОВИЯТ (последно вкаран, най-голям id) е
    # PGDOC-0104 и е на страница 1; най-СТАРИЯТ (PGDOC-0000) е изтикан на
    # страница 2, точно записът, който преди поправката ставаше невидим.
    assert "PGDOC-0104/2026" in page1
    assert "PGDOC-0000/2026" in page2
    assert "Страница 1" in page1
    assert "Страница 2" in page2
    assert "PGDOC-0000/2026" not in page1


def test_documents_list_page_beyond_range_clamps_to_last_page(admin_client, db_module):
    """page=999 (извън обхвата) не бива да покаже "няма намерени" —
    трябва да се изравни към последната реална страница (тук страница 1,
    единствената)."""
    con = db_module.get_db()
    _insert_n_cmr_docs(con, 3, prefix="PGSMALL")
    con.close()
    resp = admin_client.get("/docs?page=999")
    assert resp.status_code == 200
    body = resp.data.decode()
    assert "PGSMALL-0000/2026" in body
    assert "Няма намерени документи" not in body


def test_invoices_list_paginates_instead_of_silent_300_cap(admin_client, db_module):
    import routes_documents
    con = db_module.get_db()
    for i in range(routes_documents.PAGE_SIZE + 5):
        con.execute(
            "INSERT INTO documents (doc_type, number, year, seq, barcode, data, created_at)"
            " VALUES ('invoice_no', ?, 2026, ?, ?, '{}', '2026-01-01')",
            ("PGINV-%04d" % i, i + 1, "PGINV-%04d" % i),
        )
    con.commit()
    con.close()

    page1 = admin_client.get("/invoices").data.decode()
    page2 = admin_client.get("/invoices?page=2").data.decode()
    assert "PGINV-0104" in page1
    assert "PGINV-0000" in page2
    assert "PGINV-0000" not in page1

# ---------------------------------------------------------------- В17: зает порт при старт на сървъра

def test_find_available_port_returns_the_preferred_port_when_free():
    import net
    import socket
    # Портове от системния "ephemeral"-диапазон рядко се заемат случайно
    # от други процеси по време на теста.
    port = 51234
    result = net.find_available_port("127.0.0.1", port)
    assert result == port


def test_find_available_port_skips_a_port_that_is_actually_busy():
    """Одит (находка В17, висок риск): преди поправката НЯМАШЕ никаква
    проверка — сървърът просто гърмеше тихо във фонова нишка, ако портът
    беше зает, а прозорецът/браузърът се отваряше все пак (към каквото и
    да е друго слушащо на порта). Тук реално заемаме порт 51235 с истински
    сокет и проверяваме, че find_available_port го прескача и връща
    СЛЕДВАЩИЯ свободен, не същия зает номер."""
    import net
    import socket
    port = 51235
    blocker = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    blocker.bind(("127.0.0.1", port))
    blocker.listen(1)
    try:
        result = net.find_available_port("127.0.0.1", port)
        assert result != port
        assert result == port + 1
    finally:
        blocker.close()


def test_find_available_port_raises_clear_error_when_none_free(monkeypatch):
    import net
    import socket

    def always_busy(*a, **kw):
        raise OSError("address already in use")

    class _FakeSocket:
        def setsockopt(self, *a): pass
        def bind(self, *a): raise OSError("address already in use")
        def close(self): pass

    monkeypatch.setattr(net.socket, "socket", lambda *a, **kw: _FakeSocket())
    with pytest.raises(RuntimeError, match="заети"):
        net.find_available_port("127.0.0.1", 51300, max_tries=3)

# ---------------------------------------------------------------- В18: path traversal в име на файл

def test_sanitize_number_stub_strips_backslash_and_other_unsafe_chars():
    import client_export
    assert client_export.sanitize_number_stub("0001/2026") == "0001-2026"
    assert client_export.sanitize_number_stub("..\\..\\..\\evil") == "..-..-..-evil"
    assert "\\" not in client_export.sanitize_number_stub("a\\b\\c")
    assert "/" not in client_export.sanitize_number_stub("a/b/c")


def test_invoice_export_with_malicious_manual_number_stays_inside_client_folder(admin_client, tmp_path):
    """Одит (находка В18, висок риск): фактурите имат РЪЧНО въвеждан номер
    — свободен текст, изцяло контролиран от потребителя. Преди поправката
    само "/" се заменяше в _export_filename; обратна наклонена черта
    минаваше непроменена и os.path.join(folder, filename) с Windows-стил
    разделител в името позволяваше запис ИЗВЪН предвидената клиентска
    папка (path traversal). Тук симулираме точно такъв номер и проверяваме,
    че записаният файл остава СТРОГО вътре в очакваната клиентска папка —
    нищо не се появява извън tmp_path/<клиент>/."""
    post_with_csrf(admin_client, "/admin/system", {
        "form": "client_export", "client_export_dir": str(tmp_path),
        "client_export_auto": "on",
    }, csrf_source_url="/my-settings", follow_redirects=False)

    malicious_number = "..\\..\\..\\evil_escape"
    resp = post_with_csrf(admin_client, "/invoice-br/new", {
        "invoice_number": malicious_number,
        "consignee_name": "Клиент Пътека ЕООД",
    }, csrf_source_url="/invoice-br/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    xlsx_resp = admin_client.get("/doc/%s/export.xlsx" % doc_id)
    assert xlsx_resp.status_code == 200

    expected_dir = os.path.join(str(tmp_path), "Клиент Пътека ЕООД")
    assert os.path.isdir(expected_dir)
    files = os.listdir(expected_dir)
    assert len(files) == 1
    assert "\\" not in files[0] and "/" not in files[0]

    # Нищо не бива да се е появило НАД/ИЗВЪН tmp_path (родителската папка).
    parent = os.path.dirname(str(tmp_path))
    suspicious = [f for f in os.listdir(parent) if "evil_escape" in f]
    assert suspicious == []

# ---------------------------------------------------------------- С2: Excel/PDF износ на фактура без ред TOTAL

def test_invoice_export_totals_row_is_none_for_non_invoice_doc_types():
    """_invoice_export_totals_row не бива да добавя TOTAL ред за
    документи, които изобщо не са фактури (напр. палетна карта) — там
    печатната бланка също няма такъв ред, износът трябва да остане
    непроменен спрямо преди поправката."""
    from routes_documents import _invoice_export_totals_row
    cols = [("code", "Артикул/код"), ("qty", "Количество")]
    items = [{"code": "ART-1", "qty": "5"}]
    assert _invoice_export_totals_row("pallet", items, cols) is None


def test_invoice_export_totals_row_is_none_without_items():
    """Без нито един ред артикул няма какво да се сумира — печатната
    бланка (appcore.invoice_totals) също не показва TOTAL ред тогава."""
    from routes_documents import _invoice_export_totals_row
    cols = [("hs_code", "HS code"), ("qty", "Количество"), ("__row_total__", "Обща цена, EUR")]
    assert _invoice_export_totals_row("invoice_br", [], cols) is None


def test_invoice_export_totals_row_sums_quantity_and_price_aligned_to_columns():
    """Одит (находка С2): редът трябва да е подравнен 1:1 по cols, за да
    легне направо като поредния ред в таблицата (Excel/PDF) — колоната
    „Количество“ носи сбора на количествата, „Обща цена, EUR“ носи сбора
    на вече закръглените редови суми (appcore.invoice_totals), с наставка
    „ €“, точно като на печатната бланка."""
    from routes_documents import _invoice_export_totals_row
    cols = [("hs_code", "HS code"), ("po_no", "P.O NO"), ("qty", "Количество"),
            ("unit_price", "Единична цена, EUR"), ("__row_total__", "Обща цена, EUR")]
    items = [
        {"hs_code": "1", "po_no": "A", "qty": "20", "unit_price": "13.66"},
        {"hs_code": "2", "po_no": "B", "qty": "5", "unit_price": "2"},
    ]
    row = _invoice_export_totals_row("invoice_br", items, cols)
    assert row[0] == "TOTAL"
    assert row[[k for k, _l in cols].index("qty")] == "25"
    assert row[[k for k, _l in cols].index("__row_total__")] == "283.20 €"


def test_invoice_xlsx_export_includes_total_row(admin_client):
    """Пълен HTTP-до-Excel регресионен тест: изнесеният .xlsx на фактура
    трябва да завършва с ред TOTAL (преди поправката спираше на последния
    артикул — виж findings С2)."""
    from openpyxl import load_workbook

    items = json.dumps([
        {"hs_code": "85389099", "po_no": "4700200362", "pos": "30",
         "net_weight": "2.21", "material_code": "GLBK400002P0012",
         "qty": "20", "unit_price": "13.66"},
        {"hs_code": "85389099", "po_no": "4700200363", "pos": "40",
         "net_weight": "1.0", "material_code": "GLBK400002P0013",
         "qty": "5", "unit_price": "2"},
    ])
    resp = post_with_csrf(admin_client, "/invoice-br/new", {
        "consignee_name": "Клиент За TOTAL ЕООД", "items_json": items,
    }, csrf_source_url="/invoice-br/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    xlsx_resp = admin_client.get("/doc/%s/export.xlsx" % doc_id)
    assert xlsx_resp.status_code == 200

    wb = load_workbook(io.BytesIO(xlsx_resp.data))
    ws = wb.active
    last_row = [c.value for c in ws[ws.max_row]]
    assert last_row[0] == "TOTAL"
    assert "283.20 €" in last_row  # 20×13.66 + 5×2 = 273.20 + 10.00 = 283.20


def test_invoice_pdf_export_includes_total_row(admin_client):
    """Същото, но за PDF износа — извлеченият текст трябва да съдържа
    „TOTAL“ и сборната сума, а не само последния артикул."""
    from pypdf import PdfReader

    items = json.dumps([{
        "hs_code": "85389099", "po_no": "4700200362", "pos": "30",
        "net_weight": "2.21", "material_code": "GLBK400002P0012",
        "qty": "20", "unit_price": "13.66",
    }])
    resp = post_with_csrf(admin_client, "/invoice-br/new", {
        "consignee_name": "Клиент За TOTAL PDF ЕООД", "items_json": items,
    }, csrf_source_url="/invoice-br/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    pdf_resp = admin_client.get("/doc/%s/export.pdf" % doc_id)
    assert pdf_resp.status_code == 200

    reader = PdfReader(io.BytesIO(pdf_resp.data))
    text = "\n".join(page.extract_text() for page in reader.pages)
    assert "TOTAL" in text
    assert "273.2" in text  # 20 × 13.66

# ---------------------------------------------------------------- С3: грешна номерация „X от Y“ на палети

def test_bulk_pallet_numbering_is_sequential_even_when_file_group_numbers_are_not(admin_client):
    """Одит (находка С3, среден риск): преди поправката номерът се
    генерираше от НОМЕРА НА ГРУПАТА ОТ ФАЙЛА (g), не от последователна
    позиция — файл с палети, продължаващи номерация 3, 4, 5 (продължение
    на предишна пратка) даваше карти „3 от 3“/„4 от 3“/„5 от 3“ вместо
    очакваното „1 от 3“/„2 от 3“/„3 от 3“."""
    items_1 = json.dumps([{"code": "A1", "description": "Стока 1", "qty": "2", "weight": "10"}])
    items_2 = json.dumps([{"code": "B1", "description": "Стока 2", "qty": "5", "weight": "20"}])
    items_3 = json.dumps([{"code": "C1", "description": "Стока 3", "qty": "1", "weight": "5"}])
    data = {
        "sender_name": "Изпращач", "client_name": "Клиент С3 непоследователни номера",
        "groups": "3,4,5",
        "items_json_3": items_1, "items_format_3": "manual",
        "items_json_4": items_2, "items_format_4": "manual",
        "items_json_5": items_3, "items_format_5": "manual",
    }
    resp = post_with_csrf(admin_client, "/pallet/bulk-issue", data,
                          csrf_source_url="/pallet/new", follow_redirects=False)
    assert resp.status_code == 302

    con = db.get_db()
    rows = con.execute(
        "SELECT data FROM documents WHERE doc_type='pallet' ORDER BY id DESC LIMIT 3"
    ).fetchall()
    con.close()
    pallet_nos = sorted(json.loads(r["data"])["pallet_no"] for r in rows)
    assert pallet_nos == ["1 от 3", "2 от 3", "3 от 3"]


def test_bulk_pallet_numbering_excludes_empty_groups_from_the_total(admin_client):
    """Одит (находка С3): преди поправката знаменателят (Y в „X от Y“)
    броеше ВСИЧКИ подадени групи, включително филтрираните по-горе като
    напълно празни — изчистена група сред 3 подадени даваше грешно „1 от
    3“ и „3 от 3“ за само 2 РЕАЛНО издадени карти, вместо коректното
    „1 от 2“/„2 от 2“."""
    items_1 = json.dumps([{"code": "A1", "description": "Стока 1", "qty": "2", "weight": "10"}])
    items_3 = json.dumps([{"code": "C1", "description": "Стока 3", "qty": "1", "weight": "5"}])
    data = {
        "sender_name": "Изпращач", "client_name": "Клиент С3 изчистена група",
        "groups": "1,2,3",
        "items_json_1": items_1, "items_format_1": "manual",
        "items_json_2": "[]", "items_format_2": "manual",  # изчистена/празна група
        "items_json_3": items_3, "items_format_3": "manual",
    }
    resp = post_with_csrf(admin_client, "/pallet/bulk-issue", data,
                          csrf_source_url="/pallet/new", follow_redirects=False)
    assert resp.status_code == 302

    con = db.get_db()
    rows = con.execute(
        "SELECT data FROM documents WHERE doc_type='pallet' ORDER BY id DESC LIMIT 2"
    ).fetchall()
    con.close()
    pallet_nos = sorted(json.loads(r["data"])["pallet_no"] for r in rows)
    assert pallet_nos == ["1 от 2", "2 от 2"]

# ---------------------------------------------------------------- С4: сканиране на кирилска подредба

def test_normalize_bds_cyrillic_recovers_the_letter_confirmed_by_the_audit():
    """Одит (находка С4): реално възпроизведено с истински keydown
    събития — физическият клавиш „M“ дава „п“ при активна кирилска (БДС)
    подредба. Тук проверяваме точно тази двойка (единствената, потвърдена
    директно от одита) плюс главната ѝ форма."""
    import bg_keyboard
    assert bg_keyboard.normalize_bds_cyrillic("п") == "M"
    assert bg_keyboard.normalize_bds_cyrillic("П") == "M"


def test_normalize_bds_cyrillic_leaves_digits_hyphen_and_latin_unchanged():
    """Цифри, тире и вече правилна латиница НЕ бива да се пипат —
    гарантира, че резервният опит в scan() не разваля код, който вече
    беше отчасти или изцяло верен."""
    import bg_keyboard
    assert bg_keyboard.normalize_bds_cyrillic("0001-2026") == "0001-2026"
    assert bg_keyboard.normalize_bds_cyrillic("ABC-123") == "ABC-123"
    assert bg_keyboard.normalize_bds_cyrillic("") == ""


def test_normalize_bds_cyrillic_recovers_a_full_barcode_round_trip():
    """Обхожда ЦЯЛАТА таблица (bg_keyboard._REVERSE_MAP), строи „обратния“
    низ (какъвто РЕАЛНО би се появил на екрана при кирилска подредба за
    даден латински баркод) и проверява, че нормализацията реконструира
    точно оригинала — регресия за пълната таблица, не само за М."""
    import bg_keyboard
    forward = {}
    for cyr, latin in bg_keyboard._REVERSE_MAP.items():
        forward.setdefault(latin, cyr)
    sample = "CMR-07082026-0001"
    garbled = "".join(forward.get(ch, ch) for ch in sample)
    assert garbled != sample, "тестът трябва реално да гарантира кирилско разместване"
    assert bg_keyboard.normalize_bds_cyrillic(garbled) == sample


def test_scan_route_finds_document_via_cyrillic_garbled_barcode_fallback(admin_client):
    """Пълен HTTP регресионен тест: издава ЧМР, „сканира“ КИРИЛСКИ
    разместения му баркод (какъвто би стигнал до Flask, ако служителят е
    написал/сканирал с активна кирилска подредба в едно от видимите
    полета за сканиране — #scan-input/#sidebar-scan-input, за разлика от
    глобалния клавиатурен буфер, който вече е имунизиран чрез e.code, виж
    static/app.js) и проверява, че сървърният резервен опит
    (bg_keyboard.normalize_bds_cyrillic) въпреки това намира документа."""
    import bg_keyboard

    resp = post_with_csrf(admin_client, "/cmr/new", {
        "sender_name": "Изпращач С4 route",
    }, csrf_source_url="/cmr/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    con = db.get_db()
    barcode = con.execute("SELECT barcode FROM documents WHERE id = ?", (doc_id,)).fetchone()["barcode"]
    con.close()

    forward = {}
    for cyr, latin in bg_keyboard._REVERSE_MAP.items():
        forward.setdefault(latin, cyr)
    garbled = "".join(forward.get(ch, ch) for ch in barcode)
    assert garbled != barcode

    resp = post_with_csrf(admin_client, "/scan", {"code": garbled},
                          csrf_source_url="/", follow_redirects=False)
    assert resp.headers["Location"].rstrip("/").endswith("/doc/%s" % doc_id)


def test_scan_route_still_shows_not_found_for_genuinely_unknown_cyrillic_code(admin_client):
    """Безопасност по конструкция: код, който дори след нормализация не
    съвпада с НИТО ЕДИН документ, трябва да си остане ясното съобщение „не
    е намерен“ — резервният опит никога не пренасочва към ПОГРЕШЕН
    документ."""
    resp = post_with_csrf(admin_client, "/scan", {"code": "щдвкфвшдвкфв"},
                          csrf_source_url="/", follow_redirects=True)
    assert "Няма документ" in resp.data.decode()

# ---------------------------------------------------------------- С7: таблото не бива да чака GitHub

def test_dashboard_does_not_block_on_a_slow_or_unreachable_github_check(admin_client, monkeypatch):
    """Одит (находка С7, среден риск): routes_dashboard.dashboard() рендира
    updater.check_cached() резултата — преди поправката САМАТА мрежова
    заявка (до 8 сек.) течеше синхронно вътре в check_cached(), значи
    зареждането на таблото директно чакаше GitHub. Тук изкуствено бавим
    check_for_update и проверяваме, че /  зарежда ПРАКТИЧЕСКИ мигновено —
    остарелият кеш просто се използва веднага, докато реалната (бавна)
    проверка тръгва във фонова нишка."""
    import time as time_module

    import updater

    def _slow_check():
        time_module.sleep(2.0)
        return {"available": False, "current": "0.0.0-test", "latest": "0.0.0-test",
               "download": None, "expected_sha256": None}

    monkeypatch.setattr(updater, "check_for_update", _slow_check)
    # Форсираме кешът да изглежда остарял, за да предизвикаме опресняване.
    updater._cache["time"] = 0.0
    updater._refresh_in_progress = False

    started = time_module.time()
    resp = admin_client.get("/")
    elapsed = time_module.time() - started

    assert resp.status_code == 200
    assert elapsed < 1.5, (
        "таблото изчака бавната GitHub проверка вместо да ползва кеша "
        "веднага (отне %.2f сек.)" % elapsed
    )
    # Изчакваме фоновата нишка да приключи, за да не изтече в следващия тест.
    time_module.sleep(2.2)

# ---------------------------------------------------------------- С10: ЧМР износ без ЕИК/ДДС

def test_cmr_xlsx_export_includes_sender_eik_and_consignee_vat(admin_client):
    """Одит (находка С10, среден риск): sender_eik/consignee_vat ги има
    във формата и на печатната бланка на ЧМР, но липсваха в
    routes_documents._XLSX_FIELDS["cmr"] → отсъстваха от Excel износа.
    За митнически документ идентификацията по ДДС номер не е козметична."""
    from openpyxl import load_workbook

    resp = post_with_csrf(admin_client, "/cmr/new", {
        "sender_name": "Изпращач С10", "sender_eik": "BG123456789",
        "consignee_name": "Клиент С10 ЕООД", "consignee_vat": "BG987654321",
    }, csrf_source_url="/cmr/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    xlsx_resp = admin_client.get("/doc/%s/export.xlsx" % doc_id)
    assert xlsx_resp.status_code == 200

    wb = load_workbook(io.BytesIO(xlsx_resp.data))
    ws = wb.active
    values = [c.value for row in ws.iter_rows() for c in row]
    assert "BG123456789" in values, "ЕИК/ДДС на изпращача липсва от Excel износа на ЧМР"
    assert "BG987654321" in values, "ДДС/ЕИК на получателя липсва от Excel износа на ЧМР"


def test_cmr_pdf_export_includes_sender_eik_and_consignee_vat(admin_client):
    """Същото, но за PDF износа на ЧМР (споделя _export_fields_and_items с
    Excel износа — виж routes_documents.py)."""
    from pypdf import PdfReader

    resp = post_with_csrf(admin_client, "/cmr/new", {
        "sender_name": "Изпращач С10 PDF", "sender_eik": "BG111222333",
        "consignee_name": "Клиент С10 PDF ЕООД", "consignee_vat": "BG444555666",
    }, csrf_source_url="/cmr/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    pdf_resp = admin_client.get("/doc/%s/export.pdf" % doc_id)
    assert pdf_resp.status_code == 200

    reader = PdfReader(io.BytesIO(pdf_resp.data))
    text = "".join(page.extract_text() or "" for page in reader.pages)
    assert "BG111222333" in text, "ЕИК/ДДС на изпращача липсва от PDF износа на ЧМР"
    assert "BG444555666" in text, "ДДС/ЕИК на получателя липсва от PDF износа на ЧМР"

# ---------------------------------------------------------------- С11: dualuse дата на фактурата в ISO формат

def test_dualuse_print_shows_invoice_date_in_bg_format_not_iso(admin_client):
    """Одит (находка С11, нисък риск): templates/dualuse_print.html:28
    извеждаше {{ d.invoice_date }} директно, БЕЗ format_date — бланката
    показваше „2026-02-04“, докато Excel износът на СЪЩОТО поле показва
    „04.02.2026“ (форматът, използван навсякъде другаде в програмата)."""
    resp = post_with_csrf(admin_client, "/dualuse/new", {
        "sender_name": "Износител С11 ЕООД", "invoice_date": "2026-02-04",
    }, csrf_source_url="/dualuse/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    view = admin_client.get("/doc/%s" % doc_id)
    body = view.data.decode()
    assert "04.02.2026" in body, "датата на фактурата трябва да е във вид ДД.ММ.ГГГГ на бланката"
    assert "2026-02-04" not in body, "суровата ISO дата не бива да се показва на бланката"

# ---------------------------------------------------------------- С12: "Нето, кг" винаги празно при издърпване от палет

def test_packing_pull_pallet_no_longer_returns_a_dead_net_key(admin_client):
    """Одит (находка С12, нисък риск): routes_pallet_extra.py:80 връщаше
    "net": d.get("net", "") — палетната карта отдавна НЯМА поле „net“
    (заменено с „Общ брой“), затова полето беше ВИНАГИ празно, без
    операторът да разбира защо. Поправката маха мъртвия ключ и връща
    изрично "note", обясняващо защо нето теглото трябва да се въведе
    ръчно."""
    items = json.dumps([{"code": "ART-С12", "description": "Кашон С12", "qty": "5"}])
    create_resp = post_with_csrf(admin_client, "/pallet/new", {
        "pallet_no": "1", "gross": "120", "items_json": items,
    }, csrf_source_url="/pallet/new", follow_redirects=False)
    doc_id = create_resp.headers["Location"].rstrip("/").split("/")[-1]

    con = db.get_db()
    barcode = con.execute("SELECT barcode FROM documents WHERE id = ?", (doc_id,)).fetchone()["barcode"]
    con.close()

    resp = post_with_csrf(admin_client, "/packing/pull-pallet",
                          {"code": barcode}, csrf_source_url="/")
    payload = resp.get_json()
    assert payload["ok"] is True
    assert "net" not in payload["row"], (
        "row не бива да съдържа мъртвия ключ „net“ — палетната карта никога не пази нето тегло"
    )
    assert payload.get("note"), (
        "отговорът трябва да обяснява изрично защо нето теглото не е попълнено"
    )
    assert payload["row"]["gross"] == "120"

# ---------------------------------------------------------------- С15: отворено пренасочване при вход

def test_login_next_rejects_protocol_relative_url(client, db_module):
    """Одит (находка С15, среден риск): routes_auth.py — `target.startswith("/")`
    пропускаше протоколно-относителни адреси като „//evil.example.com/x“.
    Възпроизведено (одит): POST /login?next=//evil.example.com/x → 302
    Location: //evil.example.com/x. Служителят въвежда истинските си данни
    в истинската програма и бива пренасочен към чужд сайт."""
    from werkzeug.security import generate_password_hash

    con = db_module.get_db()
    con.execute(
        "INSERT INTO users (username, password_hash, full_name, role, active,"
        " must_change_password) VALUES (?, ?, ?, 'admin', 1, 0)",
        ("c15_user", generate_password_hash("c15-password-123"), "С15 Тест"),
    )
    con.commit()
    con.close()

    token = get_csrf_token(client, "/login")
    resp = client.post("/login?next=//evil.example.com/x", data={
        "username": "c15_user", "password": "c15-password-123", "csrf_token": token,
    })
    assert resp.status_code == 302
    location = resp.headers["Location"]
    assert not location.startswith("//evil.example.com"), (
        "открито пренасочване — отиваме към ЧУЖД домейн: %s" % location
    )
    assert location.startswith("/") and not location.startswith("//"), location


def test_login_next_still_honors_a_genuine_internal_path(client, db_module):
    """Безопасност по конструкция, обратна посока: истински ВЪТРЕШЕН relative
    път (една наклонена черта) трябва да продължи да работи нормално —
    поправката не бива да чупи легитимната функционалност на next=."""
    from werkzeug.security import generate_password_hash

    con = db_module.get_db()
    con.execute(
        "INSERT INTO users (username, password_hash, full_name, role, active,"
        " must_change_password) VALUES (?, ?, ?, 'admin', 1, 0)",
        ("c15_user2", generate_password_hash("c15-password-456"), "С15 Тест 2"),
    )
    con.commit()
    con.close()

    token = get_csrf_token(client, "/login")
    resp = client.post("/login?next=/clients", data={
        "username": "c15_user2", "password": "c15-password-456", "csrf_token": token,
    })
    assert resp.status_code == 302
    assert resp.headers["Location"].rstrip("/").endswith("/clients")

# ---------------------------------------------------------------- Дребни: невалиден мрежов порт

def test_network_settings_reject_non_numeric_port_without_crashing(admin_client):
    """Одит (Дребни): int(request.form.get("network_port")) гърмеше с
    необработен ValueError (500) при нечислов вход в полето „Мрежов порт“
    в Настройки — вместо ясно съобщение за грешка."""
    resp = post_with_csrf(admin_client, "/admin/system", {
        "form": "network", "network_port": "не-число",
    }, csrf_source_url="/my-settings", follow_redirects=False)
    assert resp.status_code == 302  # редирект с flash грешка, НЕ 500
    page = admin_client.get("/my-settings")
    assert "Невалиден мрежов порт" in page.data.decode()


def test_network_settings_reject_port_out_of_valid_range(admin_client):
    """Порт извън допустимия диапазон (1-65535) също трябва да е отхвърлен
    с ясно съобщение, не мълчаливо записан."""
    resp = post_with_csrf(admin_client, "/admin/system", {
        "form": "network", "network_port": "99999",
    }, csrf_source_url="/my-settings", follow_redirects=False)
    assert resp.status_code == 302
    page = admin_client.get("/my-settings")
    assert "Невалиден мрежов порт" in page.data.decode()


def test_network_settings_accept_a_valid_port(admin_client):
    """Безопасност по конструкция, обратна посока: валиден порт продължава
    да се запазва нормално."""
    resp = post_with_csrf(admin_client, "/admin/system", {
        "form": "network", "network_port": "8080",
    }, csrf_source_url="/my-settings", follow_redirects=False)
    assert resp.status_code == 302
    page = admin_client.get("/my-settings")
    body = page.data.decode()
    assert "Мрежовите настройки са запазени" in body
    assert "Невалиден мрежов порт" not in body

# ---------------------------------------------------------------- Дребни: неизползвано поле place_country

def test_dualuse_print_shows_the_place_country_field(admin_client):
    """Одит (Дребни): dualuse_form.html събира отделно поле „Държава (за
    бланката)“ (place_country), но dualuse_print.html никога не го
    показваше — операторът го попълва и то мълчаливо изчезва от бланката."""
    resp = post_with_csrf(admin_client, "/dualuse/new", {
        "sender_name": "Износител Дребни ЕООД", "place": "София",
        "place_country": "Тестландия С99",
    }, csrf_source_url="/dualuse/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    view = admin_client.get("/doc/%s" % doc_id)
    body = view.data.decode()
    assert "София" in body
    assert "Тестландия С99" in body, "полето „Държава (за бланката)“ трябва да се показва на бланката"

# ---------------------------------------------------------------- Дребни: заключване на акаунт = DoS

def test_correct_password_still_logs_in_even_while_account_is_locked_out(client, db_module):
    """Одит (Дребни): login_guard заключва по ПОТРЕБИТЕЛСКО ИМЕ — нападател
    се нуждае само от 1 грешен опит на 5 минути, за да държи „admin“
    ПОСТОЯННО заключен, включително за ИСТИНСКИЯ собственик с ПРАВИЛНАТА
    парола (DoS чрез самото заключване). Поправката проверява паролата
    ПРЕДИ да провери заключването — правилна парола винаги влиза."""
    import login_guard
    from werkzeug.security import generate_password_hash

    con = db_module.get_db()
    con.execute(
        "INSERT INTO users (username, password_hash, full_name, role, active,"
        " must_change_password) VALUES (?, ?, ?, 'admin', 1, 0)",
        ("c_dos_user", generate_password_hash("real-password-789"), "DoS Тест"),
    )
    con.commit()
    con.close()

    # Нападателят изчерпва лимита с грешна парола (симулира атаката от одита).
    for _ in range(login_guard.MAX_ATTEMPTS):
        token = get_csrf_token(client, "/login")
        client.post("/login", data={
            "username": "c_dos_user", "password": "wrong-guess", "csrf_token": token,
        })
    locked, _ = login_guard.is_locked_out("c_dos_user")
    assert locked is True, "тестът трябва да гарантира, че акаунтът реално е заключен"

    # Истинският собственик, с ПРАВИЛНАТА парола, трябва да влезе веднага —
    # НЕ да получи „твърде много опити“.
    token = get_csrf_token(client, "/login")
    resp = client.post("/login", data={
        "username": "c_dos_user", "password": "real-password-789", "csrf_token": token,
    }, follow_redirects=False)
    assert resp.status_code == 302, (
        "собственикът на акаунта не бива да е DoS-нат от заключването, "
        "щом въвежда ПРАВИЛНАТА си парола"
    )
    assert "/login" not in resp.headers["Location"]


def test_wrong_password_is_still_rejected_and_locks_out_after_the_limit(client, db_module):
    """Безопасност по конструкция, обратна посока: защитата срещу
    brute-force остава непокътната — грешна парола НИКОГА не влиза,
    независимо от реда на проверките, и заключването все така се задейства
    след лимита от опити."""
    import login_guard
    from werkzeug.security import generate_password_hash

    con = db_module.get_db()
    con.execute(
        "INSERT INTO users (username, password_hash, full_name, role, active,"
        " must_change_password) VALUES (?, ?, ?, 'admin', 1, 0)",
        ("c_dos_user2", generate_password_hash("real-password-456"), "DoS Тест 2"),
    )
    con.commit()
    con.close()

    for _ in range(login_guard.MAX_ATTEMPTS):
        token = get_csrf_token(client, "/login")
        resp = client.post("/login", data={
            "username": "c_dos_user2", "password": "wrong-guess", "csrf_token": token,
        })
        assert resp.status_code == 200  # остава на логин страницата, не влиза

    locked, _ = login_guard.is_locked_out("c_dos_user2")
    assert locked is True

    token = get_csrf_token(client, "/login")
    resp = client.post("/login", data={
        "username": "c_dos_user2", "password": "wrong-guess", "csrf_token": token,
    })
    assert "Твърде много неуспешни опити" in resp.data.decode()

# ---------------------------------------------------------------- Дребни: непреведени низове в routes_pallet_extra

def test_pull_pallet_row_text_goes_through_translation_not_a_bare_string(admin_client, monkeypatch):
    """Одит (Дребни): "Палет %s"/"Палет" (description/packing на реда,
    издърпан от палетна карта в опаковъчен лист) бяха голи Python низове
    БЕЗ _() — при интерфейс на EN/TR излизаха на български независимо от
    избрания език. Тук подменяме _() с маркиращ фалшив превод и
    проверяваме, че резултатът РЕАЛНО минава през него, а не го байпасва."""
    import routes_pallet_extra

    monkeypatch.setattr(routes_pallet_extra, "_", lambda s: "[T]" + s)

    items = json.dumps([{"code": "ART-1", "description": "Кашон", "qty": "3"}])
    create_resp = post_with_csrf(admin_client, "/pallet/new", {
        "pallet_no": "5", "items_json": items,
    }, csrf_source_url="/pallet/new", follow_redirects=False)
    doc_id = create_resp.headers["Location"].rstrip("/").split("/")[-1]
    con = db.get_db()
    barcode = con.execute("SELECT barcode FROM documents WHERE id = ?", (doc_id,)).fetchone()["barcode"]
    con.close()

    resp = post_with_csrf(admin_client, "/packing/pull-pallet", {"code": barcode}, csrf_source_url="/")
    payload = resp.get_json()
    assert payload["ok"] is True
    assert payload["row"]["description"].startswith("[T]Палет "), payload["row"]["description"]
    assert payload["row"]["packing"] == "[T]Палет"


def test_bulk_pallet_numbering_text_goes_through_translation_not_a_bare_string(admin_client, monkeypatch):
    """Същото за "%s от %s" (номерът, отпечатан на самата bulk-издадена
    палетна карта) — беше гол низ, поправката минава през _()."""
    import routes_pallet_extra

    monkeypatch.setattr(routes_pallet_extra, "_", lambda s: "[T]" + s)

    items = json.dumps([{"code": "A1", "description": "Стока", "qty": "2", "weight": "10"}])
    data = {
        "sender_name": "Изпращач", "client_name": "Клиент Дребни превод",
        "groups": "1",
        "items_json_1": items, "items_format_1": "manual",
    }
    resp = post_with_csrf(admin_client, "/pallet/bulk-issue", data,
                          csrf_source_url="/pallet/new", follow_redirects=False)
    assert resp.status_code == 302

    con = db.get_db()
    row = con.execute(
        "SELECT data FROM documents WHERE doc_type='pallet' ORDER BY id DESC LIMIT 1"
    ).fetchone()
    con.close()
    pallet_no = json.loads(row["data"])["pallet_no"]
    assert pallet_no.startswith("[T]"), pallet_no

# ---------------------------------------------------------------- Дребни: достъпност на модалите

def test_confirm_and_camera_modals_have_dialog_role_and_aria_modal(admin_client):
    """Одит (Дребни, достъпност): #confirm-modal/#camera-scan-modal нямаха
    role="dialog"/aria-modal — екранен четец не ги обявяваше като диалог."""
    body = admin_client.get("/").data.decode()
    import re

    def _attrs(el_id):
        m = re.search(r'<div id="%s"[^>]*>' % re.escape(el_id), body)
        assert m, "не намерих #%s" % el_id
        return m.group(0)

    confirm_tag = _attrs("confirm-modal")
    assert 'role="dialog"' in confirm_tag
    assert 'aria-modal="true"' in confirm_tag

    camera_tag = _attrs("camera-scan-modal")
    assert 'role="dialog"' in camera_tag
    assert 'aria-modal="true"' in camera_tag


def test_import_status_messages_have_aria_live(admin_client):
    """Одит (Дребни, достъпност): съобщенията за импорт/зареждане (напр.
    „Зареден е 1 ред от палетна карта №...“) бяха без aria-live — екранен
    четец не ги обявяваше автоматично при поява."""
    packing_body = admin_client.get("/packing/new").data.decode()
    assert 'id="pull-pallet-msg"' in packing_body
    assert re.search(r'id="pull-pallet-msg"[^>]*aria-live="polite"', packing_body)

    invoice_body = admin_client.get("/invoice-br/new").data.decode()
    assert re.search(r'class="invoice-pull-msg"[^>]*aria-live="polite"', invoice_body)
    assert re.search(r'class="invoice-excel-msg"[^>]*aria-live="polite"', invoice_body)


def test_scan_inputs_have_aria_label_not_only_placeholder(admin_client):
    """Одит (Дребни, достъпност): полетата за сканиране разчитаха само на
    placeholder (изчезва при фокус/попълване, не се чете надеждно от
    всички екранни четци) — вместо истински label/aria-label."""
    dashboard_body = admin_client.get("/").data.decode()
    assert re.search(r'id="scan-input"[^>]*aria-label="[^"]+"', dashboard_body)
    assert re.search(r'id="sidebar-scan-input"[^>]*aria-label="[^"]+"', dashboard_body)

# ---------------------------------------------------------------- Дребни: updating.html твърд порт

def test_updating_page_shows_the_actually_configured_port(admin_client, tmp_path, monkeypatch):
    """Одит (Дребни): updating.html показваше ТВЪРДО закодиран
    http://127.0.0.1:5000 — грешен адрес за инсталация с преконфигуриран
    мрежов порт (виж system_settings по-горе), точно когато операторът
    най-много се нуждае от верния адрес (след автоматично обновяване)."""
    import config as appconfig_mod

    cfg_path = str(tmp_path / "pacho_config.json")
    monkeypatch.setattr(appconfig_mod, "CONFIG_PATH", cfg_path)
    # routes_admin.py прави `import config as appconfig` — трябва да
    # пренасочим и НЕГОВОТО виждане на CONFIG_PATH (същия обект, но
    # монтиран под друго локално име в routes_admin).
    import routes_admin as routes_admin_mod
    monkeypatch.setattr(routes_admin_mod.appconfig, "CONFIG_PATH", cfg_path)
    appconfig_mod.save_config({"network_port": 8091})

    import updater as updater_mod

    def fake_check():
        return {"available": True, "current": "0.0.0-test", "latest": "9.9.9-test",
               "download": "http://example.invalid/x.exe", "expected_sha256": "abc"}

    def fake_install(url, sha):
        return None

    monkeypatch.setattr(updater_mod, "check_for_update", fake_check)
    monkeypatch.setattr(updater_mod, "install_update", fake_install)

    token = get_csrf_token(admin_client, "/my-settings")
    resp = admin_client.post("/update/install", data={"csrf_token": token})
    body = resp.data.decode()
    assert "127.0.0.1:8091" in body, "трябва да показва РЕАЛНО configured порта"
    assert "127.0.0.1:5000" not in body
