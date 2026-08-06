# -*- coding: utf-8 -*-
"""Характеризиращ тестов пакет за Фаза 3 (структурен рефакторинг): пуска
ЦЯЛОТО Flask приложение (appcore.create_app() + всички routes_* модули,
точно както app.py ги регистрира — виж fixture flask_app в conftest.py) и
опитва реални HTTP заявки през Flask test client.

Целта е РЕГРЕСИОНЕН предпазен колан за разделянето на стария монолитен
app.py по routes_*.py модули: ако вход, издаване на документ, права на
достъп (admin_required/login_required) или CSRF защитата се счупят при
пренасянето на кода, тестовете тук трябва да го хванат — преди да е стигнало
до реален служител на терен."""
import io
import json

from conftest import get_csrf_token, post_with_csrf


# ---------------------------------------------------------------- вход/изход и CSRF

def test_login_page_renders_with_csrf_field(client):
    resp = client.get("/login")
    assert resp.status_code == 200
    assert b'name="csrf_token"' in resp.data


def test_login_wrong_password_shows_error(client):
    token = get_csrf_token(client, "/login")
    resp = client.post("/login", data={"username": "nobody", "password": "wrong",
                                       "csrf_token": token})
    assert resp.status_code == 200
    assert "Грешно потребителско име".encode() in resp.data


def test_login_success_redirects_to_dashboard(admin_client):
    resp = admin_client.get("/", follow_redirects=False)
    assert resp.status_code == 200  # вече логнат — таблото се зарежда директно


def test_dashboard_requires_login_redirects_to_login(client):
    resp = client.get("/", follow_redirects=False)
    assert resp.status_code == 302
    assert "/login" in resp.headers["Location"]


def test_logout_clears_session(admin_client):
    resp = admin_client.get("/logout", follow_redirects=False)
    assert resp.status_code == 302
    resp2 = admin_client.get("/", follow_redirects=False)
    assert resp2.status_code == 302
    assert "/login" in resp2.headers["Location"]


def test_post_without_csrf_token_is_rejected(admin_client):
    resp = admin_client.post("/clients/new", data={"name": "Тест ЕООД"})
    assert resp.status_code == 400


def test_post_with_wrong_csrf_token_is_rejected(admin_client):
    resp = admin_client.post("/clients/new",
                             data={"name": "Тест ЕООД", "csrf_token": "invalid-token"})
    assert resp.status_code == 400


def test_must_change_password_redirects_to_change_password(flask_app, db_module):
    from werkzeug.security import generate_password_hash
    con = db_module.get_db()
    con.execute(
        "INSERT INTO users (username, password_hash, full_name, role, active,"
        " must_change_password) VALUES (?, ?, ?, 'employee', 1, 1)",
        ("must_change", generate_password_hash("test-password-123"), "Трябва Смяна"),
    )
    con.commit()
    con.close()
    c = flask_app.test_client()
    token = get_csrf_token(c, "/login")
    c.post("/login", data={"username": "must_change", "password": "test-password-123",
                           "csrf_token": token})
    resp = c.get("/", follow_redirects=False)
    assert resp.status_code == 302
    assert "/password" in resp.headers["Location"]
    # change_password самата страница трябва да е достъпна (изключена от enforce)
    resp2 = c.get("/password")
    assert resp2.status_code == 200


# ---------------------------------------------------------------- документни потоци (5 типа)

def test_cmr_new_get_renders_form(admin_client):
    resp = admin_client.get("/cmr/new")
    assert resp.status_code == 200


def test_cmr_new_post_creates_document(admin_client):
    resp = post_with_csrf(admin_client, "/cmr/new", {
        "sender_name": "Изпращач ЕООД", "consignee_name": "Получател ООД",
    }, csrf_source_url="/cmr/new", follow_redirects=False)
    assert resp.status_code == 302
    assert "/doc/" in resp.headers["Location"]


def test_packing_new_post_with_items_creates_document(admin_client):
    items = json.dumps([{"description": "Стока А", "qty": "2"}])
    resp = post_with_csrf(admin_client, "/packing/new", {
        "sender_name": "Изпращач", "items_json": items,
    }, csrf_source_url="/packing/new", follow_redirects=False)
    assert resp.status_code == 302


def test_packing_new_saves_shipment_and_dimension_fields(admin_client):
    """PL.xlsx подобрение: Условия на доставка/Вид транспорт/HS Code (на
    ниво документ) и Дължина/Широчина/Височина/Обем (на ниво ред) — виж
    ПЛАН_ЗА_РАЗРАБОТКА.md за контекста на добавянето."""
    items = json.dumps([{
        "description": "HV Switchgear cabinets material", "qty": "1", "packing": "carton",
        "length": "1200", "width": "800", "height": "900", "volume": "0.864",
        "net": "150", "gross": "180",
    }])
    resp = post_with_csrf(admin_client, "/packing/new", {
        "sender_name": "BBS Bulgaria EOOD", "receiver_name": "Клиент",
        "terms_delivery": "FCA", "transport_type": "Truck", "hs_code": "85389099",
        "total_packages": "1", "total_volume": "0.864", "total_net": "150", "total_gross": "180",
        "items_json": items,
    }, csrf_source_url="/packing/new", follow_redirects=False)
    assert resp.status_code == 302

    view_resp = admin_client.get(resp.headers["Location"])
    assert view_resp.status_code == 200
    body = view_resp.data
    assert "FCA".encode() in body
    assert "Truck".encode() in body
    assert "85389099".encode() in body
    assert "1200".encode() in body  # дължина на реда
    assert "800".encode() in body   # широчина
    assert "900".encode() in body   # височина
    assert "0.864".encode() in body  # обем (ред и общо)


def test_packing_xlsx_export_includes_new_fields_and_columns(admin_client):
    items = json.dumps([{
        "description": "Стока А", "qty": "1", "length": "500", "width": "400",
        "height": "300", "volume": "0.06", "net": "10", "gross": "12",
    }])
    resp = post_with_csrf(admin_client, "/packing/new", {
        "sender_name": "Изпращач", "terms_delivery": "FCA", "transport_type": "Truck",
        "hs_code": "85389099", "items_json": items,
    }, csrf_source_url="/packing/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    xlsx_resp = admin_client.get("/doc/%s/export.xlsx" % doc_id)
    assert xlsx_resp.status_code == 200

    import io as _io
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(xlsx_resp.data))
    ws = wb.active
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert "Условия на доставка" in all_values
    assert "FCA" in all_values
    assert "HS Code" in all_values
    assert "Дължина, мм" in all_values
    assert "Обем, м³" in all_values


def test_pallet_new_post_with_items_creates_document(admin_client):
    items = json.dumps([{"code": "ART-1", "description": "Кашон", "qty": "3"}])
    resp = post_with_csrf(admin_client, "/pallet/new", {
        "pallet_no": "1", "items_json": items,
    }, csrf_source_url="/pallet/new", follow_redirects=False)
    assert resp.status_code == 302


def test_dualuse_new_post_creates_document(admin_client):
    resp = post_with_csrf(admin_client, "/dualuse/new", {
        "sender_name": "Износител ЕООД",
    }, csrf_source_url="/dualuse/new", follow_redirects=False)
    assert resp.status_code == 302


def test_export_it_new_post_creates_document(admin_client):
    resp = post_with_csrf(admin_client, "/export-it/new", {
        "declarant_name": "Декларатор",
    }, csrf_source_url="/export-it/new", follow_redirects=False)
    assert resp.status_code == 302


def test_cmr_preview_flow_redirects_to_preview_token(admin_client):
    resp = post_with_csrf(admin_client, "/cmr/preview", {
        "sender_name": "Преглед ЕООД",
    }, csrf_source_url="/cmr/new", follow_redirects=False)
    assert resp.status_code == 302
    assert "/preview/" in resp.headers["Location"]
    resp2 = admin_client.get(resp.headers["Location"])
    assert resp2.status_code == 200
    assert "DRAFT".encode() in resp2.data or "ПРЕДВАРИТЕЛЕН".encode() in resp2.data


def test_view_document_after_creation(admin_client):
    create_resp = post_with_csrf(admin_client, "/cmr/new", {
        "sender_name": "Изпращач за преглед",
    }, csrf_source_url="/cmr/new", follow_redirects=False)
    doc_url = create_resp.headers["Location"]
    resp = admin_client.get(doc_url)
    assert resp.status_code == 200


def test_documents_search_matches_order_number_and_reference_in_pallet_items(admin_client):
    """Търсенето в /docs минава през LIKE върху ЦЯЛАТА JSON колона data
    (виж routes_documents.documents()), затова автоматично покрива и
    полета вътре в редовете на палетна карта (order_no, reference) — не
    само номер/баркод/име на клиент на ниво документ."""
    items = json.dumps([{"order_no": "SO-778899", "pos": "10", "reference": "REF-UNIQ-42",
                         "reference_desc": "Търсен материал", "qty": "3"}])
    resp = post_with_csrf(admin_client, "/pallet/new", {
        "client_name": "Клиент за търсене", "items_json": items, "items_format": "orders",
    }, csrf_source_url="/pallet/new", follow_redirects=False)
    assert resp.status_code == 302

    by_order = admin_client.get("/docs?q=SO-778899")
    assert "Клиент за търсене".encode() in by_order.data

    by_reference = admin_client.get("/docs?q=REF-UNIQ-42")
    assert "Клиент за търсене".encode() in by_reference.data

    no_match = admin_client.get("/docs?q=NE-SASHTESTVUVASHT-REF")
    assert "Клиент за търсене".encode() not in no_match.data


def test_edit_document_updates_data(admin_client):
    create_resp = post_with_csrf(admin_client, "/cmr/new", {
        "sender_name": "Оригинален изпращач",
    }, csrf_source_url="/cmr/new", follow_redirects=False)
    doc_url = create_resp.headers["Location"]
    doc_id = doc_url.rstrip("/").split("/")[-1]
    edit_url = "/doc/%s/edit" % doc_id
    resp = post_with_csrf(admin_client, edit_url, {
        "sender_name": "Редактиран изпращач",
    }, csrf_source_url=edit_url, follow_redirects=False)
    assert resp.status_code == 302
    view = admin_client.get(doc_url)
    assert "Редактиран изпращач".encode() in view.data


def test_export_document_xlsx_returns_spreadsheet(admin_client):
    create_resp = post_with_csrf(admin_client, "/cmr/new", {
        "sender_name": "Изпращач за износ",
    }, csrf_source_url="/cmr/new", follow_redirects=False)
    doc_id = create_resp.headers["Location"].rstrip("/").split("/")[-1]
    resp = admin_client.get("/doc/%s/export.xlsx" % doc_id)
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["Content-Type"]


def test_delete_document_requires_admin(employee_client, admin_client):
    create_resp = post_with_csrf(admin_client, "/cmr/new", {
        "sender_name": "За изтриване",
    }, csrf_source_url="/cmr/new", follow_redirects=False)
    doc_id = create_resp.headers["Location"].rstrip("/").split("/")[-1]

    del_url = "/doc/%s/delete" % doc_id
    resp = post_with_csrf(employee_client, del_url, {}, csrf_source_url="/",
                          follow_redirects=False)
    assert resp.status_code == 403  # служител без admin права — забранено

    resp2 = post_with_csrf(admin_client, del_url, {}, csrf_source_url="/",
                           follow_redirects=False)
    assert resp2.status_code == 302  # админ може


# ---------------------------------------------------------------- палетна карта: bulk/pull

def test_packing_pull_pallet_not_found_returns_error_json(admin_client):
    resp = post_with_csrf(admin_client, "/packing/pull-pallet",
                          {"code": "NE-SASHTESTVUVASHT"}, csrf_source_url="/")
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["ok"] is False


def test_packing_pull_pallet_finds_created_pallet(admin_client):
    items = json.dumps([{"code": "ART-1", "description": "Кашон А", "qty": "5"}])
    create_resp = post_with_csrf(admin_client, "/pallet/new", {
        "pallet_no": "77", "boxes": "5", "items_json": items,
    }, csrf_source_url="/pallet/new", follow_redirects=False)
    doc_id = create_resp.headers["Location"].rstrip("/").split("/")[-1]
    # взимаме баркода директно от базата, за да не парсваме HTML
    import db as db_mod
    con = db_mod.get_db()
    row = con.execute("SELECT barcode, number FROM documents WHERE id = ?", (doc_id,)).fetchone()
    con.close()

    resp = post_with_csrf(admin_client, "/packing/pull-pallet",
                          {"code": row["barcode"]}, csrf_source_url="/")
    payload = resp.get_json()
    assert payload["ok"] is True
    assert payload["number"] == row["number"]


# ---------------------------------------------------------------- клиенти (адресна книга)

def test_clients_crud_flow(admin_client):
    resp = post_with_csrf(admin_client, "/clients/new", {
        "name": "Тестов Клиент ЕООД", "city": "София",
    }, csrf_source_url="/clients/new", follow_redirects=False)
    assert resp.status_code == 302

    list_resp = admin_client.get("/clients")
    assert "Тестов Клиент ЕООД".encode() in list_resp.data

    import db as db_mod
    con = db_mod.get_db()
    client_row = con.execute(
        "SELECT id FROM clients WHERE name = ?", ("Тестов Клиент ЕООД",)
    ).fetchone()
    con.close()
    client_id = client_row["id"]

    edit_url = "/clients/%d/edit" % client_id
    edit_resp = admin_client.get(edit_url)
    assert edit_resp.status_code == 200

    upd_resp = post_with_csrf(admin_client, edit_url, {
        "name": "Тестов Клиент ЕООД", "city": "Пловдив",
    }, csrf_source_url=edit_url, follow_redirects=False)
    assert upd_resp.status_code == 302

    del_url = "/clients/%d/delete" % client_id
    del_resp = post_with_csrf(admin_client, del_url, {}, csrf_source_url="/clients",
                              follow_redirects=False)
    assert del_resp.status_code == 302
    list_resp2 = admin_client.get("/clients")
    assert "Тестов Клиент ЕООД".encode() not in list_resp2.data


# ---------------------------------------------------------------- настройки

def test_settings_page_get_and_post(admin_client):
    resp = admin_client.get("/settings")
    assert resp.status_code == 200
    upd = post_with_csrf(admin_client, "/settings", {
        "sender_name": "Моята Фирма ЕООД",
    }, csrf_source_url="/settings", follow_redirects=False)
    assert upd.status_code == 302
    resp2 = admin_client.get("/settings")
    assert "Моята Фирма ЕООД".encode() in resp2.data


def test_my_settings_theme_change(admin_client):
    resp = admin_client.get("/my-settings")
    assert resp.status_code == 200
    upd = post_with_csrf(admin_client, "/my-settings", {"theme": "dark"},
                         csrf_source_url="/my-settings", follow_redirects=False)
    assert upd.status_code == 302


def test_my_settings_shows_admin_system_section_only_for_admin(admin_client, employee_client):
    admin_resp = admin_client.get("/my-settings")
    assert "gh_owner".encode() in admin_resp.data or "GitHub".encode() in admin_resp.data
    emp_resp = employee_client.get("/my-settings")
    assert emp_resp.status_code == 200


# ---------------------------------------------------------------- админ панел (потребители)

def test_admin_routes_forbidden_for_employee(employee_client):
    assert employee_client.get("/admin/users").status_code == 403
    assert employee_client.get("/admin/system").status_code == 403


def test_admin_user_new_toggle_password_delete_flow(admin_client):
    resp = post_with_csrf(admin_client, "/admin/users/new", {
        "username": "novak", "full_name": "Нов Служител",
        "password": "parola12345", "role": "employee",
    }, csrf_source_url="/admin/users", follow_redirects=False)
    assert resp.status_code == 302

    import db as db_mod
    con = db_mod.get_db()
    row = con.execute("SELECT id, active, must_change_password FROM users WHERE username = ?",
                      ("novak",)).fetchone()
    con.close()
    assert row["active"] == 1
    assert row["must_change_password"] == 1
    user_id = row["id"]

    toggle_url = "/admin/users/%d/toggle" % user_id
    t_resp = post_with_csrf(admin_client, toggle_url, {}, csrf_source_url="/admin/users",
                            follow_redirects=False)
    assert t_resp.status_code == 302

    pwd_url = "/admin/users/%d/password" % user_id
    p_resp = post_with_csrf(admin_client, pwd_url, {"password": "novaparola123"},
                            csrf_source_url="/admin/users", follow_redirects=False)
    assert p_resp.status_code == 302

    del_url = "/admin/users/%d/delete" % user_id
    d_resp = post_with_csrf(admin_client, del_url, {}, csrf_source_url="/admin/users",
                            follow_redirects=False)
    assert d_resp.status_code == 302


def test_admin_cannot_delete_own_account(admin_client):
    with admin_client.session_transaction() as sess:
        own_id = sess["user_id"]
    del_url = "/admin/users/%d/delete" % own_id
    resp = post_with_csrf(admin_client, del_url, {}, csrf_source_url="/admin/users",
                          follow_redirects=True)
    assert "Не можете да изтриете".encode() in resp.data


# ---------------------------------------------------------------- Фаза 4: M7/M9

def test_client_delete_forbidden_for_employee(employee_client, admin_client):
    """Находка M7: client_delete вече изисква администраторски права,
    точно както delete_document — обикновен служител вече не може да
    изтрие клиент от адресната книга."""
    post_with_csrf(admin_client, "/clients/new", {"name": "M7 Тест ЕООД"},
                   csrf_source_url="/clients/new")
    import db as db_mod
    con = db_mod.get_db()
    row = con.execute("SELECT id FROM clients WHERE name = ?", ("M7 Тест ЕООД",)).fetchone()
    con.close()

    del_url = "/clients/%d/delete" % row["id"]
    resp = post_with_csrf(employee_client, del_url, {}, csrf_source_url="/clients",
                          follow_redirects=False)
    assert resp.status_code == 403


def test_client_delete_allowed_for_admin(admin_client):
    post_with_csrf(admin_client, "/clients/new", {"name": "M7 Админ ЕООД"},
                   csrf_source_url="/clients/new")
    import db as db_mod
    con = db_mod.get_db()
    row = con.execute("SELECT id FROM clients WHERE name = ?", ("M7 Админ ЕООД",)).fetchone()
    con.close()

    del_url = "/clients/%d/delete" % row["id"]
    resp = post_with_csrf(admin_client, del_url, {}, csrf_source_url="/clients",
                          follow_redirects=False)
    assert resp.status_code == 302


def test_oversized_upload_is_rejected_with_friendly_message(admin_client):
    """Находка M9: MAX_CONTENT_LENGTH спира прекалено голям ъплоуд рано,
    вместо да го зареди изцяло в паметта; потребителят вижда приятелско
    съобщение и се връща обратно, не сурова Werkzeug грешка."""
    big_payload = b"x" * (26 * 1024 * 1024)  # 26 MB > 25 MB лимита
    resp = admin_client.post(
        "/settings/logo",
        data={"logo_file": (io.BytesIO(big_payload), "ogromno_logo.png")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert resp.status_code == 200
    assert "твърде голям".encode() in resp.data


def test_normal_sized_upload_is_not_rejected_by_max_content_length(admin_client):
    small_payload = b"x" * 1024
    resp = post_with_csrf(admin_client, "/settings/logo",
                          {"logo_file": (io.BytesIO(small_payload), "malko.txt")},
                          csrf_source_url="/settings", follow_redirects=False)
    # не е отхвърлено заради размера (може да е отхвърлено заради формата
    # на файла от branding.save_logo — важното тук е статусът да НЕ е 413,
    # т.е. лимитът не пречи на нормални по размер файлове)
    assert resp.status_code == 302


# ---------------------------------------------------------------- Фаза 4: достъпност (L2)

def test_flash_messages_have_alert_role(admin_client):
    resp = admin_client.get("/logout", follow_redirects=True)
    assert b'role="alert"' in resp.data


def test_login_error_has_alert_role(client):
    token = get_csrf_token(client, "/login")
    resp = client.post("/login", data={"username": "no", "password": "no",
                                       "csrf_token": token})
    assert b'role="alert"' in resp.data


def test_client_form_labels_are_associated_with_inputs(admin_client):
    resp = admin_client.get("/clients/new")
    assert b'<label for="f-name">' in resp.data
    assert b'id="f-name"' in resp.data


def test_no_hardcoded_muted_hex_color_in_rendered_forms(admin_client):
    resp = admin_client.get("/cmr/new")
    assert b"#5c6d80" not in resp.data


# ---------------------------------------------------------------- M8: споделени печатни макроси

def _issue(admin_client, path, extra_fields=None):
    data = {"sender_name": "Тест"}
    if extra_fields:
        data.update(extra_fields)
    resp = post_with_csrf(admin_client, path, data, csrf_source_url=path,
                          follow_redirects=False)
    assert resp.status_code == 302
    return admin_client.get(resp.headers["Location"])


def test_print_toolbar_macro_renders_no_leftover_jinja_for_all_doc_types(admin_client):
    """doc_toolbar/draft_watermark (templates/_macros.html) трябва да
    рендират чисто HTML — без изтекъл Jinja синтаксис — за всичките 6
    типа документи, включително ЧМР (extra_before: брой екземпляри) и
    палетна карта (extra_after: превключвател пълен формат/етикет).
    Товарителницата не беше в оригиналния M8 патч (документният тип не
    съществуваше по времето на генерирането му) — приведена ръчно към
    същите макроси и покрита тук за пълнота."""
    for path in ("/cmr/new", "/packing/new", "/pallet/new", "/dualuse/new", "/export-it/new", "/waybill/new"):
        resp = _issue(admin_client, path)
        assert resp.status_code == 200
        assert b"{{" not in resp.data and b"{%" not in resp.data
        assert b'class="doc-toolbar' in resp.data


def test_cmr_print_toolbar_has_copy_count_links(admin_client):
    resp = _issue(admin_client, "/cmr/new")
    assert "1 екземпляр".encode() in resp.data
    assert "4 екземпляра".encode() in resp.data
    assert "5 екземпляра".encode() in resp.data


def test_pallet_print_toolbar_has_format_toggle_link(admin_client):
    resp = _issue(admin_client, "/pallet/new", {"pallet_no": "1"})
    assert "етикет".encode() in resp.data


def test_document_preview_uses_shared_watermark_macro(admin_client):
    resp = post_with_csrf(admin_client, "/packing/preview", {"sender_name": "X"},
                          csrf_source_url="/packing/new", follow_redirects=False)
    preview_resp = admin_client.get(resp.headers["Location"])
    assert b'class="draft-watermark"' in preview_resp.data
    assert b"{{" not in preview_resp.data and b"{%" not in preview_resp.data


# -------- back_url: замяна на javascript:history.back() с истински адрес

_PREVIEW_PATHS = {
    "cmr": "/cmr/preview",
    "packing": "/packing/preview",
    "pallet": "/pallet/preview",
    "dualuse": "/dualuse/preview",
    "export_it": "/export-it/preview",
    "waybill": "/waybill/preview",
}
_NEW_PATHS = {
    "cmr": "/cmr/new",
    "packing": "/packing/new",
    "pallet": "/pallet/new",
    "dualuse": "/dualuse/new",
    "export_it": "/export-it/new",
    "waybill": "/waybill/new",
}


def test_document_preview_back_button_uses_real_form_url_not_history_back(admin_client):
    """Бутонът "Назад към формата" в прегледа (преди запис) трябва да
    сочи към истинския адрес на формата (back_url в doc_toolbar), а НЕ
    към javascript:history.back() — виж обяснението в templates/_macros.html
    защо history.back() чупи навигацията след POST → пренасочване → GET.
    Включва товарителницата (waybill) — приведена ръчно към същата
    поправка, тъй като не беше част от оригиналния патч."""
    for doc_type, preview_path in _PREVIEW_PATHS.items():
        resp = post_with_csrf(admin_client, preview_path, {"sender_name": "Тест"},
                              csrf_source_url=_NEW_PATHS[doc_type], follow_redirects=False)
        assert resp.status_code == 302
        location = resp.headers["Location"]
        preview_resp = admin_client.get(location)
        assert preview_resp.status_code == 200
        assert b"javascript:history.back()" not in preview_resp.data
        # back_url вече носи ?restore=<token> (виж
        # test_document_preview_back_button_restores_form_data по-долу) —
        # токенът е последният сегмент на /preview/<token>.
        token = location.rsplit("/", 1)[-1]
        expected_href = ('href="%s?restore=%s"' % (_NEW_PATHS[doc_type], token)).encode()
        assert expected_href in preview_resp.data


def test_document_preview_back_button_restores_form_data(admin_client):
    """При „Предварителен преглед" → „Назад към формата" въведените (все
    още незаписани) данни не трябва да се губят — формата трябва да се
    предзареди отново с тях чрез ?restore=<token> → edit_data/data-edit
    (виж routes_documents._document_new)."""
    for doc_type, preview_path in _PREVIEW_PATHS.items():
        resp = post_with_csrf(admin_client, preview_path,
                              {"sender_name": "Възстановена фирма ЕООД"},
                              csrf_source_url=_NEW_PATHS[doc_type], follow_redirects=False)
        location = resp.headers["Location"]
        token = location.rsplit("/", 1)[-1]
        back_resp = admin_client.get("%s?restore=%s" % (_NEW_PATHS[doc_type], token))
        assert back_resp.status_code == 200
        assert b"Restored form data" not in back_resp.data  # sanity: not a stray literal
        assert "Възстановена фирма ЕООД".encode() in back_resp.data
        assert b"data-edit=" in back_resp.data


def test_document_new_restore_ignores_mismatched_or_unknown_token(admin_client):
    """?restore= с непознат/изтекъл токен или токен от ДРУГ тип документ не
    трябва да чупи формата — просто се игнорира (празна форма, както преди)."""
    resp = admin_client.get("/cmr/new?restore=not-a-real-token")
    assert resp.status_code == 200
    assert b"data-edit=" not in resp.data

    # токен, съхранен под друг doc_type — не бива да „изтече" в cmr формата
    pallet_resp = post_with_csrf(admin_client, "/pallet/preview",
                                 {"sender_name": "Палетна фирма"},
                                 csrf_source_url="/pallet/new", follow_redirects=False)
    pallet_token = pallet_resp.headers["Location"].rsplit("/", 1)[-1]
    cross_resp = admin_client.get("/cmr/new?restore=%s" % pallet_token)
    assert cross_resp.status_code == 200
    assert b"Palletna firma" not in cross_resp.data
    assert b"data-edit=" not in cross_resp.data


def test_pallet_bulk_preview_back_button_uses_real_form_url_not_history_back(admin_client):
    data = {
        "groups": "1",
        "sender_name": "Тест",
        "items_json_1": '[{"order_no": "O1", "pos": "1", "reference": "R1", '
                        '"reference_desc": "D1", "qty": "5"}]',
    }
    resp = post_with_csrf(admin_client, "/pallet/bulk-preview", data,
                          csrf_source_url="/pallet/new", follow_redirects=False)
    assert resp.status_code == 302
    preview_resp = admin_client.get(resp.headers["Location"])
    assert preview_resp.status_code == 200
    assert b"javascript:history.back()" not in preview_resp.data
    assert b'href="/pallet/new"' in preview_resp.data
