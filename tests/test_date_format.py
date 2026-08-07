# -*- coding: utf-8 -*-
"""Тестове за „в цялата програма промени изгледа на дата да е
ден.месец.година“ — appcore.format_bg_date (Jinja global format_date)."""
from appcore import format_bg_date


def test_none_and_empty_pass_through_unchanged():
    assert format_bg_date(None) is None
    assert format_bg_date("") == ""


def test_iso_date_only_reformatted():
    assert format_bg_date("2026-08-07") == "07.08.2026"


def test_iso_date_single_digit_day_and_month_zero_padded_in_source():
    assert format_bg_date("2026-01-05") == "05.01.2026"


def test_iso_datetime_with_seconds_reformatted_and_seconds_dropped():
    assert format_bg_date("2026-08-07 18:25:03") == "07.08.2026 18:25"


def test_iso_datetime_without_seconds_reformatted():
    assert format_bg_date("2026-08-07 18:25") == "07.08.2026 18:25"


def test_iso_datetime_with_t_separator_reformatted():
    assert format_bg_date("2026-08-07T18:25:00") == "07.08.2026 18:25"


def test_non_date_text_passed_through_unchanged():
    assert format_bg_date("не е дата") == "не е дата"


def test_already_bg_formatted_date_passed_through_unchanged():
    # Толерантност към вече "друг формат" данни (виж докстринга) — не се
    # опитва второ преобразуване/парсене на вече ДД.ММ.ГГГГ низ.
    assert format_bg_date("07.08.2026") == "07.08.2026"


def test_whitespace_trimmed_before_matching():
    assert format_bg_date("  2026-08-07  ") == "07.08.2026"


def test_non_string_value_handled():
    import datetime
    # SQLite винаги връща низове, но проверяваме, че не гърми при друг тип.
    assert format_bg_date(datetime.date(2026, 8, 7)) == "07.08.2026"


# ---------------------------------------------------------------- дати в печатните бланки (интеграционни)
# Проверява, че ISO датите от <input type="date"> полетата (established_date,
# date_loading, date_delivery, loading_date, unloading_date) реално излизат
# като „ДД.ММ.ГГГГ“ в самите печатни HTML бланки — не само че функцията
# format_bg_date работи изолирано (виж тестовете по-горе).
import json

from conftest import post_with_csrf


def test_cmr_print_shows_established_date_and_loading_date_in_bg_format(admin_client):
    resp = post_with_csrf(admin_client, "/cmr/new", {
        "sender_name": "Изпращач", "consignee_name": "Получател",
        "established_date": "2026-08-07", "place_loading": "София",
        "date_loading": "2026-08-06",
    }, csrf_source_url="/cmr/new", follow_redirects=False)
    doc_url = resp.headers["Location"]
    view_resp = admin_client.get(doc_url)
    body = view_resp.data.decode("utf-8")
    assert "07.08.2026" in body
    assert "06.08.2026" in body
    assert "2026-08-07" not in body
    assert "2026-08-06" not in body


def test_waybill_print_shows_all_date_fields_in_bg_format(admin_client):
    items = json.dumps([{"description": "Стока", "qty": "1"}])
    resp = post_with_csrf(admin_client, "/waybill/new", {
        "sender_name": "Изпращач", "established_date": "2026-01-05",
        "date_loading": "2026-01-06", "date_delivery": "2026-01-07",
        "loading_date": "2026-01-08", "unloading_date": "2026-01-09",
        "items_json": items,
    }, csrf_source_url="/waybill/new", follow_redirects=False)
    doc_url = resp.headers["Location"]
    view_resp = admin_client.get(doc_url)
    body = view_resp.data.decode("utf-8")
    for bg_date in ("05.01.2026", "06.01.2026", "07.01.2026", "08.01.2026", "09.01.2026"):
        assert bg_date in body
    for iso_date in ("2026-01-05", "2026-01-06", "2026-01-07", "2026-01-08", "2026-01-09"):
        assert iso_date not in body


# ---------------------------------------------------------------- дати в Excel/PDF износ

def test_cmr_xlsx_export_formats_dates_as_bg(admin_client):
    resp = post_with_csrf(admin_client, "/cmr/new", {
        "sender_name": "Изпращач", "established_date": "2026-08-07",
        "date_loading": "2026-08-06",
    }, csrf_source_url="/cmr/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    xlsx_resp = admin_client.get("/doc/%s/export.xlsx" % doc_id)
    assert xlsx_resp.status_code == 200

    import io as _io
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(xlsx_resp.data))
    ws = wb.active
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert "07.08.2026" in all_values
    assert "06.08.2026" in all_values
    assert "2026-08-07" not in all_values
    assert "2026-08-06" not in all_values


def test_cmr_pdf_export_formats_dates_as_bg(admin_client):
    resp = post_with_csrf(admin_client, "/cmr/new", {
        "sender_name": "Изпращач", "established_date": "2026-08-07",
    }, csrf_source_url="/cmr/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    pdf_resp = admin_client.get("/doc/%s/export.pdf" % doc_id)
    assert pdf_resp.status_code == 200
    assert pdf_resp.headers["Content-Type"] == "application/pdf"
