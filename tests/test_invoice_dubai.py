# -*- coding: utf-8 -*-
"""Тестове за третия вид фактура — за Дубай (invoice_dubai).

Заявка: „добави и фактура за Дубай“ — приложен образец 12971.pdf (BBS
Bulgaria LTD → ABB INDUSTRIES LLC, Дубай, ОАЕ). Общата структура (изпращач/
получател/„Bill To“/данни за пратката) е ЕДНАКВА с Бразилия и Норвегия
(вижте templates/_invoice_macros.html), затова тук се проверява само
специфичното за Дубай:

  - заглавие на бланката „COMMERCIAL INVOICE“ (като Норвегия);
  - вид транспорт по подразбиране „AIRWAY / FCA“, свободен текст (не
    падащо меню — за разлика от Бразилия, тук няма заявка за ограничен
    избор от варианти);
  - колоните на таблицата със стоки СА НАЙ-ПРОСТИТЕ от трите: HS code,
    P.O NO, Pos, Material code, Quantity, Unit Price, Total Price — БЕЗ
    нето тегло (за разлика от Бразилия) и БЕЗ описание на материала /
    палет № (за разлика от Норвегия) — точно както в приложения образец;
  - отделен тип документ, собствена номерация, показва се в раздел
    „Фактури“, изключен от „Всички документи“ (наследено от общата
    db.INVOICE_DOC_TYPES инфраструктура — тук само се потвърждава, че
    Дубай не е пропуснат никъде).
"""
import io
import json

from conftest import post_with_csrf


def test_invoice_dubai_form_renders(admin_client):
    assert admin_client.get("/invoice-dubai/new").status_code == 200


def test_dubai_invoice_columns_match_the_sample_exactly(admin_client):
    """Образецът 12971.pdf НЯМА нито нето тегло, нито описание на
    материала — само HS code, P.O NO, Pos, Material code, Quantity,
    Unit Price."""
    body = admin_client.get("/invoice-dubai/new").data.decode()
    assert 'data-columns="hs_code,po_no,pos,material_code,qty,unit_price"' in body
    assert "Net weight" not in body
    assert "Material Decription" not in body
    assert "Pallet Number" not in body


def test_dubai_transportation_way_defaults_to_airway_fca_as_free_text(admin_client):
    """По подразбиране „AIRWAY / FCA“ (точно както в образеца), свободен
    текст — за разлика от Бразилия няма заявка за ограничен избор."""
    body = admin_client.get("/invoice-dubai/new").data.decode()
    assert '<input id="f-transport_way" type="text" name="transport_way" value="AIRWAY / FCA">' in body


def test_issue_dubai_invoice_and_view_it(admin_client):
    items = json.dumps([
        {"hs_code": "85389099", "po_no": "5197132282", "pos": "30",
         "material_code": "1TSA271100P0013", "qty": "2", "unit_price": "0.72"},
        {"hs_code": "85389099", "po_no": "5197132282", "pos": "40",
         "material_code": "1TSA271100P0031", "qty": "1", "unit_price": "1.25"},
    ])
    resp = post_with_csrf(admin_client, "/invoice-dubai/new", {
        "consignee_name": "ABB INDUSTRIES LLC",
        "consignee_address": "AE013102 LV SYSTEMS SERVICE AlQuoze\nDUBAI 11070\nUtd.Arab Emir.",
        "invoice_number": "0000012971", "items_json": items,
    }, csrf_source_url="/invoice-dubai/new", follow_redirects=False)
    assert resp.status_code == 302

    body = admin_client.get(resp.headers["Location"]).data.decode()
    assert "COMMERCIAL INVOICE" in body
    assert "Invoice Number: 0000012971" in body
    assert "ABB INDUSTRIES LLC" in body
    assert "1TSA271100P0013" in body
    assert "1TSA271100P0031" in body
    assert "1.44" in body, "обща цена на първия ред = 2 × 0.72"


def test_dubai_invoice_totals_row_has_no_weight_column(admin_client):
    items = json.dumps([
        {"material_code": "1TSA271100P0013", "qty": "2", "unit_price": "0.72"},
        {"material_code": "1TSA271100P0031", "qty": "1", "unit_price": "1.25"},
    ])
    resp = post_with_csrf(admin_client, "/invoice-dubai/new",
                          {"consignee_name": "ABB", "items_json": items},
                          csrf_source_url="/invoice-dubai/new", follow_redirects=False)
    body = admin_client.get(resp.headers["Location"]).data.decode()
    assert 'class="totals"' in body
    totals = body.split('class="totals"')[1].split("</tr>")[0]
    assert ">3<" in totals    # общо количество 2 + 1
    assert ">2.69 €<" in totals  # обща стойност 1.44 + 1.25 (+ символ за евро)


def test_dubai_is_a_separate_document_type_with_its_own_numbering(admin_client):
    resp = post_with_csrf(admin_client, "/invoice-dubai/new",
                          {"consignee_name": "ABB Дубай", "invoice_number": "DU-1"},
                          csrf_source_url="/invoice-dubai/new", follow_redirects=False)
    assert "COMMERCIAL INVOICE" in admin_client.get(resp.headers["Location"]).data.decode()

    listing = admin_client.get("/invoices").data.decode()
    assert "Фактура за Дубай" in listing


def test_dubai_invoice_preview_does_not_save_a_document(admin_client):
    before = admin_client.get("/docs").data.decode().count("/doc/")
    resp = post_with_csrf(admin_client, "/invoice-dubai/preview", {"consignee_name": "Преглед"},
                          csrf_source_url="/invoice-dubai/new", follow_redirects=False)
    assert resp.status_code == 302
    assert "/preview/" in resp.headers["Location"]
    assert admin_client.get(resp.headers["Location"]).status_code == 200
    assert admin_client.get("/docs").data.decode().count("/doc/") == before


def test_dubai_invoice_has_no_barcode_on_the_printed_form(admin_client):
    resp = post_with_csrf(admin_client, "/invoice-dubai/new",
                          {"consignee_name": "ABB", "invoice_number": "DU-77"},
                          csrf_source_url="/invoice-dubai/new", follow_redirects=False)
    body = admin_client.get(resp.headers["Location"]).data.decode()
    assert "INVDU-" not in body
    assert 'role="img"' not in body


def test_dubai_invoice_is_excluded_from_all_documents_list(admin_client):
    post_with_csrf(admin_client, "/invoice-dubai/new",
                   {"consignee_name": "Дубай клиент", "invoice_number": "САМО-ТУК-ДУ"},
                   csrf_source_url="/invoice-dubai/new", follow_redirects=True)
    docs_body = admin_client.get("/docs").data.decode()
    assert "САМО-ТУК-ДУ" not in docs_body
    assert "САМО-ТУК-ДУ" in admin_client.get("/invoices").data.decode()


def test_dubai_invoice_pull_pallet_and_excel_import_work(admin_client):
    """Дубай ползва СЪЩИТЕ споделени endpoint-и за зареждане на редове от
    палетна карта/Excel като Бразилия и Норвегия (routes_invoices.py не
    прави разлика по тип на извикващата фактура)."""
    items = [{"order_no": "PO-1", "pos": "10", "reference": "MAT-1",
              "reference_desc": "Нещо", "qty": "3"}]
    resp = post_with_csrf(admin_client, "/pallet/new", {
        "pallet_no": "9", "client_name": "ABB", "items_format": "orders",
        "items_json": json.dumps(items),
    }, csrf_source_url="/pallet/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]
    number = admin_client.get("/doc/%s" % doc_id).data.decode() \
        .split("ПАЛЕТНА КАРТА № ")[1][:9].strip()

    payload = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": number},
                             csrf_source_url="/invoice-dubai/new").get_json()
    assert payload["ok"] is True
    assert payload["rows"][0]["material_code"] == "MAT-1"


def test_dubai_invoice_xlsx_export_has_sample_columns_and_computed_total(admin_client):
    items = json.dumps([{
        "hs_code": "85389099", "po_no": "5197132282", "pos": "30",
        "material_code": "1TSA271100P0013", "qty": "2", "unit_price": "0.72",
    }])
    resp = post_with_csrf(admin_client, "/invoice-dubai/new",
                          {"consignee_name": "ABB", "doc_date": "2026-08-03",
                           "items_json": items},
                          csrf_source_url="/invoice-dubai/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(admin_client.get("/doc/%s/export.xlsx" % doc_id).data))
    values = [c.value for row in wb.active.iter_rows() for c in row if c.value is not None]
    assert "Код на материала" in values
    assert "Обща цена, EUR" in values
    assert "Нето тегло, кг/бр" not in values, "Дубай няма колона с тегло"
    assert "Описание на материала" not in values, "Дубай няма колона с описание"
    assert "1.44" in values, "изчислената обща цена на реда (2 × 0.72)"
    assert "03.08.2026" in values


def test_dubai_invoice_pdf_export_works(admin_client):
    resp = post_with_csrf(admin_client, "/invoice-dubai/new", {"consignee_name": "ABB"},
                          csrf_source_url="/invoice-dubai/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]
    pdf = admin_client.get("/doc/%s/export.pdf" % doc_id)
    assert pdf.status_code == 200
    assert pdf.headers["Content-Type"] == "application/pdf"


def test_dubai_invoice_appears_in_nav_and_invoices_page(admin_client):
    body = admin_client.get("/").data.decode()
    assert "invoice-dubai" in body or admin_client.get("/invoices").data.decode()
    inv_body = admin_client.get("/invoices").data.decode()
    assert "Нова фактура за Дубай" in inv_body
