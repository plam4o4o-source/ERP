# -*- coding: utf-8 -*-
"""Тестове за раздел „Фактури“ — фактура за Бразилия (invoice_br) и за
Норвегия (invoice_no).

Заявка: „добави раздел фактури, в него да има Фактура за Бразилия и
Фактура за Норвегия... направи ги отделни за съответните държави“ +
„фактурата да може да се зарежда, както се зареждат палетните карти в
опаковъчния лист, по същия начин, но само със съответните данни, които са
необходими за фактура“.

Двата образеца имат РАЗЛИЧНИ колони (потвърдено от приложените файлове и
изрично избрано от потребителя „точно като образците“):
  Бразилия: HS code, P.O NO, Pos, Net weight, Material code, Quantity,
            Unit Price, Total Price  — БЕЗ описание
  Норвегия: HS code, Material Decription, Pallet Number, P.O NO, Pos,
            Material code, Quantity, Unit Price, Total Price — БЕЗ тегло
Точно това разминаване е и най-лесното за счупване при бъдеща промяна,
затова му е отделено най-много внимание тук.
"""
import io
import json

from openpyxl import Workbook

import materials
from conftest import post_with_csrf

_CATALOG = [
    ("GLBK400002P0012", "C-PROFILE 3   1150MM", 2.21),
    ("1TFL151621P0550", "transverse section 06  folded", 2.74259375),
    ("GLBK400001P0200", "C-PROFILE 2    200MM", 0.383),
]


def _load_catalog(client):
    wb = Workbook()
    ws = wb.active
    ws.append(["ABB part ID", "Description", "Net weight\n[KG/pc]"])
    for row in _CATALOG:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)  # wb.save оставя позицията в края — без това качва празен файл
    post_with_csrf(client, "/materials/import",
                   {"excel_file": (buf, "kg.xlsx")},
                   csrf_source_url="/materials", content_type="multipart/form-data")


def _issue_pallet_with_orders(client, rows, pallet_no="7"):
    """Издава палетна карта във формат „поръчки“ (Order No, Pos,
    Reference, Reference Desc, Open Qty) — точно формата, от който
    фактурата после дърпа редовете си."""
    items = [{"order_no": o, "pos": p, "reference": r, "reference_desc": rd, "qty": q}
             for o, p, r, rd, q in rows]
    resp = post_with_csrf(client, "/pallet/new", {
        "pallet_no": pallet_no, "client_name": "ABB", "items_format": "orders",
        "items_json": json.dumps(items),
    }, csrf_source_url="/pallet/new", follow_redirects=False)
    return resp.headers["Location"].rstrip("/").split("/")[-1]


def _pallet_number(client, doc_id):
    return client.get("/doc/%s" % doc_id).data.decode().split("ПАЛЕТНА КАРТА № ")[1][:9].strip()


# ---------------------------------------------------------------- форми и издаване

def test_invoice_forms_render_for_both_countries(admin_client):
    assert admin_client.get("/invoice-br/new").status_code == 200
    assert admin_client.get("/invoice-no/new").status_code == 200


def test_brazil_invoice_columns_match_the_sample_exactly(admin_client):
    """Образецът за Бразилия НЯМА колона с описание, но ИМА нето тегло."""
    body = admin_client.get("/invoice-br/new").data.decode()
    assert 'data-columns="hs_code,po_no,pos,net_weight,material_code,qty,unit_price"' in body
    assert "Net weight" in body
    assert "Pallet Number" not in body


def test_norway_invoice_columns_match_the_sample_exactly(admin_client):
    """Образецът за Норвегия НЯМА нето тегло, но ИМА описание и палет №."""
    body = admin_client.get("/invoice-no/new").data.decode()
    assert ('data-columns="hs_code,description,pallet_no,po_no,pos,material_code,'
            'qty,unit_price"') in body
    assert "Material Decription" in body
    assert "Pallet Number" in body


def test_invoice_lookup_fills_weight_for_brazil_and_description_for_norway(admin_client):
    """Кое поле се попълва автоматично от справочника зависи от това коя
    колона изобщо съществува в съответния образец."""
    assert 'data-lookup-fill="net_weight"' in admin_client.get("/invoice-br/new").data.decode()
    assert 'data-lookup-fill="description"' in admin_client.get("/invoice-no/new").data.decode()


def test_issue_brazil_invoice_and_view_it(admin_client):
    items = json.dumps([{
        "hs_code": "85389099", "po_no": "4700200362", "pos": "30",
        "net_weight": "2.21", "material_code": "GLBK400002P0012",
        "qty": "20", "unit_price": "13.66",
    }])
    resp = post_with_csrf(admin_client, "/invoice-br/new", {
        "consignee_name": "ABB ELETRIFICACAO LTDA",
        "consignee_address": "Sorocaba - SP\nBrasil",
        "transport_way": "AIRFREIGHT / FCA", "items_json": items,
    }, csrf_source_url="/invoice-br/new", follow_redirects=False)
    assert resp.status_code == 302

    body = admin_client.get(resp.headers["Location"]).data.decode()
    assert "INVOICE" in body
    assert "ABB ELETRIFICACAO LTDA" in body
    assert "GLBK400002P0012" in body
    assert "273.2" in body, "обща цена на реда = 20 × 13.66"


def test_issue_norway_invoice_and_view_it(admin_client):
    items = json.dumps([{
        "hs_code": "85389099", "description": "C-Profile", "pallet_no": "1",
        "po_no": "4561347665", "pos": "20", "material_code": "GLBK400002P0012",
        "qty": "100", "unit_price": "3.93",
    }])
    resp = post_with_csrf(admin_client, "/invoice-no/new", {
        "consignee_name": "ABB ELECTRIFICATION NORWAY AS",
        "confirmation_number": "CN-1", "items_json": items,
    }, csrf_source_url="/invoice-no/new", follow_redirects=False)
    assert resp.status_code == 302

    body = admin_client.get(resp.headers["Location"]).data.decode()
    assert "COMMERCIAL INVOICE" in body
    assert "ABB ELECTRIFICATION NORWAY AS" in body
    assert "393" in body, "обща цена на реда = 100 × 3.93"


def test_the_two_invoice_types_are_stored_as_separate_document_types(admin_client):
    """Всяка държава е отделен тип документ — вижда се по заглавието на
    бланката и по вида в списъка с издадени фактури. (Номерата вече се
    въвеждат ръчно, затова не се сверяват автоматично генерирани префикси.)"""
    br = post_with_csrf(admin_client, "/invoice-br/new",
                        {"consignee_name": "BR", "invoice_number": "BR-1"},
                        csrf_source_url="/invoice-br/new", follow_redirects=False)
    no = post_with_csrf(admin_client, "/invoice-no/new",
                        {"consignee_name": "NO", "invoice_number": "NO-1"},
                        csrf_source_url="/invoice-no/new", follow_redirects=False)
    assert "INVOICE" in admin_client.get(br.headers["Location"]).data.decode()
    assert "COMMERCIAL INVOICE" in admin_client.get(no.headers["Location"]).data.decode()

    listing = admin_client.get("/invoices").data.decode()
    assert "Фактура за Бразилия" in listing
    assert "Фактура за Норвегия" in listing


def test_invoice_preview_does_not_save_a_document(admin_client):
    before = admin_client.get("/docs").data.decode().count("/doc/")
    resp = post_with_csrf(admin_client, "/invoice-br/preview", {"consignee_name": "Преглед"},
                          csrf_source_url="/invoice-br/new", follow_redirects=False)
    assert resp.status_code == 302
    assert "/preview/" in resp.headers["Location"]
    assert admin_client.get(resp.headers["Location"]).status_code == 200
    assert admin_client.get("/docs").data.decode().count("/doc/") == before


def test_editing_an_issued_invoice_keeps_its_item_rows(admin_client):
    """Регресия: кои типове имат редове се четеше от изброен списък в
    routes_documents, който не включваше фактурите — при редакция редовете
    им щяха да изчезнат тихо."""
    items = json.dumps([{"material_code": "GLBK400002P0012", "qty": "5", "unit_price": "2"}])
    resp = post_with_csrf(admin_client, "/invoice-br/new",
                          {"consignee_name": "ABB", "items_json": items},
                          csrf_source_url="/invoice-br/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    edit_body = admin_client.get("/doc/%s/edit" % doc_id).data.decode()
    assert "GLBK400002P0012" in edit_body

    post_with_csrf(admin_client, "/doc/%s/edit" % doc_id,
                   {"consignee_name": "ABB (променен)", "items_json": items},
                   csrf_source_url="/doc/%s/edit" % doc_id, follow_redirects=False)
    body = admin_client.get("/doc/%s" % doc_id).data.decode()
    assert "ABB (променен)" in body
    assert "GLBK400002P0012" in body, "редовете трябва да оцелеят след редакция"


# ---------------------------------------------------------------- сметките (мерна единица)
# Самата аритметика се проверява директно върху функциите, които ползват и
# бланката, и Excel/PDF износа — вместо върху целия HTML, където кратки
# числа като „30“ съвпадат случайно и тестът минава без да проверява нищо.

def test_invoice_row_total_is_quantity_times_unit_price():
    from appcore import invoice_row_total
    assert invoice_row_total({"qty": "20", "unit_price": "13.66"}) == "273.2"
    assert invoice_row_total({"qty": "10", "unit_price": "4.53"}) == "45.3"


def test_invoice_row_total_is_empty_without_quantity_or_price():
    from appcore import invoice_row_total
    assert invoice_row_total({"qty": "20"}) == ""
    assert invoice_row_total({"unit_price": "5"}) == ""
    assert invoice_row_total({"qty": "", "unit_price": ""}) == ""
    assert invoice_row_total({}) == ""


def test_invoice_row_total_accepts_decimal_comma():
    """Операторите редовно пишат „13,66“ — стойността идва от свободно
    текстово поле, не от числов вход."""
    from appcore import invoice_row_total
    assert invoice_row_total({"qty": "20", "unit_price": "13,66"}) == "273.2"


def test_invoice_row_weight_is_net_weight_times_quantity():
    from appcore import invoice_row_weight
    assert invoice_row_weight({"qty": "20", "net_weight": "4.51"}) == "90.2"
    assert invoice_row_weight({"qty": "10", "net_weight": "0.2"}) == "2"
    assert invoice_row_weight({"qty": "10"}) == ""


def test_invoice_totals_sums_quantity_price_and_weight():
    from appcore import invoice_totals
    totals = invoice_totals([
        {"qty": "10", "unit_price": "4.53", "net_weight": "0.2"},
        {"qty": "20", "unit_price": "13.66", "net_weight": "4.51"},
    ])
    assert totals == {"qty": "30", "price": "318.5", "weight": "92.2"}


def test_invoice_totals_counts_only_rows_that_have_the_needed_values():
    """Ред без цена (обичайно веднага след зареждане от палетна карта)
    влиза в общото количество, но не и в общата стойност — и не прави
    цялата сума празна."""
    from appcore import invoice_totals
    totals = invoice_totals([
        {"qty": "10", "unit_price": "2"},
        {"qty": "5", "unit_price": ""},
    ])
    assert totals["qty"] == "15"
    assert totals["price"] == "20"
    assert totals["weight"] == ""


def test_invoice_totals_tolerates_empty_and_broken_rows():
    from appcore import invoice_totals
    assert invoice_totals([]) == {"qty": "", "price": "", "weight": ""}
    assert invoice_totals(None) == {"qty": "", "price": "", "weight": ""}
    assert invoice_totals([None, "не е ред", {"qty": "абв"}]) == \
        {"qty": "", "price": "", "weight": ""}


# ---------------------------------------------------------------- суми на бланката

def _totals_row(body):
    """Изрязва РЕДА С ОБЩИТЕ СУМИ от готовата бланка (<tr class="totals">),
    за да не се проверява срещу целия HTML, където кратките числа съвпадат
    случайно."""
    assert 'class="totals"' in body, "бланката трябва да има ред с общи суми"
    return body.split('class="totals"')[1].split("</tr>")[0]


def test_brazil_invoice_totals_row_sums_quantity_price_and_weight(admin_client):
    items = json.dumps([
        {"material_code": "A", "net_weight": "0.2", "qty": "10", "unit_price": "4.53"},
        {"material_code": "B", "net_weight": "4.51", "qty": "20", "unit_price": "13.66"},
    ])
    resp = post_with_csrf(admin_client, "/invoice-br/new",
                          {"consignee_name": "ABB", "items_json": items},
                          csrf_source_url="/invoice-br/new", follow_redirects=False)
    totals = _totals_row(admin_client.get(resp.headers["Location"]).data.decode())
    assert ">30<" in totals      # общо количество 10 + 20
    assert ">318.5<" in totals   # обща стойност 45.3 + 273.2
    assert ">92.2<" in totals    # общо тегло 2.0 + 90.2


def test_norway_invoice_totals_row_has_no_weight_column(admin_client):
    items = json.dumps([{"material_code": "A", "qty": "100", "unit_price": "3.93"}])
    resp = post_with_csrf(admin_client, "/invoice-no/new",
                          {"consignee_name": "ABB", "items_json": items},
                          csrf_source_url="/invoice-no/new", follow_redirects=False)
    totals = _totals_row(admin_client.get(resp.headers["Location"]).data.decode())
    assert ">100<" in totals
    assert ">393<" in totals


# ---------------------------------------------------------------- зареждане от палетна карта

def test_pull_pallet_loads_every_row_separately_not_one_summary(admin_client):
    """Ядрото на заявката: за разлика от опаковъчния лист (един обобщен
    ред), фактурата получава ВСЕКИ ред поотделно."""
    _load_catalog(admin_client)
    doc_id = _issue_pallet_with_orders(admin_client, [
        ("4700200362", "30", "GLBK400002P0012", "C-Profile", "20"),
        ("4700200362", "40", "1TFL151621P0550", "Steel Parts", "44"),
    ])
    number = _pallet_number(admin_client, doc_id)

    payload = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": number},
                             csrf_source_url="/invoice-br/new").get_json()
    assert payload["ok"] is True
    assert payload["count"] == 2
    assert [r["material_code"] for r in payload["rows"]] == \
        ["GLBK400002P0012", "1TFL151621P0550"]


def test_pull_pallet_maps_pallet_columns_onto_invoice_columns(admin_client):
    """Order No → P.O NO, Pos → Pos, Reference → Material code,
    Reference Desc → описание, Open Qty → Quantity."""
    _load_catalog(admin_client)
    doc_id = _issue_pallet_with_orders(
        admin_client, [("4700200362", "30", "GLBK400002P0012", "C-Profile", "20")],
        pallet_no="3")
    number = _pallet_number(admin_client, doc_id)

    row = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": number},
                         csrf_source_url="/invoice-br/new").get_json()["rows"][0]
    assert row["po_no"] == "4700200362"
    assert row["pos"] == "30"
    assert row["material_code"] == "GLBK400002P0012"
    assert row["description"] == "C-Profile"
    assert row["qty"] == "20"
    assert row["pallet_no"] == "3"
    assert row["hs_code"] == "85389099"
    assert row["unit_price"] == "", "цената се въвежда ръчно, не идва от палетната карта"


def test_pull_pallet_fills_net_weight_from_the_materials_catalog(admin_client):
    """Точно това е „от файла с килограмите автоматично да се извличат
    съответните килограми във фактурата“."""
    _load_catalog(admin_client)
    doc_id = _issue_pallet_with_orders(admin_client, [
        ("PO-1", "10", "GLBK400002P0012", "C-Profile", "20"),
        ("PO-1", "20", "GLBK400001P0200", "C-Profile 2", "400"),
    ])
    number = _pallet_number(admin_client, doc_id)

    payload = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": number},
                             csrf_source_url="/invoice-br/new").get_json()
    assert [r["net_weight"] for r in payload["rows"]] == ["2.21", "0.383"]
    assert payload["matched"] == 2


def test_pull_pallet_keeps_rows_whose_code_is_not_in_the_catalog(admin_client):
    """Непознат код НЕ бива да изхвърля реда — просто остава без тегло за
    ръчно попълване, а отговорът казва колко са намерени."""
    _load_catalog(admin_client)
    doc_id = _issue_pallet_with_orders(admin_client, [
        ("PO-1", "10", "GLBK400002P0012", "C-Profile", "20"),
        ("PO-1", "20", "НЯМА-ТАКЪВ-КОД", "Непознат материал", "5"),
    ])
    number = _pallet_number(admin_client, doc_id)

    payload = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": number},
                             csrf_source_url="/invoice-br/new").get_json()
    assert payload["count"] == 2
    assert payload["matched"] == 1
    assert payload["rows"][1]["material_code"] == "НЯМА-ТАКЪВ-КОД"
    assert payload["rows"][1]["net_weight"] == ""


def test_pull_pallet_prefers_pallet_description_over_catalog_description(admin_client):
    """Описанието от палетната карта е уточнено за конкретната пратка —
    справочникът е само резервният източник."""
    _load_catalog(admin_client)
    doc_id = _issue_pallet_with_orders(
        admin_client, [("PO-1", "10", "GLBK400002P0012", "Уточнено описание", "1")])
    number = _pallet_number(admin_client, doc_id)

    row = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": number},
                         csrf_source_url="/invoice-br/new").get_json()["rows"][0]
    assert row["description"] == "Уточнено описание"


def test_pull_pallet_falls_back_to_catalog_description_when_pallet_has_none(admin_client):
    _load_catalog(admin_client)
    doc_id = _issue_pallet_with_orders(
        admin_client, [("PO-1", "10", "GLBK400002P0012", "", "1")])
    number = _pallet_number(admin_client, doc_id)

    row = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": number},
                         csrf_source_url="/invoice-br/new").get_json()["rows"][0]
    assert row["description"] == "C-PROFILE 3   1150MM"


def test_pull_pallet_reports_a_document_that_is_not_a_pallet_card(admin_client):
    resp = post_with_csrf(admin_client, "/cmr/new", {"consignee_name": "X"},
                          csrf_source_url="/cmr/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]
    number = admin_client.get("/doc/%s" % doc_id).data.decode().split("ЧМР № ")[1][:9].strip()

    payload = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": number},
                             csrf_source_url="/invoice-br/new").get_json()
    assert payload["ok"] is False
    assert "не е палетна карта" in payload["error"]


def test_pull_pallet_reports_unknown_number(admin_client):
    payload = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": "9999/1900"},
                             csrf_source_url="/invoice-br/new").get_json()
    assert payload["ok"] is False
    assert "Няма документ" in payload["error"]


def test_pull_pallet_reports_empty_code(admin_client):
    payload = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": "  "},
                             csrf_source_url="/invoice-br/new").get_json()
    assert payload["ok"] is False


# ---------------------------------------------------------------- Excel/PDF износ

def test_brazil_invoice_xlsx_export_has_sample_columns_and_computed_totals(admin_client):
    items = json.dumps([{
        "hs_code": "85389099", "po_no": "4700200362", "pos": "30",
        "net_weight": "4.51", "material_code": "GLBK400002P0012",
        "qty": "20", "unit_price": "13.66",
    }])
    resp = post_with_csrf(admin_client, "/invoice-br/new",
                          {"consignee_name": "ABB", "doc_date": "2026-08-07",
                           "items_json": items},
                          csrf_source_url="/invoice-br/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(admin_client.get("/doc/%s/export.xlsx" % doc_id).data))
    values = [c.value for row in wb.active.iter_rows() for c in row if c.value is not None]
    assert "Нето тегло, кг/бр" in values
    assert "Код на материала" in values
    assert "Обща цена, EUR" in values
    assert "273.2" in values, "изчислената обща цена на реда (20 × 13.66)"
    assert "90.2" in values, "изчисленото общо тегло на реда (4.51 × 20)"
    assert "07.08.2026" in values, "датата излиза във вида ДД.ММ.ГГГГ"


def test_norway_invoice_xlsx_export_has_no_weight_column(admin_client):
    items = json.dumps([{"description": "C-Profile", "pallet_no": "1",
                         "material_code": "X", "qty": "2", "unit_price": "3"}])
    resp = post_with_csrf(admin_client, "/invoice-no/new",
                          {"consignee_name": "ABB", "items_json": items},
                          csrf_source_url="/invoice-no/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(admin_client.get("/doc/%s/export.xlsx" % doc_id).data))
    values = [c.value for row in wb.active.iter_rows() for c in row if c.value is not None]
    assert "Палет №" in values
    assert "Описание на материала" in values
    assert "Нето тегло, кг/бр" not in values
    assert "6" in values, "изчислената обща цена на реда (2 × 3)"


def test_invoice_pdf_export_works_for_both_countries(admin_client):
    for url in ("/invoice-br/new", "/invoice-no/new"):
        resp = post_with_csrf(admin_client, url, {"consignee_name": "ABB"},
                              csrf_source_url=url, follow_redirects=False)
        doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]
        pdf = admin_client.get("/doc/%s/export.pdf" % doc_id)
        assert pdf.status_code == 200
        assert pdf.headers["Content-Type"] == "application/pdf"


def test_invoices_appear_in_the_documents_list(admin_client):
    post_with_csrf(admin_client, "/invoice-br/new", {"consignee_name": "Бразилски клиент"},
                   csrf_source_url="/invoice-br/new", follow_redirects=False)
    body = admin_client.get("/docs").data.decode()
    assert "Фактура за Бразилия" in body


def test_invoice_totals_match_the_sum_of_the_printed_row_totals():
    """Не беше в обхвата на самия патч: сумите под таблицата се трупаха от
    СУРОВИТЕ произведения, докато всеки ред се ОТПЕЧАТВА закръглен. При
    10 реда по 0.005 всеки ред излиза „0.01“ (видим сбор 0.10), а общата
    сума показваше „0.05“ — търговска фактура, която сама си противоречи и
    отива така към клиент и митница. Сумите вече се трупат от точно тези
    стойности, които се виждат на редовете."""
    from appcore import invoice_row_total, invoice_row_weight, invoice_totals

    for items in (
        [{"qty": "1", "unit_price": "0.005"}] * 10,
        [{"qty": "1", "unit_price": "0.125"}] * 3,
        [{"qty": "3", "unit_price": "10.005"}, {"qty": "7", "unit_price": "2.335"}],
        [{"qty": "7", "unit_price": "1.115", "net_weight": "0.0005"}] * 4,
    ):
        totals = invoice_totals(items)
        shown_price = sum(float(invoice_row_total(i)) for i in items if invoice_row_total(i))
        assert abs(shown_price - float(totals["price"])) < 1e-9, (
            "общата стойност (%s) не отговаря на сбора на показаните редове (%s)"
            % (totals["price"], shown_price)
        )
        shown_weight = [invoice_row_weight(i) for i in items]
        if any(shown_weight):
            total_w = sum(float(w) for w in shown_weight if w)
            assert abs(total_w - float(totals["weight"])) < 1e-9, (
                "общото тегло (%s) не отговаря на сбора на показаните редове (%s)"
                % (totals["weight"], total_w)
            )

# ---------------------------------------------------------------- банкови данни от настройките

def test_company_bank_details_load_into_the_invoice_form(admin_client):
    """Заявка: „във фирма изпращач добави IBAN-а на фирмата; да се зарежда
    във фактурите“."""
    post_with_csrf(admin_client, "/settings", {
        "sender_name": "BBS Bulgaria Ltd.",
        "sender_iban": "BG26BPBI817014Y2307201",
        "sender_swift": "BPBIBGSF",
        "sender_bank": "Postbank Gabrovo-Bulgaria",
    }, csrf_source_url="/settings", follow_redirects=True)

    for url in ("/invoice-br/new", "/invoice-no/new"):
        body = admin_client.get(url).data.decode()
        assert "BG26BPBI817014Y2307201" in body
        assert "BPBIBGSF" in body
        assert "Postbank Gabrovo-Bulgaria" in body


def test_invoice_bank_line_skips_missing_parts():
    from appcore import invoice_bank_line
    assert invoice_bank_line({"sender_iban": "BG1"}) == "IBAN : BG1"
    assert invoice_bank_line({"sender_iban": "BG1", "sender_swift": "SW"}) == \
        "IBAN : BG1    SWIFT : SW"
    assert invoice_bank_line({"sender_bank": "Банка"}) == "/ Банка /"
    assert invoice_bank_line({}) == ""
    assert invoice_bank_line(None) == ""


# ---------------------------------------------------------------- ръчен номер, без баркод

def test_invoice_number_is_taken_from_the_form_not_generated(admin_client):
    resp = post_with_csrf(admin_client, "/invoice-br/new",
                          {"consignee_name": "ABB", "invoice_number": "0000012955"},
                          csrf_source_url="/invoice-br/new", follow_redirects=False)
    body = admin_client.get(resp.headers["Location"]).data.decode()
    assert "Invoice Number: 0000012955" in body


def test_invoice_number_falls_back_to_generated_when_left_empty(admin_client):
    """Полето е задължително във формата, но ако все пак дойде празно,
    документът не бива да остане без номер изобщо."""
    resp = post_with_csrf(admin_client, "/invoice-br/new",
                          {"consignee_name": "ABB", "invoice_number": "   "},
                          csrf_source_url="/invoice-br/new", follow_redirects=False)
    body = admin_client.get(resp.headers["Location"]).data.decode()
    assert "Invoice Number:" in body
    assert "/2026" in body or "/20" in body


def test_invoices_have_no_barcode_on_the_printed_form(admin_client):
    """Заявка: „без баркод на фактурите“ — както в приложените образци."""
    resp = post_with_csrf(admin_client, "/invoice-br/new",
                          {"consignee_name": "ABB", "invoice_number": "BR-77"},
                          csrf_source_url="/invoice-br/new", follow_redirects=False)
    body = admin_client.get(resp.headers["Location"]).data.decode()
    assert "INVBR-" not in body, "вътрешният баркод не бива да се показва"
    # Иконите на страницата също са SVG, затова се хващаме за подписа на
    # САМИЯ баркод: barcode128.code128_svg винаги слага role="img" с
    # aria-label равен на кодирания текст (виж barcode128.py).
    assert 'role="img"' not in body, "на бланката на фактурата няма баркод изобщо"


def test_duplicate_invoice_number_warns_but_still_saves(admin_client):
    for _i in range(2):
        resp = post_with_csrf(admin_client, "/invoice-br/new",
                              {"consignee_name": "ABB", "invoice_number": "DUP-1"},
                              csrf_source_url="/invoice-br/new", follow_redirects=True)
    assert "вече има издаден документ с номер DUP-1" in resp.data.decode()
    assert admin_client.get("/invoices").data.decode().count("DUP-1") >= 2


def test_editing_an_invoice_can_correct_its_number(admin_client):
    """Ръчно въведеният номер трябва да може да се поправи — иначе сгрешен
    номер остава завинаги."""
    resp = post_with_csrf(admin_client, "/invoice-br/new",
                          {"consignee_name": "ABB", "invoice_number": "ГРЕШЕН-1"},
                          csrf_source_url="/invoice-br/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]

    post_with_csrf(admin_client, "/doc/%s/edit" % doc_id,
                   {"consignee_name": "ABB", "invoice_number": "ВЕРЕН-2"},
                   csrf_source_url="/doc/%s/edit" % doc_id, follow_redirects=False)
    body = admin_client.get("/doc/%s" % doc_id).data.decode()
    assert "Invoice Number: ВЕРЕН-2" in body
    assert "ГРЕШЕН-1" not in body
    assert "ВЕРЕН-2" in admin_client.get("/invoices").data.decode()


def test_invoice_preview_shows_the_typed_number(admin_client):
    resp = post_with_csrf(admin_client, "/invoice-br/preview",
                          {"consignee_name": "ABB", "invoice_number": "ПРЕГЛЕД-9"},
                          csrf_source_url="/invoice-br/new", follow_redirects=False)
    body = admin_client.get(resp.headers["Location"]).data.decode()
    assert "ПРЕГЛЕД-9" in body


# ---------------------------------------------------------------- условие за доставка

def test_terms_of_delivery_offers_fca_and_dap_with_fca_default(admin_client):
    for url in ("/invoice-br/new", "/invoice-no/new"):
        body = admin_client.get(url).data.decode()
        block = body.split('name="terms_delivery"')[1].split("</select>")[0]
        assert '<option value="FCA" selected>' in block
        assert '<option value="DAP">' in block


# ---------------------------------------------------------------- адресна книга за фактури

def _add_invoice_client(client, **over):
    data = {
        "name": "ABB Бразилия — Sorocaba",
        "delivery_name": "ABB ELETRIFICACAO LTDA",
        "delivery_address": "Rod.Sen.Jose Ermirio de Moraes, KM 11\nSorocaba - SP\nBrasil",
        "delivery_phone": "+55 11 97613-8155",
        "billing_name": "ABB ELETRIFICACAO LTDA - CNPJ 33.449.988/0001-20",
        "billing_address": "Fakturamottak\n18087-125 - Sorocaba - SP",
        "billing_phone": "+55 15 3330-6465",
        "notes": "",
    }
    data.update(over)
    return post_with_csrf(client, "/invoices/clients/new", data,
                          csrf_source_url="/invoices/clients/new", follow_redirects=False)


def test_invoice_address_book_is_empty_at_first(admin_client):
    assert "Адресната книга за фактури е празна" in \
        admin_client.get("/invoices/clients").data.decode()


def test_invoice_address_book_stores_billing_and_delivery_separately(admin_client):
    assert _add_invoice_client(admin_client).status_code == 302
    body = admin_client.get("/invoices/clients").data.decode()
    assert "ABB ELETRIFICACAO LTDA" in body
    assert "CNPJ 33.449.988/0001-20" in body
    assert "Sorocaba - SP" in body


def test_invoice_address_book_entries_reach_the_invoice_form(admin_client):
    _add_invoice_client(admin_client)
    body = admin_client.get("/invoice-br/new").data.decode()
    assert "invoice-client-select" in body
    assert "ABB Бразилия — Sorocaba" in body
    assert "ABB ELETRIFICACAO LTDA" in body, "данните се вграждат за попълване от JS"


def test_invoice_address_book_entry_can_be_edited(admin_client):
    _add_invoice_client(admin_client)
    con = __import__("db").get_db()
    entry_id = con.execute("SELECT id FROM invoice_clients").fetchone()["id"]
    con.close()

    post_with_csrf(admin_client, "/invoices/clients/%d/edit" % entry_id,
                   {"name": "Преименуван", "delivery_name": "Нов получател",
                    "billing_name": "Нов платец"},
                   csrf_source_url="/invoices/clients/%d/edit" % entry_id,
                   follow_redirects=False)
    body = admin_client.get("/invoices/clients").data.decode()
    assert "Преименуван" in body
    assert "Нов получател" in body


def test_invoice_address_book_delete_requires_admin(employee_client, admin_client):
    _add_invoice_client(admin_client)
    con = __import__("db").get_db()
    entry_id = con.execute("SELECT id FROM invoice_clients").fetchone()["id"]
    con.close()
    resp = post_with_csrf(employee_client, "/invoices/clients/%d/delete" % entry_id, {},
                          csrf_source_url="/invoices/clients", follow_redirects=False)
    assert resp.status_code in (302, 403)
    assert "ABB Бразилия" in admin_client.get("/invoices/clients").data.decode()


def test_invoice_address_book_is_separate_from_the_general_one(admin_client):
    """Двете адресни книги не се смесват (изрично избрано от потребителя)."""
    _add_invoice_client(admin_client)
    post_with_csrf(admin_client, "/clients/new", {"name": "Общ клиент ЕООД"},
                   csrf_source_url="/clients/new", follow_redirects=False)
    assert "Общ клиент ЕООД" not in admin_client.get("/invoices/clients").data.decode()
    assert "ABB Бразилия" not in admin_client.get("/clients").data.decode()


# ---------------------------------------------------------------- Excel импорт на редове

def _orders_xlsx(rows, headers=("Order No", "Pos", "Reference", "Reference Desc",
                                "Open Qty", "")):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_invoice_excel_import_loads_rows_with_weight_from_catalog(admin_client):
    """Заявка: „зареждането на материалите във фактурата могат да се
    зареждат и от Excel файл, както е в палетната карта“ — СЪЩИЯТ формат."""
    _load_catalog(admin_client)
    payload = post_with_csrf(
        admin_client, "/invoice/import-items",
        {"excel_file": (_orders_xlsx([
            ("4700200362", "30", "GLBK400002P0012", "C-Profile", 20, 1),
            ("4700200362", "40", "GLBK400001P0200", "C-Profile 2", 400, 1),
        ]), "poruchki.xlsx")},
        csrf_source_url="/invoice-br/new",
        content_type="multipart/form-data").get_json()

    assert payload["ok"] is True
    assert payload["count"] == 2
    assert payload["matched"] == 2
    assert [r["material_code"] for r in payload["rows"]] == \
        ["GLBK400002P0012", "GLBK400001P0200"]
    assert [r["net_weight"] for r in payload["rows"]] == ["2.21", "0.383"]
    assert payload["rows"][0]["po_no"] == "4700200362"
    assert payload["rows"][0]["pos"] == "30"
    assert payload["rows"][0]["qty"] == "20"


def test_invoice_excel_import_reads_a_price_column_when_present(admin_client):
    """Изрично избрано от потребителя: ако файлът има колона с цена, тя се
    зарежда; иначе остава празна за ръчно въвеждане."""
    _load_catalog(admin_client)
    payload = post_with_csrf(
        admin_client, "/invoice/import-items",
        {"excel_file": (_orders_xlsx(
            [("PO-1", "10", "GLBK400002P0012", "C-Profile", 20, 13.66)],
            headers=("Order No", "Pos", "Reference", "Reference Desc",
                     "Open Qty", "Unit Price")), "s_ceni.xlsx")},
        csrf_source_url="/invoice-br/new",
        content_type="multipart/form-data").get_json()
    assert payload["rows"][0]["unit_price"] == "13.66"


def test_invoice_excel_import_leaves_price_empty_when_column_is_missing(admin_client):
    _load_catalog(admin_client)
    payload = post_with_csrf(
        admin_client, "/invoice/import-items",
        {"excel_file": (_orders_xlsx(
            [("PO-1", "10", "GLBK400002P0012", "C-Profile", 20, 1)]), "bez_ceni.xlsx")},
        csrf_source_url="/invoice-br/new",
        content_type="multipart/form-data").get_json()
    assert payload["rows"][0]["unit_price"] == ""


def test_invoice_excel_import_rejects_unrecognised_file(admin_client):
    payload = post_with_csrf(
        admin_client, "/invoice/import-items",
        {"excel_file": (_orders_xlsx([("а", "б")], headers=("К1", "К2")), "грешен.xlsx")},
        csrf_source_url="/invoice-br/new",
        content_type="multipart/form-data").get_json()
    assert payload["ok"] is False
    assert "разпознаваеми колони" in payload["error"]


def test_invoice_excel_import_reports_missing_file(admin_client):
    payload = post_with_csrf(admin_client, "/invoice/import-items", {},
                             csrf_source_url="/invoice-br/new",
                             content_type="multipart/form-data").get_json()
    assert payload["ok"] is False


def test_invoice_excel_import_skips_empty_trailing_rows(admin_client):
    _load_catalog(admin_client)
    payload = post_with_csrf(
        admin_client, "/invoice/import-items",
        {"excel_file": (_orders_xlsx([
            ("PO-1", "10", "GLBK400002P0012", "C-Profile", 20, 1),
            ("", "", "", "", "", ""),
        ]), "s_prazni.xlsx")},
        csrf_source_url="/invoice-br/new",
        content_type="multipart/form-data").get_json()
    assert payload["count"] == 1


# ---------------------------------------------------------------- отделен списък с фактури

def test_issued_invoices_appear_only_in_the_invoices_list(admin_client):
    """Заявка: „в раздела Фактури да има издадени документи и само там да
    се появяват издадените фактури; фактури да не отиват в [Всички
    документи]“."""
    # follow_redirects=True консумира flash съобщението („Фактура … е
    # издадена“), което иначе би се показало на СЛЕДВАЩАТА страница и би
    # съдържало номера/клиента там, където точно проверяваме, че ги няма.
    post_with_csrf(admin_client, "/invoice-br/new",
                   {"consignee_name": "Бразилски клиент", "invoice_number": "САМО-ТУК-1"},
                   csrf_source_url="/invoice-br/new", follow_redirects=True)

    docs_body = admin_client.get("/docs").data.decode()
    assert "САМО-ТУК-1" not in docs_body
    assert "Бразилски клиент" not in docs_body
    assert "САМО-ТУК-1" in admin_client.get("/invoices").data.decode()


def test_documents_list_type_filter_cannot_show_invoices(admin_client):
    """Дори при изрично ?type=invoice_br общият списък не показва фактури."""
    post_with_csrf(admin_client, "/invoice-br/new",
                   {"consignee_name": "Скрит клиент", "invoice_number": "СКРИТА-1"},
                   csrf_source_url="/invoice-br/new", follow_redirects=True)
    body = admin_client.get("/docs?type=invoice_br").data.decode()
    assert "СКРИТА-1" not in body
    assert "Скрит клиент" not in body


def test_invoices_are_excluded_from_dashboard(admin_client):
    """Заявка: „и от таблото/историята на клиента“."""
    post_with_csrf(admin_client, "/invoice-br/new",
                   {"consignee_name": "Табло Клиент", "invoice_number": "ТАБЛО-1"},
                   csrf_source_url="/invoice-br/new", follow_redirects=True)
    body = admin_client.get("/").data.decode()
    assert "ТАБЛО-1" not in body
    assert "Табло Клиент" not in body
    assert "Статистика за текущия месец" in body, "таблото се е заредило нормално"


def test_invoices_are_excluded_from_client_history(admin_client):
    post_with_csrf(admin_client, "/clients/new", {"name": "История ЕООД"},
                   csrf_source_url="/clients/new", follow_redirects=False)
    con = __import__("db").get_db()
    client_id = con.execute("SELECT id FROM clients WHERE name = ?",
                            ("История ЕООД",)).fetchone()["id"]
    con.close()

    post_with_csrf(admin_client, "/invoice-br/new",
                   {"consignee_name": "История ЕООД", "invoice_number": "ИСТ-1"},
                   csrf_source_url="/invoice-br/new", follow_redirects=True)
    body = admin_client.get("/clients/%d/edit" % client_id).data.decode()
    assert "ИСТ-1" not in body
    assert "Все още няма издадени документи" in body


def test_invoices_list_filters_by_type_and_search(admin_client):
    post_with_csrf(admin_client, "/invoice-br/new",
                   {"consignee_name": "Бразилия ООД", "invoice_number": "BR-100"},
                   csrf_source_url="/invoice-br/new", follow_redirects=True)
    post_with_csrf(admin_client, "/invoice-no/new",
                   {"consignee_name": "Норвегия АС", "invoice_number": "NO-200"},
                   csrf_source_url="/invoice-no/new", follow_redirects=True)

    only_br = admin_client.get("/invoices?type=invoice_br").data.decode()
    assert "BR-100" in only_br
    assert "NO-200" not in only_br

    searched = admin_client.get("/invoices?q=Норвегия").data.decode()
    assert "NO-200" in searched
    assert "BR-100" not in searched


def test_invoices_list_has_an_edit_button(admin_client):
    """Заявка: „да има бутон за редакция на фактура“."""
    resp = post_with_csrf(admin_client, "/invoice-br/new",
                          {"consignee_name": "ABB", "invoice_number": "РЕД-1"},
                          csrf_source_url="/invoice-br/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]
    body = admin_client.get("/invoices").data.decode()
    assert "/doc/%s/edit" % doc_id in body
    assert "Редактирай" in body


def test_deleting_an_invoice_returns_to_the_invoices_list(admin_client):
    resp = post_with_csrf(admin_client, "/invoice-br/new",
                          {"consignee_name": "ABB", "invoice_number": "ИЗТР-1"},
                          csrf_source_url="/invoice-br/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]
    deleted = post_with_csrf(admin_client, "/doc/%s/delete" % doc_id, {},
                             csrf_source_url="/invoices", follow_redirects=False)
    assert deleted.headers["Location"].endswith("/invoices")


def test_other_documents_still_appear_in_the_general_list(admin_client):
    """Изключването на фактурите не бива да е засегнало останалите типове."""
    post_with_csrf(admin_client, "/cmr/new", {"consignee_name": "Обикновен клиент"},
                   csrf_source_url="/cmr/new", follow_redirects=False)
    body = admin_client.get("/docs").data.decode()
    assert "Обикновен клиент" in body
    assert "ЧМР товарителница" in body


# ---------------------------------------------------------------- суфикси след тире
# Заявка: „референция 1TGB110025P1204-RAS да се вмъкне като 1TGB110025P1204,
# за да може да се заредят автоматично килограмите от файла“.

def test_pull_pallet_strips_dash_suffix_and_inserts_the_catalog_code(admin_client):
    _load_catalog(admin_client)
    doc_id = _issue_pallet_with_orders(
        admin_client, [("PO-1", "10", "GLBK400002P0012-RAS", "C-Profile", "20")])
    number = _pallet_number(admin_client, doc_id)

    payload = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": number},
                             csrf_source_url="/invoice-br/new").get_json()
    row = payload["rows"][0]
    assert row["material_code"] == "GLBK400002P0012", \
        "суфиксът -RAS се маха и се вмъква кодът от справочника"
    assert row["net_weight"] == "2.21", "килограмите идват автоматично"
    assert payload["matched"] == 1


def test_pull_pallet_keeps_unknown_suffixed_reference_untouched(admin_client):
    """Код, който го няма в справочника нито цял, нито орязан, се вмъква
    ТОЧНО както е в палетната карта — не гадаем и не режем на сляпо."""
    _load_catalog(admin_client)
    doc_id = _issue_pallet_with_orders(
        admin_client, [("PO-1", "10", "НЕПОЗНАТ-RAS", "Нещо", "5")])
    number = _pallet_number(admin_client, doc_id)

    row = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": number},
                         csrf_source_url="/invoice-br/new").get_json()["rows"][0]
    assert row["material_code"] == "НЕПОЗНАТ-RAS"
    assert row["net_weight"] == ""


def test_invoice_excel_import_strips_dash_suffix_too(admin_client):
    _load_catalog(admin_client)
    payload = post_with_csrf(
        admin_client, "/invoice/import-items",
        {"excel_file": (_orders_xlsx([
            ("PO-1", "10", "GLBK400002P0012-RAS", "C-Profile", 20, 1),
        ]), "s_sufiks.xlsx")},
        csrf_source_url="/invoice-br/new",
        content_type="multipart/form-data").get_json()
    assert payload["rows"][0]["material_code"] == "GLBK400002P0012"
    assert payload["rows"][0]["net_weight"] == "2.21"
    assert payload["matched"] == 1


# ---------------------------------------------------------------- HS code по подразбиране

def test_invoice_tables_carry_the_default_hs_code_for_new_rows(admin_client):
    """Заявка: „в фактурите по подразбиране винаги да се поставя
    автоматично HS code 85389099“ — таблицата носи data-row-defaults, от
    който app.js попълва HS кода и на началния празен ред, и на всеки
    ред от „+ Добави ред“ (виж initItemsTable). Проверено и в реален
    браузър (tests/test_e2e_smoke.py)."""
    for url in ("/invoice-br/new", "/invoice-no/new"):
        body = admin_client.get(url).data.decode()
        assert 'data-row-defaults=\'{"hs_code": "85389099"}\'' in body, url


# ---------------------------------------------------------------- една поръчка = една фактура
# Заявка: „във фактури един номер на поръчка да бъде на една фактура
# (пример: 4700201619 една фактура за всички материали с този номер
# поръчка, 4700223566 друга фактура)“. При източник с няколко поръчки
# зареждането първо пита коя (choose_po), зарежда САМО нейните редове и
# казва кои остават за отделни фактури; при издаване на смесена фактура
# излиза предупреждение.

def test_pull_pallet_with_multiple_orders_asks_which_one_to_load(admin_client):
    _load_catalog(admin_client)
    doc_id = _issue_pallet_with_orders(admin_client, [
        ("4700201619", "10", "GLBK400002P0012", "C-Profile", "20"),
        ("4700201619", "20", "GLBK400001P0200", "C-Profile 2", "5"),
        ("4700223566", "10", "1TFL151621P0550", "Секция", "7"),
    ])
    number = _pallet_number(admin_client, doc_id)

    payload = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": number},
                             csrf_source_url="/invoice-br/new").get_json()
    assert payload["ok"] is True
    assert payload["choose_po"] is True
    assert "rows" not in payload, "нищо не се зарежда, докато не се избере поръчка"
    assert payload["pos"] == [{"po_no": "4700201619", "count": 2},
                              {"po_no": "4700223566", "count": 1}]


def test_pull_pallet_with_chosen_order_loads_only_its_rows(admin_client):
    _load_catalog(admin_client)
    doc_id = _issue_pallet_with_orders(admin_client, [
        ("4700201619", "10", "GLBK400002P0012", "C-Profile", "20"),
        ("4700223566", "10", "1TFL151621P0550", "Секция", "7"),
        ("4700201619", "20", "GLBK400001P0200", "C-Profile 2", "5"),
    ])
    number = _pallet_number(admin_client, doc_id)

    payload = post_with_csrf(admin_client, "/invoice/pull-pallet",
                             {"code": number, "po_no": "4700201619"},
                             csrf_source_url="/invoice-br/new").get_json()
    assert payload["count"] == 2
    assert all(r["po_no"] == "4700201619" for r in payload["rows"])
    assert payload["loaded_po"] == "4700201619"
    assert payload["remaining"] == [{"po_no": "4700223566", "count": 1}]
    assert payload["matched"] == 2, "matched се брои върху ЗАРЕДЕНИТЕ редове"


def test_pull_pallet_with_single_order_loads_directly_as_before(admin_client):
    """Една поръчка в картата — никакъв избор, поведението е като досега."""
    _load_catalog(admin_client)
    doc_id = _issue_pallet_with_orders(admin_client, [
        ("4700201619", "10", "GLBK400002P0012", "C-Profile", "20"),
        ("4700201619", "20", "GLBK400001P0200", "C-Profile 2", "5"),
    ])
    number = _pallet_number(admin_client, doc_id)
    payload = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": number},
                             csrf_source_url="/invoice-br/new").get_json()
    assert "choose_po" not in payload
    assert payload["count"] == 2


def test_pull_pallet_rows_without_order_number_form_their_own_group(admin_client):
    """Редове без P.O NO не се губят и не се разпределят мълчаливо — те са
    собствена група „(редове без поръчка №)“ в избора."""
    _load_catalog(admin_client)
    doc_id = _issue_pallet_with_orders(admin_client, [
        ("4700201619", "10", "GLBK400002P0012", "C-Profile", "20"),
        ("", "20", "GLBK400001P0200", "Без поръчка", "5"),
    ])
    number = _pallet_number(admin_client, doc_id)

    payload = post_with_csrf(admin_client, "/invoice/pull-pallet", {"code": number},
                             csrf_source_url="/invoice-br/new").get_json()
    assert payload["choose_po"] is True
    assert {"po_no": "", "count": 1} in payload["pos"]

    blank = post_with_csrf(admin_client, "/invoice/pull-pallet",
                           {"code": number, "po_no": ""},
                           csrf_source_url="/invoice-br/new").get_json()
    assert blank["count"] == 1
    assert blank["rows"][0]["material_code"] == "GLBK400001P0200"


def test_invoice_excel_import_with_multiple_orders_asks_and_filters(admin_client):
    _load_catalog(admin_client)
    xlsx_rows = [
        ("4700201619", "10", "GLBK400002P0012", "C-Profile", 20, 1),
        ("4700223566", "10", "1TFL151621P0550", "Секция", 7, 1),
    ]
    ask = post_with_csrf(
        admin_client, "/invoice/import-items",
        {"excel_file": (_orders_xlsx(xlsx_rows), "dve_porachki.xlsx")},
        csrf_source_url="/invoice-br/new",
        content_type="multipart/form-data").get_json()
    assert ask["choose_po"] is True
    assert [p["po_no"] for p in ask["pos"]] == ["4700201619", "4700223566"]

    picked = post_with_csrf(
        admin_client, "/invoice/import-items",
        {"excel_file": (_orders_xlsx(xlsx_rows), "dve_porachki.xlsx"),
         "po_no": "4700223566"},
        csrf_source_url="/invoice-br/new",
        content_type="multipart/form-data").get_json()
    assert picked["count"] == 1
    assert picked["rows"][0]["po_no"] == "4700223566"
    assert picked["remaining"] == [{"po_no": "4700201619", "count": 1}]


def test_issuing_an_invoice_with_mixed_orders_warns_but_still_saves(admin_client):
    items = json.dumps([
        {"material_code": "A", "po_no": "4700201619", "qty": "1", "unit_price": "2"},
        {"material_code": "B", "po_no": "4700223566", "qty": "1", "unit_price": "3"},
    ])
    resp = post_with_csrf(admin_client, "/invoice-br/new",
                          {"consignee_name": "ABB", "invoice_number": "СМЕС-1",
                           "items_json": items},
                          csrf_source_url="/invoice-br/new", follow_redirects=True)
    body = resp.data.decode()
    assert "2 различни поръчки" in body
    assert "4700201619" in body and "4700223566" in body
    assert "СМЕС-1" in admin_client.get("/invoices").data.decode(), \
        "предупреждение, не забрана — фактурата все пак се издава"


def test_issuing_an_invoice_with_one_order_does_not_warn(admin_client):
    items = json.dumps([
        {"material_code": "A", "po_no": "4700201619", "qty": "1", "unit_price": "2"},
        {"material_code": "B", "po_no": "4700201619", "qty": "2", "unit_price": "3"},
        {"material_code": "C", "po_no": "", "qty": "1", "unit_price": "1"},
    ])
    resp = post_with_csrf(admin_client, "/invoice-br/new",
                          {"consignee_name": "ABB", "invoice_number": "ЕДНА-1",
                           "items_json": items},
                          csrf_source_url="/invoice-br/new", follow_redirects=True)
    assert "различни поръчки" not in resp.data.decode()
