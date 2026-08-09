# -*- coding: utf-8 -*-
"""Тестове за подобренията на опаковъчния лист по образеца PL.xlsx —
заявка: „подобри опаковъчния лист, като използваш прикачения; добави лице
за контакти да се въвеждат в адресната книга и да се вмъква автоматично
както за клиент, и за изпращач“.

Обхванато тук:
  - „Фирма изпращач“ (Настройки) има ново поле „Лице за контакт“
    (sender_contact), което се зарежда автоматично в блока Изпращач на
    формата (с резервен вариант МОЛ/sender_person, ако е празно), заедно
    с телефона и имейла на фирмата.
  - Получателят има полета лице за контакт/телефон/имейл, попълвани от
    адресната книга (клиентската картичка вече пази „Лице за контакт“) —
    самото JS попълване е покрито с e2e тест в реален браузър.
  - Печатната бланка показва Contact person / Phone No / e-mail и за
    двете страни (само попълнените редове), по образеца.
  - Колоните на таблицата следват образеца: Вид опаковка / Kind of colli
    е ПЪРВА, после Описание на материала.
  - FCA / Truck / 85389099 са предварително попълнени (редактируеми)
    стойности по подразбиране — каквито стоят фиксирани в образеца."""
import json

from conftest import post_with_csrf


def _save_sender_settings(admin_client, **over):
    data = {
        "sender_name": "BBS Bulgaria EOOD",
        "sender_address": "47 Georgi Dimitrov Str.",
        "sender_city": "5334 Yavorets",
        "sender_phone": "+359 888 111 222",
        "sender_email": "salesbg@bbsmetal.com.tr",
        "sender_person": "МОЛ Лице",
        "sender_contact": "Plamen Hristov",
    }
    data.update(over)
    return post_with_csrf(admin_client, "/settings", data,
                          csrf_source_url="/settings", follow_redirects=False)


def _add_client_with_contact(admin_client):
    return post_with_csrf(admin_client, "/clients/new", {
        "name": "ABB Norway AS", "address": "Snarøyveien 30", "city": "Fornebu",
        "postcode": "1360", "country": "Norway",
        "contact": "Kari Nordmann", "phone": "+47 22 87 20 00",
        "email": "kari@abb.no",
    }, csrf_source_url="/clients/new", follow_redirects=False)


# ---------------------------------------------------------------- настройки

def test_settings_page_has_sender_contact_field_and_saves_it(admin_client):
    assert 'name="sender_contact"' in admin_client.get("/settings").data.decode()
    assert _save_sender_settings(admin_client).status_code == 302
    body = admin_client.get("/settings").data.decode()
    assert 'value="Plamen Hristov"' in body


# ---------------------------------------------------------------- формата

def test_packing_form_prefills_sender_contact_phone_email_from_settings(admin_client):
    _save_sender_settings(admin_client)
    body = admin_client.get("/packing/new").data.decode()
    block = body.split('name="sender_contact"')[1].split(">")[0]
    assert 'value="Plamen Hristov"' in block
    assert 'name="sender_phone"' in body
    assert 'name="sender_email"' in body
    assert "+359 888 111 222" in body
    assert "salesbg@bbsmetal.com.tr" in body


def test_packing_form_sender_contact_falls_back_to_mol_when_empty(admin_client):
    """Празно „Лице за контакт“ в Настройки → зарежда се МОЛ, вместо
    полето да остане празно (за инсталации, попълнили само МОЛ)."""
    _save_sender_settings(admin_client, sender_contact="")
    body = admin_client.get("/packing/new").data.decode()
    block = body.split('name="sender_contact"')[1].split(">")[0]
    assert 'value="МОЛ Лице"' in block


def test_packing_form_has_receiver_contact_fields(admin_client):
    body = admin_client.get("/packing/new").data.decode()
    for name in ("receiver_contact", "receiver_phone", "receiver_email"):
        assert 'name="%s"' % name in body


def test_packing_form_defaults_match_the_sample(admin_client):
    """FCA / Truck / 85389099 — фиксираните стойности от образеца PL.xlsx,
    предварително попълнени (но редактируеми) в нова форма."""
    body = admin_client.get("/packing/new").data.decode()
    assert 'name="terms_delivery" value="FCA"' in body
    assert 'name="transport_type" value="Truck"' in body
    assert 'name="hs_code" value="85389099"' in body


def test_client_json_for_autofill_includes_contact_phone_and_email(admin_client):
    """Адресната книга подава лице за контакт/телефон/имейл на JS
    попълването (bindClientSelect) — данните трябва да присъстват в
    data-clients на формата."""
    _add_client_with_contact(admin_client)
    body = admin_client.get("/packing/new").data.decode()
    assert "Kari Nordmann" in body
    assert "+47 22 87 20 00" in body
    assert "kari@abb.no" in body


# ---------------------------------------------------------------- печатната бланка

def _issue_packing_with_contacts(admin_client):
    items = json.dumps([{
        "packing": "carton / pallet", "description": "HV Switchgear cabinets material",
        "qty": "2", "length": "1200", "width": "800", "height": "900",
        "volume": "0.864", "net": "150", "gross": "180",
    }])
    resp = post_with_csrf(admin_client, "/packing/new", {
        "sender_name": "BBS Bulgaria EOOD", "sender_address": "47 Georgi Dimitrov Str.",
        "sender_contact": "Plamen Hristov", "sender_phone": "+359 888 111 222",
        "sender_email": "salesbg@bbsmetal.com.tr",
        "receiver_name": "ABB Norway AS", "receiver_contact": "Kari Nordmann",
        "receiver_phone": "+47 22 87 20 00", "receiver_email": "kari@abb.no",
        "items_json": items,
    }, csrf_source_url="/packing/new", follow_redirects=False)
    assert resp.status_code == 302
    return resp.headers["Location"]


def test_print_shows_contact_person_for_both_parties(admin_client):
    body = admin_client.get(_issue_packing_with_contacts(admin_client)).data.decode()
    assert "Contact person: <b>Plamen Hristov</b>" in body
    assert "Contact person: <b>Kari Nordmann</b>" in body
    assert "Phone No: <b>+359 888 111 222</b>" in body
    assert "Phone No: <b>+47 22 87 20 00</b>" in body
    assert "e-mail: <b>salesbg@bbsmetal.com.tr</b>" in body
    assert "e-mail: <b>kari@abb.no</b>" in body


def test_print_hides_contact_rows_when_not_filled(admin_client):
    """Празните контактни полета не оставят висящи етикети на бланката."""
    resp = post_with_csrf(admin_client, "/packing/new", {
        "sender_name": "Изпращач", "receiver_name": "Клиент",
        "sender_contact": "", "sender_phone": "", "sender_email": "",
        "receiver_contact": "", "receiver_phone": "", "receiver_email": "",
    }, csrf_source_url="/packing/new", follow_redirects=False)
    body = admin_client.get(resp.headers["Location"]).data.decode()
    assert "Contact person:" not in body
    assert "Phone No:" not in body


def test_print_goods_table_has_kind_of_colli_first_like_the_sample(admin_client):
    body = admin_client.get(_issue_packing_with_contacts(admin_client)).data.decode()
    head = body.split('<table class="goods">')[1].split("</tr>")[0]
    packing_pos = head.index("Kind of colli")
    description_pos = head.index("Material Description")
    qty_pos = head.index("Qty, pcs")
    assert packing_pos < description_pos < qty_pos
    # Стойността на реда също излиза в новия ред: опаковка преди описание.
    row = body.split("</table>")[0]
    assert row.index("carton / pallet") < row.index("HV Switchgear cabinets material")


# ---------------------------------------------------------------- Excel износ

def test_xlsx_export_includes_contact_fields_for_both_parties(admin_client):
    import io
    from openpyxl import load_workbook

    doc_url = _issue_packing_with_contacts(admin_client)
    doc_id = doc_url.rstrip("/").split("/")[-1]
    resp = admin_client.get("/doc/%s/export.xlsx" % doc_id)
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.data))
    cells = [str(c.value) for ws in wb.worksheets for row in ws.iter_rows() for c in row
             if c.value is not None]
    text = "\n".join(cells)
    assert "Plamen Hristov" in text
    assert "Kari Nordmann" in text
    assert "Лице за контакт (изпращач)" in text
    assert "Лице за контакт (получател)" in text
    # Колоната „Вид опаковка“ е преди „Описание на материала“ в износа.
    assert cells.index("Вид опаковка") < cells.index("Описание на материала")
