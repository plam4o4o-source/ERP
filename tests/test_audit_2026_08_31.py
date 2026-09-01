# -*- coding: utf-8 -*-
"""Регресионни тестове за осмия одит (31.08.2026) — находки №1–№20.

Всяка функция тук е ЗАКЛЮЧВАЩА за конкретна находка: преди поправката
пада, след нея минава. Проверено е и в двете посоки с реално изпълнение
(`git stash` на съответния файл), не само по прочит.

Находки №2, №3, №9 и №16 живеят в tests/test_e2e_smoke.py — те са
изцяло браузърни (клавиш Enter, HTML5 валидация, реално пречупване на
текст върху A4 страница) и Python HTTP клиент не може да ги възпроизведе.
"""
import io
import json
import os
import subprocess
import sys
import tempfile
import textwrap
import threading
import zipfile
from datetime import date

import pytest
from werkzeug.security import generate_password_hash

from conftest import post_with_csrf

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


# --------------------------------------------------------------- №1 + №13
def _add_admin(db_module, username):
    con = db_module.get_db()
    uid = con.execute(
        "INSERT INTO users (username, password_hash, full_name, role, active,"
        " must_change_password) VALUES (?, ?, ?, 'admin', 1, 0)",
        (username, generate_password_hash("test-password-123"), username)).lastrowid
    con.commit()
    con.close()
    return uid


def _login(flask_app, username):
    client = flask_app.test_client()
    post_with_csrf(client, "/login",
                   {"username": username, "password": "test-password-123"},
                   csrf_source_url="/login")
    return client


def test_two_admins_cannot_deactivate_each_other_into_zero_admins(
        flask_app, db_module):
    """Находка №1 (критична): проверката „не съм аз“ пазеше само от
    самозаключване, но не и КОНКУРЕНТНО. Двама админа, деактивиращи се
    взаимно едновременно, оставяха НУЛА активни администратора — а
    db.init_db засява „admin“ само при ПРАЗНА таблица users, значи
    възстановяването е ръчна редакция на .db файла."""
    # Изчистваме подразбиращия се „admin“, за да е сцената точно две лица.
    con = db_module.get_db()
    con.execute("DELETE FROM users")
    con.commit()
    con.close()
    a_id, b_id = _add_admin(db_module, "admin_a"), _add_admin(db_module, "admin_b")
    a_client, b_client = _login(flask_app, "admin_a"), _login(flask_app, "admin_b")

    barrier = threading.Barrier(2)

    def toggle(client, target_id):
        barrier.wait()
        post_with_csrf(client, "/admin/users/%d/toggle" % target_id,
                       {"expected_active": "1"}, csrf_source_url="/admin/users")

    threads = [threading.Thread(target=toggle, args=(a_client, b_id)),
               threading.Thread(target=toggle, args=(b_client, a_id))]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    con = db_module.get_db()
    active_admins = con.execute(
        "SELECT COUNT(*) AS c FROM users WHERE role = 'admin' AND active = 1"
    ).fetchone()["c"]
    con.close()
    assert active_admins >= 1, (
        "находка №1: двамата администратори се заключиха взаимно — останаха "
        "0 активни администратора")


def test_double_click_on_deactivate_does_not_cancel_itself(flask_app, db_module):
    """Находка №13: `active = 1 - active` без очаквано състояние. Два
    еднакви POST-а (двойно щракване, повторно изпращане по бавна мрежа)
    обръщаха флага два пъти и „Деактивирай“ тихо се самоотменяше —
    уволненият служител запазваше достъп."""
    _add_admin(db_module, "admin_c")
    admin = _login(flask_app, "admin_c")
    con = db_module.get_db()
    uid = con.execute(
        "INSERT INTO users (username, password_hash, full_name, role, active,"
        " must_change_password) VALUES ('sluzhitel', 'x', 'С', 'employee', 1, 0)"
    ).lastrowid
    con.commit()
    con.close()

    for _ in range(2):
        post_with_csrf(admin, "/admin/users/%d/toggle" % uid,
                       {"expected_active": "1"}, csrf_source_url="/admin/users")

    con = db_module.get_db()
    active = con.execute("SELECT active FROM users WHERE id = ?", (uid,)).fetchone()["active"]
    con.close()
    assert active == 0, (
        "находка №13: второто щракване върна служителя активен — "
        "деактивирането се самоотмени")


# --------------------------------------------------------------------- №4
def test_duplicate_invoice_number_on_new_keeps_the_entered_data(admin_client):
    """Находка №4 (критична): `IntegrityError` при зает номер изхвърляше
    оператора на празна форма — целият въведен документ се губеше.
    Сега въведеното се пази и формата се възстановява през `?restore=`."""
    post_with_csrf(admin_client, "/invoice-br/new",
                   {"consignee_name": "ABB", "invoice_number": "Ф-1"},
                   csrf_source_url="/invoice-br/new", follow_redirects=True)
    con_data = {"consignee_name": "Съвсем друг клиент", "invoice_number": "Ф-1"}
    resp = post_with_csrf(admin_client, "/invoice-br/new", con_data,
                          csrf_source_url="/invoice-br/new", follow_redirects=False)
    location = resp.headers.get("Location", "")
    assert "restore=" in location, (
        "находка №4: при зает номер няма токен за възстановяване — "
        "въведеното се губи (Location=%r)" % location)
    body = admin_client.get(location).data.decode()
    assert "Съвсем друг клиент" in body, (
        "находка №4: възстановената форма не съдържа въведеното")


# --------------------------------------------------------------------- №5
def test_negative_net_gross_and_volume_are_warned_about(admin_client):
    """Находка №5 (критична): предупреждението за отрицателни стойности
    гледаше само `qty`/`unit_price` — отрицателно нето/бруто/обем минаваше
    мълчаливо, макар appcore да ги изхвърля от сборовете."""
    items = [{"packing": "кашон", "description": "стока", "qty": "1",
              "net": "-5", "gross": "-6", "volume": "-1"}]
    resp = post_with_csrf(admin_client, "/packing/new", {
        "receiver_name": "Получател",
        "items_json": json.dumps(items, ensure_ascii=False),
    }, csrf_source_url="/packing/new", follow_redirects=True)
    body = resp.data.decode()
    assert "отрицателно" in body or "отрицателна" in body, (
        "находка №5: отрицателно нето/бруто/обем мина без нито едно "
        "предупреждение")


# --------------------------------------------------------------------- №6
def test_failed_install_is_not_retried_for_the_same_version(monkeypatch, tmp_path):
    """Находка №6 (критична): при неуспешна инсталация програмата се
    рестартираше, откриваше СЪЩАТА нова версия и пробваше пак — безкраен
    цикъл рестарт↔сваляне, който състояние в паметта не може да прекъсне
    (то умира заедно с процеса). Затова маркерът е ФАЙЛ."""
    import updater
    monkeypatch.setattr(updater, "_failed_marker_path",
                        lambda: str(tmp_path / "pacho_update_failed.txt"))
    updater.clear_failed_install_marker()
    (tmp_path / "pacho_update_failed.txt").write_text("3.99.0\n", encoding="utf-8")
    assert updater.read_failed_install_version() == "3.99.0"

    called = []
    monkeypatch.setattr(updater, "is_frozen_windows", lambda: True)
    monkeypatch.setattr(updater, "_install_update_locked",
                        lambda *a, **k: called.append(1))
    with pytest.raises(RuntimeError):
        updater.install_update("http://example/x.exe", version="3.99.0")
    assert not called, (
        "находка №6: провалилата се версия се сваля отново — цикълът остава")

    # Ръчният бутон („Обнови сега“) съзнателно пробва пак — маркерът не
    # бива да заключи админа завинаги.
    updater.install_update("http://example/x.exe", version="3.99.0",
                           ignore_failed_marker=True)
    assert called == [1]

    updater.clear_failed_install_marker()
    assert updater.read_failed_install_version() is None


# --------------------------------------------------------------------- №7
def _zip_bomb(member_size=300 * 1024 * 1024):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("xl/worksheets/sheet1.xml", b"\0" * member_size)
    return buf.getvalue()


def test_zip_bomb_is_rejected_before_it_is_decompressed():
    """Находка №7 (висока): Excel импортът викаше `load_workbook` направо
    върху каченото. 204 KB архив се разгъваше до 713 MB в паметта —
    достатъчно да убие процеса на всички в мрежов режим."""
    import appcore
    bomb = _zip_bomb()
    assert len(bomb) < 2 * 1024 * 1024, "пробата трябва да е малка на диска"
    with pytest.raises(appcore.XlsxTooLargeError):
        appcore.ensure_xlsx_within_limits(bomb)
    # Нормален файл минава без възражение.
    ok = io.BytesIO()
    with zipfile.ZipFile(ok, "w") as zf:
        zf.writestr("xl/worksheets/sheet1.xml", b"<x/>")
    appcore.ensure_xlsx_within_limits(ok.getvalue())
    # Не-zip вход не бива да гърми тук (магическите байтове се проверяват
    # другаде) — функцията просто се отдръпва.
    appcore.ensure_xlsx_within_limits("не съм zip".encode("utf-8"))


# --------------------------------------------------------------------- №8
def test_pdf_export_normalizes_decimal_separator(admin_client):
    """Находка №8 (висока, МОЯ непокрита половина): поправката от седмия
    одит стигна до печатните шаблони, но не и до PDF износа — там
    запетайните стойности от Excel импорта излизаха сурови върху
    официален документ."""
    items = [{"code": "A1", "description": "стока", "qty": "2,5", "weight": "3,7"}]
    post_with_csrf(admin_client, "/pallet/new", {
        "client_name": "Клиент",
        "items_json": json.dumps(items, ensure_ascii=False),
        "gross": "9,9", "height": "14,5",
    }, csrf_source_url="/pallet/new", follow_redirects=True)
    resp = admin_client.get("/doc/1/export.pdf")
    assert resp.status_code == 200
    import pypdf
    text = "".join(p.extract_text() or ""
                   for p in pypdf.PdfReader(io.BytesIO(resp.data)).pages)
    for comma in ("2,5", "3,7", "9,9", "14,5"):
        assert comma not in text, (
            "находка №8: суровата запетайна стойност %s влезе в PDF-а" % comma)
    for dotted in ("2.5", "3.7"):
        assert dotted in text, "липсва нормализирана стойност %s в PDF-а" % dotted


# -------------------------------------------------------------------- №10
def test_xlsx_money_columns_use_two_decimals(admin_client):
    """Находка №10 (висока): паричните колони ползваха маската за
    КОЛИЧЕСТВА „0.###“ — цена 1.20 се показваше като „1.2“, а 0.0125 като
    „0.013“, тоест Excel износът на фактура показваше друга цена от
    бланката."""
    items = [{"description": "стока", "qty": "1000", "unit_price": "1.20"}]
    post_with_csrf(admin_client, "/invoice-br/new", {
        "consignee_name": "ABB", "invoice_number": "EX-1",
        "items_json": json.dumps(items, ensure_ascii=False),
    }, csrf_source_url="/invoice-br/new", follow_redirects=True)
    resp = admin_client.get("/doc/1/export.xlsx")
    assert resp.status_code == 200
    import openpyxl
    ws = openpyxl.load_workbook(io.BytesIO(resp.data)).active
    formats = {c.number_format for row in ws.iter_rows() for c in row
               if c.number_format in ("0.00", "0.###")}
    assert "0.00" in formats, (
        "находка №10: нито една клетка не е с паричен формат — цените пак "
        "ползват маската за количества")


# -------------------------------------------------------------------- №11
def test_db_path_is_validated_before_it_is_saved(tmp_path):
    """Находка №11 (висока): пътят до базата влизаше в pacho_config.json
    само след .strip(). Печатна грешка в ИМЕТО → нова празна база с
    admin/admin123 и „изчезнали“ данни; печатна грешка в ПАПКАТА →
    приложението пада в резервен режим, където страницата „Настройки“
    вече не съществува."""
    import config as appconfig
    import sqlite3 as _sqlite3

    real_db = tmp_path / "real.db"
    _sqlite3.connect(str(real_db)).close()
    not_a_db = tmp_path / "notes.txt"
    not_a_db.write_text("здравей", encoding="utf-8")

    assert appconfig.validate_db_path("")[0] is None
    assert appconfig.validate_db_path(str(real_db))[0] is None
    assert appconfig.validate_db_path(str(tmp_path / "нема" / "a.db"))[0]
    assert appconfig.validate_db_path(str(tmp_path))[0]
    assert appconfig.validate_db_path(str(not_a_db))[0]
    # Несъществуващ файл — само с ИЗРИЧНО потвърждение.
    new_path = str(tmp_path / "pacho_logistik.db")
    assert appconfig.validate_db_path(new_path)[0]
    assert appconfig.validate_db_path(new_path, allow_new=True)[0] is None


def test_settings_form_refuses_a_broken_db_path(admin_client, monkeypatch):
    """Същата находка, но през интерфейса: невалидният път НЕ бива да
    стига до save_config."""
    import config as appconfig
    saved = []
    monkeypatch.setattr(appconfig, "save_config", lambda v: saved.append(v))
    resp = post_with_csrf(admin_client, "/admin/system", {
        "form": "network", "network_port": "5000",
        "db_path": os.path.join(tempfile.gettempdir(), "нема-такава-папка", "x.db"),
    }, csrf_source_url="/my-settings", follow_redirects=True)
    assert not saved, "находка №11: счупеният път беше записан в конфигурацията"
    assert "не съществува" in resp.data.decode()


# -------------------------------------------------------------------- №12
def test_only_one_instance_can_hold_the_lock():
    """Находка №12 (висока): нямаше НИЩО, което да спре второ копие на
    програмата. Бавен първи старт (разопаковане + антивирус) → потребителят
    щраква пак → два процеса пускат миграции, два обновяващи цикъла се
    бият за `<exe>.new` и два таймера за архив."""
    d = tempfile.mkdtemp()
    child = os.path.join(d, "child.py")
    with open(child, "w", encoding="utf-8") as f:
        f.write(textwrap.dedent("""
            import sys, time
            sys.path.insert(0, %r)
            import single_instance
            print("ЗАКЛЮЧИ" if single_instance.acquire(directory=%r) else "ОТКАЗАН",
                  flush=True)
            time.sleep(float(sys.argv[1]))
        """) % (ROOT, d))
    first = subprocess.Popen([sys.executable, child, "5"],
                             stdout=subprocess.PIPE, text=True)
    try:
        assert first.stdout.readline().strip() == "ЗАКЛЮЧИ"
        second = subprocess.run([sys.executable, child, "0"],
                                capture_output=True, text=True, timeout=60)
        assert second.stdout.strip() == "ОТКАЗАН", (
            "находка №12: второто копие се стартира успоредно с първото")
    finally:
        first.kill()
        first.wait()
    # След като първият процес си отиде (тук — убит рязко, без почистване),
    # катинарът трябва да е свободен: ОС го освобождава, за разлика от
    # маркерен файл, който би блокирал програмата завинаги.
    third = subprocess.run([sys.executable, child, "0"],
                           capture_output=True, text=True, timeout=60)
    assert third.stdout.strip() == "ЗАКЛЮЧИ", (
        "находка №12: катинарът остана държан след рязко спиране")


def test_port_probe_is_not_blind_to_a_listening_socket():
    """Втората половина на находка №12: `SO_REUSEADDR` върху пробната
    сокет е СЛЯП на Windows (позволява bind върху активно слушащ сокет).
    Проверката тук е, че кодът вече не слага тази опция на Windows."""
    import inspect
    import net
    source = inspect.getsource(net.find_available_port)
    assert 'os.name == "nt"' in source, (
        "находка №12: пробата на порта пак е еднаква за Windows и POSIX")
    assert "SO_EXCLUSIVEADDRUSE" in source

    # И че на тази (POSIX) машина поведението е непроменено: зает порт се
    # разпознава като зает.
    import socket
    sock = socket.socket()
    sock.bind(("127.0.0.1", 0))
    port = sock.getsockname()[1]
    sock.listen(1)
    try:
        assert net.find_available_port("127.0.0.1", port) != port
    finally:
        sock.close()


# -------------------------------------------------------------------- №14
def test_secret_key_creation_is_atomic(tmp_path):
    """Находка №14 (висока): `exists` → `token_hex` → `open(...,"w")` са
    три отделни стъпки. При „.exe в споделената папка“ два компютъра при
    първи старт пишеха различни ключове; единият подписваше бисквитки с
    ключ, който във файла вече го няма — след негов рестарт всички негови
    потребители изхвърчаха по средата на форма, без съобщение."""
    secret_path = str(tmp_path / "secret.txt")
    child = str(tmp_path / "c.py")
    with open(child, "w", encoding="utf-8") as f:
        f.write(textwrap.dedent("""
            import sys, time
            sys.path.insert(0, %r)
            import db
            db.SECRET_PATH = %r
            start = float(sys.argv[1])
            while time.time() < start:
                pass
            print(db.get_secret_key())
        """) % (ROOT, secret_path))
    import time as _time
    start = _time.time() + 3
    procs = [subprocess.Popen([sys.executable, child, str(start)],
                              stdout=subprocess.PIPE, text=True) for _ in range(8)]
    keys = [p.communicate(timeout=120)[0].strip() for p in procs]
    on_disk = open(secret_path, encoding="utf-8").read().strip()
    assert len(set(keys)) == 1, (
        "находка №14: %d различни ключа при едновременен пръв старт"
        % len(set(keys)))
    assert all(k == on_disk for k in keys), (
        "находка №14: процес подписва с ключ, който във файла вече го няма")


# -------------------------------------------------------------------- №15
class _CommitHook:
    """Прокси около sqlite връзката, което изпълнява „админът нулира
    паролата“ точно СЛЕД първия commit — прозорецът на находка №15."""

    def __init__(self, con, on_first_commit):
        self._con = con
        self._hook = on_first_commit
        self._fired = False

    def __getattr__(self, name):
        return getattr(self._con, name)

    def commit(self):
        self._con.commit()
        if not self._fired:
            self._fired = True
            self._hook()


def test_session_epoch_is_read_inside_the_writing_transaction(
        flask_app, db_module, monkeypatch):
    """Находка №15 (висока): епохата се четеше СЛЕД commit. Ако точно
    междувременно администратор нулира паролата на същия потребител,
    четенето връщаше ЧУЖДАТА нова епоха и я записваше в бисквитката —
    епохите съвпадаха и сесията НЕ се прекратяваше, тоест целта на
    админския маршрут („служителят е забравил да излезе на споделен
    компютър“) тихо се проваляше."""
    con = db_module.get_db()
    uid = con.execute(
        "INSERT INTO users (username, password_hash, full_name, role, active,"
        " must_change_password) VALUES (?, ?, ?, 'employee', 1, 0)",
        ("ivan", generate_password_hash("staro-parola-123"), "Иван")).lastrowid
    con.commit()
    con.close()

    user_client = flask_app.test_client()
    post_with_csrf(user_client, "/login",
                   {"username": "ivan", "password": "staro-parola-123"},
                   csrf_source_url="/login")

    real_get_db = db_module.get_db
    fired = []

    def admin_resets():
        if fired:
            return
        fired.append(1)
        c = real_get_db()
        c.execute("UPDATE users SET password_hash = ?,"
                  " session_epoch = session_epoch + 1 WHERE id = ?",
                  (generate_password_hash("admin-nulira-789"), uid))
        c.commit()
        c.close()

    monkeypatch.setattr(db_module, "get_db",
                        lambda *a, **k: _CommitHook(real_get_db(*a, **k), admin_resets))
    post_with_csrf(user_client, "/password", {
        "current": "staro-parola-123",
        "new": "novata-parola-456",
        "repeat": "novata-parola-456",
    }, csrf_source_url="/password")
    monkeypatch.setattr(db_module, "get_db", real_get_db)

    assert fired, "проверката е невалидна, ако админският нулиращ запис не е минал"
    con = db_module.get_db()
    in_db = con.execute("SELECT session_epoch FROM users WHERE id = ?",
                        (uid,)).fetchone()["session_epoch"]
    con.close()
    with user_client.session_transaction() as sess:
        in_cookie = sess.get("session_epoch")
    assert in_cookie != in_db, (
        "находка №15: бисквитката получи ЧУЖДАТА епоха (%s) — админското "
        "прекратяване на сесията се провали" % in_cookie)
    assert user_client.get("/").status_code != 200, (
        "находка №15: сесията оцеля админското нулиране на паролата")


# --------------------------------------------------------------- №17 + №18
def _draft_card():
    return {
        "doc_date": "2026-08-31", "client_name": "ABB", "client_address": "ул. 1",
        "client_city": "София", "client_country": "BG",
        "pallet_no": "P1", "packaging_type": "кашон",
        "pallet_type": "120x80", "height": "180,0", "gross": "1200,500",
        "items_format": "std",
        "items": [{"code": "К-1", "description": "Описание",
                   "qty": "12,000", "weight": "3,250"}],
    }


def test_bulk_preview_shows_the_same_numbers_as_the_printed_card(flask_app):
    """Находка №18 (ниска): маршрутът обещава „точно както ще изглеждат
    при печат“, а прегледът показваше СУРОВИТЕ десетични (запетая от
    Excel импорта), докато издадената карта ги нормализира."""
    from flask import render_template
    with flask_app.test_request_context("/"):
        body = render_template("pallet_bulk_preview.html",
                               drafts=[_draft_card()], token="t", has_logo=False)
    for comma in ("180,0", "1200,500", "12,000", "3,250"):
        assert comma not in body, (
            "находка №18: суровата стойност %s се показва в прегледа" % comma)


def test_bulk_templates_repeat_the_header_and_the_card_number(flask_app, admin_client):
    """Находка №17 (ниска): поправката от 19.08 (заглавен ред в <thead> +
    номер на картата в <tfoot>, за да е разпознаваем лист 2) стигна до
    pallet_print.html, но не и до ДВАТА групови шаблона — а груповият път
    е именно този след Excel импорт на много карти."""
    from flask import render_template
    with flask_app.test_request_context("/"):
        preview = render_template("pallet_bulk_preview.html",
                                  drafts=[_draft_card()], token="t", has_logo=False)
    assert "<thead>" in preview and "print-ident" in preview

    items = [{"code": "К-1", "description": "стока", "qty": "1", "weight": "2"}]
    post_with_csrf(admin_client, "/pallet/new", {
        "client_name": "ABB",
        "items_json": json.dumps(items, ensure_ascii=False),
    }, csrf_source_url="/pallet/new", follow_redirects=True)
    body = admin_client.get("/pallet/bulk-print?ids=1").data.decode()
    assert "<thead>" in body, (
        "находка №17: груповият печат няма заглавен ред в <thead> — карта "
        "с много редове губи имената на колоните на лист 2")
    assert "print-ident" in body, (
        "находка №17: груповият печат не повтаря номера на картата")


# -------------------------------------------------------------------- №19
def test_error_redirect_loop_between_two_pages_is_stopped(flask_app, db_module):
    """Находка №19 (ниска): защитата сравняваше `urlsplit(target).path ==
    request.path`, тоест хващаше само цикъла A→A. Браузърът пази Referer
    през 302, затова две страници, които гърмят една заради друга, се
    въртяха A→B→A→B до ERR_TOO_MANY_REDIRECTS, без нито едно видяно
    съобщение."""
    import appcore

    @flask_app.route("/tmp-loop-a")
    @appcore.login_required
    def _tmp_loop_a():
        raise RuntimeError("A гърми")

    @flask_app.route("/tmp-loop-b")
    @appcore.login_required
    def _tmp_loop_b():
        raise RuntimeError("B гърми")

    _add_admin(db_module, "admin_loop")
    client = _login(flask_app, "admin_loop")

    hops, url, referrer = 0, "/tmp-loop-a", "http://localhost/tmp-loop-b"
    while hops < 20:
        resp = client.get(url, headers={"Referer": referrer})
        if resp.status_code != 302:
            break
        hops += 1
        referrer = "http://localhost" + url
        url = resp.headers["Location"]
        if url.startswith("http"):
            url = "/" + url.split("/", 3)[3]
    assert hops <= appcore.MAX_ERROR_HOPS, (
        "находка №19: цикълът A→B→A не спря — %d скока" % hops)
    assert resp.status_code == 500

    # И обратното: единична, несвързана грешка НЕ бива да се влошава до
    # статична страница — броячът се нулира от всяка успешна страница.
    for _ in range(6):
        assert client.get("/tmp-loop-a",
                          headers={"Referer": "http://localhost/"}).status_code == 302
        assert client.get("/").status_code == 200


# -------------------------------------------------------------------- №20
def test_duplicate_number_check_uses_the_documents_own_year(admin_client, db_module):
    """Находка №20 (ниска): проверката ползваше `date.today().year`, а
    уникалният индекс е върху ЗАПИСАНАТА година на документа, която
    редакцията не променя. Редакция на документ от миналата година →
    предупреждението мълчи, после IntegrityError."""
    for num in ("СТАР-1", "СТАР-2"):
        post_with_csrf(admin_client, "/invoice-br/new",
                       {"consignee_name": "ABB", "invoice_number": num},
                       csrf_source_url="/invoice-br/new", follow_redirects=True)
    con = db_module.get_db()
    con.execute("UPDATE documents SET year = ?", (date.today().year - 1,))
    con.commit()
    ids = [r["id"] for r in con.execute("SELECT id FROM documents ORDER BY id")]
    con.close()

    resp = post_with_csrf(admin_client, "/doc/%d/edit" % ids[1],
                          {"consignee_name": "ABB", "invoice_number": "СТАР-1"},
                          follow_redirects=True)
    assert "вече има издаден документ с номер СТАР-1" in resp.data.decode(), (
        "находка №20: проверката за зает номер гледа текущата година вместо "
        "годината на самия документ — предупреждението мълчи")


# --------------------------------------------------------------------- №9
def test_every_printed_value_box_breaks_long_words():
    """Находка №9 (висока, МОЯ непокрита половина): поправката ми от
    седмия одит сложи `overflow-wrap`/`word-break` само на палетната кутия
    и решетката на ЧМР. На останалите бланки дълъг непрекъснат низ (име на
    фирма или адрес, залепен без интервали при копиране от Excel) излизаше
    извън A4 листа и се отрязваше при печат.

    Тестът е СТАТИЧЕН нарочно: браузърният тест в test_e2e_smoke.py
    доказва ефекта върху реален лист, но само за кутиите, чиито размери
    позволяват преливането да се измери. Тук се заключва самото правило —
    за ВСИЧКИ шест кутии наведнъж, за да не остане пак непокрита половина."""
    css = open(os.path.join(ROOT, "static", "style.css"), encoding="utf-8").read()
    boxes = [".plt-client-box .val", ".pbox .val", ".pkl .party .val",
             ".inv .party .val", ".inv .party .name", ".tbox .val"]
    for selector in boxes:
        start = css.find(selector + " {")
        assert start != -1, "правилото за %s липсва в style.css" % selector
        block = css[start:css.find("}", start)]
        assert "overflow-wrap: break-word" in block and "word-break: break-word" in block, (
            "находка №9: %s няма пречупване на дълги думи — дълъг низ "
            "излиза извън A4 листа" % selector)
