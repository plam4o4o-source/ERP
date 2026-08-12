# -*- coding: utf-8 -*-
"""End-to-end „дим“ тестове с истински headless браузър (Playwright) —
заявка: „направи всичко което предлагаш“ (CI подобрения: Playwright в CI).

За разлика от останалите тестове (Flask test client — вика view функциите
директно, без реален HTTP/браузър), тук стартираме ИСТИНСКИ работещ сървър
(werkzeug make_server, фонова нишка) и го управляваме с истински Chromium
през Playwright — единственият начин да хванем проблеми, които test client-ът
структурно не може: чупещ се JavaScript (напр. initItemsTable в
client_form.html/pallet_form.html), CSS оформление, и — най-важното тук —
дали печатните шаблони РЕАЛНО изглеждат като документ при емулация на
print media (виж ПЛАН_ЗА_РАЗРАБОТКА.md реда за ръчната Playwright проверка
на печатните шаблони, направена еднократно от архитекта — тук е
автоматизирана и тръгва при всеки push).

Изключени по подразбиране от бързия `pytest` (виж pytest.ini/marker "e2e")
— пускат се изрично в CI job "e2e" (.github/workflows/ci.yml), защото
изискват изтеглен браузър и са значително по-бавни от unit/интеграционните
тестове."""
import threading

import pytest

pytestmark = pytest.mark.e2e

_XLSX_HEADERS = ["Due Date", "Order No", "Pos", "Project", "Reference",
                 "Reference Desc", "Open Qty", "Unit", "Stock", ""]
_XLSX_ROW = ["2026-09-01", "E2E-ORD-1", "10", "PRJ-1", "REF-1",
            "Материал за Е2Е тест", 6, "PCS", "WH1", 1]

playwright_sync_api = pytest.importorskip("playwright.sync_api")
sync_playwright = playwright_sync_api.sync_playwright


@pytest.fixture
def live_server(flask_app, db_module):
    """Истински работещ HTTP сървър (werkzeug), в фонова нишка, срещу
    СЪЩОТО пълно приложение и временна база, каквито ползва `flask_app`
    (виж conftest.py) — само транспортът е реален TCP/HTTP, не Flask test
    client. Портът е 0 (случаен свободен), за да не се сблъска с друг
    паралелно пуснат тест/процес."""
    from werkzeug.security import generate_password_hash
    from werkzeug.serving import make_server

    con = db_module.get_db()
    con.execute(
        "INSERT INTO users (username, password_hash, full_name, role, active,"
        " must_change_password) VALUES (?, ?, ?, 'admin', 1, 0)",
        ("e2e_admin", generate_password_hash("e2e-test-password-123"), "E2E Тест"),
    )
    con.commit()
    con.close()

    # НЕ пипаме CSRF защитата — за разлика от Flask test client тестовете
    # (виж conftest.post_with_csrf), тук истински браузър зарежда истинска
    # страница с вече вградения `csrf_token()` в скрито поле на формата и
    # го изпраща естествено при submit; appcore._check_csrf е собствена
    # проверка, не през Flask-WTF, затова WTF_CSRF_ENABLED не важи за нея.
    server = make_server("127.0.0.1", 0, flask_app)
    port = server.server_port
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        yield "http://127.0.0.1:%d" % port
    finally:
        server.shutdown()
        thread.join(timeout=5)


@pytest.fixture
def page(live_server):
    with sync_playwright() as p:
        browser = p.chromium.launch()
        context = browser.new_context()
        pg = context.new_page()
        yield pg
        context.close()
        browser.close()


def _login(page, base_url):
    page.goto(base_url + "/login")
    page.fill('input[name="username"]', "e2e_admin")
    page.fill('input[name="password"]', "e2e-test-password-123")
    page.click('button[type="submit"]')
    page.wait_for_url(base_url + "/")


def test_login_and_dashboard_loads(page, live_server):
    _login(page, live_server)
    assert "Табло" in page.title()
    assert page.locator("text=Статистика за текущия месец").count() == 1


def test_issue_cmr_end_to_end_and_barcode_renders(page, live_server):
    """Пълен път: вход → форма за ново ЧМР → издаване → печатна страница —
    точно потокът, който Playwright ръчно провери еднократно при
    първоначалния одит (виж ПЛАН_ЗА_РАЗРАБОТКА.md), сега автоматизиран."""
    _login(page, live_server)
    page.goto(live_server + "/cmr/new")
    page.fill('input[name="sender_name"]', "Изпращач ЕООД")
    page.fill('input[name="consignee_name"]', "Е2Е Тест Клиент ЕООД")
    page.click('button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")
    assert "ЧМР" in page.content()
    assert "Е2Е Тест Клиент ЕООД" in page.content()
    # Баркодът е вграден SVG (barcode128.code128_svg през appcore.barcode_filter,
    # виж appcore.py) — реално рендериран от браузъра, не само присъстващ в HTML.
    assert page.locator("svg rect").count() > 0


def test_print_media_emulation_renders_document(page, live_server):
    """Емулация на print media (истинската проверка, направена ръчно от
    архитекта при одита) — печатната бланка трябва да остане видима и без
    екранните елементи (лентата с бутони е class="no-print")."""
    _login(page, live_server)
    page.goto(live_server + "/cmr/new")
    page.fill('input[name="sender_name"]', "Изпращач ЕООД")
    page.fill('input[name="consignee_name"]', "Печатен Тест ЕООД")
    page.click('button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")
    page.emulate_media(media="print")
    assert page.locator(".cmr").first.is_visible()
    assert not page.locator(".doc-toolbar").first.is_visible()


def test_client_history_card_renders_in_browser(page, live_server):
    """Регресия за v3.33.0 (История на документите от картата на клиента)
    — реално рендерирана в браузър, не само през test client."""
    _login(page, live_server)
    page.goto(live_server + "/clients/new")
    page.fill('input[name="name"]', "Е2Е Браузър Клиент ЕООД")
    page.click('button[type="submit"]')
    page.wait_for_url(live_server + "/clients")
    row = page.locator("tr", has_text="Е2Е Браузър Клиент ЕООД")
    row.get_by_text("Редакция").click()
    assert page.locator("text=Последни документи на този клиент").count() == 1
    assert page.locator("text=Все още няма издадени документи").count() == 1


def test_cmr_load_place_from_sender_button_fills_place_loading(page, live_server):
    """Заявка: „товарен пункт да се зарежда от фирма изпращач, но да има
    опция и ръчно въвеждане“ — бутон „Зареди от изпращача“ до поле 4
    „Товарен пункт“ (JS, initCmrPlaces в app.js) копира текущите стойности
    на поле 1 „Изпращач“ в текстовото поле place_loading, САМО при
    натискане (не автоматично), и полето остава свободно за ръчна промяна
    след това. Чисто клиентско JS поведение — Flask test client не
    изпълнява истински JavaScript, затова е нужен реален браузър тук."""
    _login(page, live_server)
    page.goto(live_server + "/cmr/new")
    page.fill('input[name="sender_name"]', "Товарач ЕООД")
    page.fill('input[name="sender_address"]', "ул. Складова 5")
    page.fill('input[name="sender_city"]', "4000 Пловдив")
    page.fill('input[name="sender_country"]', "България")

    # Преди натискане на бутона полето за товарен пункт не е пипано.
    assert page.locator('input[name="place_loading"]').input_value() == ""

    page.click("#load-place-from-sender-btn")
    assert page.locator('input[name="place_loading"]').input_value() == \
        "Товарач ЕООД — ул. Складова 5, 4000 Пловдив, България"

    # Ръчно въвеждане/промяна остава напълно свободно и след бутона.
    page.fill('input[name="place_loading"]', "Друг склад, ръчно въведен")
    assert page.locator('input[name="place_loading"]').input_value() == "Друг склад, ръчно въведен"


def test_pallet_bulk_import_preview_keeps_loaded_data(page, live_server, tmp_path):
    """Регресия за реален бъг: при импорт на Excel файл в палетна карта и
    последващ „Предварителен преглед“, заредените редове изчезваха и се
    показваше „Няма палетни карти за преглед“ — заявка: „при зареждане на
    файл в палетна карта и избор на преглед, изтрива се заредената
    информация“.

    Причината беше чисто клиентска (JS): initItemsTable-ът в
    pallet_bulk_review.html не задаваше table.dataset.hiddenField, а
    handler-ът при submit четеше единствено него (с твърдо закодиран
    резервен вариант "items_json", който не съществува тук — истинските
    полета са "items_json_1", "items_json_2" и т.н.) — Flask test client
    тестовете не изпълняват JS и затова не хващаха това, изисква се
    истински браузър (виж static/app.js, initItemsTable)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(_XLSX_HEADERS)
    ws.append(_XLSX_ROW)
    xlsx_path = tmp_path / "poruchki.xlsx"
    wb.save(str(xlsx_path))

    _login(page, live_server)
    page.goto(live_server + "/pallet/new")
    page.set_input_files('input[name="excel_file"]', str(xlsx_path))
    page.click('button:has-text("Зареди и раздели по палети")')
    page.wait_for_url(live_server + "/pallet/bulk-import" + "*", timeout=10000)

    # Редът от файла трябва да се вижда в самата таблица за преглед/редакция
    # (стойността е зададена през JS .value, не HTML value= атрибут, затова
    # четем я през input_value(), не CSS [value=...] селектор).
    first_input = page.locator("#pallet-items-1 tbody tr").first.locator("input").first
    assert first_input.input_value() == "E2E-ORD-1"

    page.fill('input[name="client_name"]', "Клиент Е2Е Bulk")
    page.click('button:has-text("Предварителен преглед")')
    page.wait_for_load_state("networkidle")

    body = page.content()
    assert "Няма палетни карти за преглед" not in body
    assert "E2E-ORD-1" in body
    assert "Материал за Е2Е тест" in body


def test_pallet_label_barcode_fits_within_label_width(page, live_server):
    """Не беше в обхвата на патча: v3.38.0 удължи баркода (вече съдържа и
    пълната дата — „PAL-07082026-0001“ вместо „PAL-2026-0001“), което го
    прави ~22% по-широк, а етикетният формат е само 100мм широк — затова тук
    измерваме реално в браузър, че баркодът се събира в листа и няма да се
    отреже при печат.

    Заглавният баркод е защитен от `.plt-head-barcode svg { max-width:100% }`
    в style.css. (Одит-подобна заявка от 12.08.2026: долният голям баркод
    — .plt-big-barcode — беше премахнат изцяло от всички палетни бланки,
    „баркод/QR само горе, където си беше“ — виж
    test_pallet_print_has_no_barcode_at_the_bottom_of_the_card в
    tests/test_document_layout_2026_08.py; тук вече проверяваме само
    ЕДИНСТВЕНИЯ останал баркод, горния.)"""
    _login(page, live_server)
    page.goto(live_server + "/pallet/new")
    page.fill('input[name="client_name"]', "Клиент Етикет")
    # Страницата има ДВЕ форми (импортът от Excel е отделна) — издаването
    # е това на главната форма, затова селекторът е с #main-doc-form.
    page.click('#main-doc-form button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")

    page.goto(page.url + "?format=label")
    page.emulate_media(media="print")

    assert page.query_selector(".plt-big-barcode") is None

    box = page.eval_on_selector(
        ".plt-head-barcode svg",
        "el => { const r = el.getBoundingClientRect();"
        " const p = el.closest('.print-page').getBoundingClientRect();"
        " return {svg: r.width, page: p.width}; }",
    )
    assert box["svg"] <= box["page"], (
        "баркодът (%.0fpx) прелива извън етикета (%.0fpx) и ще се отреже при печат"
        % (box["svg"], box["page"])
    )
# ---------------------------------------------------------------- фактури
# Автоматичното попълване от справочника материали и зареждането на всички
# редове от палетна карта са чисто клиентски (JS + fetch) — Flask test
# client не изпълнява JavaScript, затова истинският браузър е ЕДИНСТВЕНИЯТ
# начин да се провери, че реално работят на екрана.

def _load_materials_catalog(page, live_server, tmp_path, rows):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["ABB part ID", "Description", "Net weight\n[KG/pc]"])
    for row in rows:
        ws.append(list(row))
    path = tmp_path / "kg.xlsx"
    wb.save(str(path))

    page.goto(live_server + "/materials")
    page.set_input_files('input[name="excel_file"]', str(path))
    page.click('button:has-text("Зареди справочника")')
    page.wait_for_load_state("networkidle")


def test_invoice_material_code_autofills_net_weight_from_catalog(page, live_server, tmp_path):
    """Заявка: „от файла с килограмите автоматично да се извличат
    съответните килограми във фактурата“ — при въвеждане на код на
    материала полето „Net weight“ се попълва само."""
    _login(page, live_server)
    _load_materials_catalog(page, live_server, tmp_path,
                            [("GLBK400002P0012", "C-PROFILE 3   1150MM", 2.21)])

    page.goto(live_server + "/invoice-br/new")
    row = page.locator("#invoice-br-items tbody tr").first
    code_input = row.locator('input[data-field="material_code"]')
    weight_input = row.locator('input[data-field="net_weight"]')
    assert weight_input.input_value() == ""

    code_input.fill("GLBK400002P0012")
    code_input.blur()  # попълването се задейства при change, не при всеки натиснат клавиш
    page.wait_for_function(
        """() => document.querySelector('#invoice-br-items tbody tr input[data-field="net_weight"]').value !== ''""",
        timeout=5000)
    assert weight_input.input_value() == "2.21"


def test_invoice_autofill_never_overwrites_a_manually_typed_value(page, live_server, tmp_path):
    """Ръчно въведеното тегло е за конкретната пратка и трябва да
    надделява над справочника — иначе операторът би го губил при всяко
    поправяне на кода."""
    _login(page, live_server)
    _load_materials_catalog(page, live_server, tmp_path,
                            [("GLBK400002P0012", "C-PROFILE 3   1150MM", 2.21)])

    page.goto(live_server + "/invoice-br/new")
    row = page.locator("#invoice-br-items tbody tr").first
    row.locator('input[data-field="net_weight"]').fill("9.99")
    code_input = row.locator('input[data-field="material_code"]')
    code_input.fill("GLBK400002P0012")
    code_input.blur()
    page.wait_for_timeout(600)
    assert row.locator('input[data-field="net_weight"]').input_value() == "9.99"


def test_invoice_loads_all_rows_from_an_issued_pallet_card(page, live_server, tmp_path):
    """Заявка: „фактурата да може да се зарежда, както се зареждат
    палетните карти в опаковъчния лист“ — но с ВСИЧКИ редове поотделно и с
    тегло, изтеглено от справочника."""
    _login(page, live_server)
    _load_materials_catalog(page, live_server, tmp_path, [
        ("GLBK400002P0012", "C-PROFILE 3   1150MM", 2.21),
        ("GLBK400001P0200", "C-PROFILE 2    200MM", 0.383),
    ])

    # Издаваме палетна карта във формат „поръчки“ по реалния път — импорт
    # на справка за поръчки от Excel (форматът „поръчки“ се получава само
    # оттам, виж pallet_form.html), после масово издаване.
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Order No", "Pos", "Reference", "Reference Desc", "Open Qty", ""])
    ws.append(["4700200362", "30", "GLBK400002P0012", "C-Profile", 20, 1])
    ws.append(["4700200362", "40", "GLBK400001P0200", "C-Profile 2", 400, 1])
    orders_path = tmp_path / "poruchki_invoice.xlsx"
    wb.save(str(orders_path))

    page.goto(live_server + "/pallet/new")
    page.set_input_files('input[name="excel_file"]', str(orders_path))
    page.click('button:has-text("Зареди и раздели по палети")')
    page.wait_for_url(live_server + "/pallet/bulk-import*", timeout=10000)
    page.fill('input[name="client_name"]', "ABB")
    page.click('button:has-text("Издай всички палетни карти")')
    page.wait_for_load_state("networkidle")

    number = page.locator("table.list tbody tr td, table.list tr td").first.inner_text().strip()
    assert "/" in number, "очаква се номер на издадена палетна карта, а не %r" % number

    # Зареждаме я във фактурата.
    page.goto(live_server + "/invoice-br/new")
    page.fill("#f-pull-invoice-code", number)
    page.click(".invoice-pull-btn")
    page.wait_for_function(
        """() => document.querySelectorAll('#invoice-br-items tbody tr').length >= 3""",
        timeout=8000)

    # Първият ред е празният начален — двата заредени идват след него.
    loaded = page.locator("#invoice-br-items tbody tr")
    codes = [loaded.nth(i).locator('input[data-field="material_code"]').input_value()
             for i in range(loaded.count())]
    assert "GLBK400002P0012" in codes
    assert "GLBK400001P0200" in codes

    idx = codes.index("GLBK400002P0012")
    filled = loaded.nth(idx)
    assert filled.locator('input[data-field="po_no"]').input_value() == "4700200362"
    assert filled.locator('input[data-field="pos"]').input_value() == "30"
    assert filled.locator('input[data-field="qty"]').input_value() == "20"
    assert filled.locator('input[data-field="net_weight"]').input_value() == "2.21", \
        "теглото трябва да е изтеглено автоматично от справочника"


def test_invoice_address_book_selection_fills_both_address_blocks(page, live_server):
    """Заявка: „в раздел Фактури добави адресна книга; да съдържа данните за
    фактуриране на клиентите и също да има адрес за доставка“ — един избор
    попълва И блока Consignee, И блока Bill To. Чисто клиентско JS
    поведение (bindInvoiceClientSelect в app.js)."""
    _login(page, live_server)
    page.goto(live_server + "/invoices/clients/new")
    page.fill('input[name="name"]', "ABB Бразилия — Sorocaba")
    page.fill('input[name="delivery_name"]', "ABB ELETRIFICACAO LTDA")
    page.fill('textarea[name="delivery_address"]', "Rod.Sen. KM 11\nSorocaba - SP")
    page.fill('input[name="delivery_phone"]', "+55 11 97613-8155")
    page.fill('input[name="billing_name"]', "ABB ELETRIFICACAO LTDA - CNPJ")
    page.fill('textarea[name="billing_address"]', "Fakturamottak\n18087-125 Sorocaba")
    page.fill('input[name="billing_phone"]', "+55 15 3330-6465")
    page.click('button[type="submit"]')
    page.wait_for_url(live_server + "/invoices/clients")

    page.goto(live_server + "/invoice-br/new")
    page.select_option("#f-invoice-client-select", label="ABB Бразилия — Sorocaba")

    assert page.locator('input[name="consignee_name"]').input_value() == "ABB ELETRIFICACAO LTDA"
    assert "Sorocaba - SP" in page.locator('textarea[name="consignee_address"]').input_value()
    assert page.locator('input[name="consignee_phone"]').input_value() == "+55 11 97613-8155"
    assert page.locator('input[name="billto_name"]').input_value() == "ABB ELETRIFICACAO LTDA - CNPJ"
    assert "Fakturamottak" in page.locator('textarea[name="billto_address"]').input_value()
    assert page.locator('input[name="billto_phone"]').input_value() == "+55 15 3330-6465"


def test_invoice_excel_import_adds_rows_without_losing_typed_data(page, live_server, tmp_path):
    """Заявка: „зареждането на материалите във фактурата могат да се
    зареждат и от Excel файл“ — файлът се качва през fetch, за да НЕ се
    губи вече въведеното в останалите полета на формата (номер, получател).
    Точно това се проверява тук."""
    from openpyxl import Workbook

    _login(page, live_server)
    _load_materials_catalog(page, live_server, tmp_path,
                            [("GLBK400002P0012", "C-PROFILE 3   1150MM", 2.21)])

    wb = Workbook()
    ws = wb.active
    ws.append(["Order No", "Pos", "Reference", "Reference Desc", "Open Qty", "Unit Price"])
    ws.append(["4700200362", "30", "GLBK400002P0012", "C-Profile", 20, 13.66])
    path = tmp_path / "invoice_items.xlsx"
    wb.save(str(path))

    page.goto(live_server + "/invoice-br/new")
    page.fill('input[name="invoice_number"]', "0000012955")
    page.fill('input[name="consignee_name"]', "ABB ELETRIFICACAO LTDA")

    page.set_input_files("#f-invoice-excel-invoice-br-items", str(path))
    page.click(".invoice-excel-btn")
    page.wait_for_function(
        """() => document.querySelectorAll('#invoice-br-items tbody tr').length >= 2""",
        timeout=8000)

    rows = page.locator("#invoice-br-items tbody tr")
    codes = [rows.nth(i).locator('input[data-field="material_code"]').input_value()
             for i in range(rows.count())]
    assert "GLBK400002P0012" in codes
    loaded = rows.nth(codes.index("GLBK400002P0012"))
    assert loaded.locator('input[data-field="qty"]').input_value() == "20"
    assert loaded.locator('input[data-field="unit_price"]').input_value() == "13.66"
    assert loaded.locator('input[data-field="net_weight"]').input_value() == "2.21"

    # Вече въведеното НЕ е загубено — това е причината да е през fetch.
    assert page.locator('input[name="invoice_number"]').input_value() == "0000012955"
    assert page.locator('input[name="consignee_name"]').input_value() == "ABB ELETRIFICACAO LTDA"


def test_issued_invoice_is_absent_from_general_documents_list(page, live_server):
    """Заявка: „само там да се появяват издадените фактури“ — проверено в
    реален браузър от край до край (издаване → двата списъка)."""
    _login(page, live_server)
    page.goto(live_server + "/invoice-br/new")
    page.fill('input[name="invoice_number"]', "E2E-INV-1")
    page.fill('input[name="consignee_name"]', "Е2Е Фактурен Клиент")
    page.click('#main-doc-form button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")

    page.goto(live_server + "/invoices")
    assert "E2E-INV-1" in page.content()

    page.goto(live_server + "/docs")
    body = page.content()
    assert "E2E-INV-1" not in body
    assert "Е2Е Фактурен Клиент" not in body


def test_manual_pallet_entry_uses_order_columns_and_feeds_invoices(page, live_server, tmp_path):
    """Заявка: „палетна карта да съдържа в съдържание на палета Order No,
    Pos, Reference, Reference Desc, Qty“ — при РЪЧНО въвеждане. Реалният
    браузър е нужен, защото имената на полетата (вкл. items_format=orders)
    се задават от JS (initPalletMultiCard/renumber) чак при зареждане на
    страницата — test client не би хванал счупване там. Проверява и
    най-ценната последица: ръчно въведената карта се влива във фактура с
    Reference → Material code и тегло от справочника."""
    _login(page, live_server)
    _load_materials_catalog(page, live_server, tmp_path,
                            [("GLBK400002P0012", "C-PROFILE 3   1150MM", 2.21)])

    page.goto(live_server + "/pallet/new")
    tr = page.locator("#pallet-items tbody tr").first
    for field, value in (("order_no", "4700200362"), ("pos", "30"),
                         ("reference", "GLBK400002P0012"),
                         ("reference_desc", "C-Profile"), ("qty", "20")):
        tr.locator('input[data-field="%s"]' % field).fill(value)
    page.fill('input[name="client_name"]', "Ръчен Е2Е Клиент")
    page.click('#main-doc-form button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")

    body = page.content()
    assert "Order No" in body
    assert "GLBK400002P0012" in body
    # НЕ през text= локатор — той е нечувствителен към регистъра и хваща
    # първо flash съобщението „Палетна карта № … е издадена“, чиято опашка
    # разваля номера. Заглавието на бланката е с главни букви и е уникално
    # в суровия HTML.
    number = body.split("ПАЛЕТНА КАРТА № ")[1][:9].strip()

    # Ръчната карта се зарежда във фактура точно като импортираната.
    page.goto(live_server + "/invoice-br/new")
    page.fill("#f-pull-invoice-code", number)
    page.click(".invoice-pull-btn")
    page.wait_for_function(
        """() => Array.from(document.querySelectorAll(
              '#invoice-br-items tbody tr input[data-field="material_code"]'))
              .some(i => i.value === 'GLBK400002P0012')""",
        timeout=8000)
    rows = page.locator("#invoice-br-items tbody tr")
    codes = [rows.nth(i).locator('input[data-field="material_code"]').input_value()
             for i in range(rows.count())]
    loaded = rows.nth(codes.index("GLBK400002P0012"))
    assert loaded.locator('input[data-field="po_no"]').input_value() == "4700200362"
    assert loaded.locator('input[data-field="qty"]').input_value() == "20"
    assert loaded.locator('input[data-field="net_weight"]').input_value() == "2.21", \
        "теглото идва от справочника — ръчната карта е пълноценен формат orders"


def test_invoice_new_rows_get_hs_code_prefilled(page, live_server):
    """Заявка: „в фактурите по подразбиране винаги да се поставя
    автоматично HS code 85389099“ — и началният празен ред, и ред от
    „+ Добави ред“. Попълва се от JS (initItemsTable/data-row-defaults),
    затова е нужен реален браузър."""
    _login(page, live_server)
    page.goto(live_server + "/invoice-br/new")

    first = page.locator("#invoice-br-items tbody tr").first
    assert first.locator('input[data-field="hs_code"]').input_value() == "85389099"
    # Останалите колони на празния ред са си празни — подразбирането важи
    # само за HS кода.
    assert first.locator('input[data-field="material_code"]').input_value() == ""

    page.click('[data-add-row="invoice-br-items"]')
    rows = page.locator("#invoice-br-items tbody tr")
    assert rows.count() == 2
    assert rows.nth(1).locator('input[data-field="hs_code"]').input_value() == "85389099"


def test_invoice_pull_with_two_orders_shows_picker_and_loads_one(page, live_server, tmp_path):
    """Заявка: „във фактури един номер на поръчка да бъде на една фактура“
    — палетна карта с 2 поръчки НЕ се излива наведнъж: появява се избор
    (чист JS, renderInvoicePoChoice), зареждат се само редовете на
    избраната поръчка, а съобщението казва коя остава за отделна фактура."""
    _login(page, live_server)
    _load_materials_catalog(page, live_server, tmp_path, [
        ("GLBK400002P0012", "C-PROFILE 3   1150MM", 2.21),
        ("1TFL151621P0550", "transverse section 06  folded", 2.74),
    ])

    # Ръчна палетна карта с редове от ДВЕ поръчки.
    page.goto(live_server + "/pallet/new")
    rows = [("4700201619", "10", "GLBK400002P0012", "C-Profile", "20"),
            ("4700223566", "10", "1TFL151621P0550", "Секция", "7")]
    for i, (order, pos, ref, desc, qty) in enumerate(rows):
        if i:
            page.click('[data-add-row="pallet-items"]')
        tr = page.locator("#pallet-items tbody tr").nth(i)
        for field, value in (("order_no", order), ("pos", pos), ("reference", ref),
                             ("reference_desc", desc), ("qty", qty)):
            tr.locator('input[data-field="%s"]' % field).fill(value)
    page.fill('input[name="client_name"]', "Двупоръчков Клиент")
    page.click('#main-doc-form button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")
    number = page.content().split("ПАЛЕТНА КАРТА № ")[1][:9].strip()

    page.goto(live_server + "/invoice-br/new")
    page.fill("#f-pull-invoice-code", number)
    page.click(".invoice-pull-btn")

    # Появява се изборът на поръчка, нищо още не е заредено.
    picker = page.locator(".invoice-pull-msg select")
    picker.wait_for(timeout=8000)
    options = picker.locator("option")
    assert options.count() == 2
    assert "4700201619" in options.nth(0).inner_text()
    assert "4700223566" in options.nth(1).inner_text()
    material_inputs = page.locator(
        '#invoice-br-items tbody tr input[data-field="material_code"]')
    values = [material_inputs.nth(i).input_value() for i in range(material_inputs.count())]
    assert all(v == "" for v in values), "преди избора не се зарежда нищо"

    # Избираме втората поръчка — зареждат се САМО нейните редове.
    picker.select_option("4700223566")
    page.click(".invoice-pull-msg button")
    page.wait_for_function(
        """() => Array.from(document.querySelectorAll(
              '#invoice-br-items tbody tr input[data-field="material_code"]'))
              .some(i => i.value === '1TFL151621P0550')""",
        timeout=8000)
    values = [material_inputs.nth(i).input_value() for i in range(material_inputs.count())]
    assert "1TFL151621P0550" in values
    assert "GLBK400002P0012" not in values, "другата поръчка НЕ се зарежда тук"

    msg = page.locator(".invoice-pull-msg").inner_text()
    assert "4700201619" in msg, "съобщението казва коя поръчка остава за отделна фактура"


def test_transportation_way_dropdown_defaults_to_airfreight_and_can_switch_to_maritime(page, live_server):
    """Заявка: „за фактурите за Бразилия да се добави Transportation Way:
    AIRFREIGHT / FCA, Transportation Way: Maritime / FCA. По подразбиране
    да излиза AIRFREIGHT / FCA.“ Проверено в реален браузър, защото самият
    избор (стойността, реално подадена при submit) минава през нативен
    <select>, не просто присъствие на текст в HTML-а."""
    _login(page, live_server)
    page.goto(live_server + "/invoice-br/new")
    select = page.locator('select[name="transport_way"]')
    assert select.input_value() == "AIRFREIGHT / FCA"

    select.select_option("Maritime / FCA")
    page.fill('input[name="invoice_number"]', "МОРСКИ-1")
    page.fill('input[name="consignee_name"]', "Морски Клиент ЕООД")
    page.click('#main-doc-form button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")
    assert "Transportation Way:" in page.content()
    assert "Maritime / FCA" in page.content()


def test_editing_brazil_invoice_with_a_non_standard_transport_way_shows_it_selected(page, live_server):
    """Стара/нестандартна стойност на „Вид транспорт“ (извън двата
    варианта в менюто — напр. фактура, издадена преди то да съществува)
    не бива тихо да изчезне при отваряне за редакция: JS-ът
    (injectAndSelectOption в app.js) добавя стойността като допълнителна
    опция в менюто и я маркира като избрана, вместо селектът да остане
    без видим избор."""
    _login(page, live_server)
    page.goto(live_server + "/invoice-br/new")
    token = page.locator('#main-doc-form input[name="csrf_token"]').input_value()
    # Пращаме директно през HTTP (заобикаляйки самото падащо меню в UI) —
    # симулира стойност, записана преди менюто да съществува.
    page.request.post(live_server + "/invoice-br/new", form={
        "csrf_token": token,
        "invoice_number": "ЛЕГАСИ-1",
        "consignee_name": "Легаси Клиент ЕООД",
        "transport_way": "SEAROUTE / DAP",
    })

    page.goto(live_server + "/invoices")
    row = page.locator("tr", has_text="Легаси Клиент ЕООД")
    row.get_by_text("Редактирай").click()
    select = page.locator('select[name="transport_way"]')
    select.wait_for(timeout=8000)
    assert select.input_value() == "SEAROUTE / DAP"
    assert select.locator('option[value="SEAROUTE / DAP"]').count() == 1


def test_scanning_the_qr_public_link_opens_the_document_without_login(page, live_server, db_module):
    """Заявка: „всеки, който сканира с телефон баркода на някой от
    документите, да му се зареди директно документа, без да има нужда от
    домейна, който е в програмата“ + уточнение „само документа, нищо друго
    да не вижда“.

    Вместо реално OCR/QR декодиране на картинката (излишна тежка
    зависимост само за теста), взима public_token директно от базата —
    точно каквото QR картинката носи кодирано (виж qr_code.qr_png_data_uri)
    — и го отваря в СЪВСЕМ НОВ браузърен контекст, БЕЗ бисквитки от
    логнатата сесия по-горе — точно сценарият на чужд телефон, никога не
    влизал в програмата."""
    _login(page, live_server)
    page.goto(live_server + "/cmr/new")
    page.fill('input[name="sender_name"]', "Изпращач ЕООД")
    page.fill('input[name="consignee_name"]', "QR Е2Е Клиент ЕООД")
    page.click('button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")

    # QR-ът трябва да се вижда РЕАЛНО в браузъра — валиден data: URI, не
    # счупена картинка (проверка, която суров HTML тест не покрива).
    qr_img = page.locator(".doc-qr img")
    assert qr_img.count() == 1
    assert qr_img.get_attribute("src").startswith("data:image/png;base64,")

    con = db_module.get_db()
    row = con.execute(
        "SELECT public_token FROM documents WHERE data LIKE ?",
        ("%QR Е2Е Клиент ЕООД%",),
    ).fetchone()
    con.close()
    assert row and row["public_token"]

    context = page.context.browser.new_context()
    try:
        anon_page = context.new_page()
        anon_page.goto(live_server + "/p/" + row["public_token"])
        assert "QR Е2Е Клиент ЕООД" in anon_page.content()
        assert "ЧМР" in anon_page.content()
        assert anon_page.locator("#sidebar").count() == 0
        assert anon_page.locator(".doc-toolbar").count() == 0
    finally:
        context.close()


def test_unknown_qr_token_shows_404_without_login(page, live_server):
    context = page.context.browser.new_context()
    try:
        anon_page = context.new_page()
        resp = anon_page.goto(live_server + "/p/no-such-token-exists")
        assert resp.status == 404
    finally:
        context.close()


def test_packing_client_select_fills_contact_person_phone_and_email(page, live_server):
    """Заявка (PL.xlsx подобрение): „лице за контакти да се въвеждат в
    адресната книга и да се вмъква автоматично“ — изборът на клиент от
    адресната книга в опаковъчния лист попълва И лицето за контакт,
    телефона и имейла му (bindClientSelect в app.js — чисто клиентско JS,
    изисква реален браузър)."""
    _login(page, live_server)
    page.goto(live_server + "/clients/new")
    page.fill('input[name="name"]', "Е2Е Контакт Клиент АД")
    page.fill('input[name="city"]', "Осло")
    page.fill('input[name="contact"]', "Ola Nordmann")
    page.fill('input[name="phone"]', "+47 900 00 000")
    page.fill('input[name="email"]', "ola@example.no")
    page.click('button[type="submit"]')
    page.wait_for_url(live_server + "/clients")

    page.goto(live_server + "/packing/new")
    select = page.locator('select.client-select[data-target="receiver"]')
    option_value = select.locator("option", has_text="Е2Е Контакт Клиент АД").get_attribute("value")
    select.select_option(option_value)

    assert page.locator('input[name="receiver_name"]').input_value() == "Е2Е Контакт Клиент АД"
    assert page.locator('input[name="receiver_contact"]').input_value() == "Ola Nordmann"
    assert page.locator('input[name="receiver_phone"]').input_value() == "+47 900 00 000"
    assert page.locator('input[name="receiver_email"]').input_value() == "ola@example.no"


def test_login_scene_truck_plane_and_ship_really_move(page, live_server):
    """Заявка: „синия фон на началния екран при входа... анимиран, движещ
    камион, летящ самолет, кораб който плава“ + „подобри анимациите...
    направи го реалистично“. Проверява се в реален браузър, че CSS
    анимациите на РЕАЛИСТИЧНАТА сцена (подразбиране) наистина ВЪРВЯТ:
    изчисленият transform на всеки от трите елемента се ПРОМЕНЯ между две
    измервания, а колелата на камиона реално се въртят (не просто че
    класовете присъстват в HTML-а — това го покрива
    tests/test_login_scene.py)."""
    page.goto(live_server + "/login")
    before = {}
    for sel in (".rs-truck", ".rs-plane", ".rs-ship"):
        el = page.locator(sel)
        assert el.count() == 1, sel
        assert el.evaluate("e => getComputedStyle(e).animationName") != "none", sel
        before[sel] = el.evaluate("e => getComputedStyle(e).transform")
    wheel = page.locator(".rs-wheel").first
    before[".rs-wheel"] = wheel.evaluate("e => getComputedStyle(e).transform")
    page.wait_for_timeout(700)
    for sel, old in before.items():
        now = page.locator(sel).first.evaluate("e => getComputedStyle(e).transform")
        assert now != old, "%s не се движи" % sel

    # Формата за вход остава използваема ВЪРХУ анимацията (картата е с
    # по-висок z-index) — реален вход през сцената.
    page.fill('input[name="username"]', "e2e_admin")
    page.fill('input[name="password"]', "e2e-test-password-123")
    page.click('button[type="submit"]')
    page.wait_for_url(live_server + "/")


def test_switching_to_the_classic_login_scene_from_settings(page, live_server):
    """Заявка: „запази този и добави опция да може да се сменя в
    настройките“ — администраторът избира „Класическа“ от Системни
    настройки и входният екран показва старите силуети (и те се движат)."""
    _login(page, live_server)
    page.goto(live_server + "/my-settings")
    page.check('input[name="login_scene"][value="classic"]')
    page.click('form:has(input[name="login_scene"]) button[type="submit"]')
    page.wait_for_url(live_server + "/my-settings")

    context = page.context.browser.new_context()
    try:
        anon = context.new_page()
        anon.goto(live_server + "/login")
        truck = anon.locator(".scene-truck")
        assert truck.count() == 1
        assert anon.locator(".rs-truck").count() == 0
        t1 = truck.evaluate("e => getComputedStyle(e).transform")
        anon.wait_for_timeout(600)
        assert truck.evaluate("e => getComputedStyle(e).transform") != t1, \
            "класическият камион също се движи"
    finally:
        context.close()


def test_success_toast_appears_and_auto_dismisses(page, live_server):
    """Заявка: „съобщенията да излизат анимирано и да са по-забележими“ —
    успешното запазване показва зелен toast горе вдясно, който се скрива
    САМ след ~5 секунди (JS таймер + прогрес-линийка, initToasts в
    app.js). Проверено в реален браузър — Flask test client не изпълнява
    JS и не може да види скриването."""
    _login(page, live_server)
    page.goto(live_server + "/settings")
    page.fill('input[name="sender_name"]', "Тоуст Фирма ЕООД")
    page.click('form:has(input[name="sender_name"]) button[type="submit"]')
    toast = page.locator(".toast-success")
    toast.wait_for(timeout=8000)
    assert "Данните на фирмата изпращач са запазени." in toast.inner_text()
    assert toast.locator(".toast-bar").count() == 1, "прогрес-линийката на автоскриването"
    # Скрива се сам — без никакво действие от потребителя.
    toast.wait_for(state="detached", timeout=9000)


def test_error_toast_stays_until_manually_closed(page, live_server):
    """Грешката НЕ изчезва сама — остава, докато операторът не я затвори с
    ✕ (грешка не бива да се скрие, преди да е прочетена)."""
    _login(page, live_server)
    page.goto(live_server + "/password")
    page.fill('input[name="current"]', "грешна-парола")
    page.fill('input[name="new"]', "новапарола123")
    page.fill('input[name="repeat"]', "новапарола123")
    page.click('button[type="submit"]')
    toast = page.locator(".toast-error")
    toast.wait_for(timeout=8000)
    # Изчакваме по-дълго от таймера за автоскриване — грешката е още там.
    page.wait_for_timeout(6000)
    assert toast.count() == 1
    toast.locator(".toast-close").click()
    toast.wait_for(state="detached", timeout=3000)


def test_delete_confirm_modal_cancel_keeps_and_ok_deletes(page, live_server):
    """Стилизираният модал замества браузърния confirm(): „Отказ“ оставя
    фактурата, „Потвърди“ я изтрива наистина. Изисква реален браузър —
    целият диалог е клиентско JS (initConfirmModal в app.js)."""
    _login(page, live_server)
    page.goto(live_server + "/invoice-br/new")
    page.fill('input[name="invoice_number"]', "МОДАЛ-Е2Е-1")
    page.fill('input[name="consignee_name"]', "Модален Клиент ЕООД")
    page.click('#main-doc-form button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")

    page.goto(live_server + "/invoices")
    modal = page.locator("#confirm-modal")
    assert not modal.is_visible(), "модалът е скрит, докато не потрябва"

    row = page.locator("tr", has_text="МОДАЛ-Е2Е-1")
    row.get_by_text("Изтрий").click()
    modal.wait_for(state="visible", timeout=5000)
    assert "МОДАЛ-Е2Е-1" in page.locator("#confirm-modal-text").inner_text()

    page.click("#confirm-modal-cancel")
    modal.wait_for(state="hidden", timeout=5000)
    assert page.locator("tr", has_text="МОДАЛ-Е2Е-1").count() == 1, \
        "„Отказ“ не изтрива нищо"

    row.get_by_text("Изтрий").click()
    modal.wait_for(state="visible", timeout=5000)
    page.click("#confirm-modal-ok")
    page.wait_for_url(live_server + "/invoices*")
    assert "МОДАЛ-Е2Е-1" not in page.locator("table").inner_text() \
        if page.locator("table").count() else True


def test_added_row_animates_and_autofill_flashes_green(page, live_server):
    """Микро-анимациите: нов ред от „+ Добави ред“ носи класа .row-new
    (плавно появяване), а поле, попълнено от адресната книга, получава
    .autofilled (кратко зелено премигване) — чисто клиентско JS/CSS."""
    _login(page, live_server)

    # Клиент за автопопълването.
    page.goto(live_server + "/clients/new")
    page.fill('input[name="name"]', "Анимиран Клиент АД")
    page.fill('input[name="city"]', "Габрово")
    page.click('button[type="submit"]')
    page.wait_for_url(live_server + "/clients")

    page.goto(live_server + "/packing/new")
    # Начални редове — БЕЗ анимация (зареждане на формата не е събитие).
    assert page.locator("#packing-items tbody tr.row-new").count() == 0
    page.click('[data-add-row="packing-items"]')
    assert page.locator("#packing-items tbody tr.row-new").count() == 1

    select = page.locator('select.client-select[data-target="receiver"]')
    option_value = select.locator("option", has_text="Анимиран Клиент АД").get_attribute("value")
    select.select_option(option_value)
    name_input = page.locator('input[name="receiver_name"]')
    assert name_input.input_value() == "Анимиран Клиент АД"
    assert "autofilled" in (name_input.get_attribute("class") or "")


def test_dubai_invoice_live_totals_box_has_no_weight_line(page, live_server):
    """Заявка: „добави и фактура за Дубай“ — образецът (12971.pdf) няма
    колона с нето тегло, затова живата сума под таблицата (bindInvoiceTotals
    в app.js) не бива да показва ред „Общо нето тегло“ — за разлика от
    Бразилия, където той се показва (виж hasWeight в bindInvoiceTotals,
    задвижен от data-columns на таблицата). Проверено в реален браузър,
    защото сумата се смята изцяло на клиента при въвеждане."""
    _login(page, live_server)
    page.goto(live_server + "/invoice-dubai/new")

    first = page.locator("#invoice-dubai-items tbody tr").first
    assert first.locator('input[data-field="hs_code"]').input_value() == "85389099"
    first.locator('input[data-field="qty"]').fill("2")
    first.locator('input[data-field="unit_price"]').fill("0.72")

    totals = page.locator('.invoice-totals[data-table="invoice-dubai-items"]')
    page.wait_for_function(
        """(el) => el.innerText.indexOf('1.44') >= 0""", arg=totals.element_handle(),
        timeout=8000)
    totals_text = totals.inner_text()
    assert "Обща стойност" in totals_text
    assert "1.44" in totals_text
    assert "Общо нето тегло" not in totals_text


def test_invoice_live_js_total_matches_final_printed_total_for_rounding_edge_case(page, live_server):
    """Одит (находка К6, критична): преди поправката живата сума под
    таблицата (JS, бинарен float) и крайната сума в издадената/разпечатана
    фактура (Python, closures върху float) можеха да се разминат за
    определени комбинации количество×цена — класически пример е
    7 × 0.145 = 1.015, което „училищно” закръгляне НАГОРЕ дава 1.02, а
    двоичен float пресмята малко под .5 и закръгля НАДОЛУ до 1.01.

    Тук минаваме през ЦЕЛИЯ път, а не само единия край: попълваме 7 реда с
    количество 7 и единична цена 0.145 в реален браузър, четем сумата,
    която JS показва НА ЕКРАНА (multiplyDecimalScaled/formatScaledSum в
    app.js), издаваме фактурата и после четем сумата от СЪРВЪРНО
    РЕНДИРАНАТА разпечатка (invoice_totals/_fmt_money в appcore.py) —
    двете трябва да съвпадат точно, и двете трябва да са „7.14“
    (7 реда × 1.02 = 7.14), не „7.07“ (грешната сума при стария бъг)."""
    _login(page, live_server)
    page.goto(live_server + "/invoice-dubai/new")
    page.fill('input[name="invoice_number"]', "E2E-ROUND-1")
    page.fill('input[name="consignee_name"]', "Е2Е Закръгляне ООД")

    rows = page.locator("#invoice-dubai-items tbody tr")
    for _ in range(6):
        page.click('[data-add-row="invoice-dubai-items"]')
    assert rows.count() == 7

    for i in range(7):
        row = rows.nth(i)
        row.locator('input[data-field="qty"]').fill("7")
        row.locator('input[data-field="unit_price"]').fill("0.145")

    totals = page.locator('.invoice-totals[data-table="invoice-dubai-items"]')
    page.wait_for_function(
        """(el) => el.innerText.indexOf('7.14') >= 0""", arg=totals.element_handle(),
        timeout=8000)
    live_total_text = totals.inner_text()
    assert "7.14" in live_total_text, (
        "живата JS сума трябва да е 7.14 (закръглено на ред), не %r" % live_total_text
    )

    page.click('#main-doc-form button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")

    printed_text = page.locator("table.goods").inner_text()
    assert "7.14" in printed_text, (
        "печатната сума трябва да съвпада точно с живата JS сума (7.14), не %r"
        % printed_text
    )


def test_cmr_print_does_not_silently_clip_long_goods_description(page, live_server):
    """Одит (находка К5, критична): преди поправката `.cmr-page` при печат
    имаше ИЗРИЧНА `height: 268mm` + `overflow: hidden` — съдържание над
    физическата височина на страницата (напр. много редове в поле 9 „Вид
    на стоката“, свободен текст без ограничение на дължината) биваше
    МЪЛЧАЛИВО изрязано от разпечатката, без пренасяне на втора страница и
    без никакво предупреждение. Тук издаваме ЧМР с дълго, многоредово
    съдържание на полето и проверяваме — чрез РЕАЛЕН рендиран PDF на
    Chromium (page.pdf(), same engine потребителят вижда при „Печат/PDF“)
    — че ПОСЛЕДНИЯТ ред от текста присъства някъде в PDF-а (не само
    първите редове, каквито се събираха на една страница преди
    поправката)."""
    pdf_module = pytest.importorskip("pypdf")
    _login(page, live_server)
    page.goto(live_server + "/cmr/new")
    page.fill('input[name="sender_name"]', "Изпращач ЕООД")
    page.fill('input[name="consignee_name"]', "Дълъг Товар ЕООД")
    long_goods = "\n".join("Ред %02d — палет с материали за тест на препълване" % i
                           for i in range(1, 41))  # 40 реда, много над 1 страница
    page.fill('textarea[name="goods"]', long_goods)
    page.click('button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")

    page.emulate_media(media="print")
    pdf_bytes = page.pdf(print_background=True, prefer_css_page_size=True)
    reader = pdf_module.PdfReader(__import__("io").BytesIO(pdf_bytes))
    full_text = "\n".join(p.extract_text() or "" for p in reader.pages)

    # pypdf.extract_text() понякога вмъква допълнителни интервали заради
    # кернинга на шрифта (артефакт на извличането, не на самото съдържание)
    # — сравняваме без интервали изобщо, за да е стабилно.
    compact = full_text.replace(" ", "").replace("\n", "")
    assert reader.pages, "PDF-ът трябва да съдържа поне една страница"
    assert len(reader.pages) > 1, (
        "40-редовото поле трябваше да прелее на повече от 1 страница — "
        "ако е само 1, съдържанието вероятно пак се изрязва тихо")
    assert "Ред01" in compact
    assert "Ред40" in compact, (
        "последният ред от полето „Вид на стоката“ липсва от разпечатката — "
        "съдържанието е било изрязано вместо пренесено на следваща страница")


def test_editing_cmr_with_a_translated_packing_value_shows_it_selected(page, live_server):
    """Одит (находка В8, висок риск): "Вид на опаковката" (ЧМР) няма
    value= на <option>-ите — стойността е самият (локализиран) текст.
    Документ, записан с "Палети" (БГ текст), не бива тихо да изглежда
    изчистен при отваряне за редакция само защото текущият рендиран текст
    на опцията се различава — injectAndSelectOption в app.js трябва да
    добави точната записана стойност като опция и да я маркира избрана,
    точно както вече е доказано за transport_way по-горе в този файл."""
    _login(page, live_server)
    page.goto(live_server + "/cmr/new")
    token = page.locator('#main-doc-form input[name="csrf_token"]').input_value()
    # Симулира стойност, записана в друг език/формат от текущо рендирания
    # <option> текст (напр. записана преди превод/на друг избран език).
    page.request.post(live_server + "/cmr/new", form={
        "csrf_token": token,
        "consignee_name": "Легаси ЧМР Клиент",
        "packing": "Bulk cargo (legacy)",
    })

    page.goto(live_server + "/docs")
    row = page.locator("tr", has_text="Легаси ЧМР Клиент")
    row.get_by_text("Редактирай").click()
    select = page.locator('select[name="packing"]')
    select.wait_for(timeout=8000)
    assert select.input_value() == "Bulk cargo (legacy)"
    assert select.locator('option[value="Bulk cargo (legacy)"]').count() == 1


def test_pallet_label_format_prints_on_a_single_page(page, live_server):
    """Одит (находка В9, висок риск): .print-page.label-format имаше
    min-height: 150mm в базовото (без @media) правило — по-висока
    специфичност (две класи) от .print-page { min-height: 0 } в @media
    print, затова min-height:150mm се прилагаше И при печат. Кутията се
    раздуваше до ТОЧНО физическата височина на етикета (150мм, зададена
    през @page label), а такова изравняване "косъм под ръба" кара
    Chromium-ския печатен движок да добави ВТОРА, почти празна страница —
    доказано тук чрез реален рендиран PDF (prefer_css_page_size=True, за
    да се спази @page label, не подразбиращия се Letter формат на
    page.pdf()) + pypdf. Реалното съдържание е само ~128мм — спокойно се
    събира на един лист, щом min-height не го раздува изкуствено."""
    pdf_module = pytest.importorskip("pypdf")
    _login(page, live_server)
    page.goto(live_server + "/pallet/new")
    page.fill('input[name="client_name"]', "Клиент Етикет Едностраничен")
    page.click('#main-doc-form button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")
    page.goto(page.url + "?format=label")
    page.emulate_media(media="print")
    pdf_bytes = page.pdf(print_background=True, prefer_css_page_size=True)
    reader = pdf_module.PdfReader(__import__("io").BytesIO(pdf_bytes))
    assert len(reader.pages) == 1, (
        "етикетният формат (100×150мм) трябва да се събира на ЕДНА страница, "
        "получени са %d" % len(reader.pages)
    )


_SCAN_KEY_MAP = {
    "-": ("Minus", False),
}
for _d in "0123456789":
    _SCAN_KEY_MAP[_d] = ("Digit" + _d, False)
for _ch in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
    _SCAN_KEY_MAP[_ch] = ("Key" + _ch, True)


def _dispatch_scanner_keystrokes(page, text, garbled_key_for_letters="Ъ"):
    """Симулира ФИЗИЧЕСКИ баркод скенер (клавиатурна емулация), докато е
    активна кирилска подредба на Windows — вижте одит С4. Диспечира
    истински KeyboardEvent-и с ПРАВИЛНИЯ `code` (физическа позиция,
    независима от подредбата — точно каквото праща реален скенер), но с
    НАРОЧНО ГРЕШЕН (кирилски) `key` за буквите — `.key` е точно това,
    което Windows би превел под кирилска подредба, а `.code` е това, което
    e2e тестът тук доказва, че app.js РЕАЛНО ползва (виж initGlobalScan/
    codeToChar в static/app.js)."""
    events = []
    for ch in text:
        code, is_letter = _SCAN_KEY_MAP[ch]
        key = garbled_key_for_letters if is_letter else ch
        events.append({"code": code, "key": key, "shiftKey": is_letter})
    page.evaluate(
        """(events) => {
            for (const e of events) {
                document.dispatchEvent(new KeyboardEvent('keydown', {
                    code: e.code, key: e.key, shiftKey: e.shiftKey, bubbles: true
                }));
            }
            document.dispatchEvent(new KeyboardEvent('keydown', {
                code: 'Enter', key: 'Enter', bubbles: true
            }));
        }""",
        events,
    )


def test_global_scan_buffer_ignores_cyrillic_key_and_uses_physical_code(page, live_server):
    """Одит (находка С4, среден риск): глобалният клавиатурен буфер за
    сканиране (initGlobalScan в static/app.js) градеше низа от `e.key` —
    стойност, зависима от активната подредба на клавиатурата. При активна
    кирилска (БДС) подредба физическите клавиши на скенера (винаги
    латиница/US подредба) даваха кирилски букви вместо очакваните
    латински — документът оставаше ненамираем. Тук се издава реален ЧМР
    документ, после се „сканира“ РЕАЛНИЯТ му баркод чрез клавишни
    събития с ПРАВИЛЕН `.code`, но НАРОЧНО ГРЕШЕН (кирилски) `.key` —
    точно каквото Windows би дал при кирилска подредба — и се проверява,
    че браузърът въпреки това стига до правилния документ (буферът е
    построен само от `.code`, никога не поглежда `.key` за буквите)."""
    _login(page, live_server)
    page.goto(live_server + "/cmr/new")
    page.fill('input[name="sender_name"]', "Изпращач С4 сканиране")
    page.fill('input[name="consignee_name"]', "Клиент С4 сканиране")
    page.click('#main-doc-form button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")
    doc_url = page.url
    # Взимаме РЕАЛНО издадения баркод от списъка с документи, вместо да
    # разчитаме на CSS клас в конкретния печатен шаблон (може да липсва).
    page.goto(live_server + "/docs")
    row = page.locator("tr", has_text="Клиент С4 сканиране")
    barcode_text = row.locator("td").nth(2).inner_text().strip()

    # НАРОЧНО не сме на таблото ("/") — там #scan-input автоматично взима
    # фокус (виж "if (scan) scan.focus();" в app.js) и isEditableFocus()
    # правилно предава сканирането на СВОЕТО директно поле (сървърната
    # нормализация в bg_keyboard.py, вижте другите тестове тук), НЕ на
    # глобалния буфер — точно затова той изобщо се задейства само когато
    # НИКОЕ поле не е на фокус, какъвто е случаят на списъка с документи.
    _dispatch_scanner_keystrokes(page, barcode_text)
    page.wait_for_url(doc_url, timeout=8000)


def test_global_scan_buffer_ignores_held_down_repeat_events(page, live_server):
    """Одит (находка С4, свързано): auto-repeat (задържан клавиш) НИКОГА
    не идва от реален скенер — трябва да се игнорира изцяло, вместо да
    замърсява буфера с повторени символи. НАРОЧНО на /docs (не таблото
    „/“) — там #scan-input автоматично взима фокус и целият глобален
    буфер бездруго бездейства, което би направило теста лъжливо-успешен
    дори без поправката (виж коментара в теста за С4 по-горе)."""
    _login(page, live_server)
    page.goto(live_server + "/docs")
    page.evaluate(
        """() => {
            for (let i = 0; i < 10; i++) {
                document.dispatchEvent(new KeyboardEvent('keydown', {
                    code: 'KeyA', key: 'a', shiftKey: false, repeat: true, bubbles: true
                }));
            }
            document.dispatchEvent(new KeyboardEvent('keydown', {
                code: 'Enter', key: 'Enter', bubbles: true
            }));
        }"""
    )
    page.wait_for_timeout(300)
    assert page.url == live_server + "/docs", (
        "задържан (repeat) клавиш не биваше да натрупа/изпрати буфера за сканиране"
    )


def test_global_scan_is_suppressed_while_confirm_modal_is_open(page, live_server):
    """Одит (находка С4, свързано): isModalOpen() проверяваше само
    камерния модал — сканиране, докато е отворен диалогът за
    потвърждение (#confirm-modal), не биваше блокирано и можеше да
    отведе страницата другаде „под“ отворения диалог."""
    _login(page, live_server)
    page.goto(live_server + "/cmr/new")
    page.fill('input[name="sender_name"]', "Изпращач С4 модал")
    page.fill('input[name="consignee_name"]', "Клиент С4 модал")
    page.click('#main-doc-form button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")

    page.goto(live_server + "/docs")
    row = page.locator("tr", has_text="Клиент С4 модал")
    row.get_by_text("Изтрий").click()
    modal = page.locator("#confirm-modal")
    modal.wait_for(state="visible", timeout=5000)

    docs_url = page.url
    _dispatch_scanner_keystrokes(page, "0001-0000")
    page.wait_for_timeout(300)
    assert page.url == docs_url, (
        "сканиране, докато диалогът за потвърждение е отворен, не биваше да "
        "отвежда страницата другаде"
    )
    assert modal.is_visible(), "диалогът за потвърждение не биваше да изчезне"


def test_issue_button_is_blocked_from_a_second_click_before_the_response_arrives(page, live_server):
    """Одит (находка С5, среден риск): нито една форма за издаване нямаше
    data-busy/деактивиране на бутона — реално възпроизведено с два еднакви
    POST-а, всеки изял пореден номер (/doc/1 и /doc/2), а вторият остава
    като дупка в номерацията, ако не се изтрие от админ.

    За да проверим НАДЕЖДНО, БЕЗ да се състезаваме с реалната навигация
    (Playwright изчаква и click(), и evaluate() до края на последвалата
    навигация, което би „погълнало“ всеки опит да хванем прозореца между
    двата клика), тук изрично СПИРАМЕ самото изпращане на формата с ВТОРИ
    'submit' слушател (capture-фаза, регистриран СЛЕД initBusyForms — виж
    app.js, извиква се СЛЕД него, значи .btn-busy вече е добавен, преди
    да спрем реалното изпращане) — остава ни стабилен прозорец, в който
    Playwright-ски (истински, hit-testing) втори клик реално да провери
    дали .btn-busy (pointer-events:none) го блокира."""
    _login(page, live_server)
    page.goto(live_server + "/cmr/new")
    page.fill('input[name="sender_name"]', "Изпращач С5 двоен клик")
    page.fill('input[name="consignee_name"]', "Клиент С5 двоен клик")
    page.evaluate(
        "document.getElementById('main-doc-form')"
        ".addEventListener('submit', function (e) { e.preventDefault(); })"
    )

    btn = page.locator('#main-doc-form button[type="submit"]').first
    btn.click()
    page.wait_for_selector("#main-doc-form button.btn-busy", timeout=2000)
    with pytest.raises(Exception):
        btn.click(timeout=800)


def test_issue_button_targets_the_actually_clicked_submit_button_not_always_the_first(page, live_server):
    """Одит (находка С5, свързано): формата има ВТОРИ submit бутон
    („Предварителен преглед“, различен formaction) — ако data-busy винаги
    маркира ПЪРВИЯ бутон в DOM реда (form.querySelector), реално
    натиснатият „Преглед“ бутон би останал напълно кликаем (вижте
    e.submitter поправката в initBusyForms, static/app.js)."""
    _login(page, live_server)
    page.goto(live_server + "/cmr/new")
    page.fill('input[name="sender_name"]', "Изпращач С5 преглед")
    page.fill('input[name="consignee_name"]', "Клиент С5 преглед")
    page.evaluate(
        "document.getElementById('main-doc-form')"
        ".addEventListener('submit', function (e) { e.preventDefault(); })"
    )

    preview_btn = page.locator('#main-doc-form button[formaction*="preview"]').first
    issue_btn = page.locator('#main-doc-form button[type="submit"]').first
    preview_btn.click()
    page.wait_for_selector("#main-doc-form button.btn-busy", timeout=2000)
    assert "btn-busy" in (preview_btn.get_attribute("class") or ""), (
        "точно натиснатият бутон (Преглед) трябва да получи .btn-busy"
    )
    assert "btn-busy" not in (issue_btn.get_attribute("class") or ""), (
        "бутонът, който НЕ е бил натиснат (Издай), не биваше да се маркира"
    )


def test_invoice_material_lookup_failure_visibly_marks_the_field(page, live_server):
    """Одит (находка С6, среден риск): bindInvoiceMaterialLookup имаше
    НАПЪЛНО празен .catch — при мрежова грешка полето „Net weight“ просто
    оставаше празно, неразличимо от „проверихме, материалът наистина няма
    зададено тегло“. Тук провалваме нарочно самата заявка към справочника
    (route.abort) и проверяваме, че полето получава ВИДИМ маркер
    (.lookup-failed + подсказка), а не остава мълчаливо празно."""
    _login(page, live_server)
    page.goto(live_server + "/invoice-br/new")
    page.route("**/materials/lookup*", lambda route: route.abort())

    row = page.locator("#invoice-br-items tbody tr").first
    code_input = row.locator('input[data-field="material_code"]')
    weight_input = row.locator('input[data-field="net_weight"]')
    code_input.fill("GLBK400002P0012")
    code_input.blur()

    page.wait_for_selector(
        '#invoice-br-items tbody tr input[data-field="net_weight"].lookup-failed', timeout=5000)
    assert weight_input.input_value() == ""
    assert weight_input.get_attribute("title"), (
        "полето трябва да получи обяснителна подсказка при неуспешна проверка"
    )


def test_fetch_calls_show_session_expired_message_not_generic_error(page, live_server):
    """Одит (находка С6, среден риск): при изтекла сесия fetch следва
    пренасочването към /login (обикновен 200 с HTML), r.json() гърми със
    SyntaxError и попада в общия .catch — потребителят вижда неясното
    „Грешка при заявката.“ вместо да разбере, че трябва да влезе отново.
    Тук симулираме точно тази верига (302 → реалната /login страница) през
    route interception на „Зареди от издадена палетна карта“ (фактура) и
    проверяваме, че показаното съобщение е КОНКРЕТНО за изтекла сесия, не
    генеричното „Грешка при заявката.“."""
    _login(page, live_server)
    page.goto(live_server + "/invoice-br/new")

    def _redirect_to_login(route):
        route.fulfill(status=302, headers={"Location": "/login"})

    page.route("**/invoice/pull-pallet", _redirect_to_login)

    page.fill(".invoice-pull-code", "0001/2026")
    page.click(".invoice-pull-btn")

    msg = page.locator(".invoice-pull-msg")
    page.wait_for_function(
        """() => document.querySelector('.invoice-pull-msg').textContent.trim() !== '' &&
                document.querySelector('.invoice-pull-msg').textContent.trim() !== 'Търсене…'""",
        timeout=5000)
    assert "изтекла" in msg.inner_text(), (
        "трябва да покаже конкретното съобщение за изтекла сесия, не генерично"
    )
    assert "Грешка при заявката" not in msg.inner_text()


def test_cancelling_confirm_modal_leaves_the_button_clickable_again(page, live_server):
    """Одит (находка С8, среден риск): формата „Изтегли от GitHub“
    (my_settings.html) има И data-busy, И data-confirm — .btn-busy се
    слагаше при ВСЯКО подаване, включително прихванатото от диалога за
    потвърждение, и НИКОГА не се махаше при „Отказ“ — бутонът оставаше
    навечно некликаем (pointer-events:none) до презареждане на
    страницата. Тук кликаме, отказваме, и проверяваме, че бутонът реално
    приема втори клик."""
    _login(page, live_server)
    page.goto(live_server + "/my-settings")
    # Спираме реалната навигация след потвърждение — тестът проверява
    # само поведението на модала/бутона, не самото изтегляне от GitHub.
    page.evaluate(
        "document.querySelectorAll('form[data-confirm]').forEach("
        "function (f) { f.addEventListener('submit', function (e) { e.preventDefault(); }); })"
    )

    btn = page.locator(
        'form[action$="/system/pull-now"] button[type="submit"]').first
    btn.click()
    modal = page.locator("#confirm-modal")
    modal.wait_for(state="visible", timeout=5000)
    assert "btn-busy" in (btn.get_attribute("class") or "")

    page.click("#confirm-modal-cancel")
    modal.wait_for(state="hidden", timeout=5000)
    assert "btn-busy" not in (btn.get_attribute("class") or ""), (
        ".btn-busy трябваше да се махне при Отказ — бутонът остава завинаги некликаем иначе"
    )

    # Реален втори клик трябва отново да отвори диалога (доказва, че
    # бутонът действително е кликаем, не само че CSS класът липсва).
    btn.click()
    modal.wait_for(state="visible", timeout=5000)

# ---------------------------------------------------------------- Дребни в JS


def test_client_select_blank_option_clears_stale_previous_client_data(page, live_server):
    """Одит (Дребни): bindClientSelect правеше `parseInt(select.value,10)`
    → NaN при празната опция „— изберете клиент —“ → клиент не се намира
    → кодът просто спираше, БЕЗ да изчисти полетата — данните на
    ПРЕДИШНО избрания клиент оставаха видимо попълнени, макар падащото
    меню да показва „не е избран“, с риск да се издаде документ с грешен
    клиент."""
    _login(page, live_server)
    page.goto(live_server + "/clients/new")
    page.fill('input[name="name"]', "Е2Е Изчистване Клиент ЕООД")
    page.fill('input[name="city"]', "Пловдив")
    page.fill('input[name="phone"]', "+359 88 000 0000")
    page.click('button[type="submit"]')
    page.wait_for_url(live_server + "/clients")

    page.goto(live_server + "/packing/new")
    select = page.locator('select.client-select[data-target="receiver"]')
    option_value = select.locator("option", has_text="Е2Е Изчистване Клиент ЕООД").get_attribute("value")
    select.select_option(option_value)
    assert page.locator('input[name="receiver_name"]').input_value() == "Е2Е Изчистване Клиент ЕООД"

    select.select_option("")
    assert page.locator('input[name="receiver_name"]').input_value() == "", (
        "връщането на падащото меню към „— изберете клиент —“ трябва да "
        "изчисти полетата, не да остави данните на предишния клиент"
    )
    assert page.locator('input[name="receiver_phone"]').input_value() == ""


def test_invoice_client_select_blank_option_clears_stale_previous_client_data(page, live_server):
    """Същото за bindInvoiceClientSelect (адресната книга на фактурите)."""
    _login(page, live_server)
    page.goto(live_server + "/invoices/clients/new")
    page.fill('input[name="name"]', "Е2Е Фактура Изчистване ООД")
    page.fill('input[name="delivery_name"]', "Delivery Co")
    page.fill('input[name="delivery_phone"]', "+1 555 0100")
    page.click('button[type="submit"]')
    page.wait_for_url(live_server + "/invoices/clients")

    page.goto(live_server + "/invoice-br/new")
    page.select_option("#f-invoice-client-select", label="Е2Е Фактура Изчистване ООД")
    assert page.locator('input[name="consignee_name"]').input_value() == "Delivery Co"

    page.select_option("#f-invoice-client-select", "")
    assert page.locator('input[name="consignee_name"]').input_value() == "", (
        "връщането към празния избор трябва да изчисти consignee_name, "
        "не да остави данните на предишния клиент"
    )
    assert page.locator('input[name="consignee_phone"]').input_value() == ""


def test_invoice_total_net_weight_shows_dash_not_zero_when_nothing_entered(page, live_server):
    """Одит (Дребни): invoiceFmt(0, ...) връща "0" (не ""), затова „Общо
    нето тегло“ показваше „0 кг“ дори когато НИЩО не е въведено —
    подвеждащо, изглежда като потвърдено нулево тегло."""
    _login(page, live_server)
    page.goto(live_server + "/invoice-br/new")
    totals = page.locator('.invoice-totals')
    assert "Общо нето тегло" in totals.inner_text()
    assert "0 кг" not in totals.inner_text()
    assert "—" in totals.inner_text()

    # С попълнен ред (количество + тегло) сумата вече трябва да се покаже.
    page.fill('input[data-field="qty"]', "3")
    page.fill('input[data-field="net_weight"]', "2")
    page.locator('input[data-field="net_weight"]').blur()
    page.wait_for_timeout(150)
    assert "Общо нето тегло: <b>6" in totals.inner_html()


def test_dynamic_item_row_inputs_have_aria_label_from_column_header(page, live_server):
    """Одит (Дребни, достъпност): полетата на динамичните редове се
    създаваха без name/id/aria-label — екранен четец ги обявяваше като
    голи текстови полета, без връзка към заглавието на колоната."""
    _login(page, live_server)
    page.goto(live_server + "/packing/new")
    first_row_inputs = page.locator("#packing-items tbody tr").first.locator("input")
    count = first_row_inputs.count()
    assert count > 0
    for i in range(count):
        label = first_row_inputs.nth(i).get_attribute("aria-label")
        assert label, "динамичното поле на ред %d няма aria-label" % i


def test_camera_modal_closes_on_escape_key(page, live_server):
    """Одит (Дребни, достъпност): камерният модал се затваряше само с ✕
    бутона — клавиатурен потребител нямаше как да излезе с Escape (за
    разлика от #confirm-modal, който вече поддържа Escape/клик по фона)."""
    _login(page, live_server)
    page.goto(live_server + "/")
    page.click("#camera-scan-btn")
    modal = page.locator("#camera-scan-modal")
    modal.wait_for(state="visible", timeout=5000)

    page.keyboard.press("Escape")
    modal.wait_for(state="hidden", timeout=5000)


def test_camera_modal_closes_on_backdrop_click(page, live_server):
    """Същото, но за клик по фона зад модала (извън диалоговия правоъгълник)."""
    _login(page, live_server)
    page.goto(live_server + "/")
    page.click("#camera-scan-btn")
    modal = page.locator("#camera-scan-modal")
    modal.wait_for(state="visible", timeout=5000)

    modal.click(position={"x": 2, "y": 2})
    modal.wait_for(state="hidden", timeout=5000)


def test_item_row_numeric_zero_value_is_not_eaten_by_falsy_check(page, live_server, db_module):
    """Одит (Дребни): `item[col] || rowDefaults[col] || ""` в initItemsTable
    изяжда числовата 0 — ако запазен ред съдържа ЧИСЛОВА (не низова) 0 за
    поле (напр. количество), редакцията показва подразбиращата се
    стойност вместо истинската 0. Нормалният път на приложението винаги
    сериализира стойностите като низове ("0" е truthy в JS), затова тук
    директно инжектираме ИСТИНСКА JSON числова 0 в записа — сценарият,
    срещу който защитата трябва да работи, независимо дали днес има жив
    път до него."""
    import json as json_module

    _login(page, live_server)
    page.goto(live_server + "/pallet/new")
    tr = page.locator("#pallet-items tbody tr").first
    for field, value in (("order_no", "ORD-0"), ("reference", "ART-0"),
                         ("reference_desc", "Стартов ред"), ("qty", "1")):
        tr.locator('input[data-field="%s"]' % field).fill(value)
    page.fill('input[name="client_name"]', "Е2Е нула Клиент")
    page.click('#main-doc-form button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")

    con = db_module.get_db()
    row = con.execute(
        "SELECT id, data FROM documents WHERE data LIKE ?", ("%Е2Е нула Клиент%",)
    ).fetchone()
    doc_id = row["id"]
    data = json_module.loads(row["data"])
    # "orders" формат (подразбиращият се за нова палетна карта) — качваме
    # ЧИСЛОВА (не низова) 0 за qty, каквато нормалният JS/HTTP път никога
    # не би произвел сам (виж докстринга по-горе).
    data["items"] = [{"order_no": "ORD-0", "reference": "ART-0",
                      "reference_desc": "Нулев ред", "qty": 0}]
    con.execute("UPDATE documents SET data = ? WHERE id = ?", (json_module.dumps(data), doc_id))
    con.commit()
    con.close()

    page.goto(live_server + "/doc/%d/edit" % doc_id)
    qty_input = page.locator('input[data-field="qty"]').first
    qty_input.wait_for(timeout=5000)
    assert qty_input.input_value() == "0", (
        "числовата 0, запазена в редa, не биваше да се замени мълчаливо с "
        "подразбиращата се стойност/празно поле"
    )


def test_editing_a_document_then_previewing_then_going_back_keeps_editing_it(page, live_server):
    """Заявка: „при връщане назад от преглед за печат въведената информация
    се губи“ — виж tests/test_preview_edit_restore.py за пълното обяснение
    на причината и unit-ниво покритие; тук е пълният реален браузърен цикъл
    (типово нещо, което test client не може да провери): редактираш
    издаден документ → „Предварителен преглед“ → „Назад към формата“ →
    остава на /doc/<id>/edit (не /cmr/new) с редактираните (все още
    незаписани) стойности → „Запази промените“ действително обновява ТОЗИ
    документ, без да създава дубликат."""
    _login(page, live_server)
    page.goto(live_server + "/cmr/new")
    page.fill('input[name="consignee_name"]', "Оригинален Клиент")
    page.click('#main-doc-form button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")
    doc_id = page.url.rstrip("/").split("/")[-1]

    page.goto(live_server + "/doc/%s/edit" % doc_id)
    page.fill('input[name="consignee_name"]', "РЕДАКТИРАН УНИКАЛЕН КЛИЕНТ")
    page.click('button[formaction*="preview"]')
    page.wait_for_url("**/preview/*")
    assert "РЕДАКТИРАН УНИКАЛЕН КЛИЕНТ" in page.content()

    page.click("text=Назад към формата")
    page.wait_for_load_state("networkidle")
    assert "/edit" in page.url, (
        "landed on %r instead of staying in edit mode — the edit context "
        "(which document is being updated) is lost" % page.url
    )
    assert page.locator('input[name="consignee_name"]').input_value() == "РЕДАКТИРАН УНИКАЛЕН КЛИЕНТ"
    save_btn_text = page.locator('#main-doc-form button[type="submit"]').first.text_content()
    assert "Запази промените" in save_btn_text

    # complete the loop: actually save, confirm the SAME document (doc_id)
    # was updated, not a duplicate new one created.
    page.click('#main-doc-form button[type="submit"]')
    page.wait_for_url(live_server + "/doc/*")
    assert page.url.rstrip("/").split("/")[-1] == doc_id, (
        "saving after restore created/redirected to a DIFFERENT document "
        "than the one being edited"
    )
    assert "РЕДАКТИРАН УНИКАЛЕН КЛИЕНТ" in page.content()
    page.goto(live_server + "/docs")
    # first row inside tbody is a header row, not data — only ONE real
    # document should exist overall, no accidental duplicate.
    data_rows = page.locator("table.list tbody tr").filter(has_text="РЕДАКТИРАН УНИКАЛЕН КЛИЕНТ")
    assert data_rows.count() == 1, "a duplicate document was created (expected exactly 1)"
