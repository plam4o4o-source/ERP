# -*- coding: utf-8 -*-
"""Регресионни тестове за „валута само евро“ на паричните полета на
товарителницата за вътрешен превоз (transport_price/extra_costs) — заявка:
„направи всичко което предлагаш и да остане валута само евро“. Вижте
appcore.format_eur_amount за самото форматиране (споделено между формата,
печатния шаблон, Excel и PDF износа)."""
import io

from pypdf import PdfReader

from conftest import post_with_csrf

from appcore import format_eur_amount


# ---------------------------------------------------------------- appcore.format_eur_amount (unit)

def test_format_eur_amount_appends_euro_sign():
    assert format_eur_amount("500") == "500 €"
    assert format_eur_amount("500.50") == "500.50 €"


def test_format_eur_amount_tolerates_already_tagged_values():
    # Стари документи, въведени преди тази промяна — не пренаписваме насила.
    assert format_eur_amount("500 €") == "500 €"
    assert format_eur_amount("500 EUR") == "500 EUR"
    assert format_eur_amount("500 eur") == "500 eur"


def test_format_eur_amount_handles_empty_and_none():
    assert format_eur_amount(None) == ""
    assert format_eur_amount("") == ""
    assert format_eur_amount("   ") == ""  # само whitespace се третира като липсващо


def test_format_eur_amount_strips_surrounding_whitespace_value():
    assert format_eur_amount("  500  ") == "500 €"


# ---------------------------------------------------------------- форма (waybill_form.html)

def test_waybill_form_labels_show_eur_unit(admin_client):
    resp = admin_client.get("/waybill/new")
    body = resp.data.decode()
    assert "Превозна цена, EUR" in body
    assert "Допълнителни разходи, EUR" in body
    # старите означения без единица вече не трябва да съществуват самостоятелно
    assert "label for=\"f-transport_price\">Превозна цена<" not in body
    assert "label for=\"f-extra_costs\">Допълнителни разходи<" not in body


# ---------------------------------------------------------------- печат (waybill_print.html)

def _issue_waybill(client, extra=None):
    data = {"sender_name": "Изпращач", "transport_price": "1200", "extra_costs": "50"}
    data.update(extra or {})
    resp = post_with_csrf(client, "/waybill/new", data, csrf_source_url="/waybill/new",
                          follow_redirects=False)
    assert resp.status_code == 302, resp.data
    return resp.headers["Location"].rstrip("/").split("/")[-1]


def test_waybill_print_shows_amounts_with_euro_sign(admin_client):
    doc_id = _issue_waybill(admin_client)
    resp = admin_client.get("/doc/%s" % doc_id)
    body = resp.data.decode()
    assert "1200 €" in body
    assert "50 €" in body
    assert "EUR" in body  # заглавието на клетката вече показва (EUR)


def test_waybill_print_handles_missing_amounts(admin_client):
    doc_id = _issue_waybill(admin_client, {"transport_price": "", "extra_costs": ""})
    resp = admin_client.get("/doc/%s" % doc_id)
    assert resp.status_code == 200  # не гърми при празни стойности


# ---------------------------------------------------------------- Excel износ

def test_waybill_xlsx_export_labels_and_formats_amounts(admin_client):
    from openpyxl import load_workbook

    doc_id = _issue_waybill(admin_client)
    resp = admin_client.get("/doc/%s/export.xlsx" % doc_id)
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    rows = {row[0].value: row[1].value for row in ws.iter_rows() if row[0].value}
    assert rows.get("Превозна цена, EUR") == "1200 €"
    assert rows.get("Допълнителни разходи, EUR") == "50 €"


# ---------------------------------------------------------------- PDF износ

def test_waybill_pdf_export_shows_euro_amounts(admin_client):
    doc_id = _issue_waybill(admin_client)
    resp = admin_client.get("/doc/%s/export.pdf" % doc_id)
    assert resp.status_code == 200
    reader = PdfReader(io.BytesIO(resp.data))
    text = "\n".join(page.extract_text() for page in reader.pages)
    assert "1200 €" in text
    assert "50 €" in text
    assert "Превозна цена, EUR" in text


def test_format_eur_amount_does_not_double_tag_legacy_bgn_values():
    """Не беше в обхвата на самия патч: докстрингът обещаваше толерантност
    към стари стойности в лева, но кодът добавяше „€“ и към тях — стара
    товарителница със „500 лв.“ се показваше като „500 лв. €“, тоест с две
    валути наведнъж. Полето беше свободен текст преди v3.31.0, затова
    такива стойности реално съществуват в вече издадени документи."""
    assert format_eur_amount("500 лв.") == "500 лв."
    assert format_eur_amount("500 лв") == "500 лв"
    assert format_eur_amount("500 ЛВ.") == "500 ЛВ."
    assert format_eur_amount("500 BGN") == "500 BGN"
    assert format_eur_amount("500 bgn") == "500 bgn"


def test_format_eur_amount_still_tags_plain_numbers():
    """Поправката за старите левови стойности не бива да е отменила
    основното поведение — обикновено число пак получава „€“."""
    assert format_eur_amount("1200.50") == "1200.50 €"
    assert format_eur_amount("0") == "0 €"
