# -*- coding: utf-8 -*-
"""Регресионни тестове за одит №4 (19.08.2026) — находки №14, №25, №26,
№27, №28, №29, №38, №39, №40.

Обединяващата тема е „входът отвън“: качен Excel файл (импорт на поръчки,
на редове за фактура и на справочника материали), адресната книга, която
се вгражда във всяка форма, и копието на износа в клиентската папка. И
трите пътя досега мълчаха точно там, където операторът има нужда да знае:
файл се приемаше наполовина без предупреждение, форма теглеше мегабайти
без причина, а провалено копие изглеждаше като успешно.
"""
import io
import os

import pytest
from openpyxl import Workbook

import client_export
import materials
import routes_invoices
import routes_pallet_extra
from conftest import post_with_csrf

# ---------------------------------------------------------------- помощни

_ORDER_HEADER = ["Order No", "Pos", "Reference", "Reference Desc", "Open Qty", ""]


def _orders_xlsx(rows, header=None, top_rows=(), sheets=None):
    """.xlsx със справка за поръчки (форматът и на двата импорта).

    `top_rows` са декоративни редове НАД заглавния; `sheets` (по избор) е
    списък от (име, редове) за файл с няколко листа — виж находка №29."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in top_rows:
        ws.append(list(r))
    ws.append(list(header or _ORDER_HEADER))
    for r in rows:
        ws.append(list(r))
    for name, extra in (sheets or []):
        ws2 = wb.create_sheet(name)
        for r in extra:
            ws2.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _catalog_xlsx(rows, headers=("ABB part ID", "Description", "Net weight\n[KG/pc]"),
                  top_rows=()):
    wb = Workbook()
    ws = wb.active
    for r in top_rows:
        ws.append(list(r))
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class _CountingSheet:
    """Изкуствен лист, който БРОИ колко реда реално са прочетени.

    Единственият надежден начин да се докаже находка №14: истинският
    дефект не е в резултата (той винаги е бил правилен — първите 5000
    реда), а в това КОЛКО се чете, за да се стигне до него."""

    def __init__(self, total_rows, header=None):
        self.total_rows = total_rows
        self.header = list(header or _ORDER_HEADER)
        self.rows_read = 0

    def iter_rows(self, values_only=True):
        yield tuple(self.header)
        self.rows_read += 1
        for i in range(self.total_rows):
            self.rows_read += 1
            yield ("ORD-%d" % i, "10", "REF-%d" % i, "Материал %d" % i, 1, 1)


# ================================================================ №14
# Таванът от 5000 реда не пазеше нито паметта, нито времето — целият файл
# се материализираше с list(ws.iter_rows()) ПРЕДИ рязането.

_LIMIT = (routes_pallet_extra._HEADER_SCAN_ROWS
          + routes_pallet_extra._MAX_IMPORT_DATA_ROWS + 1)


@pytest.mark.parametrize("parse,module", [
    (routes_pallet_extra._parse_order_export, routes_pallet_extra),
    (routes_invoices._parse_invoice_items_xlsx, routes_invoices),
])
def test_import_reads_only_the_rows_it_can_actually_use(parse, module):
    """Одит (19.08.2026, находка №14): преди поправката и двата разбора
    правеха `list(ws.iter_rows(...))` — тоест изчерпваха ЦЕЛИЯ лист (при
    файл от 9 MB с 300 000 реда: 27.8 сек и +148 MB RSS), след което
    любезно съобщаваха, че са заредили само първите 5000."""
    sheet = _CountingSheet(200000)
    parsed, warnings = parse(sheet)
    assert parsed, "разборът трябва да върне редове както преди"
    # Заглавие + таван на данните + 1 ред за откриване на орязването,
    # плюс най-много един допълнителен `next()` при проверката.
    assert sheet.rows_read <= _LIMIT + 1, (
        "прочетени са %d реда — целият лист пак влиза в паметта" % sheet.rows_read)


@pytest.mark.parametrize("parse", [
    routes_pallet_extra._parse_order_export,
    routes_invoices._parse_invoice_items_xlsx,
])
def test_truncation_warning_is_raised_exactly_when_the_file_is_longer(parse):
    """Предупреждението за орязване вече се вдига по НЕИЗЧЕРПАН итератор,
    а не по дължината на прочетен цял списък — поведението пред оператора
    трябва да е идентично на старото."""
    limit = routes_pallet_extra._MAX_IMPORT_DATA_ROWS

    exact = _CountingSheet(limit)
    _parsed, warnings = parse(exact)
    assert not any("повече от" in w for w in warnings), (
        "точно %d реда НЕ е орязване" % limit)

    one_more = _CountingSheet(limit + 1)
    _parsed, warnings = parse(one_more)
    assert any("повече от" in w for w in warnings)


def test_pallet_import_still_loads_at_most_the_row_cap(admin_client):
    """Резултатът пред оператора е непроменен: 5001 реда → 5000 внесени +
    предупреждение."""
    rows = [("ORD-%d" % i, "10", "REF-%d" % i, "Материал", 1, 1) for i in range(5001)]
    resp = post_with_csrf(
        admin_client, "/pallet/bulk-import", {"excel_file": (_orders_xlsx(rows), "big.xlsx")},
        csrf_source_url="/pallet/new", content_type="multipart/form-data")
    body = resp.data.decode()
    assert "повече от 5000 реда" in body
    assert "5000 реда общо" in body


# ================================================================ №29
# Импортът четеше само първия лист.

def test_pallet_import_finds_the_data_sheet_behind_a_decorative_one(admin_client):
    """Одит (19.08.2026, находка №29): файл с декоративен лист „Инфо“
    отпред и валидни колони на втория лист отказваше с „Файлът не съдържа
    разпознаваеми колони“ — при напълно валиден файл."""
    wb = Workbook()
    info = wb.active
    info.title = "Инфо"
    info.append(["Справка за поръчки", None])
    info.append(["Генерирана на", "19.08.2026"])
    data = wb.create_sheet("Данни")
    data.append(list(_ORDER_HEADER))
    data.append(["ORD-77", "10", "REF-77", "Материал А", 6, 1])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = post_with_csrf(
        admin_client, "/pallet/bulk-import", {"excel_file": (buf, "poruchki.xlsx")},
        csrf_source_url="/pallet/new", content_type="multipart/form-data")
    body = resp.data.decode()
    assert "не съдържа разпознаваеми колони" not in body
    assert "ORD-77" in body
    # …и е ясно ОТ КОЙ лист са данните.
    assert "Данни" in body


def test_invoice_import_finds_the_data_sheet_behind_a_decorative_one(admin_client):
    wb = Workbook()
    info = wb.active
    info.title = "Инфо"
    info.append(["Справка", None])
    data = wb.create_sheet("Данни")
    data.append(list(_ORDER_HEADER))
    data.append(["ORD-88", "20", "REF-88", "Материал Б", 3, ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = post_with_csrf(
        admin_client, "/invoice/import-items", {"excel_file": (buf, "poruchki.xlsx")},
        csrf_source_url="/invoice-br/new", content_type="multipart/form-data")
    payload = resp.get_json()
    assert payload["ok"] is True
    assert payload["count"] == 1
    assert any("Данни" in w for w in payload.get("warnings", []))


# ================================================================ №40
# Съобщението при изместен заглавен ред не обясняваше причината.

def test_header_below_the_scan_window_says_where_the_header_must_be(admin_client):
    """Одит (19.08.2026, находка №40): заглавие на ред 15 даваше ТОЧНО
    същото съобщение като напълно грешен файл."""
    top = [["Декоративен ред %d" % i] for i in range(14)]
    resp = post_with_csrf(
        admin_client, "/pallet/bulk-import",
        {"excel_file": (_orders_xlsx([("ORD-1", "10", "REF-1", "Мат", 1, 1)], top_rows=top),
                        "poruchki.xlsx")},
        csrf_source_url="/pallet/new", content_type="multipart/form-data",
        follow_redirects=True)
    body = resp.data.decode()
    assert "не съдържа разпознаваеми колони" in body
    assert "първите 10 реда" in body


def test_invoice_import_error_also_explains_the_header_window(admin_client):
    top = [["Декоративен ред %d" % i] for i in range(14)]
    resp = post_with_csrf(
        admin_client, "/invoice/import-items",
        {"excel_file": (_orders_xlsx([("ORD-1", "10", "REF-1", "Мат", 1, "")], top_rows=top),
                        "poruchki.xlsx")},
        csrf_source_url="/invoice-br/new", content_type="multipart/form-data")
    assert "първите 10 реда" in resp.get_json()["error"]


# ================================================================ №39
# Формули без кеширана стойност се внасяха като празни.

def _formula_orders_xlsx():
    """Файл, какъвто ГЕНЕРИРА библиотека (не самият Excel): формулите
    нямат кеширана стойност, затова `data_only=True` ги чете като None."""
    wb = Workbook()
    ws = wb.active
    ws.append(list(_ORDER_HEADER))
    ws.append(["ORD-1", "10", "REF-1", "Материал А", "=2+3", 1])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_pallet_import_warns_about_formulas_without_stored_values(admin_client):
    """Одит (19.08.2026, находка №39): редът се внасяше с `qty: ''`, а
    съобщението гласеше „Открити са 1 палетни карти (1 реда общо)“ — на
    вид напълно успешен импорт."""
    resp = post_with_csrf(
        admin_client, "/pallet/bulk-import",
        {"excel_file": (_formula_orders_xlsx(), "generated.xlsx")},
        csrf_source_url="/pallet/new", content_type="multipart/form-data")
    body = resp.data.decode()
    assert "формули без запазени стойности" in body


def test_invoice_import_warns_about_formulas_without_stored_values(admin_client):
    resp = post_with_csrf(
        admin_client, "/invoice/import-items",
        {"excel_file": (_formula_orders_xlsx(), "generated.xlsx")},
        csrf_source_url="/invoice-br/new", content_type="multipart/form-data")
    payload = resp.get_json()
    assert any("формули без запазени стойности" in w for w in payload.get("warnings", []))


def test_a_normal_file_never_gets_the_formula_warning(admin_client):
    """Обратната посока — предупреждението не бива да се появява при
    обикновен файл с истински числа."""
    resp = post_with_csrf(
        admin_client, "/pallet/bulk-import",
        {"excel_file": (_orders_xlsx([("ORD-1", "10", "REF-1", "Мат", 6, 1)]), "ok.xlsx")},
        csrf_source_url="/pallet/new", content_type="multipart/form-data")
    assert "формули без запазени стойности" not in resp.data.decode()


# ================================================================ №38
# Справочникът материали не беше получил защитите на находка №18.

def test_catalog_import_caps_the_number_of_rows():
    rows = [("CODE-%05d" % i, "Материал %d" % i, 1.0) for i in range(5100)]
    stats = {}
    entries = materials.parse_catalog_xlsx(_catalog_xlsx(rows), stats=stats)
    assert len(entries) == materials._MAX_IMPORT_DATA_ROWS
    assert stats["truncated"] is True


def test_catalog_import_reports_the_header_row_when_it_is_not_the_first():
    stats = {}
    entries = materials.parse_catalog_xlsx(
        _catalog_xlsx([("X-1", "Нещо", 1.5)],
                      top_rows=[[], ["Ценоразпис H2'26", None, None]]),
        stats=stats)
    assert entries == [("X-1", "Нещо", "1.5")]
    assert stats["header_row"] == 3


def test_catalog_import_counts_duplicate_codes_inside_the_file():
    """Одит (19.08.2026, находка №38): дублиран код тихо презаписваше
    първия с ДРУГО тегло, а съобщението рапортуваше „1 нови и 1
    обновени“, все едно става дума за два различни материала."""
    stats = {}
    materials.parse_catalog_xlsx(_catalog_xlsx([
        ("DUP-1", "Материал", 1.0),
        ("DUP-1", "Материал (нова версия)", 2.0),
        ("dup-1", "Същият, с малки букви", 3.0),
    ]), stats=stats)
    assert stats["duplicate_codes"] == 2


def test_catalog_import_warns_about_merged_cells(con):
    wb = Workbook()
    ws = wb.active
    ws.append(["ABB part ID", "Description", "Net weight\n[KG/pc]"])
    ws.append(["X-1", "Нещо", 1.5])
    ws.merge_cells("A5:B5")
    buf = io.BytesIO()
    wb.save(buf)
    stats = {}
    materials.parse_catalog_xlsx(buf.getvalue(), stats=stats)
    assert stats["merged_cells"] is True


def test_catalog_import_screen_shows_the_warnings(admin_client):
    """Числата от разбора реално стигат до оператора (а не само до
    вътрешния речник)."""
    payload = _catalog_xlsx([
        ("W-1", "Тегло като текст", "nan"),
        ("W-2", "Дублиран", 1.0),
        ("w-2", "Дублиран с малки букви", 2.0),
    ])
    resp = post_with_csrf(
        admin_client, "/materials/import",
        {"excel_file": (io.BytesIO(payload), "catalog.xlsx")},
        csrf_source_url="/materials", content_type="multipart/form-data",
        follow_redirects=True)
    body = resp.data.decode()
    assert "неразпознато тегло" in body
    assert "вече се среща по-горе" in body


# ================================================================ №28а
# „nan“ като ТЕКСТ и отрицателно тегло влизаха сурови в справочника.

@pytest.mark.parametrize("raw", ["nan", "NaN", "N/A", "n/a", "—", "-", "inf",
                                 "#VALUE!", "няма данни"])
def test_non_numeric_weight_never_reaches_the_catalog(raw):
    """Одит (19.08.2026, находка №28а): поправката на находка №26 покри
    само `float('nan')`. ТЕКСТОВА клетка със същото съдържание падаше в
    последния клон към суровия `_cellstr` и се записваше БУКВАЛНО — после
    се попълва автоматично в поле „Net weight“ на фактура, вижда се на
    официалната бланка и се отхвърля при сумирането."""
    entries = materials.parse_catalog_xlsx(_catalog_xlsx([("X-1", "Нещо", raw)]))
    assert entries == [("X-1", "Нещо", "")]


def test_negative_weight_is_rejected_and_counted():
    stats = {}
    entries = materials.parse_catalog_xlsx(
        _catalog_xlsx([("X-1", "Нещо", -2.5), ("X-2", "Друго", 1.25)]), stats=stats)
    assert entries == [("X-1", "Нещо", ""), ("X-2", "Друго", "1.25")]
    assert stats["bad_weights"] == 1


def test_valid_weights_are_untouched_by_the_new_filter():
    """Регресионна застраховка: филтърът не бива да „изяде“ нищо валидно
    (включително много малки тегла, идващи от Excel като научна нотация —
    находка №26 от 16.08)."""
    entries = dict((c, w) for c, _d, w in materials.parse_catalog_xlsx(_catalog_xlsx([
        ("A-1", "цяло", 3),
        ("A-2", "дробно", 2.74259375),
        ("A-3", "много малко", 8.7135e-05),
        ("A-4", "с запетая като текст", "1,25"),
        ("A-5", "нула", 0),
    ])))
    assert entries == {"A-1": "3", "A-2": "2.742594", "A-3": "0.000087",
                       "A-4": "1.25", "A-5": "0"}


# ================================================================ №28б
# lookup и lookup_many връщаха различно тегло за един и същ материал.

def test_codes_differing_only_by_case_stay_one_material(con):
    """Одит (19.08.2026, находка №28б): „abc-77“ и „ABC-77“ живееха като
    ДВА реда с различно тегло; `lookup` връщаше точното съвпадение, а
    `lookup_many` сливаше вариантите (печелеше последният) — ръчно
    въведеният код и същият код през Excel импорт даваха РАЗЛИЧНИ
    килограми в ЕДНА И СЪЩА фактура."""
    stats = {}
    materials.replace_catalog(con, [("ABC-77", "Материал", "1.5")])
    added, updated = materials.replace_catalog(
        con, [("abc-77", "Същият материал", "9.75")], stats=stats)

    assert materials.count(con) == 1, "разлика само в регистъра не бива да прави втори ред"
    assert (added, updated) == (0, 1)
    assert stats["case_conflicts"] == 1

    one = materials.lookup(con, "abc-77")
    many = materials.lookup_many(con, ["abc-77"])["abc-77"]
    assert one["net_weight"] == many["net_weight"] == "9.75"
    assert materials.lookup(con, "ABC-77")["net_weight"] == "9.75"


def test_database_refuses_a_second_row_differing_only_by_case(con):
    """Защитата е и на ниво база (db._m010) — не само в кода за зареждане,
    за да не може друг път да я заобиколи."""
    import sqlite3

    con.execute("INSERT INTO materials (code, description, net_weight)"
                " VALUES ('QQ-1', 'Материал', '1')")
    with pytest.raises(sqlite3.IntegrityError):
        con.execute("INSERT INTO materials (code, description, net_weight)"
                    " VALUES ('qq-1', 'Дубликат', '2')")


def test_existing_case_duplicates_are_collapsed_by_the_migration(tmp_db_path, monkeypatch):
    """Съществуваща база отпреди миграцията: двата варианта се сливат в
    най-скоро обновения, вместо програмата да откаже да стартира."""
    import sqlite3

    import db as db_mod
    monkeypatch.setattr(db_mod, "DB_PATH", tmp_db_path)
    monkeypatch.setattr(db_mod, "SECRET_PATH", tmp_db_path + ".secret")

    raw = sqlite3.connect(tmp_db_path)
    raw.executescript(
        "CREATE TABLE materials (code TEXT PRIMARY KEY, description TEXT NOT NULL DEFAULT '',"
        " net_weight TEXT NOT NULL DEFAULT '', updated_at TEXT NOT NULL DEFAULT '');"
        "INSERT INTO materials VALUES ('ABC-77', 'старо', '1.5', '2026-01-01 10:00:00');"
        "INSERT INTO materials VALUES ('abc-77', 'ново', '9.75', '2026-08-01 10:00:00');")
    raw.commit()
    raw.close()

    db_mod.init_db()
    con = db_mod.get_db()
    try:
        rows = con.execute("SELECT code, net_weight FROM materials").fetchall()
        assert len(rows) == 1
        assert rows[0]["net_weight"] == "9.75", "пази се най-скоро обновеният ред"
    finally:
        con.close()


# ================================================================ №27
# Два различни клиента → една папка.

def test_long_client_names_do_not_collapse_into_one_folder():
    """Одит (19.08.2026, находка №27): `cleaned[:120]` режеше сляпо — два
    реални клиента с дълги имена, различаващи се чак НАКРАЯ, получаваха
    ИДЕНТИЧНА папка и износите им се смесваха."""
    common = ("Дружество с ограничена отговорност за международна спедиция, "
              "логистика и митническо представителство ЕКСПРЕС ТРАНС ГРУП БЪЛГАРИЯ")
    a = client_export.sanitize_client_folder_name(common + " — клон Пловдив")
    b = client_export.sanitize_client_folder_name(common + " — клон Варна")
    assert len(common + " — клон Пловдив") > 120
    assert a != b
    assert len(a) <= 120 and len(b) <= 120


def test_short_client_names_are_not_touched_by_the_hash():
    """Обичайното име НЕ бива да получава хеш — папките остават четими."""
    assert client_export.sanitize_client_folder_name("ACME ООД") == "ACME ООД"


@pytest.mark.parametrize("name", ["PRN.txt", "prn.txt", "AUX.2", "CON.log", "nul.dat"])
def test_reserved_windows_names_with_an_extension_are_rejected(name):
    """Одит (19.08.2026, находка №27): проверката сравняваше ЦЕЛИЯ низ, а
    Windows отказва `PRN.txt` също толкова, колкото и `PRN` — `mkdir`
    гърмеше и копието се губеше тихо (виж находка №26)."""
    assert client_export.sanitize_client_folder_name(name) == "Без_име"


def test_folder_lookup_is_case_insensitive(tmp_path):
    """Одит (19.08.2026, находка №27): „фирма ООД“ и „ФИРМА ООД“ даваха
    ДВЕ папки на Linux и ЕДНА на Windows — тоест износът на един и същ
    клиент се озоваваше на различни места според машината."""
    base = str(tmp_path)
    first = client_export.client_export_path(base, "фирма ООД", "a.xlsx")
    second = client_export.client_export_path(base, "ФИРМА ООД", "b.xlsx")
    assert os.path.dirname(first) == os.path.dirname(second)
    assert len(os.listdir(base)) == 1


# ================================================================ №26
# Провалено копие в клиентската папка беше невидимо за оператора.

def test_export_status_distinguishes_disabled_from_failed(tmp_path):
    blocking = tmp_path / "не_е_папка.txt"
    blocking.write_text("файл", encoding="utf-8")
    disabled = client_export.save_client_export_status(
        {}, "pallet", {"client_name": "X"}, "f.xlsx", b"x")
    failed = client_export.save_client_export_status(
        {"client_export_auto": "on", "client_export_dir": str(blocking)},
        "pallet", {"client_name": "X"}, "f.xlsx", b"x")
    ok = client_export.save_client_export_status(
        {"client_export_auto": "on", "client_export_dir": str(tmp_path)},
        "pallet", {"client_name": "X"}, "f.xlsx", b"x")
    assert disabled == client_export.EXPORT_SKIPPED
    assert failed == client_export.EXPORT_FAILED
    assert ok == client_export.EXPORT_OK


def _issue_pallet(admin_client):
    resp = post_with_csrf(admin_client, "/pallet/new", {
        "client_name": "Клиент За Копие",
        "items_json": '[{"order_no": "ORD-1", "pos": "10", "reference": "REF-1",'
                      ' "reference_desc": "Материал", "qty": "2"}]',
        "items_format": "orders",
    }, csrf_source_url="/pallet/new")
    assert resp.status_code == 302
    return int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])


def test_failed_client_folder_copy_is_flashed_to_the_operator(admin_client, db_module,
                                                              tmp_path):
    """Одит (19.08.2026, находка №26): върнатата стойност се игнорираше и
    на двете места; единствената следа беше ред в лог файл, който
    потребител на .exe никога не отваря. Свалянето УСПЯВА, така че
    операторът остава убеден, че копието е и на общия диск."""
    doc_id = _issue_pallet(admin_client)

    blocking = tmp_path / "не_е_папка.txt"
    blocking.write_text("файл, не директория", encoding="utf-8")
    con = db_module.get_db()
    db_module.save_settings(con, {"client_export_auto": "on",
                                  "client_export_dir": str(blocking)})
    con.commit()
    con.close()

    resp = admin_client.get("/doc/%d/export.xlsx" % doc_id)
    assert resp.status_code == 200, "самото сваляне остава непокътнато"
    assert resp.data[:2] == b"PK"

    # Отговорът е файл, не HTML — съобщението се показва при следващата
    # страница, точно както всички останали известия.
    body = admin_client.get("/docs").data.decode()
    assert "копието в клиентската папка НЕ беше записано" in body


def test_successful_client_folder_copy_says_nothing(admin_client, db_module, tmp_path):
    doc_id = _issue_pallet(admin_client)
    # Отделна поддиректория: самата tmp_path съдържа и файла на тестовата
    # база (виж conftest.tmp_db_path).
    target = tmp_path / "клиентски_копия"
    target.mkdir()
    con = db_module.get_db()
    db_module.save_settings(con, {"client_export_auto": "on",
                                  "client_export_dir": str(target)})
    con.commit()
    con.close()

    assert admin_client.get("/doc/%d/export.xlsx" % doc_id).status_code == 200
    body = admin_client.get("/docs").data.decode()
    assert "НЕ беше записано" not in body
    assert os.listdir(str(target)) == ["Клиент За Копие"]


# ================================================================ №25
# Всяка форма теглеше пълния списък клиенти (~2 MB HTML).

def _add_clients(db_module, count, prefix="Клиент"):
    con = db_module.get_db()
    for i in range(count):
        con.execute(
            "INSERT INTO clients (name, alias, address, city, postcode, country, eik,"
            " vat, phone, email, contact) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            ("%s фирма номер %04d ООД" % (prefix, i), "KL%04d" % i,
             "ул. Тестова %d, бл. 5, вх. А" % i, "София", "1000", "България",
             "20%09d" % i, "BG20%09d" % i, "+359 2 %07d" % i,
             "office%04d@example.com" % i, "Иван Иванов %d" % i))
    con.commit()
    con.close()


def test_clients_list_is_paginated_and_searchable(admin_client, db_module):
    """Одит (19.08.2026, находка №25): адресната книга беше единственият
    голям списък БЕЗ пагинация и БЕЗ сървърно търсене — при 5 000 записа
    583 ms и 5 891 KB HTML за едно отваряне."""
    from routes_clients import PAGE_SIZE

    _add_clients(db_module, PAGE_SIZE + 20)

    first = admin_client.get("/clients")
    assert first.status_code == 200
    assert first.data.decode().count("/clients/") < PAGE_SIZE * 6  # груба горна граница
    assert "фирма номер 0000 ООД" in first.data.decode()
    assert "фирма номер 0119 ООД" not in first.data.decode(), "втората страница не се рендира"

    second = admin_client.get("/clients?page=2")
    assert "фирма номер 0119 ООД" in second.data.decode()

    found = admin_client.get("/clients?q=0042").data.decode()
    assert "фирма номер 0042 ООД" in found
    assert "фирма номер 0043 ООД" not in found


def test_clients_search_is_case_insensitive_for_cyrillic(admin_client, db_module):
    _add_clients(db_module, 3, prefix="Възлов")
    body = admin_client.get("/clients?q=%D0%B2%D1%8A%D0%B7%D0%BB%D0%BE%D0%B2").data.decode()
    assert "Възлов фирма номер 0000 ООД" in body


def test_form_embeds_at_most_the_client_limit(admin_client, db_module):
    """Формата вече не носи цялата адресна книга — при 5 000 клиента това
    бяха над 2 MB HTML при ВСЯКО отваряне."""
    from appcore import CLIENT_EMBED_LIMIT

    _add_clients(db_module, CLIENT_EMBED_LIMIT + 50)
    body = admin_client.get("/cmr/new").data.decode()
    assert body.count("<option value=\"") <= CLIENT_EMBED_LIMIT + 40  # + типове опаковка и др.
    assert "data-clients-total=\"%d\"" % (CLIENT_EMBED_LIMIT + 50) in body \
        or "data-clients-total='%d'" % (CLIENT_EMBED_LIMIT + 50) in body
    assert "фирма номер 0000 ООД" in body, "първите клиенти остават вградени"
    assert "фирма номер 0349 ООД" not in body


def test_small_address_book_is_still_embedded_whole(admin_client, db_module):
    """При типична инсталация (десетки клиенти) НИЩО не се променя —
    целият списък си остава вграден и автодовършването работи без мрежа."""
    _add_clients(db_module, 25)
    body = admin_client.get("/cmr/new").data.decode()
    for i in range(25):
        assert "фирма номер %04d ООД" % i in body


def test_client_lookup_endpoint_returns_full_data_for_autocomplete(admin_client, db_module):
    """Автодовършването е ключова функция: намереният по AJAX клиент носи
    ТОЧНО същите полета като вградения, включително пунктовете за
    разтоварване (иначе ЧМР формата би загубила избора им)."""
    import db as db_mod

    _add_clients(db_module, 5)
    con = db_module.get_db()
    client_id = con.execute("SELECT id FROM clients ORDER BY id LIMIT 1").fetchone()["id"]
    db_mod.save_unload_points(con, client_id, [
        {"label": "Склад 2", "address": "ул. Складова 1", "city": "Пловдив",
         "postcode": "4000", "country": "България"}])
    con.commit()
    con.close()

    payload = admin_client.get("/clients/lookup?q=KL0000").get_json()
    assert payload["ok"] is True
    assert len(payload["clients"]) == 1
    entry = payload["clients"][0]
    for field in ("id", "name", "address", "city", "postcode", "country",
                  "eik", "vat", "phone", "email", "contact", "unload_points"):
        assert field in entry
    assert entry["unload_points"][0]["label"] == "Склад 2"


def test_client_lookup_requires_login(client):
    assert client.get("/clients/lookup?q=x").status_code in (302, 401)


def test_invoice_clients_list_is_paginated_and_searchable(admin_client, db_module):
    con = db_module.get_db()
    for i in range(120):
        con.execute("INSERT INTO invoice_clients (name, delivery_name, billing_name)"
                    " VALUES (?, ?, ?)",
                    ("Запис %04d" % i, "Доставка %04d" % i, "Фактуриране %04d" % i))
    con.commit()
    con.close()

    first = admin_client.get("/clients" if False else "/invoices/clients").data.decode()
    assert "Запис 0000" in first
    assert "Запис 0119" not in first
    assert "Запис 0119" in admin_client.get("/invoices/clients?page=2").data.decode()

    found = admin_client.get("/invoices/clients?q=0042").data.decode()
    assert "Запис 0042" in found
    assert "Запис 0043" not in found
