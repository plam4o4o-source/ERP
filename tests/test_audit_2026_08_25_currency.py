# -*- coding: utf-8 -*-
"""Одит (25.08.2026, находка №11): валутата „EURO“ трябва да е еднаква на
бланката и в износа.

Печатната бланка показва „Currency: EURO“ (d.currency or 'EURO'), а Excel/PDF
износът пишеше суровото поле — за фактура без изрична валута (изчистено поле
или стара фактура отпреди полето) бланката казваше „EURO“, а Excel клетката
„Валута“ оставаше празна. Всички суми са фиксирано в евро, затова
подразбиращата се валута е EURO навсякъде.
"""
import io
import json

import openpyxl
from conftest import post_with_csrf


def _issue_br_invoice(admin_client, extra=None):
    data = {"consignee_name": "Клиент",
            "items_json": json.dumps([{"description": "Стока", "qty": "1",
                                       "unit_price": "2", "net_weight": "1"}],
                                     ensure_ascii=False)}
    if extra:
        data.update(extra)
    r = post_with_csrf(admin_client, "/invoice-br/new", data,
                       csrf_source_url="/invoice-br/new", follow_redirects=False)
    assert r.status_code == 302
    return int(r.headers["Location"].rstrip("/").rsplit("/", 1)[-1])


def _currency_cell(admin_client, doc_id):
    xr = admin_client.get("/doc/%d/export.xlsx" % doc_id)
    assert xr.status_code == 200
    ws = openpyxl.load_workbook(io.BytesIO(xr.data)).active
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Валута":
                return ws.cell(row=cell.row, column=cell.column + 1).value
    raise AssertionError("полето „Валута“ липсва в износа")


def test_excel_currency_defaults_to_euro_when_empty(admin_client):
    """Празна валута → Excel показва „EURO“, както бланката."""
    doc_id = _issue_br_invoice(admin_client)  # без подадено currency
    assert _currency_cell(admin_client, doc_id) == "EURO"
    # И самата бланка го показва.
    body = admin_client.get("/doc/%d" % doc_id).get_data(as_text=True)
    assert "EURO" in body


def test_excel_currency_keeps_an_explicit_value(admin_client):
    """Ако операторът е въвел изрична валута, тя се пази (не се заменя)."""
    doc_id = _issue_br_invoice(admin_client, {"currency": "USD"})
    assert _currency_cell(admin_client, doc_id) == "USD"
