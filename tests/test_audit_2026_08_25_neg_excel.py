# -*- coding: utf-8 -*-
"""Одит (25.08.2026, находка №12): отрицателен ред и Excel SUM.

Всички суми в проекта изключват отрицателните редове като невалидни
(находка С1). Excel износът обаче записваше отрицателното количество като
ИСТИНСКО число в колоната — значи получателят, който сумира колоната с
=SUM(), получаваше различна сума от отпечатания TOTAL (който изключва реда):
количества 10 и −3 → печатен TOTAL 10, а Excel SUM 7. Сега отрицателната
стойност остава ВИДИМА като текст (както на бланката), но извън числовата
сума — печатният TOTAL и всеки =SUM() по колоната съвпадат.
"""
import io
import json

import openpyxl
from conftest import post_with_csrf


def _export_br_invoice_with_negative_row(admin_client):
    items = json.dumps([
        {"description": "Стока А", "qty": "10", "unit_price": "2", "net_weight": "1"},
        {"description": "Връщане", "qty": "-3", "unit_price": "2", "net_weight": "1"},
    ], ensure_ascii=False)
    r = post_with_csrf(admin_client, "/invoice-br/new",
                       {"consignee_name": "Клиент", "items_json": items},
                       csrf_source_url="/invoice-br/new", follow_redirects=False)
    assert r.status_code == 302
    doc_id = int(r.headers["Location"].rstrip("/").rsplit("/", 1)[-1])
    xr = admin_client.get("/doc/%d/export.xlsx" % doc_id)
    assert xr.status_code == 200
    return openpyxl.load_workbook(io.BytesIO(xr.data)).active


def _col_index(ws, header_text):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == header_text:
                return cell.row, cell.column
    raise AssertionError("колоната %r не е намерена" % header_text)


def _total_row_index(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "TOTAL":
                return cell.row
    raise AssertionError("TOTAL редът не е намерен")


def test_negative_qty_is_not_a_numeric_cell(admin_client):
    ws = _export_br_invoice_with_negative_row(admin_client)
    header_row, qty_col = _col_index(ws, "Количество")
    total_row = _total_row_index(ws)

    # Числови количества по РЕДОВЕТЕ (между заглавието и TOTAL реда).
    data_numeric = []
    negative_seen = False
    for r in range(header_row + 1, total_row):
        v = ws.cell(row=r, column=qty_col).value
        if isinstance(v, (int, float)):
            data_numeric.append(v)
        elif isinstance(v, str) and v.strip() == "-3":
            negative_seen = True

    assert negative_seen, "отрицателният ред трябва да остане ВИДИМ като текст"
    assert all(v >= 0 for v in data_numeric), "в колоната остана отрицателно ЧИСЛО"
    # =SUM() по редовите числови клетки дава 10 (само редът с 10), не 7.
    assert sum(data_numeric) == 10, "=SUM() по редовете трябва да е 10, не 7"

    # И съвпада с отпечатания TOTAL на количеството.
    total_qty = ws.cell(row=total_row, column=qty_col).value
    assert total_qty == 10, "печатният TOTAL за количеството трябва да е 10"
