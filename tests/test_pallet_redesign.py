# -*- coding: utf-8 -*-
"""Регресионни тестове за преработката на палетната карта — заявка:

„в палетна карта замени брой кашони със вид опаковка (палет, кашон,
опаковка), тип палет като се избере друг да се отварят прозорци за
въвеждане ръчно на размерите, нетно тегло замени със общ брой - сумата
количество от палетната карта, в тип палет остави само размерите без
текст, премахни Импорт от ексел, добави бутон за добавяне на следваща
палетна карта“

Покрива: премахнатия единичен Excel импорт (форма+route+примерен файл),
„Вид опаковка“ вместо „Брой кашони“, изчислимо „Общ брой“ вместо ръчно
„Нето, кг“, опростените „Тип палет“ опции (само размери), ръчно въведен
(„Друг“, през модал в браузъра — тук проверено директно през POST, тъй
като HTTP тестовете не изпълняват JS) тип палет, и новия „+ Добави
следваща палетна карта“ композитор (bulk-issue с items_format != orders).
Самият JS/модал е допълнително проверен ръчно с Playwright (headless
Chromium), виж CHANGELOG.md."""
import io
import json

import pytest
from openpyxl import Workbook

from conftest import post_with_csrf


def _issue_pallet(admin_client, extra_fields=None):
    data = {"sender_name": "Тест ЕООД", "client_name": "Клиент ООД"}
    if extra_fields:
        data.update(extra_fields)
    resp = post_with_csrf(admin_client, "/pallet/new", data, csrf_source_url="/pallet/new",
                          follow_redirects=False)
    assert resp.status_code == 302, resp.data[:300]
    return admin_client.get(resp.headers["Location"])


# ---------------------------------------------------------------- премахнат единичен Excel импорт

def test_pallet_import_route_removed(admin_client):
    resp = post_with_csrf(admin_client, "/pallet/import", {}, csrf_source_url="/pallet/new")
    assert resp.status_code == 404


def test_pallet_sample_route_removed(admin_client):
    resp = admin_client.get("/pallet/sample.xlsx")
    assert resp.status_code == 404


def test_pallet_form_has_no_single_excel_import_section(admin_client):
    resp = admin_client.get("/pallet/new")
    body = resp.data.decode()
    assert "Импорт от Excel" not in body
    assert "Импорт от справка за поръчки" in body  # bulk импорта си остава


# ---------------------------------------------------------------- „Вид опаковка“ вместо „Брой кашони“

def test_pallet_form_has_packaging_type_not_boxes(admin_client):
    resp = admin_client.get("/pallet/new")
    body = resp.data.decode()
    assert 'name="packaging_type"' in body
    assert 'name="boxes"' not in body
    assert 'name="net"' not in body
    assert "Вид опаковка" in body
    assert "Общ брой" in body


def test_pallet_new_saves_and_prints_packaging_type(admin_client):
    items = json.dumps([{"code": "ART-1", "description": "Стока", "qty": "3"},
                        {"code": "ART-2", "description": "Стока 2", "qty": "4"}])
    resp = _issue_pallet(admin_client, {
        "packaging_type": "Кашон / Box", "items_json": items,
    })
    body = resp.data.decode()
    assert "Кашон / Box" in body
    assert "ВИД ОПАКОВКА" in body


# ---------------------------------------------------------------- „Общ брой“ (изчислимо)

def _issue_pallet_via_bulk_composer(admin_client, items_json):
    """Общ брой отпадна от печатната бланка (заявка 12.08.2026, второ
    уточнение — pallet_print.html вече не показва .pallet-total-qty), но
    appcore.pallet_total_qty() продължава да се вика и показва на
    списъчната страница pallet_bulk_result.html — единственото останало
    място, където сървърен рендер (без JS) може да провери изчислената
    стойност. Тук минаваме през същия „+ Добави следваща палетна карта“
    композитор (bulk-issue с groups=1), който static/app.js ползва и за
    ЕДИНСТВЕНА карта, щом бъде добавена втора и после премахната — вижте
    test_pallet_bulk_issue_via_composer_creates_multiple_documents по-долу
    за пълния 2-карти вариант."""
    data = {
        "sender_name": "Тест ЕООД", "client_name": "Клиент ООД",
        "groups": "1",
        "pallet_type_1": "120×80", "packaging_type_1": "Палет / Pallet",
        "gross_1": "10",
        "items_json_1": items_json, "items_format_1": "manual",
    }
    resp = post_with_csrf(admin_client, "/pallet/bulk-issue", data,
                          csrf_source_url="/pallet/new", follow_redirects=False)
    assert resp.status_code == 302
    return admin_client.get(resp.headers["Location"])


def test_pallet_total_qty_is_computed_from_items_not_stored_field(admin_client):
    items = json.dumps([{"code": "A", "description": "X", "qty": "2.5"},
                        {"code": "B", "description": "Y", "qty": "3"}])
    resp = _issue_pallet_via_bulk_composer(admin_client, items)
    body = resp.data.decode()
    assert resp.status_code == 200
    assert "5.5" in body  # 2.5 + 3


def test_pallet_total_qty_ignores_non_numeric_rows(admin_client):
    items = json.dumps([{"code": "A", "description": "X", "qty": "abc"},
                        {"code": "B", "description": "Y", "qty": "4"}])
    resp = _issue_pallet_via_bulk_composer(admin_client, items)
    body = resp.data.decode()
    assert resp.status_code == 200
    assert ">4<" in body or ">4 " in body or "\n4\n" in body or "4</div>" in body


def test_appcore_pallet_total_qty_helper_directly():
    from appcore import pallet_total_qty
    assert pallet_total_qty([{"qty": "1"}, {"qty": "2"}]) == "3"
    assert pallet_total_qty([{"qty": "1.5"}, {"qty": "1.5"}]) == "3"
    assert pallet_total_qty([{"qty": "2"}, {"qty": "abc"}]) == "2"
    assert pallet_total_qty([]) == ""
    assert pallet_total_qty(None) == ""
    assert pallet_total_qty([{"qty": ""}]) == ""


# ---------------------------------------------------------------- „Тип палет“ — размери, „Друг“ модал

def test_pallet_type_options_are_dimensions_only(admin_client):
    resp = admin_client.get("/pallet/new")
    body = resp.data.decode()
    assert "EUR / EPAL" not in body
    assert "Индустриален" not in body
    assert "Половин палет" not in body
    assert ">120×80<" in body
    assert 'value="__other__"' in body


def test_pallet_type_custom_dimensions_from_modal_round_trip(admin_client):
    """Самият модал е JS/браузър UI (проверено ръчно с Playwright) — тук
    се проверява само крайният резултат: свободна стойност за pallet_type
    (каквато модалът генерира, напр. "150×100") се записва и печата
    нормално, без validate-грешка."""
    resp = _issue_pallet(admin_client, {"pallet_type": "150×100"})
    assert "150×100" in resp.data.decode()


# ---------------------------------------------------------------- „+ Добави следваща палетна карта“

def test_pallet_bulk_issue_via_composer_creates_multiple_documents(admin_client):
    """Симулира точно това, което JS композиторът (initPalletMultiCard в
    static/app.js) подава при 2 ръчно попълнени карти: суфиксирани полета
    + групи + items_format_N="manual" (НЕ "orders" — редовете тук са
    код/описание/кол./тегло, не поръчки)."""
    items_1 = json.dumps([{"code": "A1", "description": "Стока 1", "qty": "2", "weight": "10"}])
    items_2 = json.dumps([{"code": "B1", "description": "Стока 2", "qty": "5", "weight": "20"}])
    data = {
        "sender_name": "Изпращач", "client_name": "Клиент композитор",
        "doc_date": "2026-08-06", "notes": "Обща бележка",
        "groups": "1,2",
        "pallet_type_1": "120×80", "packaging_type_1": "Палет / Pallet",
        "gross_1": "12", "height_1": "90",
        "items_json_1": items_1, "items_format_1": "manual",
        "pallet_type_2": "80×60", "packaging_type_2": "Кашон / Box",
        "gross_2": "22", "height_2": "70",
        "items_json_2": items_2, "items_format_2": "manual",
    }
    resp = post_with_csrf(admin_client, "/pallet/bulk-issue", data,
                          csrf_source_url="/pallet/new", follow_redirects=False)
    assert resp.status_code == 302
    result_resp = admin_client.get(resp.headers["Location"])
    assert result_resp.status_code == 200
    body = result_resp.data.decode()
    assert "Клиент композитор" in body
    assert "Палет / Pallet" in body
    assert "Кашон / Box" in body
    # „Общ брой“ вместо старото „Нето“
    assert "Кашони" not in body
    assert "Нето" not in body

    # номерацията "N от M" се генерира на сървъра (виж _collect_bulk_pallet_drafts)
    import db as db_mod
    con = db_mod.get_db()
    rows = con.execute(
        "SELECT data FROM documents WHERE doc_type='pallet' ORDER BY id DESC LIMIT 2"
    ).fetchall()
    con.close()
    pallet_nos = sorted(json.loads(r["data"])["pallet_no"] for r in rows)
    assert pallet_nos == ["1 от 2", "2 от 2"]

    # редовете от втората карта НЕ трябва да мигрират в първата и обратно
    datas = [json.loads(r["data"]) for r in rows]
    formats = set(d.get("items_format") for d in datas)
    assert "orders" not in formats  # композиторът подава "manual", не "orders"


def test_pallet_bulk_issue_composer_items_do_not_show_order_columns_on_print(admin_client):
    items_1 = json.dumps([{"code": "X1", "description": "Проверка колони", "qty": "1"}])
    data = {
        "client_name": "Клиент", "groups": "1",
        "packaging_type_1": "Опаковка / Package",
        "items_json_1": items_1, "items_format_1": "manual",
    }
    resp = post_with_csrf(admin_client, "/pallet/bulk-issue", data,
                          csrf_source_url="/pallet/new", follow_redirects=False)
    assert resp.status_code == 302
    import db as db_mod
    con = db_mod.get_db()
    row = con.execute(
        "SELECT id FROM documents WHERE doc_type='pallet' ORDER BY id DESC LIMIT 1"
    ).fetchone()
    con.close()
    view_resp = admin_client.get("/doc/%d" % row["id"])
    body = view_resp.data.decode()
    assert "Проверка колони" in body
    assert "Поръчка №" not in body  # orders-форматът НЕ трябва да се появи


# ---------------------------------------------------------------- Excel износ

def test_pallet_xlsx_export_has_packaging_type_and_computed_total(admin_client):
    items = json.dumps([{"code": "A", "description": "X", "qty": "3"},
                        {"code": "B", "description": "Y", "qty": "4"}])
    resp = _issue_pallet(admin_client, {"packaging_type": "Палет / Pallet", "items_json": items})
    # doc id от Location преди последния admin_client.get (viev_document) —
    # вместо това вземаме директно от базата, по-надеждно.
    import db as db_mod
    con = db_mod.get_db()
    row = con.execute("SELECT id FROM documents WHERE doc_type='pallet' ORDER BY id DESC LIMIT 1").fetchone()
    con.close()

    xlsx_resp = admin_client.get("/doc/%d/export.xlsx" % row["id"])
    assert xlsx_resp.status_code == 200
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_resp.data))
    ws = wb.active
    all_values = [cell.value for r in ws.iter_rows() for cell in r if cell.value is not None]
    assert "Вид опаковка" in all_values
    assert "Палет / Pallet" in all_values
    assert "Общ брой" in all_values
    assert 7 in all_values or "7" in all_values  # 3 + 4
    assert "Брой кашони" not in all_values
    assert "Нето, кг" not in all_values


def test_pallet_bulk_import_keeps_only_the_five_selected_columns(admin_client):
    """Заявка: „палетна карта да зарежда само информацията от следните
    колони и да съдържа само тях... Order No, Pos, Reference, Reference
    Desc (да се зарежда информацията когато я има във файла), Open Qty,
    друго не променяй“ — останалите колони на файла (Due Date, Project,
    Unit, Stock) трябва да се ИГНОРИРАТ, дори да присъстват."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Due Date", "Order No", "Pos", "Project", "Reference",
              "Reference Desc", "Open Qty", "Unit", "Stock", ""])
    ws.append(["2026-09-01", "ORD-1", "10", "PRJ-1", "REF-1", "Материал А", 6, "PCS", "WH1", 1])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = post_with_csrf(
        admin_client, "/pallet/bulk-import", {"excel_file": (buf, "poruchki.xlsx")},
        csrf_source_url="/pallet/new", follow_redirects=False, content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    body = resp.data.decode()
    # Запазените 5 колони — присъстват.
    assert "ORD-1" in body
    assert "REF-1" in body
    assert "Материал А" in body
    assert '"qty": "6"' in body
    # Игнорираните колони — тяхното СЪДЪРЖАНИЕ не бива да се появи никъде
    # (нито в data-items JSON, нито в самата таблица за преглед).
    assert "PRJ-1" not in body
    assert "PCS" not in body
    assert "WH1" not in body
    assert "2026-09-01" not in body


def test_parse_order_export_plus_separated_group_numbers_duplicates_row_into_each_card():
    """Не беше в обхвата на самия патч, тъй като предполагаше тази логика
    вече съществуваща (реконструирана лично по спецификацията в
    докстринговете на _parse_order_export/CHANGELOG.md — виж
    _parse_group_numbers в routes_pallet_extra.py). Ред с „1+3“ в
    групиращата колона трябва да отиде И в двете карти; ред с обикновен
    единичен номер — само в своята."""
    from openpyxl import Workbook
    from routes_pallet_extra import _parse_order_export

    wb = Workbook()
    ws = wb.active
    ws.append(["Due Date", "Order No", "Pos", "Project", "Reference",
              "Reference Desc", "Open Qty", "Unit", "Stock", ""])
    ws.append(["2026-09-01", "ORD-1", "10", "PRJ-1", "REF-1", "Споделен материал", 6, "PCS", "WH1", "1+3"])
    ws.append(["2026-09-02", "ORD-2", "20", "PRJ-1", "REF-2", "Само карта 2", 4, "PCS", "WH2", "2"])
    ws.append(["2026-09-03", "ORD-3", "30", "PRJ-1", "REF-3", "Само карта 3", 8, "PCS", "WH1", "3"])

    groups = _parse_order_export(ws)
    assert groups is not None
    assert sorted(groups.keys()) == [1, 2, 3]
    assert len(groups[1]) == 1
    assert groups[1][0]["reference_desc"] == "Споделен материал"
    assert len(groups[2]) == 1
    assert groups[2][0]["reference_desc"] == "Само карта 2"
    assert len(groups[3]) == 2  # споделеният ред + собствения ѝ ред
    descs = sorted(it["reference_desc"] for it in groups[3])
    assert descs == ["Само карта 3", "Споделен материал"]


def test_parse_group_numbers_helper_directly():
    from routes_pallet_extra import _parse_group_numbers
    assert _parse_group_numbers("1") == [1]
    assert _parse_group_numbers("1+3+4") == [1, 3, 4]
    assert _parse_group_numbers(None) == [1]
    assert _parse_group_numbers("") == [1]
    assert _parse_group_numbers("abc") == [1]
    assert _parse_group_numbers(2) == [2]  # числова клетка, не низ


# ---------------------------------------------------------------- печат на всички карти наведнъж

def test_pallet_bulk_result_has_print_all_button(admin_client):
    """Заявка: „запазят картите да може да се принтират директно всичкия
    брой карти“ — бутон „Разпечатай всички карти наведнъж“ на страницата
    след масово издаване, сочещ към новия pallet_bulk_print route."""
    items_1 = json.dumps([{"code": "A1", "description": "Стока 1", "qty": "2"}])
    data = {
        "client_name": "Клиент Печат Всички", "groups": "1",
        "items_json_1": items_1, "items_format_1": "manual",
    }
    resp = post_with_csrf(admin_client, "/pallet/bulk-issue", data,
                          csrf_source_url="/pallet/new", follow_redirects=False)
    assert resp.status_code == 302
    result_resp = admin_client.get(resp.headers["Location"])
    body = result_resp.data.decode()
    assert "Разпечатай всички карти наведнъж" in body
    assert "/pallet/bulk-print?ids=" in body


def test_pallet_bulk_print_shows_all_cards_in_one_page(admin_client):
    items_1 = json.dumps([{"code": "A1", "description": "Първа карта стока", "qty": "2"}])
    items_2 = json.dumps([{"code": "B1", "description": "Втора карта стока", "qty": "5"}])
    data = {
        "client_name": "Клиент Двете Карти", "groups": "1,2",
        "items_json_1": items_1, "items_format_1": "manual",
        "items_json_2": items_2, "items_format_2": "manual",
    }
    resp = post_with_csrf(admin_client, "/pallet/bulk-issue", data,
                          csrf_source_url="/pallet/new", follow_redirects=False)
    assert resp.status_code == 302
    result_resp = admin_client.get(resp.headers["Location"])
    body = result_resp.data.decode()

    import re
    m = re.search(r'/pallet/bulk-print\?ids=([\d,]+)', body)
    assert m, body
    print_resp = admin_client.get("/pallet/bulk-print?ids=%s" % m.group(1))
    assert print_resp.status_code == 200
    print_body = print_resp.data.decode()
    assert "Първа карта стока" in print_body
    assert "Втора карта стока" in print_body
    # И двете карти са в ЕДНА страница (два .print-page блока), не отделни.
    assert print_body.count('class="print-page') == 2


def test_pallet_bulk_print_with_no_matching_docs_redirects_with_message(admin_client):
    resp = admin_client.get("/pallet/bulk-print?ids=999999", follow_redirects=True)
    assert resp.status_code == 200
    assert "Няма намерени документи за печат" in resp.data.decode()


def test_manually_issued_pallet_card_prints_with_the_five_order_columns(admin_client):
    """Заявка: „палетна карта да съдържа в съдържание на палета Order No,
    Pos, Reference, Reference Desc, Qty“ — и при РЪЧНО въвеждане, не само
    при импорт от справка. Формата вече подава items_format=orders (виж
    pallet_form.html), затова и печатната бланка трябва да покаже петте
    колони на формат „orders“, а не старите Артикул/код и Тегло."""
    items = json.dumps([{"order_no": "4700200362", "pos": "30",
                         "reference": "GLBK400002P0012",
                         "reference_desc": "C-Profile", "qty": "20"}])
    resp = post_with_csrf(admin_client, "/pallet/new", {
        "client_name": "Ръчен Клиент", "items_format": "orders",
        "items_json": items,
    }, csrf_source_url="/pallet/new", follow_redirects=False)
    body = admin_client.get(resp.headers["Location"]).data.decode()
    assert "Order No" in body
    assert "Reference Desc" in body
    assert "GLBK400002P0012" in body
    assert "Item code" not in body, "старите колони не бива да се показват"


# ---------------------------------------------------------------- без пореден № на редовете (документа)

def test_pallet_card_document_has_no_row_number_column_orders_format(admin_client):
    """Заявка: „палетна карта на документа да няма пореден номер на
    материлите в началото махни колоната“ — печатната бланка на издадена
    палетна карта (pallet_print.html) вече не показва колона „№“
    (пореден номер на реда) пред стоковите редове, за нито един от двата
    формата (orders / generic)."""
    items = json.dumps([{"order_no": "4700200362", "pos": "30",
                         "reference": "GLBK400002P0012",
                         "reference_desc": "C-Profile", "qty": "20"}])
    resp = post_with_csrf(admin_client, "/pallet/new", {
        "client_name": "Клиент Без Пореден Номер", "items_format": "orders",
        "items_json": items,
    }, csrf_source_url="/pallet/new", follow_redirects=False)
    body = admin_client.get(resp.headers["Location"]).data.decode()
    assert "<th>Поръчка № / Order No</th>" in body
    assert '<td class="c">1</td>' not in body
    assert "GLBK400002P0012" in body


def test_pallet_card_document_has_no_row_number_column_generic_format(admin_client):
    items = json.dumps([{"code": "A1", "description": "Профил", "qty": "5", "weight": "12"}])
    resp = post_with_csrf(admin_client, "/pallet/new", {
        "client_name": "Клиент Общ Формат", "items_json": items,
    }, csrf_source_url="/pallet/new", follow_redirects=False)
    body = admin_client.get(resp.headers["Location"]).data.decode()
    assert "Item code" in body
    assert '<td class="c">1</td>' not in body
    assert "Профил" in body


def test_pallet_bulk_print_has_no_row_number_column(admin_client):
    """Същата проверка, но за груповия печат на много вече издадени карти
    наведнъж (pallet_bulk_print.html) — идентична бланка, повторена по-долу
    в кода, затова се проверява отделно."""
    items = json.dumps([{"order_no": "4700200362", "pos": "30",
                         "reference": "GLBK400002P0012",
                         "reference_desc": "C-Profile", "qty": "20"}])
    resp = post_with_csrf(admin_client, "/pallet/new", {
        "client_name": "Клиент Груп Печат", "items_format": "orders",
        "items_json": items,
    }, csrf_source_url="/pallet/new", follow_redirects=False)
    doc_id = resp.headers["Location"].rstrip("/").split("/")[-1]
    body = admin_client.get("/pallet/bulk-print?ids=%s" % doc_id).data.decode()
    assert "GLBK400002P0012" in body
    assert '<td class="c">1</td>' not in body
