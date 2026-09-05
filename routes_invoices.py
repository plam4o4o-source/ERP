# -*- coding: utf-8 -*-
"""Фактури — отделен тип за Бразилия (invoice_br) и за Норвегия
(invoice_no), плюс зареждането на редове от вече издадена палетна карта.

Издаването/прегледът минават през СЪЩИТЕ generic функции като останалите
документи (appcore.DOCUMENT_FLOWS → routes_documents._document_new/
_document_preview) — тук са само тънките wrapper-и с точните endpoint
имена, и специфичното за фактурите: изтеглянето на редове от палетна
карта с автоматично попълване на теглото/описанието от справочника
материали.

Защо отделни типове, а не един с поле „държава“: самите бланки се
различават по заглавие И по колони на таблицата със стоките (Бразилия:
Net weight, без описание; Норвегия: Material Description + Pallet Number,
без тегло; Дубай: нито едното, нито другото — виж приложените образци).
Общ тип би означавал разклонения във всяка бланка, всеки износ и всяка
форма; отделните типове следват вече установения в програмата модел
„един тип = една бланка“.
"""
import io
import itertools
import json
import zipfile

from flask import (abort, flash, jsonify, redirect, render_template, request,
                   url_for)
from flask_babel import gettext as _

import applog
import db
import invoice_clients_module
import materials
from appcore import (XlsxTooLargeError, admin_required, ensure_xlsx_within_limits, get_db, login_required,
                     paginate_documents, safe_json_data)
from routes_documents import PAGE_SIZE, _document_new, _document_preview

# Одит (16.08.2026, находка №18, средна): огледално на routes_pallet_extra.
# _HEADER_SCAN_ROWS/_MAX_IMPORT_DATA_ROWS/_xlsx_has_merged_cells — вижте
# коментарите там за пълния разказ.
_HEADER_SCAN_ROWS = 10
_MAX_IMPORT_DATA_ROWS = 5000


def _read_limited_rows(ws):
    """Огледално на routes_pallet_extra._read_limited_rows — виж там за
    пълния разказ (одит 19.08.2026, находка №14: `list(ws.iter_rows(...))`
    материализираше ЦЕЛИЯ лист в паметта, а таванът от 5000 реда се
    прилагаше чак СЛЕД това — 27 878 ms и +148 MB за файл от 9 MB)."""
    it = ws.iter_rows(values_only=True)
    limit = _HEADER_SCAN_ROWS + _MAX_IMPORT_DATA_ROWS + 1
    rows = list(itertools.islice(it, limit))
    exhausted = next(it, None) is None if len(rows) == limit else True
    return rows, exhausted


def _xlsx_has_formulas(file_bytes):
    """Огледално на routes_pallet_extra._xlsx_has_formulas (одит
    19.08.2026, находка №39: формулна клетка без кеширана стойност се чете
    като None при `data_only=True` и редът се внася празен, а съобщението
    изглежда напълно успешно)."""
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            for name in zf.namelist():
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                    xml = zf.read(name)
                    if b"<f>" in xml or b"<f " in xml:
                        return True
    except Exception:
        return False
    return False


def _header_row_hint():
    """Одит (19.08.2026, находка №40) — огледално на routes_pallet_extra."""
    return _("Заглавният ред трябва да е в първите %d реда на листа.") % _HEADER_SCAN_ROWS


def _formula_hint():
    """Одит (19.08.2026, находка №39) — огледално на routes_pallet_extra."""
    return _("Файлът съдържа формули без запазени стойности — отворете го и го "
             "запишете от Excel, след което опитайте отново.")


def _parse_sheets(wb, parser):
    """Огледално на routes_pallet_extra._parse_sheets — одит (19.08.2026,
    находка №29): импортът четеше САМО първия лист, така че файл с
    декоративен лист „Инфо“ отпред и лист „Данни“ с валидните колони
    отказваше с „Файлът не съдържа разпознаваеми колони“."""
    sheets = list(wb.worksheets)
    for ws in sheets:
        parsed, warnings = parser(ws)
        if parsed:
            if len(sheets) > 1:
                warnings.insert(0, _("Данните са прочетени от лист „%(sheet)s“ "
                                     "(файлът съдържа %(count)d листа).")
                                % {"sheet": ws.title, "count": len(sheets)})
            return parsed, warnings, ws.title
    return None, [], None


def _xlsx_has_merged_cells(file_bytes):
    """Огледално на routes_pallet_extra._xlsx_has_merged_cells — виж там
    за пълния разказ защо проверката чете суровия XML директно, а не
    минава през openpyxl (read_only режимът, ползван по-долу за пестене
    на памет, изобщо не излага merged_cells)."""
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            for name in zf.namelist():
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                    if b"<mergeCell " in zf.read(name):
                        return True
    except Exception:
        return False
    return False


def register(app):
    app.add_url_rule("/invoice-br/new", "invoice_br_new", invoice_br_new,
                     methods=["GET", "POST"])
    app.add_url_rule("/invoice-br/preview", "invoice_br_preview", invoice_br_preview,
                     methods=["POST"])
    app.add_url_rule("/invoice-no/new", "invoice_no_new", invoice_no_new,
                     methods=["GET", "POST"])
    app.add_url_rule("/invoice-no/preview", "invoice_no_preview", invoice_no_preview,
                     methods=["POST"])
    app.add_url_rule("/invoice-dubai/new", "invoice_dubai_new", invoice_dubai_new,
                     methods=["GET", "POST"])
    app.add_url_rule("/invoice-dubai/preview", "invoice_dubai_preview", invoice_dubai_preview,
                     methods=["POST"])
    app.add_url_rule("/invoice/pull-pallet", "invoice_pull_pallet", invoice_pull_pallet,
                     methods=["POST"])
    app.add_url_rule("/invoice/import-items", "invoice_import_items",
                     invoice_import_items, methods=["POST"])
    app.add_url_rule("/invoices", "invoices_list", invoices_list)
    app.add_url_rule("/invoices/clients", "invoice_clients_list", invoice_clients_list)
    # Одит (05.09.2026, находка №12): сървърно автодовършване за формите —
    # огледално на /clients/lookup за общата адресна книга.
    app.add_url_rule("/invoices/clients/lookup", "invoice_clients_lookup",
                     invoice_clients_lookup)
    app.add_url_rule("/invoices/clients/new", "invoice_client_edit", invoice_client_edit,
                     methods=["GET", "POST"])
    app.add_url_rule("/invoices/clients/<int:entry_id>/edit", "invoice_client_edit",
                     invoice_client_edit, methods=["GET", "POST"])
    app.add_url_rule("/invoices/clients/<int:entry_id>/delete", "invoice_client_delete",
                     invoice_client_delete, methods=["POST"])


@login_required
def invoice_br_new():
    return _document_new("invoice_br")


@login_required
def invoice_br_preview():
    return _document_preview("invoice_br")


@login_required
def invoice_no_new():
    return _document_new("invoice_no")


@login_required
def invoice_no_preview():
    return _document_preview("invoice_no")


@login_required
def invoice_dubai_new():
    return _document_new("invoice_dubai")


@login_required
def invoice_dubai_preview():
    return _document_preview("invoice_dubai")


#: Тарифният код е един и същ за всички редове в приложените образци и
#: рядко се сменя — попълва се по подразбиране на всеки изтеглен ред, за
#: да не се преписва на ръка, но си остава редактируем във формата.
DEFAULT_HS_CODE = "85389099"


def _split_rows_by_po(rows, requested_po):
    """Групира готовите редове за фактура по номер на поръчка (P.O NO) —
    заявка: „във фактури един номер на поръчка да бъде на една фактура
    (пример: 4700201619 една фактура за всички материали с този номер
    поръчка, 4700223566 друга фактура)“.

    Палетната карта/Excel файлът често съдържат редове от НЯКОЛКО поръчки
    наведнъж, а една фактура трябва да покрива ЕДНА. Затова:

    - източник с 0 или 1 различни поръчки → редовете се зареждат направо,
      както досега (връща (None, rows));
    - няколко поръчки И операторът още не е избрал (requested_po is None)
      → връща ({"choose_po": True, "pos": [...]}, None) — формата показва
      избор на поръчка (виж renderInvoicePoChoice в app.js), НИЩО не се
      зарежда още;
    - избрана поръчка (requested_po, може и празен низ — групата „редове
      без поръчка №“) → връща само НЕЙНИТЕ редове + списък „remaining“ с
      останалите поръчки, които операторът трябва да издаде на ОТДЕЛНИ
      фактури (показва се в съобщението под бутона).

    Редовете с празен P.O NO са собствена група — не се разпределят
    мълчаливо към някоя поръчка, нито се губят."""
    order = []
    counts = {}
    for r in rows:
        po = (r.get("po_no") or "").strip()
        if po not in counts:
            counts[po] = 0
            order.append(po)
        counts[po] += 1

    if requested_po is None and len(order) <= 1:
        return None, rows
    if requested_po is None:
        return {"choose_po": True,
                "pos": [{"po_no": po, "count": counts[po]} for po in order]}, None
    wanted = requested_po.strip()
    filtered = [r for r in rows if (r.get("po_no") or "").strip() == wanted]
    remaining = [{"po_no": po, "count": counts[po]} for po in order if po != wanted]
    return {"loaded_po": wanted, "remaining": remaining}, filtered


@login_required
def invoice_pull_pallet():
    """Издърпва ВСИЧКИ редове на вече издадена палетна карта, преобразувани
    в редове за фактура — заявка: „фактурата да може да се зарежда, както
    се зареждат палетните карти в опаковъчния лист, но само със
    съответните данни, които са необходими за фактура“.

    За разлика от packing_pull_pallet (издърпва ЕДИН обобщен ред, защото
    опаковъчният лист описва колети, не отделни артикули), фактурата има
    нужда от всеки материал поотделно — колоните на палетната карта във
    формат „поръчки“ съвпадат почти едно към едно с тези на фактурата:

        Order No       → P.O NO
        Pos            → Pos
        Reference      → Material code
        Reference Desc → Material Description
        Open Qty       → Quantity

    Липсващото нето тегло (и описанието, ако картата няма попълнено) се
    допълва от справочника материали по кода — това е самото „автоматично
    извличане на килограмите“ от заявката. Кодове, които ги няма в
    справочника, просто остават с празно тегло за ръчно попълване; редът
    НЕ се пропуска. Единичната цена винаги остава празна — тя се въвежда
    ръчно (виж отговора на въпроса при заданието).
    """
    code = (request.form.get("code") or "").strip()
    if not code:
        return {"ok": False, "error": _("Въведете номер или баркод на палетна карта.")}

    con = get_db()
    row = con.execute(
        "SELECT * FROM documents WHERE doc_type = 'pallet' AND (barcode = ? OR number = ?)"
        " ORDER BY id DESC LIMIT 1",
        (code, code),
    ).fetchone()
    if row is None:
        other = con.execute(
            "SELECT doc_type FROM documents WHERE barcode = ? OR number = ?"
            " ORDER BY id DESC LIMIT 1",
            (code, code),
        ).fetchone()
        if other is not None:
            # Одит (19.08.2026, находка №13): заглавието на типа е маркирано
            # с db.N_() и се превежда ТУК, на мястото на показване.
            title = _(db.DOC_TYPES.get(other["doc_type"], {}).get("title", other["doc_type"]))
            return {"ok": False,
                    "error": _("Намереният документ не е палетна карта (%s).") % title}
        return {"ok": False, "error": _("Няма документ с номер/баркод „%s“.") % code}

    d = safe_json_data(row["data"])
    # Одит (01.09.2026, девети одит, находка №6): непокритата третина на
    # находка №3 (29.08). Тя добави филтъра при ВХОДА (appcore.parse_items) и
    # втора защита в ИЗНОСА (_export_fields_and_items) — изрично „за вече
    # записани документи с развален ред“. Този маршрут чете ТОЧНО такива вече
    # записани данни и вика it.get() без проверка: карта, издадена преди
    # поправката (или пипната ръчно в .db), дава AttributeError → 500 →
    # операторът вижда само „Грешка при заявката“ и картата остава
    # непрехвърляема във фактура завинаги, без втора врата като при износа.
    items = [it for it in (d.get("items") or []) if isinstance(it, dict)]
    if not items:
        return {"ok": False,
                "error": _("Палетна карта № %s няма редове за прехвърляне.") % row["number"]}

    orders_format = d.get("items_format") == "orders"
    # Кодът на материала е "reference" при формат „поръчки“ и "code" при
    # обикновения формат — и в двата случая това е ключът към справочника.
    codes = [(it.get("reference") if orders_format else it.get("code")) or "" for it in items]
    found = materials.lookup_many(con, codes)

    pallet_no = d.get("pallet_no") or ""
    rows = []
    for it, material_code in zip(items, codes):
        entry = found.get(material_code)
        if orders_format:
            description = it.get("reference_desc") or ""
            po_no, pos, qty = it.get("order_no") or "", it.get("pos") or "", it.get("qty") or ""
        else:
            description = it.get("description") or ""
            po_no, pos, qty = "", "", it.get("qty") or ""
        rows.append({
            "hs_code": DEFAULT_HS_CODE,
            "po_no": po_no,
            "pos": pos,
            # Кодът във фактурата е КАНОНИЧНИЯТ от справочника, когато е
            # намерен там — заявка: „референция 1TGB110025P1204-RAS да се
            # вмъкне като 1TGB110025P1204“. Справките за поръчки понякога
            # добавят суфикс след тире към ABB кода; lookup_many го маха
            # (виж materials.code_candidates) и тук записваме изчистения
            # код, за да съвпада с ценоразписа и занапред. Ненамерен код
            # остава ТОЧНО както е в палетната карта — не гадаем.
            "material_code": entry["code"] if entry else material_code,
            # Описанието от палетната карта има предимство пред това от
            # справочника (операторът може да го е уточнил за конкретната
            # пратка); справочникът е резервният източник.
            "description": description or (entry["description"] if entry else ""),
            "pallet_no": pallet_no,
            "qty": qty,
            "net_weight": entry["net_weight"] if entry else "",
            "unit_price": "",
            # служебен флаг за броенето на "matched" СЛЕД филтрирането по
            # поръчка — маха се преди отговора (pop по-долу).
            "_hit": entry is not None,
        })

    # Една поръчка = една фактура (виж _split_rows_by_po): при няколко
    # поръчки в картата операторът първо избира коя да зареди.
    requested_po = request.form["po_no"] if "po_no" in request.form else None
    extra, filtered = _split_rows_by_po(rows, requested_po)
    if filtered is None:
        return dict({"ok": True, "number": row["number"]}, **extra)

    matched = sum(1 for r in filtered if r.pop("_hit"))
    result = {
        "ok": True,
        "number": row["number"],
        "count": len(filtered),
        "matched": matched,
        "rows": filtered,
    }
    if extra:
        result.update(extra)
    return result


# ---------------------------------------------------------------- импорт от Excel
# Заявка: „зареждането на материалите във фактурата могат да се зареждат и
# от Excel файл; давам ти пример, както е в палетната карта — по същия
# начин да се вмъкват и за фактури“.
#
# Ползва СЪЩИЯ файлов формат като импорта в палетната карта (справка за
# поръчки: Order No, Pos, Reference, Reference Desc, Open Qty), затова един
# и същ файл върши работа и на двете места. Разликата е, че палетната карта
# РАЗДЕЛЯ редовете по номер на палет в последната колона, а фактурата ги
# взима всичките наред — тя описва една пратка, не отделни палети.
#
# Допълнително чете и колона с единична цена, ако файлът има такава
# (изрично избрано от потребителя) — иначе цената остава празна за ръчно
# въвеждане.

_PRICE_HEADERS = ("unit price", "price", "unit price (euro)", "unit price(euro)",
                  "единична цена", "цена")


def _cellstr(v):
    """Клетка към низ, без излишно „.0“ за цели числа, записани като float
    (същата помощна функция като в routes_pallet_extra/materials)."""
    if v is None:
        return ""
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        # Одит (02.09.2026, десети одит, находка №4): дробният float се
        # връщаше СУРОВ (`str(v)`), тоест точно както IEEE754 го пази.
        # Клетка с формула в Excel („Open Qty“ = разлика от две числа)
        # редовно държи 2.9000000000000004, а `fmt_num(value)` с
        # decimals=None НАРОЧНО пази въведената точност — тоест този запис
        # с 16 знака се отпечатваше буквално в колоната за количество на
        # палетна карта / търговска фактура, документ за клиента и за
        # митницата. Отделно беше и спусъкът на находка №1: един такъв ред
        # разваляше цялата жива сума на екрана.
        # `materials._weight_cell` решава същия проблем за теглата от
        # 19.08.2026 („%.6f“ + отрязване на нулите) — тук е същото, но с
        # предпазна клауза: ако закръглянето би превърнало ненулева
        # стойност в „0“ (напр. 8.7e-09), се връща суровият запис, за да
        # остане редът разпознат от `unparsable_item_rows` и операторът да
        # получи предупреждение, вместо тихо да види количество нула.
        text = ("%.6f" % v).rstrip("0").rstrip(".")
        if text in ("", "0", "-0") and v != 0:
            return str(v).strip()
        return text or "0"
    return str(v).strip()


def _parse_invoice_items_xlsx(ws):
    """Чете справка за поръчки и връща (rows, warnings): `rows` е списък
    от редове за фактура, или None ако колоните не се разпознават;
    `warnings` е списък от низове за flash (орязване/заглавен ред не на
    позиция 0 — виж находка №18).

    Колоните се търсят по ЗАГЛАВИЕ (не по позиция) — както в палетната
    карта. Ред без нито един попълнен от интересните ни полета се
    пропуска (файловете редовно имат празни редове най-отдолу)."""
    warnings = []
    # Одит (19.08.2026, находка №14): вместо `list(ws.iter_rows(...))` —
    # виж _read_limited_rows по-горе (целият файл влизаше в паметта ПРЕДИ
    # рязането на 5000 реда).
    rows, exhausted = _read_limited_rows(ws)
    if not rows:
        return None, warnings

    # Одит (16.08.2026, находка №18): огледално на routes_pallet_extra.
    # _parse_order_export — сканира първите _HEADER_SCAN_ROWS реда за
    # истинския заглавен ред, вместо сляпо да предполага позиция 0 (чест
    # допълнителен ред отгоре при реални ERP/BI износи).
    def find_col_in(header_lower, *names):
        for name in names:
            for i, h in enumerate(header_lower):
                if h == name:
                    return i
        return None

    header_idx = 0
    header = [_cellstr(c).lower() for c in (rows[0] or [])]
    for idx in range(min(_HEADER_SCAN_ROWS, len(rows))):
        candidate = [_cellstr(c).lower() for c in (rows[idx] or [])]
        if (find_col_in(candidate, "order no", "order number", "orderno") is not None
                and find_col_in(candidate, "open qty", "qty", "quantity") is not None):
            header_idx, header = idx, candidate
            break
    if header_idx > 0:
        warnings.append(_("Заглавният ред е открит на ред %d от файла (пропуснати са "
                          "%d реда над него) — проверете дали разпознатите данни са "
                          "правилни.") % (header_idx + 1, header_idx))

    data_rows = rows[header_idx + 1:]
    # Одит (19.08.2026, находка №14): орязването се разпознава по неизчерпан
    # итератор, а не по дължината на списък с целия лист.
    if len(data_rows) > _MAX_IMPORT_DATA_ROWS or not exhausted:
        warnings.append(_("Файлът съдържа повече от %d реда данни — заредени са само "
                          "първите %d, останалите са пропуснати.")
                        % (_MAX_IMPORT_DATA_ROWS, _MAX_IMPORT_DATA_ROWS))
        data_rows = data_rows[:_MAX_IMPORT_DATA_ROWS]

    def find_col(*names):
        return find_col_in(header, *names)

    col_order = find_col("order no", "order number", "orderno")
    col_pos = find_col("pos", "position")
    col_ref = find_col("reference")
    col_desc = find_col("reference desc", "reference description", "ref desc",
                        "description", "material description")
    col_qty = find_col("open qty", "qty", "quantity")
    col_price = find_col(*_PRICE_HEADERS)
    if col_order is None or col_qty is None:
        return None, warnings

    def cell(row, idx):
        return _cellstr(row[idx]) if idx is not None and idx < len(row) else ""

    out = []
    for row in data_rows:
        values = {
            "po_no": cell(row, col_order),
            "pos": cell(row, col_pos),
            "material_code": cell(row, col_ref),
            "description": cell(row, col_desc),
            "qty": cell(row, col_qty),
            "unit_price": cell(row, col_price),
        }
        if not any(values.values()):
            continue
        out.append(values)
    return (out or None), warnings


@login_required
def invoice_import_items():
    """Зарежда редове във фактурата от качен Excel файл и ги връща като
    JSON (формата ги добавя в таблицата, без да губи вече въведеното).

    Нето теглото се допълва от справочника материали по кода — точно
    както при зареждането от палетна карта."""
    from openpyxl import load_workbook

    file = request.files.get("excel_file")
    if not file or not file.filename:
        return {"ok": False, "error": _("Изберете Excel файл (.xlsx).")}
    file_bytes = file.read()
    # Одит (31.08.2026, находка №7): виж ensure_xlsx_within_limits — таван на
    # разархивирания размер преди всяко четене на архива.
    try:
        ensure_xlsx_within_limits(file_bytes)
    except XlsxTooLargeError as exc:
        return {"ok": False, "error": str(exc)}
    # Одит (16.08.2026, находка №18): read_only=True пести памет за голям
    # качен файл — вижте _HEADER_SCAN_ROWS/_MAX_IMPORT_DATA_ROWS по-горе.
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:
        applog.log_exception("routes_invoices: неуспешно четене на качен .xlsx файл")
        return {"ok": False,
                "error": _("Файлът не може да бъде прочетен. Уверете се, че е валиден .xlsx файл.")}

    warnings = []
    if _xlsx_has_merged_cells(file_bytes):
        warnings.append(_("Файлът съдържа обединени клетки — стойности извън първата "
                          "клетка на обединен диапазон може да липсват."))

    # Одит (19.08.2026, находка №29): всички листове, не само първият.
    parsed, parse_warnings, _sheet = _parse_sheets(wb, _parse_invoice_items_xlsx)
    warnings.extend(parse_warnings)
    if not parsed:
        # Одит (19.08.2026, находки №40 и №39) — виж огледалния коментар в
        # routes_pallet_extra.pallet_bulk_import.
        error = _("Файлът не съдържа разпознаваеми колони (Order No, Pos, "
                  "Reference, Reference Desc, Open Qty) или редове за импорт.")
        error += " " + _header_row_hint()
        if _xlsx_has_formulas(file_bytes):
            error += " " + _formula_hint()
        return {"ok": False, "error": error}

    # Одит (19.08.2026, находка №39): колоната с количествата е изцяло
    # празна, а файлът съдържа формули без запазени стойности — иначе
    # отговорът изглежда напълно успешен („заредени N реда“) с нула
    # използваеми количества.
    if not any((r.get("qty") or "").strip() for r in parsed) and _xlsx_has_formulas(file_bytes):
        warnings.append(_formula_hint())

    con = get_db()
    found = materials.lookup_many(con, [r["material_code"] for r in parsed])
    rows = []
    for r in parsed:
        entry = found.get(r["material_code"])
        rows.append({
            "hs_code": DEFAULT_HS_CODE,
            "po_no": r["po_no"],
            "pos": r["pos"],
            # Каноничният код от справочника, когато е намерен — суфикси
            # като „-RAS“ се махат (виж коментара в invoice_pull_pallet и
            # materials.code_candidates); ненамерен код остава от файла.
            "material_code": entry["code"] if entry else r["material_code"],
            # Описанието от файла има предимство; справочникът е резервен.
            "description": r["description"] or (entry["description"] if entry else ""),
            "pallet_no": "",
            "qty": r["qty"],
            "net_weight": entry["net_weight"] if entry else "",
            "unit_price": r["unit_price"],
            "_hit": entry is not None,
        })

    # Една поръчка = една фактура — същият механизъм като при зареждането
    # от палетна карта (виж _split_rows_by_po): при няколко поръчки във
    # файла операторът първо избира коя да зареди, останалите отиват на
    # отделни фактури.
    requested_po = request.form["po_no"] if "po_no" in request.form else None
    extra, filtered = _split_rows_by_po(rows, requested_po)
    if filtered is None:
        return dict({"ok": True, "filename": file.filename}, **extra)

    matched = sum(1 for r in filtered if r.pop("_hit"))
    result = {"ok": True, "count": len(filtered), "matched": matched,
              "filename": file.filename, "rows": filtered}
    if warnings:
        result["warnings"] = warnings
    if extra:
        result.update(extra)
    return result


# ---------------------------------------------------------------- издадени фактури

@login_required
def invoices_list():
    """Списък САМО с издадените фактури — заявка: „в раздела Фактури да има
    издадени документи и само там да се появяват издадените фактури“.
    Общият списък „Всички документи“ ги изключва (виж
    routes_documents.documents и db.INVOICE_DOC_TYPES)."""
    doc_type = request.args.get("type", "")
    # Одит (03.09.2026, находка №9): непознат тип се НУЛИРА веднага, не само
    # за SQL филтъра. Шаблонът прави `doc_types[sel_type].title` — суровата
    # стойност от адреса стигаше дотам и вдигаше UndefinedError, тоест
    # анонимно съставим линк (/docs?type=') гарантирано сваляше 500-ка и
    # задействаше пренасочването по Referer (виж и поправката в appcore).
    # Одит (05.09.2026, находка №4): пазачът беше срещу ГРЕШНИЯ речник.
    # Копиран е дословно от routes_documents.documents, но ТОЗИ шаблон
    # получава СТЕСНЕН речник `invoice_types` (само фактурните типове, виж
    # по-долу). Шестте нефактурни типа минаваха проверката и гърмяха в
    # шаблона. Проверено с изпълнение: /invoices?type=cmr → UndefinedError
    # („'dict object' has no attribute 'cmr'“) → 302 с „Възникна неочаквана
    # грешка“; тоест точно съставимият отвън линк, който находка №9 трябваше
    # да затвори, оставаше отворен — при това с по-невинен на вид адрес от
    # този в теста ѝ.
    if doc_type not in db.INVOICE_DOC_TYPES:
        doc_type = ""
    query = request.args.get("q", "").strip()
    page = request.args.get("page", 1, type=int) or 1
    # Одит (12.08.2026, находка №20): филтър по диапазон от дати — липсваше
    # тук, макар списъкът с всички документи (routes_documents.documents)
    # да го има, при иначе визуално еднакви интерфейси за търсене (двата
    # копирани един от друг). Същата логика (по d.created_at, включва целия
    # ден на date_to) — вижте documents() за пълния разказ.
    date_from = request.args.get("from", "").strip()
    date_to = request.args.get("to", "").strip()

    where = ("WHERE d.doc_type IN (%s)"
            % ",".join("?" for _ in db.INVOICE_DOC_TYPES))  # nosec B608 -- само „?“ плейсхолдъри по брой
    params = list(db.INVOICE_DOC_TYPES)
    if doc_type in db.INVOICE_DOC_TYPES:
        where += " AND d.doc_type = ?"
        params.append(doc_type)
    if query:
        # В7: ci_contains (db._ci_contains) — вижте routes_documents.py.
        where += " AND (ci_contains(d.number, ?) OR ci_contains(d.data, ?))"
        params += [query, query]
    # Одит (16.08.2026, находка №22): sargable сравнение directno върху
    # текста на created_at (вижте пълния разказ в routes_documents.
    # documents()), вместо `date(d.created_at) >= date(?)` — обвиването на
    # КОЛОНАТА в date() пречи на idx_documents_created_at.
    if date_from:
        where += " AND d.created_at >= ?"
        params.append(date_from)
    if date_to:
        where += " AND d.created_at < date(?, '+1 day')"
        params.append(date_to)

    con = get_db()
    # В15/находка №20: истинска пагинация чрез общия appcore.paginate_documents
    # helper (вижте документацията там — преди тази поправка кодът тук беше
    # копие на почти идентичния блок в routes_documents.documents()).
    docs, page, total_pages, total_count = paginate_documents(
        con, where, params, page, page_size=PAGE_SIZE)
    return render_template(
        "invoices.html", docs=docs, metas=[safe_json_data(d["data"]) for d in docs],
        invoice_types={k: v for k, v in db.DOC_TYPES.items() if k in db.INVOICE_DOC_TYPES},
        sel_type=doc_type, q=query, date_from=date_from, date_to=date_to,
        page=page, total_pages=total_pages, total_count=total_count)


# ---------------------------------------------------------------- адресна книга за фактури

@login_required
def invoice_clients_lookup():
    """Одит (05.09.2026, находка №12): търси в адресната книга за фактури.

    ЗАЩО съществува: формите за фактура вграждаха ЦЯЛАТА книга в HTML-а —
    веднъж като <option>-и и втори път като JSON за автопопълването.
    Измерено: 500 записа = 357 KB HTML, 2 000 записа = 1 103 KB на всяко
    отваряне. Общата адресна книга получи същото лечение на 19.08 (находка
    №25); тази остана непокрита.

    Отговорът носи ПЪЛНИТЕ полета на записа, за да работи попълването на
    двата адресни блока еднакво, независимо дали записът е дошъл от
    вградения списък или от търсенето."""
    con = get_db()
    query = request.args.get("q", "").strip()
    rows = invoice_clients_module.search(con, query)
    truncated = len(rows) > invoice_clients_module.LOOKUP_LIMIT
    rows = rows[:invoice_clients_module.LOOKUP_LIMIT]
    return jsonify({"clients": [dict(r) for r in rows], "truncated": truncated})


@login_required
def invoice_clients_list():
    """Одит (19.08.2026, находка №25, средна): огледално на
    routes_clients.clients_list — списъкът рендираше ВСИЧКИ записи наведнъж
    и нямаше никакво сървърно търсене. Тук записите са едри (два пълни
    адресни блока на ред), затова липсата на пагинация тежи още повече."""
    query = request.args.get("q", "").strip()
    page = request.args.get("page", 1, type=int) or 1
    entries, page, total_pages, total_count = invoice_clients_module.paginate(
        get_db(), query, page)
    return render_template("invoice_clients.html", entries=entries, q=query,
                           page=page, total_pages=total_pages, total_count=total_count)


@login_required
def invoice_client_edit(entry_id=None):
    con = get_db()
    # Одит (01.09.2026, девети одит, находка №5): `is not None`, не truthiness.
    # `/invoices/clients/0/edit` маршрутизира с entry_id=0 (IntegerConverter е
    # \d+ без min) — при `if entry_id` нулата пада в клона „нов запис“, GET
    # даваше 200 с празна форма вместо 404, а POST → invoice_clients_module.
    # save(..., 0) → INSERT (там също `if entry_id`), тихо създаваше НОВ запис.
    # Огледалните client_edit (routes_clients) и invoice_client_delete вече
    # ползват `is not None`/дават 404 за id=0 — това беше непокритата половина.
    entry = invoice_clients_module.get(con, entry_id) if entry_id is not None else None
    if entry_id is not None and entry is None:
        abort(404)
    # Одит (01.09.2026, девети одит, находка №3): шаблонът вече чете
    # стойностите от `values`/`entry_values`, за да НЕ губи въведеното при
    # отказ от валидацията. Досега рендираше само от `entry`: при празно име
    # всичките осем полета (двата многоредови адреса, двата телефона, двете
    # имена, бележката) се връщаха празни при СЪЗДАВАНЕ, а при РЕДАКЦИЯ — със
    # старите стойности от базата, тоест редакциите на оператора тихо
    # изчезваха и изглеждаха „неприети“, без нищо да го каже.
    entry_values = dict(entry) if entry is not None else {}
    if request.method == "POST":
        if not (request.form.get("name") or "").strip():
            flash(_("Въведете име на записа."), "error")
            return render_template(
                "invoice_client_form.html", entry=entry,
                entry_values=entry_values,
                values={k: (request.form.get(k) or "")
                        for k in invoice_clients_module._FIELDS})
        invoice_clients_module.save(con, request.form, entry_id)
        flash(_("Записът в адресната книга за фактури е запазен."), "success")
        return redirect(url_for("invoice_clients_list"))
    return render_template("invoice_client_form.html", entry=entry,
                           entry_values=entry_values)


@admin_required
def invoice_client_delete(entry_id):
    # Одит (16.08.2026, находка №33): огледално на routes_clients.
    # client_delete/routes_documents.delete_document — DELETE ... WHERE
    # id=? за вече несъществуващ запис е no-op без грешка; преди тази
    # поправка операторът виждаше подвеждащото "Записът е изтрит" дори
    # когато нищо реално не е било изтрито.
    con = get_db()
    row = con.execute("SELECT id FROM invoice_clients WHERE id = ?", (entry_id,)).fetchone()
    if row is None:
        abort(404)
    invoice_clients_module.delete(con, entry_id)
    flash(_("Записът е изтрит от адресната книга за фактури."), "success")
    return redirect(url_for("invoice_clients_list"))
